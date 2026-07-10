"""European Banking Authority (EBA) connector.

The EBA exposes no machine-readable catalog/API — its statistical products are
published as downloadable files on static-HTML product pages, and the file URLs
rotate every release (they embed a release-month folder + UUID, or a per-release
numeric id). So each fetch fn re-discovers the current file URL by scraping the
relevant landing page, then downloads and parses it. Full re-pull every run
(stateless): the source has no incremental/`since` filter.

Published subsets (one raw asset per accepted collect entity):

Risk Dashboard (scraped from the Risk Dashboard product page)
  * risk-dashboard-kri                — "Key Risk Indicators" by country and
      EU/EEA aggregate, quarterly. The single latest "Data Annex
      InteractiveRiskDashboard" XLSX carries the full history on its
      "KRIs by country and EU" sheet — a tidy [Period, Country, Number, Name,
      Ratio] table.
  * risk-dashboard-credit-parameters  — IFRS9 / IRB credit-risk parameters
      (Default rate, Loss rate, PD-adjusted, LGD; N/25th/50th/75th/W.A stats)
      by country and portfolio segment. Each "Credit Risk parameters annex"
      XLSX holds ONE quarter on a single matrix sheet, so we fetch every
      quarterly file on the page and union — melting the metric×statistic
      column matrix into long [period, country, segment, metric, statistic,
      value] rows.
  * risk-dashboard-raq                — Risk Assessment Questionnaire, a
      semi-annual bank sentiment survey. The latest "Statistical annex RAQ"
      XLSX carries the full "Historical Results" block (periods as columns);
      we melt it to long [period, period_label, question, answer, share].

EU-wide Transparency Exercise (scraped from the TE product page → latest
exercise's Full_database)
  * te-credit-risk / te-market-risk / te-sovereign-exposures  — bank-by-bank
      long tables (tr_cre / tr_mrk / tr_sov .csv). Already tidy: LEI_Code, NSA,
      Period (YYYYMM), Item, Label, template-specific breakdown dims, Amount.
  * te-capital-own-funds / te-other-exposures — both derive from tr_oth.csv,
      which consolidates several templates distinguished by its `Sheet` column.
      capital-own-funds = Sheet in {Capital, Leverage}; other-exposures = the
      remaining sheets (Assets, Liabilities, P&L, Key metrics, RWA OV1). The two
      partition tr_oth cleanly with no overlap. (The old per-file tr_cap/tr_rwa/
      tr_lev CSVs no longer exist — the data was folded into tr_oth.)
  * te-data-dictionary                — the SDD.xlsx codebook mapping every
      numeric Item code to its CSV / Template / Category / Label. Reference data.

Aggregate statistical data — Supervisory Disclosure (scraped from the
supervisory-disclosure aggregate-statistical-data page)
  * supervisory-measures-penalties    — per competent authority, counts of
      supervisory measures and administrative penalties, under BOTH the CRD
      (annex 4 part 5) and IFD (annex 4 part 3) frameworks. Each workbook has
      one sheet per authority (ISO-2 code, SSM=ECB); we walk every sheet's
      [code, category, label, count] block into long rows tagged with framework
      and reporting year.

EU-wide Stress Test (scraped from the stress-test product/results page)
  * stress-test-results               — bank-by-bank projected outcomes under
      baseline vs adverse scenarios (see fetch_stress_test_results).

Discovery is robust to URL rotation because it always parses the live listing
pages; it is NOT robust to the EBA renaming the product pages, sheets, or file
naming conventions — those surface as a loud parse failure (empty result ->
node failure), which is the intended behaviour.
"""

import io
import re
import csv as _csv
import urllib.parse

import pyarrow as pa
import pyarrow.csv as pacsv
import pyarrow.compute as pc
import openpyxl

from subsets_utils import NodeSpec, get, save_raw_parquet, transient_retry

EBA_BASE = "https://www.eba.europa.eu"
RISK_DASHBOARD_PAGE = f"{EBA_BASE}/risk-and-data-analysis/risk-analysis/risk-monitoring/risk-dashboard"
TE_LANDING_PAGE = f"{EBA_BASE}/risk-and-data-analysis/risk-analysis/eu-wide-transparency-exercise"
STRESS_TEST_PAGE = f"{EBA_BASE}/risk-and-data-analysis/risk-analysis/eu-wide-stress-testing"
SUP_DISCLOSURE_PAGE = f"{EBA_BASE}/supervisory-convergence/supervisory-disclosure/aggregate-statistical-data"
KRI_SHEET = "KRIs by country and EU"

# tr_oth.csv consolidates several templates; capital-own-funds owns these sheets,
# other-exposures owns the rest.
TE_CAPITAL_SHEETS = ["Capital", "Leverage"]

# Which Transparency-Exercise Full_database CSV each single-file TE subset uses.
TE_SINGLE_FILES = {
    "eba-te-credit-risk": "tr_cre.csv",
    "eba-te-market-risk": "tr_mrk.csv",
    "eba-te-sovereign-exposures": "tr_sov.csv",
}


@transient_retry()
def _fetch_text(url: str) -> str:
    resp = get(url, timeout=(10.0, 90.0))
    resp.raise_for_status()
    return resp.text


@transient_retry()
def _fetch_bytes(url: str) -> bytes:
    resp = get(url, timeout=(15.0, 300.0))
    resp.raise_for_status()
    return resp.content


def _abs(href: str) -> str:
    """Absolute URL from a (possibly %-encoded, possibly relative) href."""
    href = href.strip()
    if href.startswith("http"):
        return href
    # re-quote spaces etc. but leave already-encoded sequences intact
    return EBA_BASE + urllib.parse.quote(href, safe="/:%?&=+,;@()")


def _xlsx_hrefs(html: str) -> list[str]:
    """All .xlsx hrefs on a page, URL-decoded so the filename is readable."""
    return [urllib.parse.unquote(h) for h in re.findall(r'href="([^"]+\.xlsx)"', html)]


# --------------------------------------------------------------------------- #
# Risk Dashboard — Key Risk Indicators
# --------------------------------------------------------------------------- #
def _latest_risk_dashboard_annex_url() -> str:
    html = _fetch_text(RISK_DASHBOARD_PAGE)
    hrefs = [h for h in _xlsx_hrefs(html) if "Data Annex InteractiveRiskDashboard" in h]
    if not hrefs:
        raise AssertionError("Risk Dashboard page exposed no Data Annex XLSX link")
    best = None
    for href in hrefs:
        m = re.search(r"Q(\d)\s+(\d{4})", href)
        rank = (int(m.group(2)), int(m.group(1))) if m else (0, 0)
        if best is None or rank > best[0]:
            best = (rank, href)
    return _abs(best[1])


def fetch_risk_dashboard_kri(node_id: str) -> None:
    asset = node_id  # spec id == asset name
    content = _fetch_bytes(_latest_risk_dashboard_annex_url())
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if KRI_SHEET not in wb.sheetnames:
        raise AssertionError(f"sheet '{KRI_SHEET}' missing; sheets={wb.sheetnames}")
    ws = wb[KRI_SHEET]

    periods, countries, codes, names, values = [], [], [], [], []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header: [Period], [Country], [Number], [Name], [Ratio]
        period, country, code, name, ratio = (list(row) + [None] * 5)[:5]
        if period is None or country is None or code is None:
            continue
        periods.append(int(period))
        countries.append(str(country))
        codes.append(str(code))
        names.append(None if name is None else str(name))
        values.append(None if ratio is None else float(ratio))

    table = pa.table(
        {
            "period": pa.array(periods, pa.int32()),
            "country": pa.array(countries, pa.string()),
            "indicator_code": pa.array(codes, pa.string()),
            "indicator_name": pa.array(names, pa.string()),
            "value": pa.array(values, pa.float64()),
        }
    )
    if table.num_rows == 0:
        raise AssertionError("KRI sheet parsed to 0 rows")
    save_raw_parquet(table, asset)


# --------------------------------------------------------------------------- #
# Risk Dashboard — Credit Risk Parameters annex (per-quarter matrix files)
# --------------------------------------------------------------------------- #
def _credit_param_quarter_key(href: str) -> tuple[int, int]:
    m = re.search(r"Q(\d)\s+(\d{4})", href)
    return (int(m.group(2)), int(m.group(1))) if m else (0, 0)


def _parse_credit_parameters(content: bytes, period: int) -> list[dict]:
    """Melt one quarter's country×segment × (metric,statistic) matrix to long."""
    ws = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True).worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header_i = next(
        (i for i, r in enumerate(rows) if any(c == "Default rate" for c in r if c)), None
    )
    if header_i is None:
        raise AssertionError("Credit Risk parameters annex: no 'Default rate' header found")
    metrics, stats = rows[header_i], rows[header_i + 1]
    colmap: dict[int, tuple[str, str]] = {}
    current_metric = None
    for j, cell in enumerate(metrics):
        if cell and str(cell).strip():
            current_metric = str(cell).strip()
        stat = stats[j] if j < len(stats) else None
        if current_metric and stat and str(stat).strip():
            colmap[j] = (current_metric, str(stat).strip())

    out: list[dict] = []
    country = None
    for r in rows[header_i + 2:]:
        if r[1] and str(r[1]).strip():
            country = str(r[1]).strip()
        if not country or r[2] is None or not str(r[2]).strip():
            continue
        segment = str(r[2]).strip()
        for j, (metric, stat) in colmap.items():
            v = r[j] if j < len(r) else None
            if v is None or v == "":
                continue
            try:
                fv = float(v)
            except (TypeError, ValueError):
                continue
            out.append(
                {
                    "period": period,
                    "country": country,
                    "segment": segment,
                    "metric": metric,
                    "statistic": stat,
                    "value": fv,
                }
            )
    return out


def fetch_risk_dashboard_credit_parameters(node_id: str) -> None:
    asset = node_id
    html = _fetch_text(RISK_DASHBOARD_PAGE)
    hrefs = sorted(
        {h for h in _xlsx_hrefs(html) if "Credit Risk parameters annex" in h},
        key=_credit_param_quarter_key,
    )
    if not hrefs:
        raise AssertionError("Risk Dashboard page exposed no Credit Risk parameters annex")
    rows: list[dict] = []
    for href in hrefs:
        year, quarter = _credit_param_quarter_key(href)
        rows.extend(_parse_credit_parameters(_fetch_bytes(_abs(href)), year * 10 + quarter))
    if not rows:
        raise AssertionError("Credit Risk parameters annex parsed to 0 rows")
    table = pa.table(
        {
            "period": pa.array([r["period"] for r in rows], pa.int32()),
            "country": pa.array([r["country"] for r in rows], pa.string()),
            "segment": pa.array([r["segment"] for r in rows], pa.string()),
            "metric": pa.array([r["metric"] for r in rows], pa.string()),
            "statistic": pa.array([r["statistic"] for r in rows], pa.string()),
            "value": pa.array([r["value"] for r in rows], pa.float64()),
        }
    )
    save_raw_parquet(table, asset)


# --------------------------------------------------------------------------- #
# Risk Assessment Questionnaire (RAQ) — Statistical Annex, historical block
# --------------------------------------------------------------------------- #
def _raq_file_key(href: str) -> tuple[int, int]:
    m = re.search(r"(Spring|Autumn)\s*(\d{4})", href)
    if not m:
        return (0, 0)
    return (int(m.group(2)), 1 if m.group(1) == "Spring" else 2)


def _raq_period_int(label: str) -> int:
    m = re.match(r"(Spring|Autumn)-(\d\d)", label)
    year = 2000 + int(m.group(2))
    return year * 10 + (1 if m.group(1) == "Spring" else 2)


def fetch_risk_dashboard_raq(node_id: str) -> None:
    asset = node_id
    html = _fetch_text(RISK_DASHBOARD_PAGE)
    candidates = [
        h for h in _xlsx_hrefs(html) if "RAQ" in h and re.search(r"(Spring|Autumn)\s*\d{4}", h)
    ]
    if not candidates:
        raise AssertionError("Risk Dashboard page exposed no RAQ statistical annex")
    href = max(candidates, key=_raq_file_key)
    wb = openpyxl.load_workbook(io.BytesIO(_fetch_bytes(_abs(href))), read_only=True, data_only=True)
    if "Results" not in wb.sheetnames:
        raise AssertionError(f"RAQ annex missing 'Results' sheet; sheets={wb.sheetnames}")
    rows = list(wb["Results"].iter_rows(values_only=True))

    period_cols: dict[int, str] | None = None
    question = None
    out: list[dict] = []
    for r in rows:
        e = r[4] if len(r) > 4 else None
        if e is None:
            continue
        text = str(e).strip()
        labels = {
            j: str(c).strip()
            for j, c in enumerate(r)
            if j >= 5 and c and re.match(r"(Spring|Autumn)-\d\d", str(c).strip())
        }
        if labels:
            period_cols = labels
        if re.match(r"^Q\d+\b", text):
            question = text
            continue
        if question is None or period_cols is None:
            continue
        for j, label in period_cols.items():
            v = r[j] if j < len(r) else None
            if v is None or v == "":
                continue
            try:
                fv = float(v)
            except (TypeError, ValueError):
                continue
            out.append(
                {
                    "period": _raq_period_int(label),
                    "period_label": label,
                    "question": question,
                    "answer": text,
                    "share": fv,
                }
            )
    if not out:
        raise AssertionError("RAQ Results parsed to 0 rows")
    table = pa.table(
        {
            "period": pa.array([r["period"] for r in out], pa.int32()),
            "period_label": pa.array([r["period_label"] for r in out], pa.string()),
            "question": pa.array([r["question"] for r in out], pa.string()),
            "answer": pa.array([r["answer"] for r in out], pa.string()),
            "share": pa.array([r["share"] for r in out], pa.float64()),
        }
    )
    save_raw_parquet(table, asset)


# --------------------------------------------------------------------------- #
# EU-wide Transparency Exercise — Full_database CSV + SDD codebook
# --------------------------------------------------------------------------- #
def _latest_te_full_database_base() -> str:
    """Return the newest exercise year's Full_database base URL (…/<id>)."""
    main = _fetch_text(TE_LANDING_PAGE)
    years = {}
    for href in re.findall(r'href="([^"]*transparency[^"]*)"', main):
        m = re.search(r"/(20\d\d)-eu-wide-transparency", href)
        if m:
            years[int(m.group(1))] = href
    if not years:
        raise AssertionError("no Transparency Exercise year pages found on landing page")
    pat = re.compile(r"(https://www\.eba\.europa\.eu/assets/TE\d+/Full_database/\d+)/[^\"']+")
    for year in sorted(years, reverse=True):
        page = _fetch_text(_abs(years[year]))
        m = pat.search(page)
        if m:
            return m.group(1)
    raise AssertionError("no Full_database link found on any TE year page")


def _read_te_csv(content: bytes) -> pa.Table:
    """Read an EBA TE CSV forcing every column to string (no inference on a
    100MB+ file); the transform casts Period/Amount downstream."""
    header_line = content.split(b"\n", 1)[0].decode("utf-8-sig")
    cols = next(_csv.reader([header_line]))
    return pacsv.read_csv(
        io.BytesIO(content),
        convert_options=pacsv.ConvertOptions(
            column_types={c: pa.string() for c in cols}, strings_can_be_null=True
        ),
    )


def _fetch_te_single(node_id: str) -> None:
    asset = node_id
    filename = TE_SINGLE_FILES[node_id]
    table = _read_te_csv(_fetch_bytes(f"{_latest_te_full_database_base()}/{filename}"))
    if table.num_rows == 0:
        raise AssertionError(f"{filename} parsed to 0 rows")
    save_raw_parquet(table, asset)


def fetch_te_credit_risk(node_id: str) -> None:
    _fetch_te_single(node_id)


def fetch_te_market_risk(node_id: str) -> None:
    _fetch_te_single(node_id)


def fetch_te_sovereign_exposures(node_id: str) -> None:
    _fetch_te_single(node_id)


def _fetch_te_oth(node_id: str, capital: bool) -> None:
    asset = node_id
    table = _read_te_csv(_fetch_bytes(f"{_latest_te_full_database_base()}/tr_oth.csv"))
    in_capital = pc.is_in(table["Sheet"], value_set=pa.array(TE_CAPITAL_SHEETS))
    table = table.filter(in_capital if capital else pc.invert(in_capital))
    if table.num_rows == 0:
        raise AssertionError(f"tr_oth.csv ({'capital' if capital else 'other'}) parsed to 0 rows")
    save_raw_parquet(table, asset)


def fetch_te_capital_own_funds(node_id: str) -> None:
    _fetch_te_oth(node_id, capital=True)


def fetch_te_other_exposures(node_id: str) -> None:
    _fetch_te_oth(node_id, capital=False)


def fetch_te_data_dictionary(node_id: str) -> None:
    asset = node_id
    ws = openpyxl.load_workbook(
        io.BytesIO(_fetch_bytes(f"{_latest_te_full_database_base()}/SDD.xlsx")),
        read_only=True,
        data_only=True,
    )["SDD"]
    rows = list(ws.iter_rows(values_only=True))
    # header is the row carrying 'CSV' + 'Item' (row 0 is a merged note)
    header_i = next(
        i for i, r in enumerate(rows) if r and "CSV" in [str(c).strip() if c else "" for c in r]
    )
    hdr = [str(c).strip() if c else "" for c in rows[header_i]]
    idx = {h: i for i, h in enumerate(hdr) if h in ("CSV", "Template", "Item", "Category", "Label")}
    csv_file, template, item, category, label = [], [], [], [], []
    for r in rows[header_i + 1:]:
        it = r[idx["Item"]] if idx.get("Item") is not None else None
        if it is None or not str(it).strip():
            continue

        def g(key):
            j = idx.get(key)
            return None if j is None or r[j] is None else str(r[j]).strip()

        csv_file.append(g("CSV"))
        template.append(g("Template"))
        item.append(str(it).strip())
        category.append(g("Category"))
        label.append(g("Label"))
    if not item:
        raise AssertionError("SDD.xlsx parsed to 0 rows")
    table = pa.table(
        {
            "csv_file": pa.array(csv_file, pa.string()),
            "template": pa.array(template, pa.string()),
            "item": pa.array(item, pa.string()),
            "category": pa.array(category, pa.string()),
            "label": pa.array(label, pa.string()),
        }
    )
    save_raw_parquet(table, asset)


# --------------------------------------------------------------------------- #
# Supervisory Disclosure — measures & administrative penalties (CRD + IFD)
# --------------------------------------------------------------------------- #
def _latest_annex4(html: str, framework: str, part: int) -> str:
    pat = (
        r'href="(/sites/default/files/[^"]*?supervisory_disclosure_'
        + framework
        + r"_-_annex_4_-_part_"
        + str(part)
        + r'(?:_rev\d+)?\.xlsx)"'
    )
    links = set(re.findall(pat, html))
    if not links:
        raise AssertionError(f"supervisory-disclosure page exposed no {framework} annex-4 part {part}")

    def date_key(h):
        m = re.search(r"/(\d{4})-(\d{2})/", h)
        return (int(m.group(1)), int(m.group(2))) if m else (0, 0)

    return _abs(max(links, key=date_key))


def _parse_supervisory(content: bytes, framework: str) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    year = None
    out: list[dict] = []
    for sheet in wb.sheetnames:
        if sheet.strip().lower() == "index":
            continue
        category = None
        for row in wb[sheet].iter_rows(values_only=True):
            r = list(row) + [None] * 6
            if year is None:
                for c in r[:5]:
                    if c:
                        m = re.search(r"year\s+(\d{4})", str(c))
                        if m:
                            year = int(m.group(1))
            code, cat_cell, label, val = r[1], r[2], r[3], r[4]
            if cat_cell and str(cat_cell).strip():
                category = str(cat_cell).strip()
            if code is None or label is None or val is None:
                continue
            code_s = str(code).strip()
            if not re.match(r"^\d{2,3}$", code_s):
                continue
            try:
                count = int(float(val))
            except (TypeError, ValueError):
                continue
            out.append(
                {
                    "framework": framework,
                    "competent_authority": sheet.strip(),
                    "item_code": code_s,
                    "category": category,
                    "item_label": str(label).strip(),
                    "count": count,
                }
            )
    for rec in out:
        rec["reporting_year"] = year
    return out


def fetch_supervisory_measures_penalties(node_id: str) -> None:
    asset = node_id
    html = _fetch_text(SUP_DISCLOSURE_PAGE)
    records: list[dict] = []
    records += _parse_supervisory(_fetch_bytes(_latest_annex4(html, "crd", 5)), "CRD")
    records += _parse_supervisory(_fetch_bytes(_latest_annex4(html, "ifd", 3)), "IFD")
    if not records:
        raise AssertionError("supervisory measures/penalties parsed to 0 rows")
    table = pa.table(
        {
            "framework": pa.array([r["framework"] for r in records], pa.string()),
            "reporting_year": pa.array([r["reporting_year"] for r in records], pa.int32()),
            "competent_authority": pa.array([r["competent_authority"] for r in records], pa.string()),
            "item_code": pa.array([r["item_code"] for r in records], pa.string()),
            "category": pa.array([r["category"] for r in records], pa.string()),
            "item_label": pa.array([r["item_label"] for r in records], pa.string()),
            "count": pa.array([r["count"] for r in records], pa.int64()),
        }
    )
    save_raw_parquet(table, asset)


# --------------------------------------------------------------------------- #
# EU-wide Stress Test — bank-by-bank results
# --------------------------------------------------------------------------- #
# The headline results (capital, P&L, REA, summary — TRA_OTH.csv) are a tidy
# long fact table keyed by bank (LEI) × Period × Item × Scenario. Each exercise
# has its own landing page /eu-wide-stress-test-<year>, linked from the main
# stress-testing page; the file lives under /assets/st<yy>/full_database/<id>/,
# both segments rotating per release, so we re-discover from HTML each time. The
# deeper per-exposure credit-risk breakdowns (TRA_CRE_IRB/STA, ~1.1M rows, a
# different grain) are not a separate accepted entity and are not fetched here.
def _latest_stress_test_tra_oth_url() -> str:
    main = _fetch_text(STRESS_TEST_PAGE)
    landings = {}
    for href in re.findall(r'href="([^"]*eu-wide-stress-test-(\d{4})[^"]*)"', main):
        landings[int(href[1])] = href[0]
    if not landings:
        raise AssertionError("no eu-wide-stress-test-<year> landing page linked from the ST page")
    page = _fetch_text(_abs(landings[max(landings)]))
    csvs = re.findall(
        r'href=["\']([^"\']*?/assets/st\d+/full_database/[^"\']+/TRA_OTH\.csv)["\']', page
    )
    if not csvs:
        raise AssertionError("stress-test landing page exposed no TRA_OTH.csv full-database link")
    return _abs(csvs[0])


def fetch_stress_test_results(node_id: str) -> None:
    asset = node_id
    table = _read_te_csv(_fetch_bytes(_latest_stress_test_tra_oth_url()))
    if table.num_rows == 0:
        raise AssertionError("TRA_OTH.csv parsed to 0 rows")
    save_raw_parquet(table, asset)


DOWNLOAD_SPECS = [
    NodeSpec(id="eba-risk-dashboard-kri", fn=fetch_risk_dashboard_kri, kind="download"),
    NodeSpec(id="eba-risk-dashboard-credit-parameters", fn=fetch_risk_dashboard_credit_parameters, kind="download"),
    NodeSpec(id="eba-risk-dashboard-raq", fn=fetch_risk_dashboard_raq, kind="download"),
    NodeSpec(id="eba-te-credit-risk", fn=fetch_te_credit_risk, kind="download"),
    NodeSpec(id="eba-te-market-risk", fn=fetch_te_market_risk, kind="download"),
    NodeSpec(id="eba-te-sovereign-exposures", fn=fetch_te_sovereign_exposures, kind="download"),
    NodeSpec(id="eba-te-capital-own-funds", fn=fetch_te_capital_own_funds, kind="download"),
    NodeSpec(id="eba-te-other-exposures", fn=fetch_te_other_exposures, kind="download"),
    NodeSpec(id="eba-te-data-dictionary", fn=fetch_te_data_dictionary, kind="download"),
    NodeSpec(id="eba-supervisory-measures-penalties", fn=fetch_supervisory_measures_penalties, kind="download"),
    NodeSpec(id="eba-stress-test-results", fn=fetch_stress_test_results, kind="download"),
]
