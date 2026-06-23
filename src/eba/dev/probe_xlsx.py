import io
from subsets_utils import get
import openpyxl

url = "https://www.eba.europa.eu/sites/default/files/2026-06/b0e0d454-4e83-4bd2-8f46-6491dfe896da/Data%20Annex%20InteractiveRiskDashboard%20Q1%202026.xlsx"
r = get(url, timeout=(10,180))
print("status", r.status_code, "bytes", len(r.content))
wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
print("sheets:", wb.sheetnames)
for sn in wb.sheetnames[:8]:
    ws = wb[sn]
    print(f"\n--- sheet '{sn}' ---")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(i, [ (c if not isinstance(c,str) else c[:24]) for c in row[:12]])
        if i >= 6: break
