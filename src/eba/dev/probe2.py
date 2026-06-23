import io
from subsets_utils import get
import openpyxl

urls = {
 "credit_params": "https://www.eba.europa.eu/sites/default/files/2026-06/3445c227-75ab-4222-b1e4-ac1ba701b854/Credit%20Risk%20parameters%20annex%20-%20Q1%202026.xlsx",
 "raq": "https://www.eba.europa.eu/sites/default/files/2026-06/a3781e48-b19d-484a-80fa-b02cfaa01344/Statistical_annex_Spring2026RAQ%20%282%29.xlsx",
 "sup_pen": "https://www.eba.europa.eu/sites/default/files/2025-11/73da9c34-a85c-46e9-8c24-2c9dea34cb10/supervisory_disclosure_crd_-_annex_4_-_part_5.xlsx",
}
for name,url in urls.items():
    try:
        r = get(url, timeout=(10,180)); 
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        print(f"\n##### {name}  ({len(r.content)} bytes)  sheets={wb.sheetnames[:12]}")
        for sn in wb.sheetnames[:3]:
            ws = wb[sn]
            print(f"  -- '{sn}' --")
            for i,row in enumerate(ws.iter_rows(values_only=True)):
                print("   ",i,[ (c if not isinstance(c,str) else c[:20]) for c in row[:10]])
                if i>=5: break
    except Exception as e:
        print(name, "ERR", type(e).__name__, e)
