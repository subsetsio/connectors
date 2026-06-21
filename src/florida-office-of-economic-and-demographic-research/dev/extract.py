import json, io, datetime
import openpyxl
from subsets_utils import get

paths = json.load(open("dev/paths.json"))
ROOT = "https://edr.state.fl.us"


def _num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _txt(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _numeric_count(row):
    return sum(1 for c in row[1:] if _num(c) is not None)


def extract_sheet(sheet_name, rows):
    # drop fully-empty trailing/leading handled implicitly
    # find first data row: col0 is a non-empty label AND >=1 numeric in cols>=1.
    # Requiring a non-empty col0 excludes stacked-header rows that carry a stray
    # year number (e.g. "2025") but have an empty label column.
    first_data = None
    for i, row in enumerate(rows):
        if len(row) and _txt(row[0]) and _numeric_count(row) >= 1:
            first_data = i
            break
    if first_data is None:
        return []

    # header rows: non-empty rows above first_data that have content in cols>=1
    header_rows = []
    for i in range(first_data):
        row = rows[i]
        if any(_txt(c) for c in row[1:]):
            header_rows.append(row)

    ncols = max(len(r) for r in rows)

    def col_label(c):
        parts = []
        for hr in header_rows:
            if c < len(hr):
                t = _txt(hr[c])
                if t and (not parts or parts[-1] != t):
                    parts.append(t)
        lbl = " / ".join(parts)
        return lbl[:200] if lbl else f"col{c}"

    # row dimension name = col0 header (joined)
    row_dim_parts = []
    for hr in header_rows:
        t = _txt(hr[0]) if len(hr) else ""
        if t and (not row_dim_parts or row_dim_parts[-1] != t):
            row_dim_parts.append(t)
    row_dim = (" / ".join(row_dim_parts) or "row_label")[:120]

    labels = {c: col_label(c) for c in range(1, ncols)}

    out = []
    for row in rows[first_data:]:
        label = _txt(row[0]) if len(row) else ""
        if not label:
            continue
        low = label.lower()
        if low.startswith(("source", "note", "notes", "*", "total all", "footnote")):
            continue
        had = False
        for c in range(1, min(len(row), ncols)):
            v = row[c]
            t = _txt(v)
            if not t or t in ("-", "--", "n/a", "na"):
                continue
            had = True
            out.append({
                "sheet": sheet_name,
                "row_dim": row_dim,
                "row_label": label[:200],
                "measure": labels.get(c, f"col{c}"),
                "value_num": _num(v),
                "value_text": t[:300],
            })
        _ = had
    return out


def extract_workbook(content):
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    records = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        records.extend(extract_sheet(sn, rows))
    wb.close()
    return records


if __name__ == "__main__":
    for eid in ["flcopops", "advaltxco", "2025-pop-estimates",
                "314-300-utility-service-tax-water", "stxcolldist2024",
                "lottrates", "legislatorssalaries",
                "fy2024transientrentals-may2025revision"]:
        r = get(ROOT + paths[eid], timeout=(10, 120))
        recs = extract_workbook(r.content)
        print("\n===", eid, "=> records:", len(recs))
        import collections
        sheets = collections.Counter(x["sheet"] for x in recs)
        print("  sheets:", dict(list(sheets.items())[:4]), "..." if len(sheets) > 4 else "")
        for x in recs[:3]:
            print("   ", {k: (v[:40] if isinstance(v, str) else v) for k, v in x.items()})
