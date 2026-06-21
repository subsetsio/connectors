import json, io
import openpyxl
from subsets_utils import get

paths = json.load(open("dev/paths.json"))
ROOT = "https://edr.state.fl.us"

samples = [
    "flcopops",
    "advaltxco",
    "2025-pop-estimates",
    "314-300-utility-service-tax-water",
    "stxcolldist2024",
    "lottrates",
    "legislatorssalaries",
    "fy2024transientrentals-may2025revision",
    "countyformation",
]

for eid in samples:
    path = paths[eid]
    url = ROOT + path
    r = get(url, timeout=(10, 120))
    r.raise_for_status()
    print("\n" + "=" * 90)
    print(eid, "| HTTP", r.status_code, "| bytes", len(r.content), "| ctype", r.headers.get("content-type"))
    try:
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    except Exception as ex:
        print("  openpyxl FAILED:", type(ex).__name__, ex)
        continue
    print("  sheets:", wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    print("  dims:", ws.max_row, "x", ws.max_column)
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(row)
        if i >= 11:
            break
    for i, row in enumerate(rows):
        cells = ["" if c is None else str(c)[:18] for c in row[:9]]
        print(f"   r{i}: " + " | ".join(cells))
    wb.close()
