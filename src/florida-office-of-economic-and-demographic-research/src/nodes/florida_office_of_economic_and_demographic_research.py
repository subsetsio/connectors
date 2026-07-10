"""Florida Office of Economic and Demographic Research (EDR) connector.

EDR has no API or catalog endpoint — it publishes data as individual Excel
workbooks at stable URLs under https://edr.state.fl.us/Content/. Each rank-
accepted entity is one workbook (one published Delta table). The workbooks are
human-formatted: multi-sheet (a sheet often encodes a year vintage), with
banner/title rows and multi-row stacked headers above the real data.

A bespoke parser per workbook is infeasible at 125 heterogeneous files, so the
fetch fn applies one robust, layout-tolerant extraction that melts every sheet
into a uniform tidy/long shape:

    sheet | row_dim | row_label | measure | value_num | value_text

`row_label` is the first-column entity (county / municipality / fiscal year),
`measure` is the (possibly multi-row) column header joined with " / ", `sheet`
carries the per-sheet vintage. This is honest, SQL-readable, and non-empty for
every workbook. The SQL transform is then a thin type-and-project pass.

Fetch shape: stateless full re-pull (shape 1). The whole corpus re-downloads
each refresh — workbooks are small (tens of KB to <1 MB) and EDR overwrites them
in place as new vintages publish, so there is no incremental filter to exploit
and a stored watermark would only risk skipping revised vintages.
"""

import datetime
import io

import openpyxl
from subsets_utils import NodeSpec, get, save_raw_ndjson, transient_retry

from constants import ENTITY_IDS, SOURCE_PATHS

SLUG = "florida-office-of-economic-and-demographic-research"
PREFIX = f"{SLUG}-"
ROOT = "https://edr.state.fl.us"

# placeholder cell values that mean "no data" in EDR workbooks
_NULLISH = {"", "-", "--", "---", "n/a", "na", "n.a.", "*", "**"}


def _num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _txt(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _numeric_count(row):
    return sum(1 for c in row[1:] if _num(c) is not None)


def _extract_sheet(sheet_name, rows):
    """Melt one worksheet into long records. Returns [] for sheets with no
    detectable data block (note/methodology/blank sheets)."""
    # First data row: col0 is a non-empty label AND at least one numeric cell
    # follows. Requiring a non-empty col0 excludes stacked-header rows that carry
    # a stray year number (e.g. "2025") in an otherwise label-less row.
    first_data = None
    for i, row in enumerate(rows):
        if len(row) and _txt(row[0]) and _numeric_count(row) >= 1:
            first_data = i
            break
    if first_data is None:
        return []

    # Header rows: every non-empty row above the data that has content beyond
    # the label column (drops col0-only banner/title rows and blank rows).
    header_rows = [r for r in rows[:first_data] if any(_txt(c) for c in r[1:])]
    ncols = max(len(r) for r in rows)

    def joined_label(col):
        parts = []
        for hr in header_rows:
            if col < len(hr):
                t = _txt(hr[col])
                if t and (not parts or parts[-1] != t):
                    parts.append(t)
        return (" / ".join(parts))[:200]

    row_dim = joined_label(0) or "row_label"
    measures = {c: (joined_label(c) or f"col{c}") for c in range(1, ncols)}

    out = []
    for row in rows[first_data:]:
        label = _txt(row[0]) if len(row) else ""
        if not label:
            continue
        low = label.lower()
        if low.startswith(("source", "note", "notes", "footnote", "total all")):
            continue
        for c in range(1, min(len(row), ncols)):
            v = row[c]
            t = _txt(v)
            if not t or t.lower() in _NULLISH:
                continue
            out.append({
                "sheet": sheet_name[:120],
                "row_dim": row_dim[:120],
                "row_label": label[:200],
                "measure": measures.get(c, f"col{c}"),
                "value_num": _num(v),
                "value_text": t[:300],
            })
    return out


@transient_retry()
def _download(url):
    resp = get(url, timeout=(10.0, 180.0))
    resp.raise_for_status()
    return resp.content


def fetch_one(node_id: str) -> None:
    asset = node_id  # the spec id IS the asset name
    entity_id = node_id[len(PREFIX):]
    path = SOURCE_PATHS[entity_id]  # KeyError here = a bug in the spec set
    url = ROOT + path

    content = _download(url)
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        records = []
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
            records.extend(_extract_sheet(sheet_name, rows))
    finally:
        workbook.close()

    if not records:
        # A union workbook that melts to nothing is a real extraction failure —
        # fail loud rather than publish an empty table.
        raise ValueError(f"{asset}: extracted 0 records from {url}")

    save_raw_ndjson(records, asset)


DOWNLOAD_SPECS = [
    NodeSpec(
        id=f"{SLUG}-{eid.lower().replace('_', '-')}",
        fn=fetch_one,
        kind="download",
    )
    for eid in ENTITY_IDS
]
