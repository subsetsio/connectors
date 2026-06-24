"""EIOPA insurance statistics connector.

Mechanism: direct bulk download of the stable S3-hosted insurance-statistics
files (research mechanism `bulk_insurance_statistics`). Each subset is backed by
one or more files under a single asset directory; URLs are name-stable and the
files are full-corpus snapshots republished in place, so the fetch is a stateless
full re-pull every run (overwrite). No incremental filter exists and none is
needed — each file is a single GET of a few MB (xlsx exposure is ~66MB).

Six subsets are clean long-format CSVs (balance sheet / own funds / premiums,
solo + group); two are XLSX-only (asset exposures, financial-stability
indicators) and are parsed to normalized rows in the fetch fn. All raw is written
as gzipped NDJSON (per-entity schemas differ; within an entity rows are
homogeneous) and cast/published by the SQL transforms.
"""

import csv
import io
import json

import openpyxl

from subsets_utils import get, transient_retry, raw_writer
from constants import (
    BASE_URL,
    CSV_FILES,
    EXPOSURE_FILE,
    FS_FILE,
)


# ---------------------------------------------------------------- HTTP helpers

@transient_retry()
def _download_text(name: str) -> str:
    resp = get(f"{BASE_URL}/{name}.csv", timeout=(10.0, 180.0))
    resp.raise_for_status()
    return resp.text


@transient_retry()
def _download_bytes(name: str) -> bytes:
    resp = get(f"{BASE_URL}/{name}.xlsx", timeout=(10.0, 300.0))
    resp.raise_for_status()
    return resp.content


def _ndjson_sink(asset: str):
    """Context manager yielding a (write_row) over a gzipped NDJSON raw asset."""
    return raw_writer(asset, "ndjson.gz", mode="wt", compression="gzip")


# ------------------------------------------------------- generic CSV long block

def _to_float(v):
    if v is None:
        return None
    v = v.strip()
    if v == "":
        return None
    try:
        return float(v)
    except ValueError:
        return None


def _to_int(v):
    if v is None:
        return None
    v = v.strip()
    if v == "":
        return None
    try:
        return int(float(v))
    except ValueError:
        return None


def _csv_field_map(fieldnames):
    """Map this file's header row to canonical field keys (header text varies
    slightly across files, e.g. the submissions column suffix)."""
    m = {}
    for h in fieldnames:
        if h is None:
            continue
        hs = h.strip()
        low = hs.lower()
        if hs in ("Reporting country", "Region"):
            m["reporting_entity"] = h
        elif hs == "Reference period":
            m["reference_period"] = h
        elif hs == "Undertaking type":
            m["undertaking_type"] = h
        elif hs == "Item":                       # premiums: metric group
            m["metric"] = h
        elif hs == "Business type":              # premiums: line of business
            m["business_type"] = h
        elif hs == "Item code":
            m["item_code"] = h
        elif hs == "Item name":
            m["item_name"] = h
        elif hs == "Value":
            m["value"] = h
        elif low.startswith("number of submissions"):
            m["n_submissions"] = h
        elif low.startswith("date of extraction"):
            m["extraction_date"] = h
    return m


def fetch_csv_block(node_id: str) -> None:
    asset = node_id
    entity = node_id[len("eiopa-"):]   # spec id is prefixed; CSV_FILES keys are bare
    files = CSV_FILES[entity]
    with _ndjson_sink(asset) as fh:
        wrote = 0
        for name, frequency in files:
            text = _download_text(name)
            reader = csv.DictReader(io.StringIO(text))
            fm = _csv_field_map(reader.fieldnames or [])
            if "reporting_entity" not in fm or "item_code" not in fm:
                raise AssertionError(
                    f"{name}: unexpected header {reader.fieldnames!r}"
                )

            def _txt(r, key):
                """Canonical field as trimmed string, or None (missing/blank)."""
                col = fm.get(key)
                if col is None:
                    return None
                v = r.get(col)
                if v is None:
                    return None
                v = v.strip()
                return v or None

            for r in reader:
                row = {
                    "reporting_entity": _txt(r, "reporting_entity"),
                    "reference_period": _txt(r, "reference_period"),
                    "frequency": frequency,
                    "undertaking_type": _txt(r, "undertaking_type"),
                    "metric": _txt(r, "metric"),
                    "business_type": _txt(r, "business_type"),
                    "item_code": _txt(r, "item_code"),
                    "item_name": _txt(r, "item_name"),
                    "value": _to_float(r.get(fm["value"])) if "value" in fm else None,
                    "n_submissions": _to_int(r.get(fm["n_submissions"]))
                    if "n_submissions" in fm else None,
                    "extraction_date": _txt(r, "extraction_date"),
                }
                fh.write(json.dumps(row) + "\n")
                wrote += 1
        if wrote == 0:
            raise AssertionError(f"{asset}: no rows parsed from {files!r}")


# ---------------------------------------------------------------- asset exposures

def fetch_exposure(node_id: str) -> None:
    asset = node_id
    content = _download_bytes(EXPOSURE_FILE)
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb["Exposure (Raw data)"]
    rows = ws.iter_rows(values_only=True)
    header = [(h.strip() if isinstance(h, str) else h) for h in next(rows)]
    idx = {name: i for i, name in enumerate(header)}

    def cell(r, name):
        i = idx.get(name)
        return r[i] if (i is not None and i < len(r)) else None

    with _ndjson_sink(asset) as fh:
        wrote = 0
        for r in rows:
            if r is None or all(c is None for c in r):
                continue
            val = cell(r, "Value (euro millions)")
            ed = cell(r, "Date of extraction (yyyymmdd)")
            row = {
                "reference_period": cell(r, "Reference period"),
                "nca_iso_code": cell(r, "NCA_ISO_CODE"),
                "reporting_country": cell(r, "Reporting country"),
                "undertaking_type": cell(r, "Undertaking type"),
                "cic_main_category": cell(r, "CIC main category"),
                "cic_sub_category": cell(r, "CIC sub-category"),
                "portfolio_type": cell(r, "Portfolio type"),
                "location_of_investment": cell(r, "Location of investment"),
                "real_estate_exposure": (cell(r, "Real estate exposure") or None),
                "type_of_real_estate_exposure": (cell(r, "Type of real estate exposure") or None),
                "value_eur_millions": float(val) if isinstance(val, (int, float)) else None,
                "extraction_date": str(int(ed)) if isinstance(ed, (int, float)) else (ed or None),
            }
            fh.write(json.dumps(row) + "\n")
            wrote += 1
        if wrote == 0:
            raise AssertionError(f"{asset}: no exposure rows parsed")


# -------------------------------------------------- financial-stability indicators

def fetch_fs_indicators(node_id: str) -> None:
    asset = node_id
    content = _download_bytes(FS_FILE)
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb["FS Indicators"]
    all_rows = list(ws.iter_rows(values_only=True))

    # Locate the header row carrying the percentile labels.
    header_i = None
    for i, r in enumerate(all_rows):
        if r and any(isinstance(c, str) and "percentile" in c.lower() for c in r):
            header_i = i
            break
    if header_i is None:
        raise AssertionError(f"{asset}: percentile header row not found")

    with _ndjson_sink(asset) as fh:
        wrote = 0
        current_item = None
        for r in all_rows[header_i + 1:]:
            if r is None:
                continue
            item = r[0] if len(r) > 0 else None
            period = r[1] if len(r) > 1 else None
            if isinstance(item, str) and item.strip():
                current_item = item.strip()
            if not (isinstance(period, str) and period.strip()):
                continue  # group separator / blank row

            def num(i):
                v = r[i] if len(r) > i else None
                return float(v) if isinstance(v, (int, float)) else None

            row = {
                "item_name": current_item,
                "reference_period": period.strip(),
                "p10": num(2),
                "p25": num(3),
                "median": num(4),
                "p75": num(5),
                "p90": num(6),
                "n_observations": int(r[7]) if (len(r) > 7 and isinstance(r[7], (int, float))) else None,
            }
            fh.write(json.dumps(row) + "\n")
            wrote += 1
        if wrote == 0:
            raise AssertionError(f"{asset}: no FS indicator rows parsed")


# ----------------------------------------------------------------- DOWNLOAD_SPECS

from subsets_utils import NodeSpec, SqlNodeSpec  # noqa: E402

DOWNLOAD_SPECS = (
    [
        NodeSpec(id=eid, fn=fetch_csv_block, kind="download")
        for eid in CSV_FILES
    ]
    + [
        NodeSpec(id="solo-asset-exposures", fn=fetch_exposure, kind="download"),
        NodeSpec(id="financial-stability-indicators", fn=fetch_fs_indicators, kind="download"),
    ]
)

# Prefix every id with the slug, per contract: f"eiopa-{entity_id}".
DOWNLOAD_SPECS = [
    NodeSpec(id=f"eiopa-{s.id}", fn=s.fn, kind="download") for s in DOWNLOAD_SPECS
]


# ----------------------------------------------------------------- TRANSFORM_SPECS

def _generic_long_sql(dep: str, premiums: bool) -> str:
    extra = ""
    if premiums:
        extra = "metric,\n            business_type,\n            "
    return f'''
        SELECT DISTINCT
            reporting_entity,
            reference_period,
            frequency,
            undertaking_type,
            {extra}item_code,
            item_name,
            CAST(value AS DOUBLE) AS value,
            TRY_CAST(n_submissions AS BIGINT) AS n_submissions,
            CAST(try_strptime(extraction_date, '%Y%m%d') AS DATE) AS extraction_date
        FROM "{dep}"
        WHERE value IS NOT NULL
    '''


_EXPOSURE_SQL = '''
    SELECT DISTINCT
        reference_period,
        nca_iso_code,
        reporting_country,
        undertaking_type,
        cic_main_category,
        cic_sub_category,
        portfolio_type,
        location_of_investment,
        real_estate_exposure,
        type_of_real_estate_exposure,
        CAST(value_eur_millions AS DOUBLE) AS value_eur_millions,
        CAST(try_strptime(extraction_date, '%Y%m%d') AS DATE) AS extraction_date
    FROM "eiopa-solo-asset-exposures"
    WHERE value_eur_millions IS NOT NULL
'''

_FS_SQL = '''
    SELECT
        item_name,
        reference_period,
        CAST(p10 AS DOUBLE)    AS p10,
        CAST(p25 AS DOUBLE)    AS p25,
        CAST(median AS DOUBLE) AS median,
        CAST(p75 AS DOUBLE)    AS p75,
        CAST(p90 AS DOUBLE)    AS p90,
        TRY_CAST(n_observations AS BIGINT) AS n_observations
    FROM "eiopa-financial-stability-indicators"
    WHERE median IS NOT NULL
'''

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"eiopa-{eid}-transform",
        deps=[f"eiopa-{eid}"],
        sql=_generic_long_sql(f"eiopa-{eid}", premiums="premiums" in eid),
    )
    for eid in CSV_FILES
] + [
    SqlNodeSpec(
        id="eiopa-solo-asset-exposures-transform",
        deps=["eiopa-solo-asset-exposures"],
        sql=_EXPOSURE_SQL,
    ),
    SqlNodeSpec(
        id="eiopa-financial-stability-indicators-transform",
        deps=["eiopa-financial-stability-indicators"],
        sql=_FS_SQL,
    ),
]
