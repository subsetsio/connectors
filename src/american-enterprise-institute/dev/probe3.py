import io, re
from subsets_utils import get
import openpyxl, pyarrow as pa, duckdb
import sys
sys.path.insert(0,'src')
from nodes.american_enterprise_institute import _discover_workbook_url, COLUMNS, SCHEMA

url=_discover_workbook_url()
content=get(url,timeout=(10,180)).content
wb=openpyxl.load_workbook(io.BytesIO(content),read_only=True,data_only=True)
ws=wb[wb.sheetnames[0]]
it=ws.iter_rows(values_only=True); next(it)
rows=[]
for raw in it:
    if raw is None or all(c is None for c in raw) or raw[0] is None: continue
    rec={c:raw[i] for i,c in enumerate(COLUMNS) if i<len(raw)}
    for c in COLUMNS: rec.setdefault(c,None)
    rows.append(rec)
t=pa.Table.from_pylist(rows,schema=SCHEMA)
print("rows",len(t))
con=duckdb.connect()
con.register("american-enterprise-institute-housing-market-indicators",t)
sql=open('src/nodes/american_enterprise_institute.py').read()
q=sql.split("sql='''")[1].split("''',")[0]
r=con.execute(q).fetch_arrow_table()
print("transform rows",len(r))
print("cols",r.column_names)
print(con.execute("SELECT year,quarter,segment,median_sale_price FROM ("+q+") WHERE metro='National' ORDER BY year DESC, quarter DESC LIMIT 4").fetchall())
print("median price median", con.execute("SELECT median(median_sale_price) FROM ("+q+")").fetchone())
