from subsets_utils import get
import re, io
import openpyxl

PAGE = "https://www.aei.org/national-and-metro-housing-market-indicators/"
r = get(PAGE, timeout=(10,120))
links = re.findall(r'href="([^"]+\.xlsx[^"]*)"', r.text)
ts = [l for l in links if "data_download" in l.lower() or "interactive" in l.lower()]
ts = [l for l in ts if "top-100-metros" not in l.lower()]
print("chosen:", ts[0])
xr = get(ts[0], timeout=(10,180))
print("xlsx status", xr.status_code, "bytes", len(xr.content))
wb = openpyxl.load_workbook(io.BytesIO(xr.content), read_only=True, data_only=True)
print("sheets", wb.sheetnames)
ws = wb[wb.sheetnames[0]]
it = ws.iter_rows(values_only=True)
hdr = next(it)
print("ncols", len(hdr))
print("HEADER", hdr)
import collections
metros=set(); segs=set(); yqs=set(); n=0
samples=[]
for row in it:
    if all(c is None for c in row): continue
    n+=1
    metros.add(row[0]); segs.add(row[3]); yqs.add(row[2])
    if n<=3: samples.append(row)
print("rows", n, "metros", len(metros), "segments", segs)
print("yq sample sorted first/last", sorted(x for x in yqs if x)[:2], sorted(x for x in yqs if x)[-2:])
for s in samples: print("SAMPLE", s)
