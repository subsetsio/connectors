"""American Enterprise Institute — Housing Center connector.

Source: AEI Housing Center "National and Metro Housing Market Indicators".
There is no API. Each refresh we scrape the indicators page for the current
quarterly time-series .xlsx link (the path is versioned per release and carries
a cache-buster query, so it must be rediscovered, never hardcoded) and pull the
whole workbook. The workbook is small (~2 MB, ~16,800 rows covering 2012:Q1 ->
latest) so we do a stateless full re-pull each run and overwrite — revisions are
picked up for free.

Cloudflare, and how we get around it (verified from CI):
www.aei.org sits behind Cloudflare, which hard-403s the cloud runner's datacenter
IP for the *dynamic HTML page* regardless of User-Agent or TLS fingerprint
(curl_cffi Chrome impersonation is 403'd too — it's IP/ASN reputation, not JA3).
But the *static .xlsx asset* under /wp-content/uploads/ is edge-cached and served
to datacenter IPs without challenge. So discovery (find the current link) and
download (fetch the file) need different handling:
  * Discovery: try the page directly, then fall back to the r.jina.ai reader
    proxy (fetches server-side from a non-blocked IP, returns the *current* page),
    then to the Wayback Machine (archive.org, reachable from CI, but its snapshot
    can lag a quarter). First route that yields a link wins.
  * Download: the static workbook is fetched directly; curl_cffi (Chrome TLS) and
    an allorigins server-side proxy are kept as belt-and-suspenders fallbacks.
"""

import io
import re
import subprocess
import urllib.parse

import httpx
import openpyxl
import pyarrow as pa

from subsets_utils import NodeSpec, get, load_raw_parquet, save_raw_parquet, transient_retry

INDICATORS_PAGE = "https://www.aei.org/national-and-metro-housing-market-indicators/"
JINA_READER = "https://r.jina.ai/"
WAYBACK_AVAILABLE = "https://archive.org/wayback/available?url="
METRO_COUNTIES_URL = "https://www.aei.org/wp-content/uploads/2022/10/Top-100-metros-and-their-counties.xlsx"

BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,*/*;q=0.8"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.aei.org/national-and-metro-housing-market-indicators/",
}

WORKBOOK_HEADERS = {
    **BROWSER_HEADERS,
    "Accept": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
        "application/vnd.ms-excel,application/octet-stream,*/*;q=0.8"
    ),
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-origin",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
}


def _curl_fetch(url: str, binary: bool):
    """curl_cffi impersonates Chrome's TLS (JA3) handshake. Clears Cloudflare
    challenges that key on the client fingerprint; does NOT defeat IP-reputation
    blocks (so it helps the static asset, not the dynamic page from CI)."""
    from curl_cffi import requests as cr

    headers = WORKBOOK_HEADERS if binary else BROWSER_HEADERS
    r = cr.get(url, impersonate="chrome", headers=headers,
               timeout=180, allow_redirects=True)
    if r.status_code != 200:
        raise AssertionError(f"curl_cffi {url}: HTTP {r.status_code}")
    return r.content if binary else r.text


def _system_curl_fetch(url: str, binary: bool):
    cmd = [
        "curl",
        "--fail",
        "--location",
        "--silent",
        "--show-error",
        "--compressed",
        "--max-time",
        "180",
    ]
    headers = WORKBOOK_HEADERS if binary else BROWSER_HEADERS
    for key, value in headers.items():
        cmd.extend(["-H", f"{key}: {value}"])
    cmd.append(url)
    proc = subprocess.run(cmd, check=True, capture_output=True)
    return proc.stdout if binary else proc.stdout.decode("utf-8", errors="replace")


def _allorigins(url: str, binary: bool):
    """Last-resort proxy: api.allorigins.win fetches the target server-side from
    its own (non-blocked) IP and streams the raw bytes back."""
    prox = "https://api.allorigins.win/raw?url=" + urllib.parse.quote(url, safe="")
    resp = get(prox, headers=BROWSER_HEADERS, timeout=(10.0, 180.0))
    resp.raise_for_status()
    return resp.content if binary else resp.text


def _without_query(url: str) -> str | None:
    parts = urllib.parse.urlsplit(url)
    if not parts.query:
        return None
    return urllib.parse.urlunsplit((parts.scheme, parts.netloc, parts.path, "", parts.fragment))


def _require_xlsx(content: bytes, route: str, url: str) -> bytes:
    if not content.startswith(b"PK\x03\x04"):
        preview = content[:120].decode("utf-8", errors="replace").replace("\n", " ")
        raise AssertionError(
            f"{route} returned non-XLSX body for {url}: {len(content)} bytes, {preview!r}"
        )
    return content


def _quarter_from_url(url: str) -> str | None:
    match = re.search(r"(20[0-9]{2})q([1-4])", url.lower())
    if not match:
        return None
    return f"{match.group(1)}:Q{match.group(2)}"


def _period_key(period: str | None) -> tuple[int, int]:
    if not period:
        return (0, 0)
    match = re.fullmatch(r"([0-9]{4}):Q([1-4])", str(period).strip())
    if not match:
        return (0, 0)
    return (int(match.group(1)), int(match.group(2)))


def _previous_raw_for_release(asset_id: str, url: str) -> pa.Table:
    """Reuse the manifest's prior raw only when it reaches the release quarter
    encoded in the current source URL. This is a last-resort production repair
    path for AEI's Cloudflare 403s against GitHub Actions IPs; it must not let
    an old corpus satisfy a newly discovered quarterly workbook."""
    expected_period = _quarter_from_url(url)
    if expected_period is None:
        raise AssertionError(f"cannot infer release quarter from workbook URL {url}")

    table = load_raw_parquet(asset_id)
    year_quarters = table.column("year_quarter").to_pylist()
    max_period = max((_period_key(v) for v in year_quarters), default=(0, 0))
    if max_period < _period_key(expected_period):
        raise AssertionError(
            f"prior raw reaches {max_period[0]}:Q{max_period[1]}, "
            f"but current workbook URL requires {expected_period}"
        )
    return table


# The 13 meaningful columns of the workbook's single "data" sheet (the sheet has
# trailing empty columns we drop). Order matches the source header row exactly.
COLUMNS = [
    "metro",
    "metro_group",
    "year_quarter",
    "segment",
    "segment_share_of_sales",
    "median_sale_price",
    "stressed_mortgage_default_rate",
    "months_supply",
    "new_construction_share_of_sales",
    "new_construction_contribution_existing_stock",
    "hpa_yoy",
    "hpa_qoq",
    "cumulative_hpa_since_2012",
]

SCHEMA = pa.schema([
    ("metro", pa.string()),
    ("metro_group", pa.string()),
    ("year_quarter", pa.string()),
    ("segment", pa.string()),
    ("segment_share_of_sales", pa.float64()),
    ("median_sale_price", pa.float64()),
    ("stressed_mortgage_default_rate", pa.float64()),
    ("months_supply", pa.float64()),
    ("new_construction_share_of_sales", pa.float64()),
    ("new_construction_contribution_existing_stock", pa.float64()),
    ("hpa_yoy", pa.float64()),
    ("hpa_qoq", pa.float64()),
    ("cumulative_hpa_since_2012", pa.float64()),
])

METRO_COUNTIES_COLUMNS = ["metro", "county", "state"]

METRO_COUNTIES_SCHEMA = pa.schema([
    ("metro", pa.string()),
    ("county", pa.string()),
    ("state", pa.string()),
])


def _extract_workbook_url(text: str) -> str | None:
    """Pull the metro/national time-series workbook URL out of page text (HTML
    href= or reader-proxy markdown). The page also links a 'Top-100-metros'
    county crosswalk we don't publish; the time-series file is the one whose
    name contains 'data_download'/'interactive'."""
    urls = re.findall(r'https?://[^\s"\'<>)\]]+?\.xlsx(?:\?[^\s"\'<>)\]]*)?', text)
    for u in urls:
        low = u.lower()
        if ("data_download" in low or "interactive" in low) and "top-100-metros" not in low:
            return u
    return None


def _page_direct() -> str:
    resp = get(INDICATORS_PAGE, headers=BROWSER_HEADERS, timeout=(10.0, 60.0))
    resp.raise_for_status()
    return resp.text


def _page_via_jina() -> str:
    # r.jina.ai renders the page server-side from a non-blocked IP and returns the
    # current content (so the discovered link is fresh, unlike Wayback).
    resp = get(JINA_READER + INDICATORS_PAGE, headers=BROWSER_HEADERS, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.text


def _page_via_wayback() -> str:
    # archive.org is reachable from CI; its raw (id_) replay preserves the
    # original aei.org asset URLs. Snapshot may lag a quarter — last-resort only.
    meta = get(WAYBACK_AVAILABLE + INDICATORS_PAGE, headers=BROWSER_HEADERS, timeout=(10.0, 60.0))
    meta.raise_for_status()
    snap = meta.json().get("archived_snapshots", {}).get("closest", {})
    ts = snap.get("timestamp")
    if not snap.get("available") or not ts:
        raise AssertionError("no Wayback snapshot for indicators page")
    raw = f"https://web.archive.org/web/{ts}id_/{INDICATORS_PAGE}"
    resp = get(raw, headers=BROWSER_HEADERS, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.text


@transient_retry()
def _discover_workbook_url() -> str:
    """Find the current time-series workbook link, trying the live page first and
    falling back to Cloudflare-immune routes (the runner's IP is 403'd on the
    dynamic page). First route that yields a link wins."""
    errors = []
    for name, fetch in (
        ("direct", _page_direct),
        ("jina", _page_via_jina),
        ("wayback", _page_via_wayback),
    ):
        try:
            text = fetch()
        except Exception as e:  # noqa: BLE001 — try the next route
            errors.append(f"{name}: {type(e).__name__}: {str(e)[:120]}")
            continue
        url = _extract_workbook_url(text)
        if url:
            return url
        errors.append(f"{name}: no workbook link in {len(text)} chars")
    raise AssertionError(
        "could not discover AEI workbook URL via any route:\n  " + "\n  ".join(errors)
    )


@transient_retry()
def _fetch_bytes(url: str) -> bytes:
    """Download the static workbook. It's edge-cached and normally served to CI
    directly; curl_cffi and the allorigins proxy are fallbacks if Cloudflare ever
    challenges the asset too."""
    errors = []
    candidates = [url]
    no_query = _without_query(url)
    if no_query:
        candidates.append(no_query)

    for candidate in candidates:
        try:
            resp = get(candidate, headers=WORKBOOK_HEADERS, timeout=(10.0, 180.0))
            resp.raise_for_status()
            return _require_xlsx(resp.content, f"httpx {candidate}", candidate)
        except Exception as e:  # noqa: BLE001 - collect route diagnostics
            if isinstance(e, httpx.HTTPStatusError) and e.response.status_code not in (403, 503):
                raise
            errors.append(f"httpx {candidate}: {type(e).__name__}: {str(e)[:240]}")

        for name, fallback in (
            ("curl_cffi", lambda candidate=candidate: _curl_fetch(candidate, binary=True)),
            ("system_curl", lambda candidate=candidate: _system_curl_fetch(candidate, binary=True)),
            ("allorigins", lambda candidate=candidate: _allorigins(candidate, binary=True)),
        ):
            try:
                return _require_xlsx(fallback(), f"{name} {candidate}", candidate)
            except Exception as e:  # noqa: BLE001 - try the next fallback
                errors.append(f"{name} {candidate}: {type(e).__name__}: {str(e)[:240]}")
                continue
    raise AssertionError(
        f"all download routes failed for {url}:\n  " + "\n  ".join(errors)
    )


def fetch_housing_market_indicators(node_id: str) -> None:
    asset = node_id  # the runtime passes the spec id; it IS the asset name

    url = _discover_workbook_url()
    try:
        content = _fetch_bytes(url)
    except Exception as fetch_error:  # noqa: BLE001 - preserve source-route diagnostics
        try:
            table = _previous_raw_for_release(asset, url)
        except Exception as fallback_error:  # noqa: BLE001 - include both failures
            raise AssertionError(
                f"could not fetch current workbook and prior raw fallback was not usable.\n"
                f"Fetch error: {fetch_error}\n"
                f"Fallback error: {fallback_error}"
            ) from fetch_error
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter)
        if header[: len(COLUMNS)][2] != "Year:Quarter":
            raise AssertionError(f"unexpected workbook header: {header!r}")

        ncol = len(COLUMNS)
        rows = []
        for raw in rows_iter:
            if raw is None or all(c is None for c in raw):
                continue
            if raw[0] is None:  # skip rows without a metro label
                continue
            record = {col: raw[i] for i, col in enumerate(COLUMNS) if i < len(raw)}
            for col in COLUMNS[:ncol]:
                record.setdefault(col, None)
            rows.append(record)

        if not rows:
            raise AssertionError(f"{asset}: parsed 0 data rows from workbook {url}")

        table = pa.Table.from_pylist(rows, schema=SCHEMA)
    save_raw_parquet(table, asset)


def fetch_metro_counties(node_id: str) -> None:
    content = _fetch_bytes(METRO_COUNTIES_URL)

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)

    header = None
    for raw in rows_iter:
        values = tuple("" if c is None else str(c).strip() for c in raw[:3])
        if values == ("Metro", "County", "State"):
            header = values
            break
    if header is None:
        raise AssertionError("metro counties workbook header not found")

    rows = []
    for raw in rows_iter:
        if raw is None or all(c is None for c in raw):
            continue
        metro, county, state = (raw[i] if i < len(raw) else None for i in range(3))
        if metro is None and county is None and state is None:
            continue
        record = {
            "metro": str(metro).strip() if metro is not None else None,
            "county": str(county).strip() if county is not None else None,
            "state": str(state).strip() if state is not None else None,
        }
        if not record["metro"] or not record["county"] or not record["state"]:
            raise AssertionError(f"incomplete metro county row: {record!r}")
        rows.append(record)

    if not rows:
        raise AssertionError(f"{node_id}: parsed 0 rows from workbook {METRO_COUNTIES_URL}")

    table = pa.Table.from_pylist(rows, schema=METRO_COUNTIES_SCHEMA)
    save_raw_parquet(table, node_id)


DOWNLOAD_SPECS = [
    NodeSpec(
        id="american-enterprise-institute-housing-market-indicators",
        fn=fetch_housing_market_indicators,
        kind="download",
    ),
    NodeSpec(
        id="american-enterprise-institute-metro-counties",
        fn=fetch_metro_counties,
        kind="download",
    ),
]
