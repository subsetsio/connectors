import openpyxl, io
for label,path in {"M5":"dev/wb_a35c9a.xlsx","S10":"dev/wb_9b2be0.xlsx"}.items():
    content=open(path,'rb').read()
    # try read_only with reset_dimensions
    wb=openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws=wb["Data"]; 
    try: ws.reset_dimensions()
    except Exception as e: print("reset err",e)
    n=sum(1 for _ in ws.iter_rows(values_only=True))
    print(f"{label} read_only+reset Data rows={n}")
    wb.close()
    # try non-read_only
    wb2=openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    grid=[list(r) for r in wb2["Data"].iter_rows(values_only=True)]
    print(f"{label} non-readonly Data rows={len(grid)}; first cells:")
    for i,r in enumerate(grid[:7]): print("  ",i,[ (str(c)[:16] if c is not None else None) for c in r[:6]])
    wb2.close()
