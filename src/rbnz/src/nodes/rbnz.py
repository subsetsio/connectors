"""Reserve Bank of New Zealand (RBNZ) statistical series connector.

RBNZ publishes ~99 statistical series as Excel workbooks under a stable media
path. One Delta table is published per series (the catalog/rank unit). Each
series has one or more option variants (frequency / long-run history /
breakdown), and each option is composed of one or more .xlsx workbook files
(historical splits). This module fetches every workbook for a series, parses it
into a uniform long format — (date, indicator, value) plus series metadata —
and publishes one table per series.

Fetch shape: stateless full re-pull (shape 1). Each workbook is small (KB-MB);
re-fetching the full corpus every run is cheap and picks up RBNZ's in-place
revisions for free. No watermark/cursor.

URL pattern (verified against the CRAN `RBNZ` package, github.com/rntq472/RBNZ):
  https://www.rbnz.govt.nz/-/media/project/sites/rbnz/files/statistics/series/{group}/{code_lower}/{stem}.xlsx
The {group} folder is a hyphenated prefix bucket; the exact mapping is fuzzy for
a few series, so the fetch tries the plausible candidates in order until one
returns a valid workbook.

Parsing: the standard RBNZ workbook has a "Series Definitions" sheet (one row of
metadata per data column) and a "Data" sheet (col A = date, remaining columns =
one numeric indicator each, header on the row above the first date). A handful
of series use special multi-sheet layouts. The parser is generic and
sheet-agnostic: for every non-metadata sheet it auto-detects the date column
(the column with the most datetime cells) and the header row, then melts each
numeric indicator column to long rows. This degrades gracefully across the
standard and special layouts without hand-coding each one.

FETCH STRATEGY (live origin first, Wayback fallback):
www.rbnz.govt.nz sits behind Cloudflare, which hard-403s datacenter IP ranges —
verified on the GitHub Actions runner, every request returns 403 regardless of
client (httpx AND curl_cffi Chrome-TLS impersonation both blocked; the block is
IP reputation, not TLS fingerprint or auth). So each workbook is fetched with a
single fast curl_cffi attempt against the live origin (authoritative + fresh,
and it succeeds from a clean IP), and on any failure we fall back to the Wayback
Machine (archive.org is NOT Cloudflare-fronted and is reachable from the cloud
runner). Wayback's latest snapshot is found via the CDX API and fetched raw via
the `id_` endpoint. This mirrors the `ioc` connector's pattern. Trade-off: from
the (blocked) CI runner the data is as fresh as Wayback's most recent crawl
(observed: within a few months); a clean-IP runner gets live, current data.
"""

import datetime
import gzip
import io

import pyarrow as pa
from curl_cffi import requests as cffi_requests

from subsets_utils import (
    NodeSpec,
    SqlNodeSpec,
    get,
    save_raw_parquet,
    transient_retry,
)

# --- entity union (the accepted series) ------------------------------------
from constants import ENTITY_IDS

# series_code -> {option_name: [workbook file stem(s)]} (historical splits within
# an option concatenate). Transcribed from the CRAN RBNZ package's seriesOptions.
SERIES_OPTIONS = {
    "B1": {"daily": ["hb1-daily-1973-1998", "hb1-daily-1999-2017", "hb1-daily"],
           "monthly": ["hb1-monthly-1973-1998", "hb1-monthly"]},
    "B2": {"dailyclose": ["hb2-daily-close-1985-2017", "hb2-daily-close"],
           "monthlyclose": ["hb2-monthly-close"]},
    "B3": {"default": ["hb3"]},
    "B4": {"NZD": ["hb4nzd"], "USD": ["hb4usd"], "EUR": ["hb4eur"]},
    "B6": {"default": ["hb6"]},
    "B7": {"default": ["hb7"]},
    "B10": {"default": ["hb10"]},
    "B13": {"default": ["hb13-1983-to-1998", "hb13-1999-to-2014", "hb13"]},
    "B20": {"default": ["hb20"]},
    "B21": {"default": ["hb21"]},
    "B25": {"default": ["hb25"]},
    "B26": {"default": ["hb26"]},
    "B27": {"default": ["hb27"]},
    "B30": {"default": ["hb30"]},
    "C5": {"default": ["hc5"]},
    "C12": {"monthend": ["hc12-month-end"], "dailyaverage": ["hc12-daily-average-balances"]},
    "C13": {"issued": ["hc13-issued-in-nz"], "used": ["hc13-used-in-nz"]},
    "C21": {"default": ["hc21"]},
    "C22": {"default": ["hc22"]},
    "C30": {"default": ["hc30"]},
    "C31": {"default": ["hc31"]},
    "C32": {"default": ["hc32"]},
    "C33": {"default": ["hc33"]},
    "C35": {"default": ["hc35"]},
    "C40": {"default": ["hc40"]},
    "C41": {"default": ["hc41"]},
    "C50": {"default": ["hc50"], "longrun": ["hc50-long-run"]},
    "C51": {"default": ["hc51"]},
    "C52": {"default": ["hc52"]},
    "C55": {"default": ["hc55"]},
    "C60": {"default": ["hc60"]},
    "C65": {"default": ["hc65"]},
    "C66": {"default": ["hc66"]},
    "C70": {"default": ["hc70"]},
    "C71": {"default": ["hc71"]},
    "D3": {"default": ["hd3"]},
    "D9": {"weekly": ["hd9-weekly"], "monthly": ["hd9-monthly"]},
    "D10": {"default": ["hd10"]},
    "D12": {"default": ["hd12"]},
    "D30": {"default": ["hd30"]},
    "D31": {"default": ["hd31"]},
    "D35": {"default": ["hd35"]},
    "E1": {"default": ["he1"]},
    "E2": {"default": ["he2"]},
    "F3": {"default": ["hf3"]},
    "F4": {"default": ["hf4"]},
    "F5": {"default": ["hf5"]},
    "H1": {"default": ["hh1"]},
    "H2": {"default": ["hh2"]},
    "H3": {"default": ["hh3"]},
    "J10": {"default": ["hj10"]},
    "J20": {"default": ["hj20"]},
    "L1": {"default": ["hl1"]},
    "L2": {"default": ["hl2"]},
    "L3": {"default": ["hl3"]},
    "M1": {"default": ["hm1"], "longrun": ["hm1-long"]},
    "M2": {"default": ["hm2"]},
    "M3": {"default": ["hm3"]},
    "M4": {"default": ["hm4"]},
    "M5": {"default": ["hm5"]},
    "M6": {"default": ["hm6"]},
    "M7": {"default": ["hm7"]},
    "M8": {"default": ["hm8"]},
    "M9": {"default": ["hm9"]},
    "M10": {"default": ["hm10"]},
    "M12": {"default": ["hm12"]},
    "M14": {"default": ["hm14"]},
    "M15": {"default": ["hm15"]},
    "R1": {"default": ["hr1"]},
    "R2": {"default": ["hr2"]},
    "R3": {"default": ["hr3"]},
    "S10": {"default": ["hs10"], "longrun": ["hs10-long-run"], "moredetail": ["hs10m"]},
    "S20": {"default": ["hs20"]},
    "S21": {"default": ["hs21"]},
    "S30": {"default": ["hs30"]},
    "S31": {"default": ["hs31"]},
    "S32": {"default": ["hs32"]},
    "S33": {"default": ["hs33"], "longrun": ["hs33-long-run"]},
    "S34": {"default": ["hs34"]},
    "S35": {"default": ["hs35"]},
    "S36": {"default": ["hs36"]},
    "S37": {"default": ["hs37"]},
    "S40": {"default": ["hs40"], "longrun": ["hs40-long-run"]},
    "S41": {"default": ["hs41"]},
    "S42": {"default": ["hs42"]},
    "S45": {"default": ["hs45"], "total": ["hs45t"]},
    "S46": {"default": ["hs46"], "average": ["hs46a"], "total": ["hs46t"]},
    "S50": {"default": ["hs50"], "longrun": ["hs50-long-run"]},
    "S51": {"default": ["hs51"]},
    "T1": {"default": ["ht1"]},
    "T4": {"default": ["ht4"]},
    "T11": {"default": ["ht11"]},
    "T21": {"default": ["ht21"]},
    "T31": {"default": ["ht31"]},
    "T40": {"default": ["ht40"]},
    "T41": {"default": ["ht41"]},
    "T42": {"default": ["ht42"]},
    "T43": {"default": ["ht43"]},
    "T44": {"default": ["ht44"]},
    "T45": {"default": ["ht45"]},
    "T46": {"default": ["ht46"]},
    "T47": {"default": ["ht47"]},
    "T48": {"default": ["ht48"]},
}

_MEDIA_BASE = "https://www.rbnz.govt.nz/-/media/project/sites/rbnz/files/statistics/series"
_IMPERSONATE = "chrome"

SCHEMA = pa.schema([
    ("series_code", pa.string()),
    ("option", pa.string()),
    ("file_stem", pa.string()),
    ("sheet", pa.string()),
    ("indicator_label", pa.string()),
    ("series_id", pa.string()),
    ("series_name", pa.string()),
    ("unit", pa.string()),
    ("date", pa.date32()),
    ("value", pa.float64()),
])

_META_SHEET_HINTS = ("series definition", "definitions", "definition")
_SKIP_SHEET_HINTS = ("series definition", "definitions", "definition", "notes",
                     "note", "disclaimer", "cover", "contents", "index", "about",
                     "table description", "check", "working", "calc")


# --- HTTP ------------------------------------------------------------------
# Latest archived snapshot in ONE request: a far-future timestamp makes Wayback
# redirect to the most recent capture. This avoids the separate CDX lookup,
# which rate-limits (503/504) hard under the 99-subprocess burst and was the
# dominant cause of node timeouts.
_WAYBACK_LATEST = "https://web.archive.org/web/29991231235959id_/"

# Transient classes for the Wayback path (httpx via subsets_utils.get).


def _try_live(url: str) -> bytes | None:
    """One fast attempt at the live origin via Chrome-TLS impersonation. RBNZ
    Cloudflare-blocks datacenter IPs, so this normally fails from CI — we do NOT
    retry (a retry storm here is what makes the run crawl); just fall through to
    Wayback. From a clean IP it returns fresh, authoritative bytes."""
    try:
        resp = cffi_requests.get(url, impersonate=_IMPERSONATE, timeout=30)
        if resp.status_code == 200 and resp.content[:2] == b"PK":
            return resp.content
    except Exception:
        pass
    return None


@transient_retry(attempts=7, min_wait=2, max_wait=30)
def _wayback_bytes(orig_url: str) -> bytes | None:
    """Latest archived copy of orig_url via a single id_ request, or None if the
    URL was never archived (Wayback 404 — wrong group folder)."""
    # Generous connect timeout: web.archive.org connections are slow to
    # establish under the concurrent-subprocess burst, and a too-tight connect
    # deadline turns a working-but-slow connection into a retry/timeout cascade.
    resp = get(_WAYBACK_LATEST + orig_url, timeout=(30.0, 180.0))
    if resp.status_code == 404:
        return None
    resp.raise_for_status()
    body = resp.content
    if body[:2] != b"PK":
        try:
            body = gzip.decompress(body)  # id_ may serve gzip-encoded bytes
        except (OSError, gzip.BadGzipFile):
            pass
    return body if body[:2] == b"PK" else None


# Exact media-path folder for series whose group is NOT just the lowercased
# first letter — verified live. D/F/R split awkwardly across 'd' and 'd-f-r';
# pinning the correct one avoids a wasted (and under-load, timeout-prone) probe
# of the wrong folder.
_GROUP_OVERRIDE = {
    "D3": "d-f-r", "D10": "d-f-r", "D12": "d-f-r",
    "D9": "d", "D30": "d", "D31": "d", "D35": "d",
    "F3": "d-f-r", "F4": "d-f-r", "F5": "d-f-r",
    "R1": "d-f-r", "R2": "d-f-r", "R3": "d-f-r",
}


def _group_candidates(code: str) -> list:
    """The media-path folder bucket(s) for a series. Returns exactly one when
    known (the common case — no wasted wrong-folder request); L/S map to 'l-s',
    everything else to its lowercased first letter."""
    if code in _GROUP_OVERRIDE:
        return [_GROUP_OVERRIDE[code]]
    letter = code[0]
    if letter in ("S", "L"):
        return ["l-s"]
    return [letter.lower()]


def _download_stem(code: str, stem: str) -> bytes:
    for grp in _group_candidates(code):
        url = f"{_MEDIA_BASE}/{grp}/{code.lower()}/{stem}.xlsx"
        content = _try_live(url)        # 1. live origin (fresh, clean-IP only)
        if content:
            return content
        content = _wayback_bytes(url)   # 2. Wayback fallback (CI-reachable)
        if content:
            return content
        # neither path produced this candidate's file — wrong group folder; the
        # real file lives under another candidate, so try the next one.
    raise FileNotFoundError(
        f"could not fetch workbook {code}/{stem} live or via Wayback from any "
        f"group candidate {_group_candidates(code)}"
    )


# --- workbook parsing ------------------------------------------------------
# RBNZ workbooks share one layout once read correctly: a "Data" sheet whose
# leading rows are stacked headers (category / sub-category name, a "Notes" row,
# a "Unit" row, a "Series Id" row — each keyed by a label in column 0), followed
# by data rows with the date in column 0 and one numeric indicator per remaining
# column. A separate "Series Definitions" sheet duplicates the per-column id /
# name / unit. We extract metadata INLINE from the Data sheet's header rows
# (robust to column count), and fall back to the definitions sheet keyed by
# series id. NOTE: openpyxl read_only mode trusts a stored worksheet dimension
# that several RBNZ files get wrong (truncating to 1 row) — we call
# reset_dimensions() to force a real scan.

_DATE_FORMATS = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m")
_NULL_TOKENS = {"", "..", "...", "-", "n/a", "na", "nan", "null"}


def _clean(v):
    if v is None:
        return None
    s = str(v).replace("\n", " ").replace("\r", " ").strip()
    s = " ".join(s.split())
    return s or None


def _as_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    # Bare year (int or numeric string) — some workbooks (e.g. F3) index rows by
    # year. Narrow range avoids mistaking ordinary integer values for years.
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if float(v).is_integer() and 1850 <= int(v) <= 2100:
            return datetime.date(int(v), 1, 1)
        return None
    if isinstance(v, str):
        s = v.strip()
        for fmt in _DATE_FORMATS:
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        if s.isdigit() and 1850 <= int(s) <= 2100:
            return datetime.date(int(s), 1, 1)
    return None


def _to_float(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if s.lower() in _NULL_TOKENS:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _read_grid(ws):
    try:
        ws.reset_dimensions()  # force a real scan; RBNZ files mis-declare dims
    except AttributeError:
        pass
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    if not grid:
        return [], 0
    ncols = max((len(r) for r in grid), default=0)
    for r in grid:
        if len(r) < ncols:
            r.extend([None] * (ncols - len(r)))
    return grid, ncols


def _load_meta_by_id(wb):
    """{series_id: {name, unit}} from the Series Definitions sheet. {} if absent."""
    sheet = next((n for n in wb.sheetnames
                  if any(h in n.lower() for h in _META_SHEET_HINTS)), None)
    if sheet is None:
        return {}
    grid, ncols = _read_grid(wb[sheet])
    if not grid:
        return {}
    header_idx = next((i for i, row in enumerate(grid)
                       if len([c for c in row if _clean(c)]) >= 2), None)
    if header_idx is None:
        return {}
    low = [(_clean(c) or "").lower() for c in grid[header_idx]]

    def find_col(*needles):
        for j, h in enumerate(low):
            if any(n in h for n in needles):
                return j
        return None

    id_col = find_col("series id", "seriesid")
    unit_col = find_col("unit")
    name_col = find_col("description", "name") or find_col("series")
    if id_col is None:
        return {}
    out = {}
    for row in grid[header_idx + 1:]:
        def cell(j):
            return _clean(row[j]) if (j is not None and j < len(row)) else None
        sid = cell(id_col)
        if sid:
            out.setdefault(sid, {"name": cell(name_col), "unit": cell(unit_col)})
    return out


def _parse_sheet(ws, meta_by_id):
    """Melt one worksheet to long records. Returns [] if no date column found."""
    grid, ncols = _read_grid(ws)
    if ncols == 0:
        return []

    # date column = the column with the most date-parseable cells
    best_col, best_count = None, 0
    for c in range(ncols):
        cnt = sum(1 for r in grid if _as_date(r[c]) is not None)
        if cnt > best_count:
            best_count, best_col = cnt, c
    if best_col is None or best_count < 3:
        return []
    date_col = best_col

    first_data = next(i for i, r in enumerate(grid) if _as_date(r[date_col]) is not None)

    # header region (rows above first data row): locate the inline label rows
    id_row = unit_row = None
    name_rows = []
    for row in grid[:first_data]:
        key = (_clean(row[date_col]) or "").lower()
        if key in ("series id", "seriesid", "series_id"):
            id_row = row
        elif key == "unit":
            unit_row = row
        elif key in ("notes", "note", "group", ""):
            if key == "" and any(_clean(c) for c in row):
                name_rows.append(row)  # unlabelled row = category/name header
        # else: a labelled row we don't recognise -> ignore
    # Fallback label row when there is no "Series Id" row (non-standard sheets
    # like F3, whose real column headers sit above a blank separator row): pick
    # the header-region row with the most filled value-column cells.
    fallback_header = None
    best_fill = 0
    for row in grid[:first_data]:
        fill = sum(1 for c in range(ncols)
                   if c != date_col and c < len(row) and _clean(row[c]))
        if fill > best_fill:
            best_fill, fallback_header = fill, row

    def col_val(row, c):
        return _clean(row[c]) if (row is not None and c < len(row)) else None

    out = []
    for c in range(ncols):
        if c == date_col:
            continue
        sid = col_val(id_row, c)
        name = " - ".join(p for p in (col_val(r, c) for r in name_rows) if p) or None
        unit = col_val(unit_row, c)
        label = sid or name or col_val(fallback_header, c)
        if not label:
            continue
        # column qualifies only if it carries numeric data
        if not any(_to_float(r[c]) is not None for r in grid[first_data:]):
            continue
        meta = meta_by_id.get(sid or label, {})
        series_name = name or meta.get("name")
        series_unit = unit or meta.get("unit")
        for r in grid[first_data:]:
            d = _as_date(r[date_col])
            if d is None:
                continue
            val = _to_float(r[c])
            if val is None:
                continue
            out.append({
                "indicator_label": label,
                "series_id": sid,
                "series_name": series_name,
                "unit": series_unit,
                "date": d,
                "value": val,
                "_sheet": ws.title,
            })
    return out


def _parse_workbook(content: bytes, code: str, option: str, stem: str) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        meta_by_id = _load_meta_by_id(wb)
        out = []
        for sheet_name in wb.sheetnames:
            if any(h in sheet_name.lower() for h in _SKIP_SHEET_HINTS):
                continue
            for rec in _parse_sheet(wb[sheet_name], meta_by_id):
                out.append({
                    "series_code": code,
                    "option": option,
                    "file_stem": stem,
                    "sheet": rec["_sheet"],
                    "indicator_label": rec["indicator_label"],
                    "series_id": rec["series_id"],
                    "series_name": rec["series_name"],
                    "unit": rec["unit"],
                    "date": rec["date"],
                    "value": rec["value"],
                })
        return out
    finally:
        wb.close()


# --- fetch -----------------------------------------------------------------
def fetch_one(node_id: str) -> None:
    asset = node_id  # the runtime passes the spec id; it IS the asset name
    code = node_id[len("rbnz-"):].upper()
    options = SERIES_OPTIONS[code]

    all_rows = []
    for option, stems in options.items():
        for stem in stems:
            content = _download_stem(code, stem)
            all_rows.extend(_parse_workbook(content, code, option, stem))

    if not all_rows:
        raise ValueError(
            f"{asset}: parsed 0 data rows from {sum(len(s) for s in options.values())} "
            f"workbook file(s) - the RBNZ layout for {code} may have changed"
        )

    table = pa.Table.from_pylist(all_rows, schema=SCHEMA)
    save_raw_parquet(table, asset)


DOWNLOAD_SPECS = [
    NodeSpec(id=f"rbnz-{eid.lower().replace('_', '-')}", fn=fetch_one, kind="download")
    for eid in ENTITY_IDS
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"{s.id}-transform",
        deps=[s.id],
        temporal="date",
        sql=f'''
            SELECT DISTINCT
                CAST(date AS DATE)     AS date,
                series_code,
                option,
                indicator_label,
                series_id,
                series_name,
                unit,
                CAST(value AS DOUBLE)  AS value
            FROM "{s.id}"
            WHERE value IS NOT NULL AND date IS NOT NULL
            ORDER BY date, indicator_label
        ''',
    )
    for s in DOWNLOAD_SPECS
]
