"""Shared transport + value-coercion helpers for the Freedom House connector.

Freedom House publishes no programmatic API — its data ships as a small fixed set
of bulk Excel workbooks, one per report family, each a full multi-decade
country-year panel hosted at a stable (date-prefixed) URL on freedomhouse.org.
Each subset module in nodes/ does its own (era-varying) Excel reshaping; the
transport and the generic cell-coercion live here once.
"""

import io
import re

from openpyxl import load_workbook

from subsets_utils import get, transient_retry

_NA = {"", "-", "–", "—", "n/a", "na", "n.a.", "..", "."}

# Worksheet-name substrings that mark a notes/index/intro sheet (not the data).
_NOTES = ("index", "intro", "understanding", "key", "notes", "distribution")


# ---------------------------------------------------------------------------
# transport
# ---------------------------------------------------------------------------


@transient_retry()
def _download_xlsx(url: str) -> bytes:
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.content


def _workbook(url: str):
    return load_workbook(io.BytesIO(_download_xlsx(url)), read_only=True, data_only=True)


def _sheet_rows(ws) -> list[list]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


# ---------------------------------------------------------------------------
# value coercion
# ---------------------------------------------------------------------------
def _txt(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.lower() in _NA:
        return None
    return s


def _num(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s.lower() in _NA:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _year(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    m = re.search(r"(19|20)\d{2}", str(v))
    return int(m.group(0)) if m else None


def _find_header_row(rows: list[list], marker: str, limit: int = 8) -> int:
    """Index of the row whose first cell equals `marker` (case-insensitive)."""
    m = marker.strip().lower()
    for i, row in enumerate(rows[:limit]):
        if row and _txt(row[0]) and _txt(row[0]).lower() == m:
            return i
    raise AssertionError(f"header row starting with {marker!r} not found in first {limit} rows")


def _data_sheet(wb, *, exclude_substrings):
    """Pick the worksheet that is NOT a notes/index/intro sheet (largest by rows
    among the survivors), so we don't depend on a year-stamped sheet name."""
    cand = [
        sn
        for sn in wb.sheetnames
        if not any(x in sn.lower() for x in exclude_substrings)
    ]
    if not cand:
        cand = list(wb.sheetnames)
    cand.sort(key=lambda sn: wb[sn].max_row, reverse=True)
    return wb[cand[0]]
