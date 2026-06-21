import io
from openpyxl import load_workbook
from subsets_utils import get


def wb_for(url):
    r = get(url, timeout=(10, 120))
    r.raise_for_status()
    return load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)


def rows(ws, n):
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        out.append(list(row))
        if i >= n:
            break
    return out


# 1. FIW all data full header
wb = wb_for("https://freedomhouse.org/sites/default/files/2025-02/All_data_FIW_2013-2024.xlsx")
ws = wb["FIW13-25"]
hdr = rows(ws, 1)[1]
print("FIW-ALL header (", len(hdr), "cols):")
print(hdr)
wb.close()

# 2. FOTN full header
wb = wb_for("https://freedomhouse.org/sites/default/files/2023-09/FOTN_2023_Country_Score_Data.xlsx")
ws = wb["FOTN 2023 score data"]
hdr = rows(ws, 1)[1]
print("\nFOTN header (", len(hdr), "cols):")
print(hdr)
wb.close()

# 3. FOTP wide structure — print first 3 header rows fully, and unique sub-headers
wb = wb_for("https://freedomhouse.org/sites/default/files/2020-02/FOTP1980-FOTP2017_Public-Data.xlsx")
ws = wb["Data"]
rr = rows(ws, 4)
print("\nFOTP r2 (Edition):", rr[2])
print("\nFOTP r3 (Years):", rr[3])
print("\nFOTP r4 (subheaders):", rr[4])
wb.close()

# 4. FIW ratings r0/r1/r2 fully (year block structure)
wb = wb_for("https://freedomhouse.org/sites/default/files/2025-02/Country_and_Territory_Ratings_and_Statuses_FIW_1973-2024.xlsx")
ws = wb["Country Ratings, Statuses "]
rr = rows(ws, 2)
print("\nRATINGS r0 (edition):", rr[0][:20])
print("RATINGS r1 (years):", rr[1][:20])
print("RATINGS r2 (sub):", rr[2][:20])
print("RATINGS ncols:", len(rr[0]))
wb.close()
