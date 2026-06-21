import io
from openpyxl import load_workbook
from subsets_utils import get

URLS = {
    "fiw-all-data": "https://freedomhouse.org/sites/default/files/2025-02/All_data_FIW_2013-2024.xlsx",
    "fiw-ratings-statuses": "https://freedomhouse.org/sites/default/files/2025-02/Country_and_Territory_Ratings_and_Statuses_FIW_1973-2024.xlsx",
    "nations-in-transit": "https://freedomhouse.org/sites/default/files/2024-05/All_Data_Nations_in_Transit_NIT_2005-2024_For_website.xlsx",
    "freedom-on-the-net": "https://freedomhouse.org/sites/default/files/2023-09/FOTN_2023_Country_Score_Data.xlsx",
    "freedom-of-the-press": "https://freedomhouse.org/sites/default/files/2020-02/FOTP1980-FOTP2017_Public-Data.xlsx",
}


def dump(name, url):
    print("=" * 80)
    print(name, url)
    r = get(url, timeout=(10, 120))
    r.raise_for_status()
    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    print("sheets:", wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"  --- sheet '{sn}'  dims={ws.max_row}x{ws.max_column}")
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(row)
            if i >= 5:
                break
        for i, row in enumerate(rows):
            print(f"    r{i}:", [c for c in row][:14])
    wb.close()


for n, u in URLS.items():
    try:
        dump(n, u)
    except Exception as e:
        print(name if False else n, "ERROR", type(e).__name__, e)
