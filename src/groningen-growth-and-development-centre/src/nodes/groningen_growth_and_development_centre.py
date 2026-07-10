"""Groningen Growth and Development Centre (GGDC) connector.

Mechanism: the DataverseNL native API (https://dataverse.nl) hosting GGDC's
official database collection. Each rank-accepted entity is one Dataverse
dataset (a GGDC database release). We resolve the dataset's file listing by
DOI, download the one tabular file we publish, and reshape it into a tidy long
table with a numeric ``value`` column. The model stage compiles one Delta-table
transform per dataset from the profiled raw.

Heterogeneous Excel/CSV layouts (PWT wide panels, Maddison long, ETD/ASUT/WIOD
matrices, the EU KLEMS release family of long CSVs) are normalised in Python
here, because the transform layer is SQL-only and cannot read xlsx/zip. The full
corpus is small (a few hundred MB across the accepted datasets) and the
Dataverse datasets are immutable
per published version, so each fetch is a stateless full re-pull (no watermark,
no incremental query — none is offered).
"""

import io

import openpyxl
import pandas as pd
import pyarrow as pa

from subsets_utils import NodeSpec, get, save_raw_parquet, transient_retry
from constants import DATASETS

SLUG = "groningen-growth-and-development-centre"
DATAVERSE = "https://dataverse.nl"


# --------------------------------------------------------------------------- #
# Dataverse access
# --------------------------------------------------------------------------- #
@transient_retry()
def _resolve_file_id(doi: str, filename: str) -> int:
    """Look up the numeric datafile id for `filename` inside the dataset `doi`."""
    resp = get(
        f"{DATAVERSE}/api/datasets/:persistentId/",
        params={"persistentId": f"doi:{doi}"},
        timeout=(10.0, 120.0),
    )
    resp.raise_for_status()
    files = resp.json()["data"]["latestVersion"]["files"]
    for f in files:
        df = f["dataFile"]
        if df.get("filename") == filename:
            return int(df["id"])
    raise AssertionError(
        f"file {filename!r} not found in {doi}; have: "
        f"{[f['dataFile'].get('filename') for f in files]}"
    )


@transient_retry()
def _download_datafile(file_id: int) -> bytes:
    """Fetch a Dataverse datafile (303-redirects to a signed store URL; GET follows)."""
    resp = get(f"{DATAVERSE}/api/access/datafile/{file_id}", timeout=(10.0, 600.0))
    resp.raise_for_status()
    return resp.content


# --------------------------------------------------------------------------- #
# Excel helpers
# --------------------------------------------------------------------------- #
def _read_sheet(content: bytes, sheet: str) -> pd.DataFrame:
    """Read one worksheet into a DataFrame, dropping fully-empty rows.

    Uses openpyxl read_only iteration (not pandas.read_excel) because several
    GGDC workbooks carry an inflated <dimension> that makes read_excel scan
    ~1M phantom rows; read_only iter_rows yields only the stored cells.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header = list(next(rows))
        ncol = len(header)
        data = [r[:ncol] for r in rows if any(c is not None for c in r)]
    finally:
        wb.close()
    return pd.DataFrame(data, columns=header)


def _coerce_year(series: pd.Series) -> pd.Series:
    """Year value -> nullable Int64 (handles 'y1965', '_1995', '2000', 1)."""
    digits = series.astype(str).str.replace(r"\D", "", regex=True)
    return pd.to_numeric(digits, errors="coerce").astype("Int64")


def _finish(df: pd.DataFrame, asset: str) -> None:
    """Coerce value to float, drop null observations, save tidy parquet."""
    df = df.copy()
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    df = df.dropna(subset=["value"])
    if "year" in df.columns:
        df["year"] = df["year"].astype("Int64")
    # Object columns -> string (None preserved) for a clean arrow conversion.
    for col in df.columns:
        if col not in ("value", "year") and df[col].dtype == object:
            df[col] = df[col].astype("string")
    table = pa.Table.from_pandas(df, preserve_index=False)
    save_raw_parquet(table, asset)


# --------------------------------------------------------------------------- #
# Parsers — each returns a tidy long frame; _finish() writes it
# --------------------------------------------------------------------------- #
def _parse_pwt(content, params, asset):
    df = _read_sheet(content, params["sheet"])
    id_vars = ["countrycode", "country", "currency_unit", "year"]
    df["year"] = _coerce_year(df["year"])
    long = df.melt(id_vars=id_vars, var_name="variable", value_name="value")
    _finish(long, asset)


def _parse_maddison(content, params, asset):
    df = _read_sheet(content, params["sheet"])
    id_vars = ["countrycode", "country", "region", "year"]
    df["year"] = _coerce_year(df["year"])
    value_vars = [c for c in ("gdppc", "pop") if c in df.columns]
    long = df.melt(id_vars=id_vars, value_vars=value_vars,
                   var_name="variable", value_name="value")
    _finish(long, asset)


def _parse_pld(content, params, asset):
    df = _read_sheet(content, params["sheet"])
    id_vars = ["countrycode", "year", "sector"]
    df["year"] = _coerce_year(df["year"])
    # Numeric measures only; the i_ppp text flag coerces to NaN and is dropped.
    value_vars = [c for c in df.columns if c not in id_vars]
    long = df.melt(id_vars=id_vars, value_vars=value_vars,
                   var_name="variable", value_name="value")
    _finish(long, asset)


def _parse_etd(content, params, asset):
    df = _read_sheet(content, params["sheet"])
    id_vars = ["country", "cnt", "var", "year"]
    df["year"] = _coerce_year(df["year"])
    sector_cols = [c for c in df.columns if c not in id_vars]
    long = df.melt(id_vars=id_vars, value_vars=sector_cols,
                   var_name="sector", value_name="value")
    _finish(long, asset)


def _parse_year_wide(content, params, asset):
    df = _read_sheet(content, params["sheet"])
    id_vars = params["id_vars"]
    year_cols = [c for c in df.columns if c not in id_vars]
    long = df.melt(id_vars=id_vars, value_vars=year_cols,
                   var_name="year", value_name="value")
    long["year"] = _coerce_year(long["year"])
    long = long.dropna(subset=["year"])
    long.columns = [str(c).lower() for c in long.columns]
    _finish(long, asset)


def _parse_asut(content, params, asset):
    """Africa Supply/Use/IOT matrices -> long (row x col industry) per table type."""
    frames = []
    for sheet, dom in (("SUPPLY", None), ("USE", "Dom_Imp"), ("IOT", None)):
        df = _read_sheet(content, sheet)
        df["year"] = _coerce_year(df["year"])
        base = ["cnt", "var", "year", "cisic4"]
        if dom and dom in df.columns:
            base.append(dom)
        ind_cols = [c for c in df.columns if c not in base]
        long = df.melt(id_vars=base, value_vars=ind_cols,
                       var_name="col_industry", value_name="value")
        long = long.rename(columns={"cisic4": "row_industry"})
        long["tabletype"] = sheet
        long["dom_imp"] = long[dom] if dom and dom in long.columns else pd.NA
        long = long[["cnt", "var", "year", "tabletype", "dom_imp",
                     "row_industry", "col_industry", "value"]]
        frames.append(long)
    _finish(pd.concat(frames, ignore_index=True), asset)


def _parse_euklems(content, params, asset):
    """EU KLEMS growth-accounts output module: already long.

    Column set drifts across releases: the industry classification is `isic3`
    in the 2007-2011 releases and `isic4` from 2012 on; releases also carry one
    or two ordering helpers (`sort_id`, `alt_sort_id`) that are not data. Keep
    iso3 / var / <industry> / year / value and normalise the industry column to
    `isic` so the release family publishes a uniform schema.
    """
    df = pd.read_csv(io.BytesIO(content))
    isic = "isic4" if "isic4" in df.columns else ("isic3" if "isic3" in df.columns else None)
    keep = ["iso3", "var", "year", "value"]
    if isic:
        keep.insert(2, isic)
    df = df[[c for c in keep if c in df.columns]]
    if isic:
        df = df.rename(columns={isic: "isic"})
    df["year"] = _coerce_year(df["year"])
    _finish(df, asset)


_PARSERS = {
    "pwt": _parse_pwt,
    "maddison": _parse_maddison,
    "pld": _parse_pld,
    "etd": _parse_etd,
    "year_wide": _parse_year_wide,
    "asut": _parse_asut,
    "euklems": _parse_euklems,
}


# --------------------------------------------------------------------------- #
# Download node — one generic fetch fn dispatched by entity id
# --------------------------------------------------------------------------- #
def fetch_one(node_id: str) -> None:
    asset = node_id  # the spec id IS the asset name
    entity_id = node_id[len(SLUG) + 1:]  # strip "<slug>-"
    cfg = DATASETS[entity_id]
    file_id = _resolve_file_id(cfg["doi"], cfg["file"])
    content = _download_datafile(file_id)
    _PARSERS[cfg["parser"]](content, cfg["params"], asset)


DOWNLOAD_SPECS = [
    NodeSpec(
        id=f"{SLUG}-{eid.lower().replace('_', '-')}",
        fn=fetch_one,
        kind="download",
    )
    for eid in DATASETS
]
