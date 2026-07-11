"""Download nodes for International IDEA."""

from __future__ import annotations

from datetime import datetime, timezone
import io

import openpyxl

from subsets_utils import NodeSpec, save_raw_ndjson

from utils import GSOD_API, request

VT_EXPORT = "https://www.idea.int/data-tools/export"


def fetch_gsod_indicators(node_id: str) -> None:
    resp = request(f"{GSOD_API}/labels")
    rows = resp.json()
    if not isinstance(rows, list) or not rows:
        raise ValueError(f"GSoD /api/labels returned no rows: {str(rows)[:200]}")
    save_raw_ndjson(rows, node_id)


def fetch_gsod_indices(node_id: str) -> None:
    year2 = datetime.now(tz=timezone.utc).year + 1
    resp = request(f"{GSOD_API}/data", params={"year1": 1900, "year2": year2})
    rows = resp.json()
    if not isinstance(rows, list) or not rows:
        raise ValueError(f"GSoD /api/data returned no rows: {str(rows)[:200]}")
    save_raw_ndjson(rows, node_id)


def fetch_voter_turnout(node_id: str) -> None:
    resp = request(
        VT_EXPORT,
        params={"type": "region_only", "themeId": 293, "world": "all", "loc": "home"},
    )
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    try:
        if "All" not in wb.sheetnames:
            raise ValueError(f"voter turnout export missing 'All' sheet: {wb.sheetnames}")

        rows = []
        sheet = wb["All"]
        sheet_rows = sheet.iter_rows(values_only=True)
        header = [str(value).strip() for value in next(sheet_rows)]
        for raw_row in sheet_rows:
            if all(value is None for value in raw_row):
                continue
            row = {}
            for index, name in enumerate(header):
                value = raw_row[index] if index < len(raw_row) else None
                row[name] = None if value is None else str(value)
            rows.append(row)
    finally:
        wb.close()

    if not rows:
        raise ValueError("voter turnout 'All' sheet produced 0 rows")
    save_raw_ndjson(rows, node_id)


DOWNLOAD_SPECS = [
    NodeSpec(id="international-idea-gsod-indicators", fn=fetch_gsod_indicators, kind="download"),
    NodeSpec(id="international-idea-gsod-indices", fn=fetch_gsod_indices, kind="download"),
    NodeSpec(id="international-idea-voter-turnout", fn=fetch_voter_turnout, kind="download"),
]
