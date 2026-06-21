"""Voter Turnout Database — fetched as the XLSX bulk export.

/data-tools/export, themeId=293. The 'All' sheet carries the full per-country,
per-election turnout rows; openpyxl can only run in Python, so the fetch fn
parses the sheet to string rows and writes ndjson for SQL.
"""

import io

import openpyxl

from subsets_utils import NodeSpec, SqlNodeSpec, save_raw_ndjson

from utils import request

VT_EXPORT = "https://www.idea.int/data-tools/export"


def fetch_voter_turnout(node_id: str) -> None:
    asset = node_id
    # The 'region_only' export type with world=all returns the full per-country,
    # per-election table on the 'All' sheet (Parliamentary/Presidential/EU sheets
    # are subsets of it). openpyxl only runs in Python, so normalise to string
    # rows here and let the SQL transform do the typing.
    resp = request(
        VT_EXPORT,
        params={"type": "region_only", "themeId": 293, "world": "all", "loc": "home"},
    )
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    try:
        if "All" not in wb.sheetnames:
            raise ValueError(f"voter turnout export missing 'All' sheet: {wb.sheetnames}")
        ws = wb["All"]
        it = ws.iter_rows(values_only=True)
        header = [str(h).strip() for h in next(it)]
        rows = []
        for raw in it:
            if all(c is None for c in raw):
                continue
            row = {}
            for i, name in enumerate(header):
                v = raw[i] if i < len(raw) else None
                row[name] = None if v is None else str(v)
            rows.append(row)
    finally:
        wb.close()
    if not rows:
        raise ValueError("voter turnout 'All' sheet produced 0 rows")
    save_raw_ndjson(rows, asset)


DOWNLOAD_SPECS = [
    NodeSpec(id="international-idea-voter-turnout", fn=fetch_voter_turnout, kind="download"),
]


TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="international-idea-voter-turnout-transform",
        deps=["international-idea-voter-turnout"],
        sql='''
            SELECT
                "Country"                                                      AS country,
                "ISO2"                                                         AS iso2,
                "ISO3"                                                         AS iso3,
                "Election Type"                                                AS election_type,
                TRY_CAST("Year" AS DATE)                                       AS election_date,
                TRY_CAST(REPLACE(CAST("Voter Turnout" AS VARCHAR), '%', '') AS DOUBLE)          AS voter_turnout_pct,
                TRY_CAST(REPLACE(CAST("Total vote" AS VARCHAR), ',', '') AS BIGINT)             AS total_vote,
                TRY_CAST(REPLACE(CAST("Registration" AS VARCHAR), ',', '') AS BIGINT)           AS registration,
                TRY_CAST(REPLACE(CAST("VAP Turnout" AS VARCHAR), '%', '') AS DOUBLE)            AS vap_turnout_pct,
                TRY_CAST(REPLACE(CAST("Voting age population" AS VARCHAR), ',', '') AS BIGINT)  AS voting_age_population,
                TRY_CAST(REPLACE(CAST("Population" AS VARCHAR), ',', '') AS BIGINT)             AS population,
                TRY_CAST(REPLACE(CAST("Invalid votes" AS VARCHAR), '%', '') AS DOUBLE)          AS invalid_votes_pct,
                "Compulsory voting"                                            AS compulsory_voting
            FROM "international-idea-voter-turnout"
            WHERE "Country" IS NOT NULL
        ''',
    ),
]
