"""OPEC connector — Monthly Oil Market Report (MOMR) statistical appendix.

The only verified machine-readable surface (see research) is the MOMR
"appendix tables" workbook, published ~monthly at
  https://www.opec.org/assets/assetdb/momr-appendix-<month>-<year>.xlsx
served directly from a static path that bypasses the Cloudflare challenge on
opec.org's HTML pages. Each workbook holds five fixed statistical sheets
("Table 11 - 1" .. "Table 11 - 5"), each a human-formatted grid (multi-row
headers, region/country row labels, reporting-period columns, interspersed
"Change" delta columns and footnotes). SQL can't parse that, so each download
fn does the xlsx -> tidy long-format work in Python (openpyxl) and saves
parquet; a thin DuckDB transform dedups and publishes one Delta table per sheet.

Coverage strategy (one download spec per of the 5 entity-union tables): each
workbook is a rolling ~2-year snapshot, so we walk the last N monthly vintages,
parse each spec's own sheet, and accumulate (report_date, section, item, period,
value) long-format. The transform keeps, per (series, period), the value from
the most recent report that carried it — assembling a continuous series from the
overlapping snapshots while preserving each value's report-vintage provenance.

Filenames are irregular (month spelled out or abbreviated; the Jan-2025 file is
appendix-tables-january-2025.xlsx) and the canonical link list lives only on a
Cloudflare-blocked index, so discovery probes candidate month-token URLs
backward from the current month — a wrong name returns a 200 HTML 404 page
(NOT a 4xx), so a hit requires content-type/URL/magic-bytes checks, not status.
No incremental query exists; this is a full re-pull of the recent vintage window
each refresh (cheap: ~18 workbooks x ~150KB per spec).
"""

import datetime as _dt
import io

import httpx
import pyarrow as pa

from subsets_utils import (
    NodeSpec,
    SqlNodeSpec,
    configure_http,
    get,
    save_raw_parquet,
    transient_retry,
)

# --- discovery tuning ------------------------------------------------------
_BASE = "https://www.opec.org/assets/assetdb/"
_TARGET_VINTAGES = 18      # monthly workbooks to accumulate per table
_MAX_MONTHS = 40           # how far back to probe before giving up
_STOP_AFTER_MISSES = 8     # consecutive month-misses => past the archive edge
_BROWSER_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
)

# entity_id -> the workbook sheet it publishes from + descriptive metadata.
_TABLES = {
    "table-11-1": {"sheet": "Table 11 - 1",
                   "title": "World oil demand and production balance, mb/d"},
    "table-11-2": {"sheet": "Table 11 - 2",
                   "title": "World oil demand and production balance - revisions vs prior issue, mb/d"},
    "table-11-3": {"sheet": "Table 11 - 3",
                   "title": "OECD oil stocks and oil on water at the end of period"},
    "table-11-4": {"sheet": "Table 11 - 4",
                   "title": "Non-DoC liquids production and DoC NGLs by country, mb/d"},
    "table-11-5": {"sheet": "Table 11 - 5",
                   "title": "World rig count, units"},
}

# month number -> filename tokens to try (irregular: full + observed abbreviations)
_MONTH_TOKENS = {
    1: ["january", "jan"], 2: ["february", "feb"], 3: ["march", "mar"],
    4: ["april", "apr"], 5: ["may"], 6: ["june", "jun"],
    7: ["july", "jul"], 8: ["august", "aug"], 9: ["september", "sept", "sep"],
    10: ["october", "oct"], 11: ["november", "nov"], 12: ["december", "dec"],
}

# --- HTTP transport --------------------------------------------------------


@transient_retry(attempts=5, max_wait=60)
def _http_get(url: str) -> httpx.Response:
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp


def _try_workbook(url: str) -> bytes | None:
    """Return workbook bytes if `url` resolves to a real xlsx, else None.

    A wrong filename on this host redirects to /404.html, which answers 403
    (observed on the cloud runner) or 200-with-an-HTML-body (observed locally)
    depending on the egress. So a "miss" must be recognised by BOTH a permanent
    403/404 status AND the final-URL / magic-byte checks — never let either kill
    the spec, since probing for a non-existent month is normal control flow.
    """
    try:
        resp = _http_get(url)
    except httpx.HTTPStatusError as e:
        if e.response.status_code in (403, 404):
            return None
        raise
    if "404" in str(resp.url):
        return None
    content = resp.content
    if not content or content[:2] != b"PK":   # xlsx is a zip -> "PK"
        return None
    return content


def _candidate_urls(year: int, month: int) -> list[str]:
    urls = [f"{_BASE}momr-appendix-{tok}-{year}.xlsx" for tok in _MONTH_TOKENS[month]]
    if (year, month) == (2025, 1):             # observed one-off prefix
        urls.append(f"{_BASE}appendix-tables-january-2025.xlsx")
    return urls


# --- xlsx parsing ----------------------------------------------------------
_NOTE_PREFIXES = ("source", "note", "totals may", "* ", "** ", "*** ", "**** ",
                  "(a) -", "the full publication")


def _classify_period(cell):
    """Map a header cell to (period_str, frequency, period_start_date) or None.

    Years -> annual; NQyy -> quarterly; datetime -> monthly. 'Change' / ratio
    columns ('2025/24', 'May/Apr') match nothing and are dropped.
    """
    if isinstance(cell, _dt.datetime):
        return (f"{cell.year:04d}-{cell.month:02d}", "monthly",
                _dt.date(cell.year, cell.month, 1))
    if isinstance(cell, _dt.date):
        return (f"{cell.year:04d}-{cell.month:02d}", "monthly",
                _dt.date(cell.year, cell.month, 1))
    s = str(cell).strip()
    if len(s) == 4 and s.isdigit():
        y = int(s)
        if 1960 <= y <= 2100:
            return (s, "annual", _dt.date(y, 1, 1))
        return None
    if len(s) == 4 and s[0] in "1234" and s[1] in "Qq" and s[2:].isdigit():
        q = int(s[0])
        year = 2000 + int(s[2:])
        return (f"{q}Q{s[2:]}", "quarterly", _dt.date(year, (q - 1) * 3 + 1, 1))
    return None


def _num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return f if f == f else None   # drop NaN
    return None


def _label_cell(row, first_period_col):
    """Leftmost non-empty text cell before the period columns -> (col_idx, text)."""
    for ci in range(0, first_period_col):
        v = row[ci]
        if isinstance(v, str) and v.strip():
            return ci, v.strip()
    return None, None


def _parse_sheet(rows):
    """Generic MOMR-appendix grid -> list of dicts {section,item,period,frequency,
    period_start,value}. Handles label-only section rows, column-indent nesting
    (e.g. region sub-rows under a stock category), and mixed annual/quarterly/
    monthly period columns."""
    # 1) header row = first row carrying >=3 period-like cells.
    header_idx = None
    period_cols = {}
    for r_i, row in enumerate(rows):
        cols = {}
        for c_i, cell in enumerate(row):
            if cell is None:
                continue
            cls = _classify_period(cell)
            if cls is not None:
                cols[c_i] = cls
        if len(cols) >= 3:
            header_idx, period_cols = r_i, cols
            break
    if header_idx is None:
        return []

    first_pcol = min(period_cols)
    out = []
    section = ""
    stack = []   # [(col_idx, label)] for column-indent nesting within a section

    for row in rows[header_idx + 1:]:
        lc, label = _label_cell(row, first_pcol)
        if label is None:
            continue
        low = label.lower()
        vals = [(c, period_cols[c]) for c in period_cols if _num(row[c]) is not None]

        if not vals:
            # label-only row: a section header (or a trailing footnote we skip)
            if any(low.startswith(p) for p in _NOTE_PREFIXES):
                continue
            section = label
            stack = []
            continue

        if any(low.startswith(p) for p in _NOTE_PREFIXES):
            continue

        # column-indent nesting: parent = labels strictly left of this one
        while stack and stack[-1][0] >= lc:
            stack.pop()
        parent = [lbl for _, lbl in stack]
        item = " / ".join(parent + [label])
        stack.append((lc, label))

        for c, (period, freq, pstart) in vals:
            out.append({
                "section": section,
                "item": item,
                "period": period,
                "frequency": freq,
                "period_start": pstart,
                "value": _num(row[c]),
            })
    return out


_SCHEMA = pa.schema([
    ("report_period", pa.string()),
    ("report_date", pa.date32()),
    ("table", pa.string()),
    ("table_title", pa.string()),
    ("section", pa.string()),
    ("item", pa.string()),
    ("period", pa.string()),
    ("frequency", pa.string()),
    ("period_start", pa.date32()),
    ("value", pa.float64()),
])


def fetch_one(node_id: str) -> None:
    configure_http(headers={"User-Agent": _BROWSER_UA})
    from openpyxl import load_workbook

    entity_id = node_id[len("opec-"):]
    meta = _TABLES[entity_id]
    sheet_name, title = meta["sheet"], meta["title"]

    now = _dt.datetime.now(_dt.timezone.utc)
    year, month = now.year, now.month

    records = []
    vintages = 0
    misses = 0
    months_tried = 0
    while (months_tried < _MAX_MONTHS and vintages < _TARGET_VINTAGES
           and misses < _STOP_AFTER_MISSES):
        content = None
        for url in _candidate_urls(year, month):
            content = _try_workbook(url)
            if content is not None:
                break
        if content is not None:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            try:
                if sheet_name in wb.sheetnames:
                    rows = list(wb[sheet_name].iter_rows(values_only=True))
                    parsed = _parse_sheet(rows)
                    rp = f"{year:04d}-{month:02d}"
                    rd = _dt.date(year, month, 1)
                    for rec in parsed:
                        rec.update({
                            "report_period": rp, "report_date": rd,
                            "table": sheet_name, "table_title": title,
                        })
                        records.append(rec)
                    vintages += 1
            finally:
                wb.close()
            misses = 0
        else:
            misses += 1
        month -= 1
        if month == 0:
            month, year = 12, year - 1
        months_tried += 1

    if not records:
        raise RuntimeError(
            f"{node_id}: no MOMR appendix vintages parsed for sheet "
            f"'{sheet_name}' (probed {months_tried} months back from "
            f"{now.year}-{now.month:02d})"
        )

    cols = {name: [r[name] for r in records] for name in _SCHEMA.names}
    table = pa.table(cols, schema=_SCHEMA)
    save_raw_parquet(table, node_id)


DOWNLOAD_SPECS = [
    NodeSpec(id=f"opec-{eid}", fn=fetch_one, kind="download")
    for eid in _TABLES
]


# --- transforms: one published Delta table per sheet -----------------------
# Dedup to the most-recent reporting vintage per (series, period); build a clean
# series label from section+item. 0 rows fails the node (correctness gate).
def _transform_sql(download_id: str) -> str:
    return f'''
        WITH ranked AS (
            SELECT
                CASE WHEN section IS NULL OR section = ''
                     THEN item ELSE section || ' - ' || item END  AS series,
                section,
                item,
                period,
                frequency,
                CAST(period_start AS DATE)  AS period_start,
                CAST(value AS DOUBLE)       AS value,
                report_period,
                CAST(report_date AS DATE)   AS report_date,
                table_title,
                row_number() OVER (
                    PARTITION BY
                        CASE WHEN section IS NULL OR section = ''
                             THEN item ELSE section || ' - ' || item END,
                        period
                    ORDER BY report_date DESC
                ) AS rn
            FROM "{download_id}"
            WHERE value IS NOT NULL AND period IS NOT NULL
        )
        SELECT series, section, item, period, frequency, period_start,
               value, report_period, report_date, table_title
        FROM ranked
        WHERE rn = 1
        ORDER BY series, period_start
    '''


TRANSFORM_SPECS = [
    SqlNodeSpec(id=f"{s.id}-transform", deps=[s.id], sql=_transform_sql(s.id))
    for s in DOWNLOAD_SPECS
]
