"""Parsing helpers for the NSIDC Sea Ice Index (G02135) connector.

The flat CSVs (daily/monthly extent, daily climatology) are simple; the
regional products ship as multi-sheet, pivoted Excel workbooks (one sheet per
sub-region x measure), so the bulk of this module is reshaping those wide
year/day layouts into long-format rows the parquet writers and SQL transforms
can read directly. Hemisphere is normalised to Arctic/Antarctic across every
table (the monthly-extent CSV encodes it as the region code N/S).
"""
from __future__ import annotations

import io
from datetime import date

import openpyxl

SENTINEL = -9999.0

HEMISPHERE = {"N": "Arctic", "S": "Antarctic"}

MONTH_NUM = {
    "January": 1, "February": 2, "March": 3, "April": 4, "May": 5, "June": 6,
    "July": 7, "August": 8, "September": 9, "October": 10, "November": 11,
    "December": 12,
}


def to_float(value) -> float | None:
    """Parse a numeric cell/token, mapping blanks and the -9999 sentinel to None."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
        value = float(value)
    value = float(value)
    if value == SENTINEL:
        return None
    return value


def parse_daily_extent(text: str, hemisphere: str) -> list[dict]:
    """Daily extent CSV: 2-line header (names + units), then one row per date.

    Columns: Year, Month, Day, Extent, Missing, Source Data (dropped).
    """
    rows: list[dict] = []
    for line in text.splitlines()[2:]:
        if not line.strip():
            continue
        parts = line.split(",")
        if len(parts) < 5 or not parts[0].strip().isdigit():
            continue
        year = int(parts[0].strip())
        month = int(parts[1].strip())
        day = int(parts[2].strip())
        rows.append({
            "hemisphere": hemisphere,
            "date": date(year, month, day),
            "year": year,
            "month": month,
            "day": day,
            "extent_million_sq_km": to_float(parts[3]),
            "missing_million_sq_km": to_float(parts[4]),
        })
    return rows


def parse_climatology(text: str, hemisphere: str) -> list[dict]:
    """Daily climatology CSV: line 0 = 'std Years = ...', line 1 = header, then rows.

    Columns: DOY, Average Extent, Std Deviation, 10th, 25th, 50th, 75th, 90th.
    """
    rows: list[dict] = []
    for line in text.splitlines()[2:]:
        if not line.strip():
            continue
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 8 or not parts[0].isdigit():
            continue
        rows.append({
            "hemisphere": hemisphere,
            "day_of_year": int(parts[0]),
            "average_extent_million_sq_km": to_float(parts[1]),
            "std_deviation_million_sq_km": to_float(parts[2]),
            "pct_10": to_float(parts[3]),
            "pct_25": to_float(parts[4]),
            "pct_50": to_float(parts[5]),
            "pct_75": to_float(parts[6]),
            "pct_90": to_float(parts[7]),
        })
    return rows


def parse_monthly_extent(text: str) -> list[dict]:
    """Monthly extent CSV: single header, one row per year for a fixed month.

    Columns: year, mo, source_dataset, region (N/S), extent, area.
    """
    rows: list[dict] = []
    for line in text.splitlines()[1:]:
        if not line.strip():
            continue
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 6 or not parts[0].isdigit():
            continue
        year = int(parts[0])
        month = int(parts[1])
        rows.append({
            "hemisphere": HEMISPHERE.get(parts[3], parts[3]),
            "year": year,
            "month": month,
            "date": date(year, month, 1),
            "source_dataset": parts[2],
            "extent_million_sq_km": to_float(parts[4]),
            "area_million_sq_km": to_float(parts[5]),
        })
    return rows


def _split_sheet_name(sheet: str) -> tuple[str, str] | None:
    """'Central-Arctic-Extent-km^2' -> ('Central-Arctic', 'extent'); None to skip."""
    base = sheet
    if base.endswith("-km^2"):
        base = base[: -len("-km^2")]
    if "-" not in base:
        return None
    region, measure = base.rsplit("-", 1)
    measure = measure.lower()
    if measure not in {"area", "extent"}:
        return None
    return region, measure


def parse_regional_daily(content: bytes, hemisphere: str) -> list[dict]:
    """Regional daily workbook: one sheet per region x measure.

    Sheet layout: header row ('month', 'day', 1978, 1979, ...); data rows carry
    a month name only on the first day of each month, a day number, then one
    value column per year. Reshaped to one row per (region, date).
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    values: dict[str, dict[tuple, float]] = {"area": {}, "extent": {}}
    try:
        for sheet in wb.sheetnames:
            split = _split_sheet_name(sheet)
            if split is None:
                continue
            region, measure = split
            it = wb[sheet].iter_rows(values_only=True)
            years = next(it)[2:]
            cur_month = None
            for row in it:
                mname = row[0]
                if mname:
                    cur_month = MONTH_NUM.get(str(mname).strip())
                day = row[1]
                if day is None or cur_month is None:
                    continue
                for offset, yr in enumerate(years):
                    if yr is None:
                        continue
                    val = to_float(row[2 + offset])
                    if val is None:
                        continue
                    values[measure][(region, int(yr), cur_month, int(day))] = val
    finally:
        wb.close()

    rows: list[dict] = []
    for region, yr, mo, day in set(values["area"]) | set(values["extent"]):
        try:
            dt = date(yr, mo, day)
        except ValueError:
            continue  # e.g. Feb 29 on a non-leap year — no real observation
        rows.append({
            "hemisphere": hemisphere,
            "region": region,
            "date": dt,
            "year": yr,
            "month": mo,
            "day": day,
            "area_sq_km": values["area"].get((region, yr, mo, day)),
            "extent_sq_km": values["extent"].get((region, yr, mo, day)),
        })
    return rows


def parse_regional_monthly(content: bytes, hemisphere: str) -> list[dict]:
    """Regional monthly workbook: one sheet per region x measure.

    Sheet layout: row 0 = month names (each spanning value + rank columns),
    row 1 = 'area'/'extent' + 'rank' labels, row 2 blank, then one row per year.
    Reshaped to one row per (region, year, month).
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    val: dict[str, dict[tuple, float]] = {"area": {}, "extent": {}}
    rank: dict[str, dict[tuple, float]] = {"area": {}, "extent": {}}
    try:
        for sheet in wb.sheetnames:
            split = _split_sheet_name(sheet)
            if split is None:
                continue
            region, measure = split
            rows_all = list(wb[sheet].iter_rows(values_only=True))
            if len(rows_all) < 4:
                continue
            col_month = {
                ci: MONTH_NUM[str(cell).strip()]
                for ci, cell in enumerate(rows_all[0])
                if cell and str(cell).strip() in MONTH_NUM
            }
            for row in rows_all[3:]:
                if row[0] is None:
                    continue
                yr = int(row[0])
                for ci, mo in col_month.items():
                    v = to_float(row[ci]) if ci < len(row) else None
                    if v is None:
                        continue
                    r = to_float(row[ci + 1]) if ci + 1 < len(row) else None
                    val[measure][(region, yr, mo)] = v
                    rank[measure][(region, yr, mo)] = r
    finally:
        wb.close()

    rows: list[dict] = []
    for region, yr, mo in set(val["area"]) | set(val["extent"]):
        rows.append({
            "hemisphere": hemisphere,
            "region": region,
            "year": yr,
            "month": mo,
            "date": date(yr, mo, 1),
            "area_sq_km": val["area"].get((region, yr, mo)),
            "area_rank": rank["area"].get((region, yr, mo)),
            "extent_sq_km": val["extent"].get((region, yr, mo)),
            "extent_rank": rank["extent"].get((region, yr, mo)),
        })
    return rows
