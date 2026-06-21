"""Shared workbook download + parsing helpers for the Atlanta Fed GDPNow connector.

Both publishable datasets (track record, forecast evolution) derive from the
same ~11MB GDPNow Excel workbook. This module owns the single HTTP fetch and the
generic cell/sheet parsing helpers shared across the subset files. It defines no
NodeSpecs — those live in the per-subset modules. The `_` prefix keeps this file
invisible to load_nodes()'s spec scanner.
"""
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

from subsets_utils import get, transient_retry


GDPNOW_URL = (
    "https://www.atlantafed.org/-/media/Project/Atlanta/FRBA/Documents/"
    "cqer/researchcq/gdpnow/GDPTrackingModelDataAndForecasts.xlsx"
)


# --- HTTP -------------------------------------------------------------------


@transient_retry()
def download_workbook() -> bytes:
    """Fetch the GDPNow workbook bytes. The /-/media/... document URL serves
    directly to an ordinary client (the HTML pages are bot-protected, the file
    is not)."""
    resp = get(GDPNOW_URL, timeout=(10.0, 180.0))
    resp.raise_for_status()
    content = resp.content
    if len(content) < 100_000:
        # The real workbook is ~11MB; anything tiny is an error/interstitial page.
        raise AssertionError(
            f"GDPNow workbook download too small ({len(content)} bytes) — "
            "expected a multi-MB xlsx"
        )
    return content


# --- helpers ----------------------------------------------------------------

def date_str(value) -> str | None:
    """Excel cell -> 'YYYY-MM-DD', or None if not a real date."""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return None


def num(value) -> float | None:
    """Excel cell -> float, or None if missing / non-numeric (skips text
    summary rows that sit among the data)."""
    if value is None or isinstance(value, (str, datetime, date)):
        return None
    try:
        f = float(value)
    except (ValueError, TypeError):
        return None
    if f != f:  # NaN
        return None
    return f


def read_sheet(content: bytes, sheet: str):
    """Yield each data row of a sheet as a {header: value} dict. Headers are
    the first row (stripped); the first occurrence of a header name wins."""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return
        cols = {}
        for idx, h in enumerate(header_row):
            if h is None:
                continue
            name = str(h).strip()
            cols.setdefault(name, idx)
        for row in rows:
            yield {name: (row[idx] if idx < len(row) else None)
                   for name, idx in cols.items()}
    finally:
        wb.close()
