import io
from subsets_utils import get
import openpyxl

NA_URL = "https://usa.visa.com/content/dam/VCOM/regional/na/us/partner-with-us/economic-insights/documents/vbei-visa-north-america-smi-data-appendix.xlsx"
GL_URL = "https://usa.visa.com/content/dam/VCOM/regional/na/us/partner-with-us/economic-insights/documents/vbei-visa-global-smi-data-appendix.xlsx"

def load(url):
    r = get(url, timeout=(10,120)); r.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)

# --- North America: discover blocks from row 1 titles ---
wb = load(NA_URL)
ws = wb[wb.sheetnames[0]]
rows = [list(r) for r in ws.iter_rows(values_only=True)]
print("NA total rows:", len(rows), "cols:", len(rows[0]))
title_row = rows[0]
blocks=[]
for ci,v in enumerate(title_row):
    if isinstance(v,str) and "(" in v and ")" in v and ci>=2:
        blocks.append((ci, v.strip()))
print("NA blocks (col, title):", blocks)
# subheader row 2
print("NA subheader r2:", [ (ci, rows[1][ci]) for ci,_ in blocks ])
# count data rows for first block
b0=blocks[0][0]
data=[r for r in rows[2:] if isinstance(r[b0], __import__("datetime").datetime)]
print("NA data rows for block0:", len(data), "first:", data[0][b0], "last:", data[-1][b0])
print("NA sample row block0:", rows[2][b0], rows[2][b0+1], rows[2][b0+2])

# --- Global ---
wb2 = load(GL_URL)
ws2 = wb2["Spending Momentum Index"]
g = [list(r) for r in ws2.iter_rows(values_only=True)]
# find header row: cell A == 'Code'
hdr_idx = next(i for i,r in enumerate(g) if r and r[0]=="Code")
print("GL header row idx(0based):", hdr_idx)
hdr = g[hdr_idx]
date_cols = [(ci,v) for ci,v in enumerate(hdr) if hasattr(v,"year")]
print("GL n date cols:", len(date_cols), "span:", date_cols[0][1], "..", date_cols[-1][1])
import datetime as dt
drows=[r for r in g[hdr_idx+1:] if r and isinstance(r[0],str) and r[0] not in ("End of Worksheet",)]
print("GL data rows:", len(drows))
for r in drows[:3]:
    print("  ", r[0], r[1], r[2], "| first val:", r[date_cols[0][0]])
print("GL distinct segs:", sorted(set(r[2] for r in drows)))
