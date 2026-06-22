"""Visa Spending Momentum Index (SMI) connector.

The public surface is two fixed-URL XLSX "data appendix" workbooks published by
Visa Business and Economic Insights, refreshed monthly in place:

  * North America — US + Canada, Headline / Discretionary / Non-Discretionary,
    each split into Seasonally-Adjusted and Non-Seasonally-Adjusted, monthly
    from Jan-2014. Laid out as side-by-side 3-column blocks (Date | SA | NSA).
  * Global — six economies (Australia, India, Japan, Brazil, Eurozone, United
    Kingdom) x All-segments / Discretionary / Non-Discretionary, single value
    per series (no SA/NSA split), monthly from Apr-2024. A wide matrix with
    months as columns.

Both files are tiny (<70KB) full-history snapshots with no incremental filter,
so the shape is stateless full re-pull: fetch the whole workbook each run,
reshape to long format, overwrite. The two workbooks have genuinely different
schemas (NA carries a seasonal_adjustment dimension and a "Headline" segment;
Global carries an ISO country code and no SA/NSA split) so they are two
independent download assets feeding two published tables.
"""

import io
import datetime as dt

import pyarrow as pa

from subsets_utils import (
    NodeSpec,
    SqlNodeSpec,
    get,
    transient_retry,
    save_raw_parquet,
)

NA_URL = (
    "https://usa.visa.com/content/dam/VCOM/regional/na/us/partner-with-us/"
    "economic-insights/documents/vbei-visa-north-america-smi-data-appendix.xlsx"
)
GLOBAL_URL = (
    "https://usa.visa.com/content/dam/VCOM/regional/na/us/partner-with-us/"
    "economic-insights/documents/vbei-visa-global-smi-data-appendix.xlsx"
)

# Geography label -> canonical name, for the parenthetical in NA block titles
# (e.g. "Headline (US)" -> segment "Headline", geography "United States").
_NA_GEO = {"US": "United States", "USA": "United States", "Canada": "Canada"}

NA_SCHEMA = pa.schema([
    ("date", pa.date32()),
    ("geography", pa.string()),
    ("spending_segment", pa.string()),
    ("seasonal_adjustment", pa.string()),
    ("index_value", pa.float64()),
])

GLOBAL_SCHEMA = pa.schema([
    ("date", pa.date32()),
    ("country_code", pa.string()),
    ("country", pa.string()),
    ("spending_segment", pa.string()),
    ("index_value", pa.float64()),
])


@transient_retry()
def _fetch_workbook(url):
    import openpyxl

    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)


def _as_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def _as_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def fetch_north_america(node_id: str) -> None:
    """North America SMI: parse the side-by-side metric blocks into long rows."""
    asset = node_id
    wb = _fetch_workbook(NA_URL)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    title_row = rows[0]
    # A block starts at every column whose title-row cell is "<segment> (<geo>)".
    blocks = []  # (date_col, segment, geography)
    for ci, cell in enumerate(title_row):
        if ci < 2 or not isinstance(cell, str) or "(" not in cell or ")" not in cell:
            continue
        segment, _, geo_raw = cell.partition("(")
        segment = segment.strip()
        geo_key = geo_raw.strip().rstrip(")").strip()
        geography = _NA_GEO.get(geo_key, geo_key)
        blocks.append((ci, segment, geography))
    if not blocks:
        raise AssertionError(f"{asset}: no metric blocks found in NA workbook title row")

    out = []
    for date_col, segment, geography in blocks:
        # Block layout: Date | Seasonally adjusted | Non-Seasonally adjusted.
        # Each block carries its own Date column, so walk them independently.
        for r in rows[2:]:
            d = _as_date(r[date_col])
            if d is None:
                continue
            for offset, adj in ((1, "Seasonally adjusted"), (2, "Non-seasonally adjusted")):
                val = _as_float(r[date_col + offset]) if date_col + offset < len(r) else None
                if val is None:
                    continue
                out.append({
                    "date": d,
                    "geography": geography,
                    "spending_segment": segment,
                    "seasonal_adjustment": adj,
                    "index_value": val,
                })

    if not out:
        raise AssertionError(f"{asset}: parsed 0 rows from NA workbook")
    table = pa.Table.from_pylist(out, schema=NA_SCHEMA)
    save_raw_parquet(table, asset)


def fetch_global(node_id: str) -> None:
    """Global SMI: melt the wide months-as-columns matrix into long rows."""
    asset = node_id
    wb = _fetch_workbook(GLOBAL_URL)
    ws = wb["Spending Momentum Index"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    # The tabular block has its header at the first row whose A cell is "Code".
    hdr_idx = next((i for i, r in enumerate(rows) if r and r[0] == "Code"), None)
    if hdr_idx is None:
        raise AssertionError(f"{asset}: could not find 'Code' header row in Global workbook")
    header = rows[hdr_idx]
    date_cols = [(ci, _as_date(v)) for ci, v in enumerate(header) if _as_date(v) is not None]
    if not date_cols:
        raise AssertionError(f"{asset}: no date columns in Global header row")

    out = []
    for r in rows[hdr_idx + 1:]:
        code = r[0] if r else None
        if not isinstance(code, str) or code.strip() == "" or code == "End of Worksheet":
            continue
        country = r[1]
        segment = r[2]
        for ci, d in date_cols:
            val = _as_float(r[ci]) if ci < len(r) else None
            if val is None:
                continue
            out.append({
                "date": d,
                "country_code": code,
                "country": country,
                "spending_segment": segment,
                "index_value": val,
            })

    if not out:
        raise AssertionError(f"{asset}: parsed 0 rows from Global workbook")
    table = pa.Table.from_pylist(out, schema=GLOBAL_SCHEMA)
    save_raw_parquet(table, asset)


DOWNLOAD_SPECS = [
    NodeSpec(id="visa-north-america-smi", fn=fetch_north_america, kind="download"),
    NodeSpec(id="visa-global-smi", fn=fetch_global, kind="download"),
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="visa-north-america-smi-transform",
        deps=["visa-north-america-smi"],
        sql='''
            SELECT
                CAST(date AS DATE)            AS date,
                geography,
                spending_segment,
                seasonal_adjustment,
                CAST(index_value AS DOUBLE)   AS index_value
            FROM "visa-north-america-smi"
            WHERE index_value IS NOT NULL
        ''',
    ),
    SqlNodeSpec(
        id="visa-global-smi-transform",
        deps=["visa-global-smi"],
        sql='''
            SELECT
                CAST(date AS DATE)            AS date,
                country_code,
                country,
                spending_segment,
                CAST(index_value AS DOUBLE)   AS index_value
            FROM "visa-global-smi"
            WHERE index_value IS NOT NULL
        ''',
    ),
]
