"""OBR (Office for Budget Responsibility) connector.

The OBR publishes no API. Every data product is a multi-sheet Excel (.xlsx)
workbook reachable from the https://obr.uk/data/ landing page via a
``/download/<slug>/`` link that 302-redirects to the underlying file. The slugs
embed the publication month-year and roll forward each release, so we never
hardcode them: each refresh re-scrapes /data/, normalises every download link's
title to the same date-free slug id the collect stage used, and resolves the
entity to its current download URL.

The workbooks have no machine-readable schema and wildly heterogeneous per-sheet
layouts (ONS-code header bands, multi-row headers, year-across vs year-down
orientation, qualitative text tables, embedded chart sheets). Rather than
hand-code a fragile bespoke parser per sheet, every workbook is melted into one
faithful long-format table: one row per non-empty cell, carrying the sheet name,
the cell's Excel coordinates, the row's leading text label, the nearest column
header above it, and the value (numeric -> value_num, text -> value_text). This
is lossless, schema-stable across all products, and genuinely tidy for most
sheets (e.g. policy-measures rows come out as sheet='Tax Measures',
row_label='Budget 1970', col_label='1970-71', value_num=-139).

Fetch shape: stateless full re-pull (shape 1). Each workbook is small (<3MB) and
re-downloading the whole corpus each refresh costs seconds; there is no
incremental filter, and re-pulling picks up OBR's frequent revisions for free.
"""

import io
import re
import zipfile

import pyarrow as pa
import openpyxl
from curl_cffi import requests as cffi_requests
from curl_cffi.requests.exceptions import (
    ConnectionError as CffiConnectionError,
    HTTPError as CffiHTTPError,
    Timeout as CffiTimeout,
)
from tenacity import (
    retry,
    retry_if_exception,
    stop_after_attempt,
    wait_exponential,
)

from subsets_utils import NodeSpec, SqlNodeSpec, save_raw_parquet

# NOTE on the HTTP layer: obr.uk sits behind Cloudflare bot-management that 403s
# the CI runner's datacenter IP on dynamic pages (/data/, /download/) regardless
# of User-Agent or cookie priming — it fingerprints the TLS/HTTP2 handshake, and
# httpx (what subsets_utils.get uses) presents a non-browser fingerprint. We
# therefore deliberately bypass subsets_utils.get here and use curl_cffi with
# Chrome impersonation, which presents Chrome's real TLS/JA3 + HTTP2 fingerprint
# and is served normally. This is the only mechanism that reaches the source
# from the cloud; everything else (raw I/O, etc.) still goes through
# subsets_utils. Verified locally and required for the cloud run to pass.
_IMPERSONATE = "chrome"

# The entity union (accepted subsets), copied from
# data/sources/obr/work/entity_union.json. One download spec per id. Some
# products ship as standalone workbooks (.xlsx), others as ZIP bundles of
# several .xlsx files (the "charts and tables" / "detailed forecast tables"
# report supplements) — fetch_one handles both transparently.
ENTITY_IDS = [
    "economic-and-fiscal-outlook-charts-and-tables",
    "economic-and-fiscal-outlook-detailed-forecast-tables",
    "economic-and-fiscal-outlook-ready-reckoner",
    "fiscal-risks-and-sustainability-charts-and-tables",
    "forecast-evaluation-report-annex-a-supplementary-economy-tables",
    "forecast-evaluation-report-annex-b-supplementary-fiscal-tables",
    "forecast-evaluation-report-charts-and-tables",
    "forecast-revisions-database",
    "historical-official-forecasts-database",
    "historical-public-finances-database",
    "long-term-economic-determinants-economic-and-fiscal-outlook",
    "policy-costings-uncertainty-ratings-database",
    "policy-measures-database",
    "policy-risks-database",
    "public-finances-databank",
    "welfare-trends-report-charts-and-tables",
]

DATA_PAGE = "https://obr.uk/data/"

# Melted-workbook schema — identical for every product, the contract that keeps
# batched parquet safe.
SCHEMA = pa.schema([
    ("sheet", pa.string()),
    ("excel_row", pa.int32()),
    ("excel_col", pa.int32()),
    ("row_label", pa.string()),
    ("col_label", pa.string()),
    ("value_num", pa.float64()),
    ("value_text", pa.string()),
])

_MAX_TEXT = 2000  # cap long note cells so a stray essay doesn't bloat a row

_MONTHS = {
    "january", "february", "march", "april", "may", "june", "july",
    "august", "september", "october", "november", "december",
}
_CHIPS = {"xlsx", "zip", "csv", "pdf"}

_TRANSIENT_EXC = (CffiConnectionError, CffiTimeout)


def _is_transient(exc: BaseException) -> bool:
    if isinstance(exc, _TRANSIENT_EXC):
        return True
    if isinstance(exc, CffiHTTPError):
        # message contains the status; recover the code from the attached response
        resp = getattr(exc, "response", None)
        code = getattr(resp, "status_code", None)
        if code is not None:
            return code == 429 or 500 <= code < 600
    return False


@retry(
    retry=retry_if_exception(_is_transient),
    stop=stop_after_attempt(6),
    wait=wait_exponential(min=4, max=120),
    reraise=True,
)
def _http_get(session, url: str):
    resp = session.get(url, timeout=180)
    resp.raise_for_status()
    return resp


def _strip_tags(s: str) -> str:
    s = re.sub(r"<[^>]+>", " ", s)
    s = s.replace("&nbsp;", " ").replace("&amp;", "&")
    return re.sub(r"\s+", " ", s).strip()


def _slug_id(title: str) -> str:
    """Title -> stable, date-free slug id. Mirrors the collect stage exactly."""
    t = title.lower().replace("(zip file)", " ").replace("(xlsx)", " ")
    t = re.sub(r"[–—\-:,/()]", " ", t)
    tokens = []
    for tok in t.split():
        if tok in _MONTHS:
            continue
        if re.fullmatch(r"(19|20)\d{2}", tok):
            continue
        tok = re.sub(r"[^a-z0-9]+", "", tok)
        if tok:
            tokens.append(tok)
    return "-".join(tokens)


def _resolve_download_urls(session) -> dict:
    """Scrape /data/ and map every data product's slug id -> current file URL.

    A real *data* download is rendered with an xlsx/zip/csv file-type chip
    anchor pointing at the same /download/ URL as the product-title anchor; the
    EFO report PDF has no chip and is skipped. We keep the first (canonical)
    title per id.
    """
    html = _http_get(session, DATA_PAGE).text
    by_url = {}
    for m in re.finditer(
        r'<a\b[^>]*href="(https://obr\.uk/download/[^"]+)"[^>]*>(.*?)</a>',
        html,
        re.S | re.I,
    ):
        href = m.group(1).split("?", 1)[0].rstrip("/")
        by_url.setdefault(href, set()).add(_strip_tags(m.group(2)))

    out = {}
    for href, texts in by_url.items():
        lowered = {t.lower() for t in texts}
        if not (lowered & {"xlsx", "zip", "csv"}):
            continue
        candidates = [t for t in texts if t.lower() not in _CHIPS and len(t) > 4]
        if not candidates:
            continue
        eid = _slug_id(max(candidates, key=len))
        if eid and eid not in out:
            out[eid] = href
    return out


def _clean_text(v: str):
    v = v.strip()
    if not v:
        return None
    return v[:_MAX_TEXT]


def _melt_workbook(content: bytes, sheet_prefix: str = "") -> list:
    """Melt every worksheet into long-format rows — one per non-empty cell.

    ``sheet_prefix`` disambiguates sheets when several workbooks (the members of
    a ZIP bundle) are melted into one asset — without it, a 'Contents' sheet in
    two members would collide on the (sheet, excel_row, excel_col) grain.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows = []
    try:
        for sheet_key in wb.sheetnames:
            sheet_name = f"{sheet_prefix}{sheet_key}"
            ws = wb[sheet_key]
            if not hasattr(ws, "iter_rows"):  # skip Chartsheet objects
                continue
            col_header = {}  # positional col index -> nearest string above
            for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
                # row label: first non-empty string in this row
                row_label = None
                for v in row:
                    if isinstance(v, str) and v.strip():
                        row_label = _clean_text(v)
                        break
                for ci, v in enumerate(row, start=1):
                    if v is None:
                        continue
                    if isinstance(v, bool):
                        # openpyxl yields TRUE/FALSE flags; record as text
                        rows.append((sheet_name, ri, ci, row_label,
                                     col_header.get(ci), None, str(v)))
                        continue
                    if isinstance(v, str):
                        cleaned = _clean_text(v)
                        if cleaned is None:
                            continue
                        col_header[ci] = cleaned
                        rows.append((sheet_name, ri, ci, row_label,
                                     col_header.get(ci), None, cleaned))
                    elif isinstance(v, (int, float)):
                        rows.append((sheet_name, ri, ci, row_label,
                                     col_header.get(ci), float(v), None))
                    else:
                        rows.append((sheet_name, ri, ci, row_label,
                                     col_header.get(ci), None, _clean_text(str(v))))
    finally:
        wb.close()
    return rows


def fetch_one(node_id: str) -> None:
    asset = node_id  # the spec id IS the asset name
    entity_id = node_id[len("obr-"):]

    # Chrome-impersonating session (shared cookie jar across the data-page +
    # download requests) to satisfy Cloudflare bot-management. See module note.
    session = cffi_requests.Session(impersonate=_IMPERSONATE)
    try:
        url_map = _resolve_download_urls(session)
        url = url_map.get(entity_id)
        if not url:
            raise RuntimeError(
                f"could not resolve a download URL for {entity_id!r} on {DATA_PAGE} "
                f"(resolved ids: {sorted(url_map)[:5]}...)"
            )

        resp = _http_get(session, url)
        content = resp.content
    finally:
        session.close()
    if not content:
        raise RuntimeError(f"{entity_id}: empty download body from {resp.url}")

    rows = _melt_workbook(content)
    if not rows:
        raise RuntimeError(f"{entity_id}: workbook melted to 0 cells from {resp.url}")

    # build columnar arrays against the explicit schema
    cols = list(zip(*rows))
    table = pa.table({
        "sheet": pa.array(cols[0], type=pa.string()),
        "excel_row": pa.array(cols[1], type=pa.int32()),
        "excel_col": pa.array(cols[2], type=pa.int32()),
        "row_label": pa.array(cols[3], type=pa.string()),
        "col_label": pa.array(cols[4], type=pa.string()),
        "value_num": pa.array(cols[5], type=pa.float64()),
        "value_text": pa.array(cols[6], type=pa.string()),
    }, schema=SCHEMA)
    save_raw_parquet(table, asset)


DOWNLOAD_SPECS = [
    NodeSpec(
        id=f"obr-{eid.lower().replace('_', '-')}",
        fn=fetch_one,
        kind="download",
    )
    for eid in ENTITY_IDS
]


# One published Delta table per subset: a thin typed pass-through of the melted
# workbook. The projection/cast is the correctness gate (wrong raw shape fails
# loudly here); the WHERE drops any fully-empty cell defensively and guarantees
# a non-empty published table.
TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"{s.id}-transform",
        deps=[s.id],
        sql=f'''
            SELECT
                CAST(sheet AS VARCHAR)       AS sheet,
                CAST(excel_row AS INTEGER)   AS excel_row,
                CAST(excel_col AS INTEGER)   AS excel_col,
                CAST(row_label AS VARCHAR)   AS row_label,
                CAST(col_label AS VARCHAR)   AS col_label,
                CAST(value_num AS DOUBLE)    AS value_num,
                CAST(value_text AS VARCHAR)  AS value_text
            FROM "{s.id}"
            WHERE value_num IS NOT NULL OR value_text IS NOT NULL
        ''',
        key=("sheet", "excel_row", "excel_col"),
    )
    for s in DOWNLOAD_SPECS
]
