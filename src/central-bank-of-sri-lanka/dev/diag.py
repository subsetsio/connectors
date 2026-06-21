import io,re,sys
from subsets_utils import get
from openpyxl import load_workbook
sys.path.insert(0,"src/nodes")
import central_bank_of_sri_lanka as M
problem=["table-1.24","table-3.13","table-3.14","table-3.16","table-3.31","table-6.8","table-6.9",
         "table-1.22","table-1.37","table-1.40","table-1.41","table-1.42","table-1.43",
         "table-2.2","table-2.3","table-2.6","table-2.10","table-2.12"]
for e in problem:
    try:
        url=M._resolve_xlsx_url(e)
        wb=load_workbook(io.BytesIO(M._get_bytes(url)),read_only=True,data_only=True)
        rows=[]
        for ws in wb.worksheets: rows.extend(M._parse_sheet(list(ws.iter_rows(values_only=True))))
        total=len(rows)
        passf=sum(1 for r in rows if r["value"] is not None and r["row_label"] and r["row_label"].strip())
        nv=sum(1 for r in rows if r["value"] is not None)
        empties=sum(1 for r in rows if not (r["row_label"] and r["row_label"].strip()))
        print(f"{e:14} total={total:4} value_nonnull={nv:4} empty_rowlabel={empties:4} PASS_FILTER={passf:4}")
    except Exception as ex:
        print(f"{e:14} ERROR {type(ex).__name__}: {ex}")
