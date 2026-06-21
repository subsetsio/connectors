import io,re,sys
from subsets_utils import get
from openpyxl import load_workbook
sys.path.insert(0,"dev"); from parser import parse_sheet
BASE="https://www.cbsl.gov.lk/en/statistics/economic-and-social-statistics"
def token(e): return f"kei_table{e.rsplit('-',1)[-1]}" if e.startswith("kei-table-") else "table"+e[len("table-"):]
def page(e): return "kei" if e.startswith("kei-") else "chapter-"+e[len("table-"):].split(".")[0]
def load(e):
    h=get(f"{BASE}/{page(e)}",timeout=60).text
    u=re.search(r'href="([^"]*?/sheets/ess_\d{4}_'+re.escape(token(e))+r'_e\.xlsx)"',h).group(1)
    if u.startswith("/"): u="https://www.cbsl.gov.lk"+u
    wb=load_workbook(io.BytesIO(get(u,timeout=90).content),read_only=True,data_only=True)
    return list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
for e in ["table-6.7","table-1.2","table-3.17"]:
    recs,o=parse_sheet(load(e))
    print(f"\n=== {e} [{o}] {len(recs)} ===")
    for r in recs[:3]: print("  ",r)
    nv=sum(1 for r in recs if r["value"] is not None)
    py=sum(1 for r in recs if r["period_year"] is not None)
    print(f"  value_nonnull={nv}/{len(recs)} period_year_set={py}/{len(recs)}")
