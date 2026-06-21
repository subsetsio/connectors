import io,re,sys,json,concurrent.futures as cf
from subsets_utils import get
from openpyxl import load_workbook
sys.path.insert(0,"src/nodes"); import central_bank_of_sri_lanka as M
ids=json.load(open("/Users/nathansnellaert/Documents/hardened/data/sources/central-bank-of-sri-lanka/work/entity_union.json"))
def work(e):
    try:
        url=M._resolve_xlsx_url(e)
        wb=load_workbook(io.BytesIO(M._get_bytes(url)),read_only=True,data_only=True)
        rows=[]
        for ws in wb.worksheets: rows.extend(M._parse_sheet(list(ws.iter_rows(values_only=True))))
        passf=sum(1 for r in rows if r["value"] is not None and r["row_label"] and r["row_label"].strip())
        return (e,len(rows),passf,None)
    except Exception as ex:
        return (e,0,0,f"{type(ex).__name__}: {ex}")
bad=[]; tot=0
with cf.ThreadPoolExecutor(max_workers=12) as ex:
    for e,n,pf,err in ex.map(work,ids):
        tot+=pf
        if err or pf<1: bad.append((e,n,pf,err))
print("total published rows:",tot)
print("FAIL (pass_filter<1 or error):",len(bad))
for b in bad: print("  ",b)
