import io,sys
from subsets_utils import get
from openpyxl import load_workbook
sys.path.insert(0,"src/nodes"); import central_bank_of_sri_lanka as M
for e in ["table-3.13","table-1.24","table-6.8","table-3.1","table-1.1"]:
    url=M._resolve_xlsx_url(e)
    wb=load_workbook(io.BytesIO(M._get_bytes(url)),read_only=True,data_only=True)
    rows=[]
    for ws in wb.worksheets: rows.extend(M._parse_sheet(list(ws.iter_rows(values_only=True))))
    print(f"=== {e}: {len(rows)} ===")
    for r in [x for x in rows if x['value'] is not None][:2]: print("  ",r)
