import io, re, json, sys
from subsets_utils import get
from openpyxl import load_workbook
sys.path.insert(0, "dev")
from parser import parse_sheet

union = json.load(open("/Users/nathansnellaert/Documents/hardened/data/sources/central-bank-of-sri-lanka/work/entity_union.json"))
ids = union if isinstance(union, list) else list(union)
print("entities:", len(ids))

def token(eid):
    if eid.startswith("kei-table-"):
        return f"kei_table{eid.rsplit('-',1)[-1]}"
    return "table" + eid[len("table-"):]

def chapter_page(eid):
    if eid.startswith("kei-"):
        return "kei"
    return "chapter-" + eid[len("table-"):].split(".")[0]

BASE="https://www.cbsl.gov.lk/en/statistics/economic-and-social-statistics"
from collections import Counter
orient=Counter(); fails=[]; small=[]; rowcounts=[]
import concurrent.futures as cf

def work(eid):
    try:
        tok=token(eid)
        page=get(f"{BASE}/{chapter_page(eid)}", timeout=60).text
        m=re.search(r'href="([^"]*?/sheets/ess_\d{4}_'+re.escape(tok)+r'_e\.xlsx)"', page)
        if not m:
            return (eid,"NO_LINK",0,None)
        url=m.group(1)
        if url.startswith("/"): url="https://www.cbsl.gov.lk"+url
        content=get(url, timeout=90).content
        wb=load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws=wb[wb.sheetnames[0]]
        rows=list(ws.iter_rows(values_only=True))
        recs,o=parse_sheet(rows)
        return (eid,o,len(recs),None)
    except Exception as e:
        return (eid,"ERR",0,f"{type(e).__name__}: {e}")

with cf.ThreadPoolExecutor(max_workers=12) as ex:
    for eid,o,n,err in ex.map(work, ids):
        orient[o]+=1; rowcounts.append((eid,n))
        if o in ("ERR","NO_LINK","none","empty") or n==0:
            fails.append((eid,o,n,err))
        elif n<5:
            small.append((eid,o,n))

print("orientations:", dict(orient))
print("total rows:", sum(n for _,n in rowcounts))
print("FAILS (%d):"%len(fails))
for f in fails: print("  ",f)
print("SMALL <5 rows (%d):"%len(small))
for s in small[:30]: print("  ",s)
