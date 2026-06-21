import io,sys
from subsets_utils import get
from openpyxl import load_workbook
sys.path.insert(0,"src/nodes"); import central_bank_of_sri_lanka as M
def dump(e,n=12):
    url=M._resolve_xlsx_url(e)
    wb=load_workbook(io.BytesIO(M._get_bytes(url)),read_only=True,data_only=True)
    ws=wb.worksheets[0]
    print("="*70,e)
    for i,row in enumerate(ws.iter_rows(values_only=True)):
        if i>=n: break
        c=list(row)
        while c and c[-1] is None: c.pop()
        print(i,[ (x[:16] if isinstance(x,str) else x) for x in c[:11]])
for e in ["table-3.13","table-6.8","table-1.24","table-3.16"]:
    dump(e)
