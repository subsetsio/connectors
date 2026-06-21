import io, re, sys
from subsets_utils import get
from openpyxl import load_workbook
sys.path.insert(0,"dev")
from parser import parse_sheet
BASE="https://www.cbsl.gov.lk/en/statistics/economic-and-social-statistics"
def token(eid):
    return f"kei_table{eid.rsplit('-',1)[-1]}" if eid.startswith("kei-table-") else "table"+eid[len("table-"):]
def page(eid):
    return "kei" if eid.startswith("kei-") else "chapter-"+eid[len("table-"):].split(".")[0]
def load(eid):
    tok=token(eid); html=get(f"{BASE}/{page(eid)}",timeout=60).text
    m=re.search(r'href="([^"]*?/sheets/ess_\d{4}_'+re.escape(tok)+r'_e\.xlsx)"',html)
    url=m.group(1)
    if url.startswith("/"): url="https://www.cbsl.gov.lk"+url
    wb=load_workbook(io.BytesIO(get(url,timeout=90).content),read_only=True,data_only=True)
    return list(wb[wb.sheetnames[0]].iter_rows(values_only=True))

# dump raw 6.7
rows=load("table-6.7")
print("=== 6.7 raw ===")
for i,r in enumerate(rows[:8]):
    c=list(r)
    while c and c[-1] is None:c.pop()
    print(i,[ (x[:16] if isinstance(x,str) else x) for x in c[:13]])

for eid in ["table-1.1","table-3.1","table-4.10","table-8.1","kei-table-1"]:
    recs,o=parse_sheet(load(eid))
    print(f"\n=== {eid} [{o}] {len(recs)} rows ===")
    for rec in recs[:4]: print("  ",rec)
