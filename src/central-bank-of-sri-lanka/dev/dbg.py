import io, re, sys
from subsets_utils import get
from openpyxl import load_workbook
sys.path.insert(0,"dev")
import parser as P
BASE="https://www.cbsl.gov.lk/en/statistics/economic-and-social-statistics"
def token(eid): return f"kei_table{eid.rsplit('-',1)[-1]}" if eid.startswith("kei-table-") else "table"+eid[len("table-"):]
def page(eid): return "kei" if eid.startswith("kei-") else "chapter-"+eid[len("table-"):].split(".")[0]
def load(eid):
    tok=token(eid); html=get(f"{BASE}/{page(eid)}",timeout=60).text
    m=re.search(r'href="([^"]*?/sheets/ess_\d{4}_'+re.escape(tok)+r'_e\.xlsx)"',html)
    url=m.group(1)
    if url.startswith("/"): url="https://www.cbsl.gov.lk"+url
    wb=load_workbook(io.BytesIO(get(url,timeout=90).content),read_only=True,data_only=True)
    return list(wb[wb.sheetnames[0]].iter_rows(values_only=True))

rows=load("table-3.1")
grid=[list(r) for r in rows]
nc=max(len(r) for r in grid)
for r in grid: r+=[None]*(nc-len(r))
P._sanitize(grid)
grid=[r for r in grid if any(c is not None for c in r)]
nrows=len(grid)
def realnum(r): return sum(1 for c in range(nc) if P._num(grid[r][c])[0] is not None and P._year_of(grid[r][c]) is None)
data_rows=[r for r in range(nrows) if realnum(r)>=1]
print("nrows",nrows,"data_rows[:5]",data_rows[:5])
for c in range(min(nc,3)):
    yc=sum(1 for r in data_rows if P._year_of(grid[r][c]) is not None)
    print("col",c,"yearcount_in_datarows",yc,"of",len(data_rows))
for i in range(5):
    print(i, grid[i][:6], "realnum=",realnum(i))
