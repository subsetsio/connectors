import io, re, sys
from subsets_utils import get
from openpyxl import load_workbook
BASE="https://www.cbsl.gov.lk/en/statistics/economic-and-social-statistics"
def token(eid):
    if eid.startswith("kei-table-"): return f"kei_table{eid.rsplit('-',1)[-1]}"
    return "table"+eid[len("table-"):]
def page(eid):
    return "kei" if eid.startswith("kei-") else "chapter-"+eid[len("table-"):].split(".")[0]
def dump(eid, nrows=10):
    tok=token(eid)
    html=get(f"{BASE}/{page(eid)}",timeout=60).text
    m=re.search(r'href="([^"]*?/sheets/ess_\d{4}_'+re.escape(tok)+r'_e\.xlsx)"',html)
    url=m.group(1)
    if url.startswith("/"): url="https://www.cbsl.gov.lk"+url
    wb=load_workbook(io.BytesIO(get(url,timeout=90).content),read_only=True,data_only=True)
    ws=wb[wb.sheetnames[0]]
    print("="*60, eid)
    for i,row in enumerate(ws.iter_rows(values_only=True)):
        if i>=nrows: break
        cells=list(row)
        while cells and cells[-1] is None: cells.pop()
        print(i,[ (c[:18] if isinstance(c,str) else c) for c in cells[:13]])
for eid in ["table-1.2","table-3.1","table-3.17","table-8.1","table-1.12","table-7.6"]:
    dump(eid)
