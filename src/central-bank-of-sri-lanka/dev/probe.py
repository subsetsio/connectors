import io
from subsets_utils import get
from openpyxl import load_workbook

URLS = {
 "table1.1": "https://www.cbsl.gov.lk/sites/default/files/cbslweb_documents/statistics/sheets/ess_2025_table1.1_e.xlsx",
 "table3.1": "https://www.cbsl.gov.lk/sites/default/files/cbslweb_documents/statistics/sheets/ess_2025_table3.1_e.xlsx",
 "table4.10":"https://www.cbsl.gov.lk/sites/default/files/cbslweb_documents/statistics/sheets/ess_2025_table4.10_e.xlsx",
 "kei_table1":"https://www.cbsl.gov.lk/sites/default/files/cbslweb_documents/statistics/sheets/ess_2025_kei_table1_e.xlsx",
}
for name,url in URLS.items():
    r = get(url, timeout=60)
    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    print("="*70)
    print(name, "sheets:", wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    print("dims:", ws.max_row, "x", ws.max_column)
    for i,row in enumerate(ws.iter_rows(values_only=True)):
        if i>=18: break
        # trim trailing Nones
        cells=list(row)
        while cells and cells[-1] is None: cells.pop()
        print(i, [ (c if not isinstance(c,str) else c[:22]) for c in cells[:14] ])
