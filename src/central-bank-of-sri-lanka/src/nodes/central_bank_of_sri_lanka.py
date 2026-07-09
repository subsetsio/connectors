"""Central Bank of Sri Lanka connector.

Source: the "Economic and Social Statistics of Sri Lanka" (ESS) annual online
release — CBSL's comprehensive set of long economic/social time series. ESS is
organised as a Key Economic Indicators (KEI) section plus eight numbered
chapters; every statistical table is published as a standalone Excel workbook
with a stable (per-release-year) URL under
  https://www.cbsl.gov.lk/sites/default/files/cbslweb_documents/statistics/sheets/
with names ess_<year>_table<chapter>.<table>_e.xlsx and ess_<year>_kei_table<n>_e.xlsx.

The ESS chapter landing pages list the per-table workbook links; the connector
resolves each table's current xlsx URL by scraping that page (so the release-year
token is discovered, never hardcoded), downloads the workbook, and "melts" it
into a uniform long format. The workbooks are wide, human-formatted spreadsheets
with header banners, unit rows, grouped multi-row headers, section sub-headers
and footnoted values, in several orientations (items-in-rows by year, years-down
by indicator, and year-less cross-tabs by country/province/district). The melt
in `_parse_sheet` normalises all of these to rows of
  (row_label, col_label, period_year, value, value_text)
so a thin DuckDB SQL transform can publish one clean numeric long table per
subset. DuckDB can only read parquet/ndjson/csv, so all Excel parsing happens
here in Python (openpyxl); raw is written as NDJSON because the per-cell payload
mixes parsed floats with footnote-bearing text.

Fetch shape: stateless full re-pull. Each workbook is small (KBs-to-MB) and the
source revises figures in place across releases, so every run re-fetches the full
workbook and overwrites raw — no watermark/cursor would safely capture in-place
revisions. Coverage is the rank-accepted entity union (182 tables).
"""

import io
import re

import pyarrow as pa
from openpyxl import load_workbook

from subsets_utils import (
    NodeSpec,
    get,
    save_raw_ndjson,
    transient_retry,
)

SLUG = "central-bank-of-sri-lanka"
BASE = "https://www.cbsl.gov.lk"
ESS = f"{BASE}/en/statistics/economic-and-social-statistics"

# Rank-accepted entity union (182 tables). Each id maps 1:1 to one ESS workbook.
from constants import ENTITY_IDS

# ---------------------------------------------------------------------------
# HTTP with bounded retry on transient failures
# ---------------------------------------------------------------------------


@transient_retry()
def _get_text(url: str) -> str:
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.text


@transient_retry()
def _get_bytes(url: str) -> bytes:
    resp = get(url, timeout=(10.0, 180.0))
    resp.raise_for_status()
    return resp.content


# ---------------------------------------------------------------------------
# URL resolution — discover each table's current xlsx link from its ESS page
# ---------------------------------------------------------------------------
def _chapter_page(entity_id: str) -> str:
    if entity_id.startswith("kei-"):
        return "kei"
    return "chapter-" + entity_id[len("table-"):].split(".")[0]


def _file_token(entity_id: str) -> str:
    # ess_<year>_<token>_e.xlsx
    if entity_id.startswith("kei-table-"):
        return "kei_table" + entity_id.rsplit("-", 1)[-1]
    return "table" + entity_id[len("table-"):]


def _resolve_xlsx_url(entity_id: str) -> str:
    page = _get_text(f"{ESS}/{_chapter_page(entity_id)}")
    token = _file_token(entity_id)
    m = re.search(
        r'href="([^"]*?/sheets/ess_\d{4}_' + re.escape(token) + r'_e\.xlsx)"',
        page,
        re.IGNORECASE,
    )
    if not m:
        raise AssertionError(
            f"{entity_id}: no xlsx link for token '{token}' on ESS "
            f"{_chapter_page(entity_id)} page (source layout may have changed)"
        )
    url = m.group(1)
    if url.startswith("/"):
        url = BASE + url
    return url


# ---------------------------------------------------------------------------
# Generic Excel -> long-format melt
#
# Handles every ESS table layout: years-across-columns (A), years-down-a-column
# (B) and year-less cross-tabs. Flattens grouped multi-row headers with
# forward-fill, carries section sub-headers into the row label, and parses
# footnoted / parenthesised / comma-separated numeric cells.
# ---------------------------------------------------------------------------
_YEAR_RE = re.compile(r'^\s*((?:19|20)\d{2})')
_NA = {"", "-", "–", "—", "n.a.", "n.a", "na", "...", "…", "n/a"}
_DECOR = re.compile(r'^\s*(\d{2}\.\s|TABLE\s+\d|KEY ECONOMIC INDICATORS)', re.IGNORECASE)


def _year_of(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, int) and 1900 <= v <= 2100:
        return v
    if isinstance(v, float) and v.is_integer() and 1900 <= v <= 2100:
        return int(v)
    if isinstance(v, str):
        m = _YEAR_RE.match(v.strip())
        if m:
            return int(m.group(1))
    return None


def _num(v):
    """Return (float|None, text|None) for a cell."""
    if v is None or isinstance(v, bool):
        return None, None
    if isinstance(v, (int, float)):
        return float(v), None
    s = str(v).strip()
    if s.lower() in _NA:
        return None, (s or None)
    paren = re.fullmatch(r'\((.*)\)', s)
    inner = paren.group(1) if paren else s
    cleaned = re.sub(r'\([^)]*\)', '', inner).replace(',', '').replace('%', '').strip()
    if re.fullmatch(r'[+-]?\d*\.?\d+', cleaned):
        f = float(cleaned)
        return (-f if paren else f), s
    return None, s


def _is_text(v):
    if v is None or isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return False
    s = str(v).strip()
    if s == "" or s.lower() in _NA:   # NA markers are neutral, not labels
        return False
    return _num(v)[0] is None


def _clean(v):
    return re.sub(r'\s+', ' ', str(v)).strip()


def _sanitize(grid):
    # Blank decorative cells (chapter banner, "TABLE x.y" marker) so forward-fill
    # of grouped headers can't smear a page title across the value columns.
    for row in grid:
        for c in range(len(row)):
            v = row[c]
            if isinstance(v, str) and _DECOR.match(v.strip()):
                row[c] = None
    return grid


def _parse_sheet(rows):
    grid = [list(r) for r in rows]
    if not grid:
        return []
    ncols0 = max((len(r) for r in grid), default=0)
    for r in grid:
        r += [None] * (ncols0 - len(r))
    _sanitize(grid)
    grid = [r for r in grid if any(c is not None for c in r)]
    if not grid:
        return []
    ncols = max(len(r) for r in grid)
    for r in grid:
        r += [None] * (ncols - len(r))
    nrows = len(grid)

    def realnum(r):
        # a genuine measurement is a number that is NOT a year; year-only rows
        # are column headers, not data.
        return sum(1 for c in range(ncols)
                   if _num(grid[r][c])[0] is not None and _year_of(grid[r][c]) is None)

    data_start = next((r for r in range(nrows) if realnum(r) >= 1), None)
    if data_start is None:
        return []

    def col_real(c):  # genuine (non-year) measurements in the data region
        return sum(1 for r in range(data_start, nrows)
                   if _num(grid[r][c])[0] is not None and _year_of(grid[r][c]) is None)

    def col_text(c):  # genuine text labels (NA markers excluded) in the data region
        return sum(1 for r in range(data_start, nrows) if _is_text(grid[r][c]))

    # The value region is the right-hand block of columns dominated by real
    # measurements; everything to its left is the stub (row-label) region. This
    # split is orientation-agnostic and survives merged year columns + secondary
    # category/period columns (e.g. Year | Month | values).
    first_val = next((c for c in range(ncols)
                      if col_real(c) > col_text(c) and col_real(c) > 0), None)
    if first_val is None:
        return []
    stub_cols = [c for c in range(first_val)
                 if any(grid[r][c] is not None for r in range(nrows))]
    value_cols = [c for c in range(first_val, ncols)
                  if any(grid[r][c] is not None for r in range(nrows))]
    if not value_cols:
        return []

    # Detect (and forward-fill) a merged period column: a stub column that, once
    # merged cells are filled downward, is a year for nearly every data row. Many
    # monthly/quarterly tables merge the year cell across its sub-rows.
    period_col = None
    ndata = max(1, nrows - data_start)
    for c in stub_cols:
        filled, last = [], None
        for r in range(data_start, nrows):
            v = grid[r][c]
            if v is not None and str(v).strip() != "":
                last = v
            filled.append(last)
        yc = sum(1 for v in filled if _year_of(v) is not None)
        if yc >= 4 and yc >= 0.8 * ndata:
            period_col = c
            for i, r in enumerate(range(data_start, nrows)):
                if grid[r][c] is None or str(grid[r][c]).strip() == "":
                    grid[r][c] = filled[i]
            break

    # header band: rows above data_start with content in a value column (skips
    # title/unit/section rows whose only content sits in the stub region).
    band = [r for r in range(data_start)
            if any(grid[r][c] is not None for c in value_cols)]
    headers = {}
    if band:
        filled = []
        for r in band:
            ff, last = [], None
            for c in range(ncols):
                if grid[r][c] is not None and str(grid[r][c]).strip() != "":
                    last = _clean(grid[r][c])
                ff.append(last)
            filled.append(ff)
        for c in value_cols:
            parts = []
            for fr in filled:
                p = fr[c]
                if p and (not parts or parts[-1] != p):
                    parts.append(p)
            headers[c] = " | ".join(parts) if parts else f"col_{c}"
    else:
        headers = {c: f"col_{c}" for c in value_cols}

    out = []
    section = None
    for r in range(data_start, nrows):
        stub_parts = [_clean(grid[r][c]) for c in stub_cols
                      if grid[r][c] is not None and str(grid[r][c]).strip() != ""]
        stub = " - ".join(stub_parts)
        recs = []
        for c in value_cols:
            val, txt = _num(grid[r][c])
            if val is None and txt is None:
                continue
            recs.append((c, val, txt))
        # section sub-header row (non-B only): label present, no values
        if period_col is None and stub and not recs:
            section = stub
            continue
        if not stub and not recs:
            continue
        row_label = f"{section} - {stub}" if (section and stub) else (stub or section or "")
        for c, val, txt in recs:
            col_label = headers.get(c, f"col_{c}")
            if period_col is not None:
                py = _year_of(grid[r][period_col])
            else:
                py = _year_of(col_label)
            out.append({
                "row_label": (row_label or "Unlabeled row"),
                "col_label": col_label,
                "period_year": py,
                "value": val,
                "value_text": txt,
            })
    return out


# ---------------------------------------------------------------------------
# Fetch
# ---------------------------------------------------------------------------
def fetch_one(node_id: str) -> None:
    """Resolve, download and melt one ESS statistical workbook to long format."""
    asset = node_id
    entity_id = node_id[len(SLUG) + 1:]
    url = _resolve_xlsx_url(entity_id)
    content = _get_bytes(url)
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows = []
    for ws in wb.worksheets:
        rows.extend(_parse_sheet(list(ws.iter_rows(values_only=True))))
    wb.close()
    if not rows:
        raise AssertionError(f"{entity_id}: parsed 0 data rows from {url}")
    save_raw_ndjson(rows, asset)


# ---------------------------------------------------------------------------
# DOWNLOAD_SPECS — one per entity in the union
# ---------------------------------------------------------------------------
DOWNLOAD_SPECS = [
    NodeSpec(id=f"{SLUG}-{eid}", fn=fetch_one, kind="download")
    for eid in ENTITY_IDS
]
