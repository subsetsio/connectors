"""Monetary Authority of Macao (AMCM) — statistical time-series connector.

Source shape: per-section/topic XLSX workbooks on cdn.amcm.gov.mo, one worksheet
per publishable table. The CDN filenames carry a volatile revision/quarter suffix,
so each fetch FIRST re-resolves the current XLSX URL via the amcm.gov.mo CMS JSON
API (one of three discovery endpoints), THEN downloads the workbook and extracts
the entity's worksheet. Workbooks are human-formatted (merged cells, multi-row
headers, trilingual labels, two orientations), so the worksheet is normalised into
a long table {series, period, year, part, date, value} here in Python — the SQL
transform is a thin type/dedup pass.

Stateless full re-pull: the whole corpus is a few MB and is re-fetched each run;
overwrite picks up revisions for free. No incremental filter exists upstream.
"""
import datetime
import io
import re

import openpyxl
import pyarrow as pa

from subsets_utils import (
    NodeSpec, SqlNodeSpec, get, save_raw_parquet, transient_retry,
)
from constants import ENTITY_META

PREFIX = "monetary-authority-of-macao-"

# CMS discovery endpoints (Api-Language selects English; default is Chinese).
_ENDPOINTS = {
    "mfs": "https://www.amcm.gov.mo/api/v1.0/page?path=research-statistics%2F"
           "statistics-page%2Fmonetary-and-financial-statistics-time-series",
    "fsi": "https://www.amcm.gov.mo/api/v1.0/cms/financial_indicators",
    "insurance": "https://www.amcm.gov.mo/api/v1.0/cms/statistic_insurances",
}
_HEADERS = {"Api-Language": "en"}

SCHEMA = pa.schema([
    ("series", pa.string()),
    ("period", pa.string()),
    ("year", pa.int32()),
    ("part", pa.string()),
    ("date", pa.string()),
    ("value", pa.float64()),
])


# --------------------------------------------------------------------------- #
# URL discovery
# --------------------------------------------------------------------------- #
@transient_retry()
def _fetch_text(url: str) -> str:
    resp = get(url, headers=_HEADERS, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.text


@transient_retry()
def _fetch_bytes(url: str) -> bytes:
    resp = get(url, headers=_HEADERS, timeout=(10.0, 180.0))
    resp.raise_for_status()
    return resp.content


def _stem(filename: str) -> str:
    """Strip extension + the volatile revision/quarter suffix to a stable stem."""
    s = filename[:-5] if filename.lower().endswith(".xlsx") else filename
    s = re.sub(r"_(?:\d{1,2}[Qq]\d{4}|\d{4}[Qq]\d{1,2}|\d{4}[Mm]\d{1,2}|\d{4})", "", s)
    s = re.sub(r"_\d+$", "", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _xlsx_links(text: str) -> list:
    out, seen = [], set()
    for m in re.findall(r'https://cdn\.amcm\.gov\.mo/[^\s"\\<>]+?\.xlsx', text):
        if m not in seen:
            seen.add(m)
            out.append(m)
    return out


def _resolve_url(endpoint: str, stem: str) -> str:
    text = _fetch_text(_ENDPOINTS[endpoint])
    for url in _xlsx_links(text):
        if _stem(url.rsplit("/", 1)[-1]) == stem:
            return url
    raise RuntimeError(f"no XLSX with stem {stem!r} found at endpoint {endpoint!r}")


def _open_sheet(xlsx_bytes: bytes, sheet: str):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    name = sheet if sheet in wb.sheetnames else next(
        (s for s in wb.sheetnames if s.strip() == sheet.strip()), None)
    if name is None:
        wb.close()
        raise RuntimeError(f"worksheet {sheet!r} not in workbook ({wb.sheetnames})")
    grid = list(wb[name].iter_rows(values_only=True))
    wb.close()
    return grid


# --------------------------------------------------------------------------- #
# Generic worksheet -> long-format parser (see module docstring)
# --------------------------------------------------------------------------- #
def _is_year(x):
    try:
        v = float(x)
    except (TypeError, ValueError):
        return False
    return (not isinstance(x, bool)) and v == int(v) and 1900 <= int(v) <= 2100


def _as_num(x):
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, str):
        s = x.strip().replace(",", "")
        if s in ("", "-", "--", "...", "..", "…", "n.a.", "n.a", "N/A", "na", "NA", "—", "–"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _q_of(x):
    if x is None:
        return None
    s = str(x).strip()
    m = re.fullmatch(r'(?i)Q\s*([1-4])', s) or re.fullmatch(r'([1-4])T', s)
    return int(m.group(1)) if m else None


def _month_of(x):
    n = _as_num(x)
    if n is not None and n == int(n) and 1 <= int(n) <= 12:
        return int(n)
    return None


def _qdate(y, q): return datetime.date(y, q * 3, 1).isoformat()
def _mdate(y, m): return datetime.date(y, m, 1).isoformat()
def _ydate(y):    return datetime.date(y, 12, 31).isoformat()


def _english(x):
    raw = str(x)
    lines = [ln.strip() for ln in raw.split("\n") if ln.strip()]
    text = lines[-1] if lines else ""
    if "\n" not in raw and " / " in text:
        text = text.split(" / ")[-1].strip()
    return re.sub(r'\s+', ' ', text).strip()


def _clean(x):
    return re.sub(r'\s+', ' ', str(x)).strip()


def _uniquify(name, used):
    if name not in used:
        used[name] = 1
        return name
    used[name] += 1
    return f"{name} ({used[name]})"


def _period_from_header(x):
    if x is None:
        return None
    if isinstance(x, datetime.datetime):
        x = x.date()
    if isinstance(x, datetime.date):
        return (x.year, f"{x.month:02d}", _mdate(x.year, x.month), f"{x.year}-{x.month:02d}")
    if _is_year(x):
        y = int(float(x)); return (y, None, _ydate(y), str(y))
    s = str(x).strip()
    m = re.fullmatch(r'(\d{1,2})/(\d{4})', s)
    if m:
        mo, y = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12:
            return (y, f"{mo:02d}", _mdate(y, mo), f"{y}-{mo:02d}")
    m = re.match(r'(\d{4})[-/](\d{1,2})(?:[-/]\d{1,2})?', s)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12:
            return (y, f"{mo:02d}", _mdate(y, mo), f"{y}-{mo:02d}")
    return None


def parse_grid(grid):
    grid = [list(r) for r in grid]
    if not grid:
        return []
    ncol = max((len(r) for r in grid), default=0)
    if ncol == 0:
        return []
    for r in grid:
        r.extend([None] * (ncol - len(r)))
    nrow = len(grid)
    col0 = [grid[r][0] for r in range(nrow)]
    year_down = sum(1 for c in col0 if _is_year(c) or isinstance(c, (datetime.date, datetime.datetime)))
    if year_down >= 4:
        return _parse_rows(grid, nrow, ncol)
    return _parse_cols(grid, nrow, ncol)


def _row_headers(grid, nrow, ncol, value_start, first_data):
    leaf_cols = [c for c in range(value_start, ncol)
                 if any(_as_num(grid[r][c]) is not None for r in range(first_data, nrow))]
    hdr_rows = list(range(first_data))
    orig_n = {r: sum(1 for c in leaf_cols if grid[r][c] is not None and _clean(grid[r][c]) != "")
              for r in hdr_rows}
    span_cap = max(1, len(leaf_cols) // 2)
    group_rows = [r for r in hdr_rows if 1 <= orig_n[r] <= span_cap]
    ff = {}
    for r in group_rows:
        last, ff[r] = None, {}
        for c in range(value_start, ncol):
            v = grid[r][c]
            if v is not None and _clean(v) != "":
                last = _clean(v)
            ff[r][c] = last
    headers, used = {}, {}
    for c in range(value_start, ncol):
        parts = []
        for r in group_rows:
            v = ff[r].get(c)
            if v:
                parts.append(_english(v))
        leaf = None
        for r in range(first_data - 1, -1, -1):
            v = grid[r][c]
            if v is not None and _clean(v) != "":
                leaf = _english(v); break
        if leaf:
            parts.append(leaf)
        comp = []
        for p in parts:
            if p and (not comp or comp[-1] != p):
                comp.append(p)
        name = " / ".join(comp) if comp else f"col{c}"
        headers[c] = _uniquify(name, used)
    return headers


def _parse_rows(grid, nrow, ncol):
    cand = [r for r in range(nrow) if any(_as_num(grid[r][c]) is not None for c in range(1, ncol))]
    if not cand:
        return []
    months = [_month_of(grid[r][1]) for r in cand] if ncol > 1 else []
    has_month = bool(months) and sum(1 for m in months if m) >= max(3, len(cand) * 0.5)
    value_start = 2 if has_month else 1
    yrrows = [r for r in range(nrow)
              if _is_year(grid[r][0]) or isinstance(grid[r][0], (datetime.date, datetime.datetime))]
    if not yrrows:
        return []
    first_data = min(yrrows)
    headers = _row_headers(grid, nrow, ncol, value_start, first_data)
    out, cur_year = [], None
    for r in range(nrow):
        v0 = grid[r][0]
        is_date = isinstance(v0, (datetime.date, datetime.datetime))
        if is_date:
            d = v0.date() if isinstance(v0, datetime.datetime) else v0
            year, part, date, period = d.year, f"{d.month:02d}", d.isoformat(), d.isoformat()
            cur_year = year
        else:
            if _is_year(v0):
                cur_year = int(float(v0))
            if has_month:
                m = _month_of(grid[r][1])
                if not m or cur_year is None:
                    continue
                year, part, date, period = cur_year, f"{m:02d}", _mdate(cur_year, m), f"{cur_year}-{m:02d}"
            else:
                if not _is_year(v0):
                    continue
                year, part, date, period = cur_year, None, _ydate(cur_year), str(cur_year)
        for c in range(value_start, ncol):
            val = _as_num(grid[r][c])
            if val is None:
                continue
            ser = headers[c]
            q = _q_of(ser)
            if q and part is None and not is_date:
                out.append({"series": "value", "period": f"{year}-Q{q}", "year": year,
                            "part": f"Q{q}", "date": _qdate(year, q), "value": val})
            else:
                out.append({"series": ser, "period": period, "year": year,
                            "part": part, "date": date, "value": val})
    return out


def _depth_text(grid, r, first_dc):
    left = right = None
    for c in range(first_dc):
        v = grid[r][c]
        if v is not None and str(v).strip() != "":
            if left is None:
                left = (c, v)
            right = (c, v)
    if right is None:
        return None
    lc, lv = left
    firstline = next((ln for ln in str(lv).split("\n") if ln.strip() != ""), "")
    spaces = len(firstline) - len(firstline.lstrip(" 　\t"))
    return lc * 1000 + spaces, _english(right[1])


def _emit_cols(grid, nrow, ncol, periods, first_dc, data_start):
    out, stack = [], []
    row_series, used = {}, {}
    for r in range(data_start, nrow):
        dt = _depth_text(grid, r, first_dc)
        if dt is None:
            continue
        depth, text = dt
        if text == "":
            continue
        while stack and stack[-1][0] >= depth:
            stack.pop()
        stack.append((depth, text))
        path = " / ".join(t for _, t in stack)
        if any(_as_num(grid[r][c]) is not None for c in periods):
            row_series[r] = _uniquify(path, used)
    for r, series in row_series.items():
        for c, pp in periods.items():
            val = _as_num(grid[r][c])
            if val is None:
                continue
            y, part, date, period = pp
            out.append({"series": series, "period": period, "year": y,
                        "part": part, "date": date, "value": val})
    return out


def _parse_cols(grid, nrow, ncol):
    q_row = None
    for r in range(min(14, nrow)):
        if sum(1 for c in range(ncol) if _q_of(grid[r][c])) >= 2:
            q_row = r; break
    if q_row is not None:
        yr_row = None
        for r in range(q_row - 1, -1, -1):
            if any(_is_year(grid[r][c]) for c in range(ncol)):
                yr_row = r; break
        if yr_row is not None:
            first_dc = min(c for c in range(ncol) if _is_year(grid[yr_row][c]))
            years, cur = {}, None
            for c in range(first_dc, ncol):
                if _is_year(grid[yr_row][c]):
                    cur = int(float(grid[yr_row][c]))
                years[c] = cur
            periods = {}
            for c in range(first_dc, ncol):
                q, y = _q_of(grid[q_row][c]), years.get(c)
                if q and y is not None:
                    periods[c] = (y, f"Q{q}", _qdate(y, q), f"{y}-Q{q}")
            if periods and first_dc >= 1:
                return _emit_cols(grid, nrow, ncol, periods, first_dc, q_row + 1)
    best_r, best_n = None, 0
    for r in range(min(16, nrow)):
        n = sum(1 for c in range(ncol) if _period_from_header(grid[r][c]))
        if n > best_n:
            best_n, best_r = n, r
    if best_r is None or best_n < 2:
        return []
    periods = {}
    for c in range(ncol):
        pp = _period_from_header(grid[best_r][c])
        if pp:
            periods[c] = pp
    first_dc = min(periods)
    if first_dc < 1:
        return []
    return _emit_cols(grid, nrow, ncol, periods, first_dc, best_r + 1)


# --------------------------------------------------------------------------- #
# Fetch — one shared function per entity (catalog connector)
# --------------------------------------------------------------------------- #
def fetch_one(node_id: str) -> None:
    asset = node_id
    meta = ENTITY_META[node_id]
    url = _resolve_url(meta["endpoint"], meta["stem"])
    grid = _open_sheet(_fetch_bytes(url), meta["sheet"])
    rows = parse_grid(grid)
    if not rows:
        raise RuntimeError(f"{node_id}: parsed 0 rows from sheet {meta['sheet']!r}")
    table = pa.Table.from_pylist(
        [{"series": r["series"], "period": r["period"], "year": int(r["year"]),
          "part": r["part"], "date": r["date"], "value": float(r["value"])} for r in rows],
        schema=SCHEMA,
    )
    save_raw_parquet(table, asset)


DOWNLOAD_SPECS = [
    NodeSpec(id=sid, fn=fetch_one, kind="download")
    for sid in ENTITY_META
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"{s.id}-transform",
        deps=[s.id],
        sql=f'''
            SELECT series, period, year, part, CAST(date AS DATE) AS date, value
            FROM "{s.id}"
            WHERE value IS NOT NULL
            QUALIFY row_number() OVER (
                PARTITION BY series, period, part ORDER BY value DESC
            ) = 1
        ''',
    )
    for s in DOWNLOAD_SPECS
]
