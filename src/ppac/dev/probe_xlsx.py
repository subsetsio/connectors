import re, io
from subsets_utils import get
XLSX_PAGES = [
 "consumption/active-domestic-customers","consumption/state-wise","consumption/state-wise-pmuy-data",
 "infrastructure/installed-refinery-capacity","natural-gas/import","natural-gas/sectoral-consumption",
 "prices/contribution-to-central-and-state-exchequer","prices/petroleum-prices-and-under-recoveries",
]
import openpyxl
for p in XLSX_PAGES:
    h=get(f"https://ppac.gov.in/{p}").text
    urls=re.findall(r'(https://ppac\.gov\.in/uploads/[^"\']+\.xlsx|https://ppac\.gov\.in/download\.php\?file=[^"\']+\.xlsx)',h)
    urls=list(dict.fromkeys(urls))
    print(f"\n######### {p}\n  #xlsx urls={len(urls)}: {[u.split('/')[-1][:50] for u in urls]}")
    if not urls: 
        # try xls
        xls=re.findall(r'(https://ppac\.gov\.in/uploads/[^"\']+\.xls)',h); print("  xls:",xls[:2]); continue
    u=urls[0]
    try:
        b=get(u).content
        wb=openpyxl.load_workbook(io.BytesIO(b),read_only=True,data_only=True)
        print(f"  file={u.split('/')[-1]}  sheets={wb.sheetnames}")
        for ws in wb.worksheets[:1]:
            print(f"  -- sheet '{ws.title}' maxrow={ws.max_row} maxcol={ws.max_column}")
            for i,row in enumerate(ws.iter_rows(values_only=True)):
                r=[c for c in row]
                while r and r[-1] is None: r.pop()
                print(f"   r{i}:",[ (str(c)[:18] if c is not None else '') for c in r[:14]])
                if i>=11: break
    except Exception as e:
        print("  ERR",type(e).__name__,e)
