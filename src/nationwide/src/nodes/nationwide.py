"""Nationwide Building Society House Price Index (NHPI) connector.

Mechanism (from research, id `bulk_xlsx`): one stable per-dataset bulk download
at https://www.nationwide.co.uk/media/hpi/download/<slug> returning a full
spreadsheet (xlsx; a couple of legacy files carry a .xls name but are xlsx
internally). No auth, no pagination, no incremental filter — full re-pull every
refresh (the corpus is ~20 small files). Each <slug> is a NodeSpec id with the
`nationwide-` prefix.

The spreadsheets carry title/metadata preamble rows above the data table and use
several layouts (UK vertical series, regional matrices with paired £/INDEX
columns, region+%% pairs, occupation panels, multi-block sheets). `parse_workbook`
normalizes any of them into a tidy long table — (date, period, series, value) —
by locating the data block (a numeric-majority row whose first cell is a period)
and reconstructing each value column's label from the header rows above it
(merged group/region names forward-filled across their span, measure tokens like
£/INDEX/%% appended, decorative title rows skipped). The SQL transform then PIVOTs
that long raw into a wide per-file Delta table (one column per series).

Stateless full re-pull: no watermark/cursor — the maintain step decides whether a
fetch fn runs; if invoked, we fetch the whole file.
"""

import datetime as dt
import io
import re

import openpyxl
import pyarrow as pa

from subsets_utils import NodeSpec, SqlNodeSpec, get, save_raw_parquet, transient_retry
from constants import ENTITY_IDS

DOWNLOAD_BASE = "https://www.nationwide.co.uk/media/hpi/download"

SCHEMA = pa.schema([
    ("date", pa.date32()),
    ("period", pa.string()),
    ("series", pa.string()),
    ("value", pa.float64()),
])

# Header cells that name a measure rather than a region/group. Lower-cased,
# exact match — descriptive headers like "INDEX Q1 1993=100" are NOT measures.
_MEASURES = {"£", "index", "price", "%", "annual change"}
_Q_MONTH = {1: 1, 2: 4, 3: 7, 4: 10}


# --------------------------------------------------------------------------- #
# Spreadsheet parsing                                                         #
# --------------------------------------------------------------------------- #
def _parse_period(v):
    """Return the period start date for a first-column cell, or None."""
    if isinstance(v, dt.datetime):
        return dt.date(v.year, v.month, 1)
    if isinstance(v, dt.date):
        return dt.date(v.year, v.month, 1)
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        iv = int(v)
        if float(v) == iv and 1900 <= iv <= 2100:
            return dt.date(iv, 1, 1)
        return None
    if not isinstance(v, str):
        return None
    s = v.strip()
    m = re.match(r"^Q([1-4])\s+(\d{4})$", s)
    if m:
        return dt.date(int(m.group(2)), _Q_MONTH[int(m.group(1))], 1)
    m = re.match(r"^(\d{4})\s+Q([1-4])$", s)
    if m:
        return dt.date(int(m.group(1)), _Q_MONTH[int(m.group(2))], 1)
    m = re.match(r"^(\d{4})$", s)
    if m and 1900 <= int(s) <= 2100:
        return dt.date(int(s), 1, 1)
    return None


def _is_noise(s):
    """Bare number or '=100'-style base marker — carries no label meaning."""
    return bool(re.match(r"^=?\s*-?\d+(\.\d+)?$", s))


def _norm(c):
    if c is None:
        return ""
    return re.sub(r"\s+", " ", str(c)).strip()


def _to_float(c):
    if c is None or isinstance(c, bool):
        return None
    if isinstance(c, (int, float)):
        return float(c)
    s = str(c).strip()
    if not s or s.upper() == "N/A":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _is_data_row(row):
    """A data row: first cell is a period AND the remainder is numeric-majority.
    This keeps a header row whose first cell holds a base-period note
    (e.g. 'Q1 1993' over a row of region names) from being read as data."""
    if not row or _parse_period(row[0]) is None:
        return False
    nums = texts = 0
    for c in row[1:]:
        if _to_float(c) is not None:
            nums += 1
        elif _norm(c):
            texts += 1
    return nums >= 1 and nums >= texts


def _parse_sheet(rows):
    data_start = None
    for i, row in enumerate(rows):
        if _is_data_row(row):
            data_start = i
            break
    if data_start is None:
        return []

    header_rows = rows[:data_start]
    width = max((len(r) for r in rows[data_start:] if r), default=0)
    region_cols = [[] for _ in range(width)]
    measure_cols = [[] for _ in range(width)]

    def add_region(c, tok):
        if tok and (not region_cols[c] or region_cols[c][-1] != tok):
            region_cols[c].append(tok)

    for hr in header_rows:
        cells = [_norm(hr[c]) if c < len(hr) else "" for c in range(width)]
        nonblank = [c for c in cells if c]
        if len(nonblank) == 1 and len(nonblank[0]) > 20:
            continue  # decorative single-cell title row
        text_tokens = [c for c in cells[1:]
                       if c and c.lower() not in _MEASURES and not _is_noise(c)]
        qualifies = len(text_tokens) >= 2  # a real multi-group/region header row
        cur = ""
        for c in range(width):
            cell = cells[c]
            is_measure = cell.lower() in _MEASURES
            is_text = bool(cell) and not is_measure and not _is_noise(cell)
            if is_text:
                cur = cell
            if c >= 1:
                if qualifies:
                    add_region(c, cur)        # spanning ff (measure-transparent)
                elif is_text:
                    add_region(c, cell)       # literal, no spanning
                if is_measure and cell not in measure_cols[c]:
                    measure_cols[c].append(cell)

    labels = {}
    for c in range(1, width):
        lbl = " ".join(region_cols[c] + measure_cols[c]).strip()
        labels[c] = lbl if lbl else f"col{c}"

    out = []
    for row in rows[data_start:]:
        if not _is_data_row(row):
            continue
        d = _parse_period(row[0])
        period = _norm(row[0])
        for c in range(1, width):
            if c >= len(row):
                break
            val = _to_float(row[c])
            if val is None:
                continue
            out.append({"date": d, "period": period,
                        "series": labels[c], "value": val})
    return out


def parse_workbook(content):
    """Parse the data sheet (the one yielding the most rows; 'Notes' skipped)
    into long-format records: {date, period, series, value}."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        best = []
        for sheet in wb.sheetnames:
            if "note" in sheet.lower():
                continue
            rows = list(wb[sheet].iter_rows(values_only=True))
            parsed = _parse_sheet(rows)
            if len(parsed) > len(best):
                best = parsed
        return best
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# Fetch                                                                        #
# --------------------------------------------------------------------------- #
@transient_retry()
def _download(url):
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.content


def fetch_one(node_id):
    slug = node_id[len("nationwide-"):]
    url = f"{DOWNLOAD_BASE}/{slug}"
    content = _download(url)
    rows = parse_workbook(content)
    if not rows:
        raise AssertionError(f"{node_id}: parsed 0 data rows from {url}")
    table = pa.Table.from_pylist(rows, schema=SCHEMA)
    save_raw_parquet(table, node_id)


DOWNLOAD_SPECS = [
    NodeSpec(
        id=f"nationwide-{eid.lower().replace('_', '-')}",
        fn=fetch_one,
        kind="download",
    )
    for eid in ENTITY_IDS
]

# One published wide Delta table per file: PIVOT the long raw so each distinct
# series becomes a column and each period a row. first(value) is a no-op
# dedup (the long raw holds one value per date+series).
TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"{spec.id}-transform",
        deps=[spec.id],
        key=("date",),
        temporal="date",
        sql=(
            f'PIVOT (SELECT date, series, value FROM "{spec.id}") '
            f"ON series USING first(value) GROUP BY date ORDER BY date"
        ),
    )
    for spec in DOWNLOAD_SPECS
]
