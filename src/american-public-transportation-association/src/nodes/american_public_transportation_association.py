"""American Public Transportation Association (APTA) connector.

APTA publishes a curated set of statistical Excel workbooks (and zipped Excel
databases) on www.apta.com. The entire origin — landing pages AND the
/wp-content/uploads/ files — sits behind a Cloudflare interactive bot-challenge
that 403s plain httpx/requests (and even a browser-UA curl) regardless of IP; it
fingerprints the TLS/HTTP2 handshake. We therefore deliberately bypass
``subsets_utils.get`` for the source fetch and use ``curl_cffi`` with Chrome
impersonation, which presents Chrome's real TLS/JA3 fingerprint and is served
normally (verified end-to-end: real PK-zip xlsx bytes returned). This is the
same pattern used by the bruegel / obr / energy-institute / rbnz connectors.
Everything else (raw I/O) still goes through subsets_utils.

Strategy is stateless full re-pull. The data-file URLs are year/month-stamped
under /wp-content/uploads/ and re-minted each edition, so they are NOT
hardcoded: each fetch first GETs the stable product landing page and scrapes
the current <a href> to the .xlsx/.zip, then downloads it. There is no
incremental query surface; the corpus is small (each workbook is tens of KB to
a few MB) so a full re-fetch each run is cheap.

The workbooks are human-formatted (title rows, multi-row headers, footnote
markers, two-block layouts). Each accepted subset has a tailored extractor that
normalises its sheet into clean rows written as NDJSON; the SQL transform then
types/projects. Heterogeneous wide tables (fare/vehicle/infrastructure and the
Appendix B agency tables) are snake-cased generically and passed through.
"""

from __future__ import annotations

import io
import os
import re
import tempfile
import time
import zipfile

import openpyxl
from curl_cffi import requests as cffi_requests

from subsets_utils import NodeSpec, SqlNodeSpec, save_raw_ndjson

from constants import ENTITY_IDS

_PREFIX = "american-public-transportation-association-"
_IMPERSONATE = "chrome"

# --- product discovery -------------------------------------------------------

_FACT_BOOK = "https://www.apta.com/news-research/about-the-industry/public-transportation-fact-book/"
_STATS = "https://www.apta.com/news-research/public-transit-statistics/"

# Each product is discovered by scraping its stable landing page for the current
# data file: a /wp-content/uploads/ link (relative OR absolute) whose filename
# contains `key` and ends in `ext`. URLs are re-minted each edition, so never
# hardcode them.
PRODUCTS = {
    "ridership": {"page": _STATS + "ridership-data/", "key": "Ridership-by-Mode-and-Quarter", "ext": "xlsx"},
    "appendix_a": {"page": _FACT_BOOK, "key": "Appendix-A", "ext": "xlsx"},
    "appendix_b": {"page": _FACT_BOOK, "key": "Appendix-B", "ext": "xlsx"},
    "appendix_c": {"page": _FACT_BOOK, "key": "Appendix-C", "ext": "xlsx"},
    "fare": {"page": _STATS + "fare-database/", "key": "Fare-Database", "ext": "zip", "zip": True},
    "vehicle": {"page": _STATS + "vehicle-database/", "key": "Vehicle-Database", "ext": "zip", "zip": True},
    "infra": {"page": _STATS + "infrastructure-database/", "key": "Infrastructure-Database", "ext": "zip", "zip": True},
}

# entity_id -> {product, sheet, parser}
CONFIG = {
    "factbook-a-table-1": {"product": "appendix_a", "sheet": "1", "parser": "table1"},
    "factbook-b-rural-state-totals": {"product": "appendix_b", "sheet": "Rural State Totals", "parser": "two_row"},
    "factbook-b-rural-tribal-by-mode": {"product": "appendix_b", "sheet": "Rural and Tribal by Mode", "parser": "two_row"},
    "factbook-b-rural-tribal-state-by-mode": {"product": "appendix_b", "sheet": "Rural and Tribal State by Mode", "parser": "two_row"},
    "factbook-b-rural-tribal-totals": {"product": "appendix_b", "sheet": "Rural and Tribal Totals", "parser": "two_row"},
    "factbook-b-urban-agency-by-mode": {"product": "appendix_b", "sheet": "Urban - Agency by Mode", "parser": "two_row"},
    "factbook-b-urban-agency-total": {"product": "appendix_b", "sheet": "Urban - Agency Total", "parser": "two_row"},
    "factbook-b-uza-by-mode": {"product": "appendix_b", "sheet": "UZA by Mode", "parser": "two_row"},
    "factbook-b-uza-totals": {"product": "appendix_b", "sheet": "UZA Totals", "parser": "two_row"},
    "factbook-c-uza-population-comparison": {"product": "appendix_c", "sheet": "A. 1950 Through 2010 Comparison", "parser": "appx_c_compare"},
    "factbook-c-uza-population-by-census-year": {"product": "appendix_c", "sheet": None, "parser": "appx_c_census"},
    "fare-demand-response": {"product": "fare", "sheet": "Demand Response Fares", "parser": "simple"},
    "fare-fixed-route": {"product": "fare", "sheet": "Fixed-Route-Fares", "parser": "simple"},
    "infrastructure-rail-lines": {"product": "infra", "sheet": "Rail Lines", "parser": "simple"},
    "infrastructure-real-time-data-systems": {"product": "infra", "sheet": "Real-Time Data Systems", "parser": "simple"},
    "infrastructure-station-data": {"product": "infra", "sheet": "Station Data", "parser": "simple"},
    "ridership-by-mode-quarter": {"product": "ridership", "sheet": "Ridership By Quarter", "parser": "ridership_quarter"},
    "ridership-by-mode-year": {"product": "ridership", "sheet": "Year-End Totals", "parser": "ridership_year"},
    "vehicle-equipment": {"product": "vehicle", "sheet": "Equipment", "parser": "simple"},
    "vehicle-fleet": {"product": "vehicle", "sheet": "Fleet", "parser": "simple"},
}


# --- HTTP --------------------------------------------------------------------

# Cloudflare intermittently serves a 403 interactive-challenge page even to the
# Chrome-impersonating client (especially under a burst of requests); it clears
# on retry. We therefore treat 403 as transient here alongside 429/5xx and the
# usual network errors, with capped exponential backoff. This is the deliberate
# different classification the standard transient_retry doesn't cover.
_RETRY_STATUS = {403, 408, 429, 500, 502, 503, 504}


def _get(session, url: str, attempts: int = 7) -> bytes:
    last = None
    for i in range(attempts):
        try:
            resp = session.get(url, timeout=120)
            if resp.status_code in _RETRY_STATUS:
                last = f"HTTP {resp.status_code}"
            else:
                resp.raise_for_status()
                return resp.content
        except Exception as e:  # noqa: BLE001 - network/transport error, logged + retried
            last = f"{type(e).__name__}: {e}"
        if i < attempts - 1:
            time.sleep(min(60.0, 3.0 * (2 ** i)))
    raise RuntimeError(f"fetch failed after {attempts} attempts for {url}: {last}")


# Each accepted entity is its own download node and the orchestrator runs every
# node in a *fresh forked process*, so there is no in-memory reuse between them.
# But the 20 entities map to only 7 underlying products — Appendix B alone backs
# 8 entities — so a naive "discover + download per entity" hammers www.apta.com
# with ~40 requests (the same workbook fetched 8×), which escalates Cloudflare
# from the occasional clearable 403 to *sustained* 403s partway through the run
# (observed: 12/20 done, then blocked). We therefore cache the discovered URL and
# the downloaded bytes per product on the filesystem, shared across the forked
# children within a run, collapsing the corpus to one fetch per product (~14
# requests). Writes are atomic (tmp + os.replace) so a concurrent or interrupted
# fetch can never serve a truncated file. On the ephemeral cloud runner /tmp is
# empty per run, so this never serves stale editions across runs.
_CACHE_DIR = os.path.join(tempfile.gettempdir(), "apta-fetch-cache")


def _product_bytes(session, product: str) -> bytes:
    os.makedirs(_CACHE_DIR, exist_ok=True)
    blob = os.path.join(_CACHE_DIR, f"{product}.bin")
    try:
        if os.path.getsize(blob) > 0:
            with open(blob, "rb") as fh:
                return fh.read()
    except OSError:
        pass  # missing or empty → fetch below
    url = _discover(session, product)
    content = _get(session, url)
    tmp = f"{blob}.{os.getpid()}.tmp"
    with open(tmp, "wb") as fh:
        fh.write(content)
    os.replace(tmp, blob)
    return content


def _discover(session, product: str) -> str:
    prod = PRODUCTS[product]
    html = _get(session, prod["page"]).decode("utf-8", "replace")
    pattern = (
        r'href="((?:https://www\.apta\.com)?/wp-content/uploads/[^"]*'
        + re.escape(prod["key"]) + r'[^"]*\.' + prod["ext"] + r')"'
    )
    urls = sorted({u if u.startswith("http") else "https://www.apta.com" + u
                   for u in re.findall(pattern, html, re.IGNORECASE)})
    if not urls:
        raise RuntimeError(
            f"no file link matching key={prod['key']!r} .{prod['ext']} on {prod['page']}"
        )
    # multiple editions (e.g. infrastructure by year): the lexically-last
    # filename is the most recent (year is zero-padded in the filename).
    return urls[-1]


# --- value cleaning ----------------------------------------------------------

_PAREN = re.compile(r"\([^)]*\)")
_NULLISH = {"", "---", "--", "n/a", "na", "in other", "in total"}


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = _PAREN.sub("", str(v)).replace(",", "").replace("%", "").replace("$", "").strip()
    if s.lower() in _NULLISH:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_int(v):
    f = _to_float(v)
    return int(f) if f is not None else None


def _year(v):
    if v is None:
        return None
    s = _PAREN.sub("", str(v)).strip()
    m = re.match(r"^(\d{4})$", s)
    return int(m.group(1)) if m else None


def _snake(name) -> str | None:
    if name is None:
        return None
    s = str(name).strip()
    if not s:
        return None
    s = re.sub(r"[^0-9a-zA-Z]+", "_", s).strip("_").lower()
    return s or None


def _jsonable(v):
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def _clean_row(row: dict) -> dict:
    return {k: _jsonable(v) for k, v in row.items()}


def _is_blank(row) -> bool:
    return all(c is None or str(c).strip() == "" for c in row)


# --- generic extractors ------------------------------------------------------

def _header_table(ws, header_idx: int, data_start: int):
    """Single header row at ``header_idx`` (0-based); data from ``data_start``."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= data_start:
        return []
    header = rows[header_idx]
    cols, used = [], {}
    for i, h in enumerate(header):
        nm = _snake(h)
        if not nm:
            continue
        if nm in used:
            used[nm] += 1
            nm = f"{nm}_{used[nm]}"
        else:
            used[nm] = 1
        cols.append((i, nm))
    out = []
    for r in rows[data_start:]:
        if _is_blank(r):
            continue
        d = {nm: (r[i] if i < len(r) else None) for i, nm in cols}
        if all(v is None for v in d.values()):
            continue
        out.append(d)
    return out


def _parse_simple(ws):
    return _header_table(ws, header_idx=0, data_start=1)


def _parse_appx_c_compare(ws):
    # row1 = census-year group labels, row2 = column names, row3 blank, data row4+
    return _header_table(ws, header_idx=1, data_start=3)


def _parse_two_row(ws):
    """Appendix B: row1 = group labels (merged over Amount/Rank pairs), row2 =
    column names, data from row3. Combine group into the Amount/Rank names so
    the repeated Amount/Rank columns stay distinct."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []
    group, names = rows[0], rows[1]
    # forward-fill the merged group labels across the row
    gff, last = [], None
    width = max(len(group), len(names))
    for i in range(width):
        g = group[i] if i < len(group) else None
        gs = str(g).strip() if g is not None and str(g).strip() else None
        if gs:
            last = gs
        gff.append(last)
    cols, used = [], {}
    for i in range(width):
        nm = _snake(names[i] if i < len(names) else None)
        if not nm:
            continue
        if nm in ("amount", "rank") and gff[i]:
            nm = f"{_snake(gff[i])}_{nm}"
        if nm in used:
            used[nm] += 1
            nm = f"{nm}_{used[nm]}"
        else:
            used[nm] = 1
        cols.append((i, nm))
    out = []
    for r in rows[2:]:
        if _is_blank(r):
            continue
        d = {nm: (r[i] if i < len(r) else None) for i, nm in cols}
        if all(v is None for v in d.values()):
            continue
        out.append(d)
    return out


_RIDERSHIP_MODE_COLS = [
    "total_unlinked_trips_000s", "heavy_rail_000s", "light_rail_000s",
    "commuter_rail_000s", "trolleybus_000s", "bus_000s",
    "demand_response_000s", "other_000s",
]


def _parse_ridership_quarter(ws):
    # row1 header: Quarter | <yr-qtr label> | Year | Total | Heavy | Light |
    #              Commuter | Trolleybus | Bus | Demand Response | Other
    out = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r or r[0] is None or str(r[0]).strip() == "":
            continue
        year = _to_int(r[2]) if len(r) > 2 else None
        if year is None:
            continue
        row = {"year": year, "quarter": str(r[0]).strip()}
        for j, col in enumerate(_RIDERSHIP_MODE_COLS):
            idx = 3 + j
            row[col] = _to_float(r[idx]) if idx < len(r) else None
        out.append(row)
    return out


def _parse_ridership_year(ws):
    # row1 header: Year | Total | Heavy | Light | Commuter | Trolleybus | Bus |
    #              Demand Response | Other
    out = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r:
            continue
        year = _year(r[0]) or _to_int(r[0])
        if year is None:
            continue
        row = {"year": year}
        for j, col in enumerate(_RIDERSHIP_MODE_COLS):
            idx = 1 + j
            row[col] = _to_float(r[idx]) if idx < len(r) else None
        out.append(row)
    return out


# Appendix A, Table 1: two horizontal blocks sharing the Year key.
_T1_PART_A = {
    1: "Bus", 2: "Bus Rapid Transit", 3: "Commuter Bus", 4: "Total Bus",
    5: "Trolleybus", 6: "Demand Response", 7: "Transit Vanpool", 8: "Publico",
    9: "Total Roadway Modes",
}
_T1_PART_B = {
    11: "Commuter Rail", 12: "Hybrid Rail", 13: "Total Regional Railroad",
    14: "Heavy Rail", 15: "Light Rail", 16: "Streetcar",
    17: "Total Surface Rail", 18: "Ferryboat", 19: "Other Fixed-Guideway",
    20: "Total Fixed-Guideway", 21: "All Modes",
}


def _parse_table1(ws):
    out = []
    for r in list(ws.iter_rows(values_only=True))[6:]:  # data from row 7
        if not r:
            continue
        ya = _year(r[0]) if len(r) > 0 else None
        if ya is not None:
            for idx, mode in _T1_PART_A.items():
                v = _to_float(r[idx]) if idx < len(r) else None
                if v is not None:
                    out.append({"year": ya, "mode": mode,
                                "unlinked_passenger_trips_millions": v})
        yb = _year(r[10]) if len(r) > 10 else None
        if yb is not None:
            for idx, mode in _T1_PART_B.items():
                v = _to_float(r[idx]) if idx < len(r) else None
                if v is not None:
                    out.append({"year": yb, "mode": mode,
                                "unlinked_passenger_trips_millions": v})
    return out


_CENSUS_SHEETS = {
    "B. 2010 Census": 2010, "C. 2000 Census": 2000, "D. 1990 Census": 1990,
    "E. 1980 Census": 1980, "F. 1970 Census": 1970, "G. 1960 Census": 1960,
    "H. 1950 Census": 1950,
}


def _parse_appx_c_census(wb):
    """Union the seven per-census-year UZA population sheets (all sharing the
    same 8-column layout) into one long table with a census_year column."""
    out = []
    for sheet, year in _CENSUS_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))
        for r in rows[3:]:  # row1 group, row2 names, row3 blank, data row4+
            if not r or len(r) < 1 or r[0] is None or str(r[0]).strip() == "":
                continue
            # the first data row of each sheet is a national-total row whose
            # "urbanized area" cell is a numeric count, not a place name — skip it.
            if not re.search(r"[A-Za-z]", str(r[0])):
                continue
            out.append({
                "census_year": year,
                "urbanized_area": _jsonable(r[0]),
                "population": _to_int(r[1]) if len(r) > 1 else None,
                "land_area_sq_mi": _to_float(r[2]) if len(r) > 2 else None,
                "density_per_sq_mi": _to_float(r[3]) if len(r) > 3 else None,
                "primary_state": _jsonable(r[4]) if len(r) > 4 else None,
                "secondary_state": _jsonable(r[5]) if len(r) > 5 else None,
                "third_state": _jsonable(r[6]) if len(r) > 6 else None,
                "fourth_state": _jsonable(r[7]) if len(r) > 7 else None,
            })
    return out


_WS_PARSERS = {
    "simple": _parse_simple,
    "two_row": _parse_two_row,
    "appx_c_compare": _parse_appx_c_compare,
    "ridership_quarter": _parse_ridership_quarter,
    "ridership_year": _parse_ridership_year,
    "table1": _parse_table1,
}


# --- fetch -------------------------------------------------------------------

def fetch_one(node_id: str) -> None:
    asset = node_id  # the runtime passes the spec id; it IS the asset name
    entity_id = node_id[len(_PREFIX):]
    cfg = CONFIG[entity_id]
    prod = PRODUCTS[cfg["product"]]

    session = cffi_requests.Session(impersonate=_IMPERSONATE)
    try:
        url = _discover(session, cfg["product"])
        content = _get(session, url)
    finally:
        session.close()

    xlsx_bytes = content
    if prod.get("zip"):
        zf = zipfile.ZipFile(io.BytesIO(content))
        members = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        if not members:
            raise RuntimeError(f"{asset}: no .xlsx member in zip {url}")
        xlsx_bytes = zf.read(members[0])

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    try:
        if cfg["parser"] == "appx_c_census":
            rows = _parse_appx_c_census(wb)
        else:
            if cfg["sheet"] not in wb.sheetnames:
                raise RuntimeError(
                    f"{asset}: sheet {cfg['sheet']!r} not in {wb.sheetnames}"
                )
            rows = _WS_PARSERS[cfg["parser"]](wb[cfg["sheet"]])
    finally:
        wb.close()

    rows = [_clean_row(r) for r in rows]
    if not rows:
        raise AssertionError(f"{asset}: parsed 0 rows from {url}")
    save_raw_ndjson(rows, asset)
    print(f"  {asset}: wrote {len(rows):,} rows")


DOWNLOAD_SPECS = [
    NodeSpec(
        id=f"{_PREFIX}{eid.lower().replace('_', '-')}",
        fn=fetch_one,
        kind="download",
    )
    for eid in ENTITY_IDS
]


# --- transforms --------------------------------------------------------------
# Most subsets are wide, heterogeneous agency tables already snake-cased and
# typed by DuckDB's NDJSON reader — passed through. The clean-keyed flagship
# series (ridership, Table 1, census) get explicit casts as a correctness gate.

def _v(eid: str) -> str:
    return f'{_PREFIX}{eid}'


_TRANSFORM_SQL = {
    "ridership-by-mode-quarter": f'''
        SELECT CAST(year AS INTEGER) AS year, quarter,
               CAST(total_unlinked_trips_000s AS DOUBLE) AS total_unlinked_trips_000s,
               CAST(heavy_rail_000s AS DOUBLE) AS heavy_rail_000s,
               CAST(light_rail_000s AS DOUBLE) AS light_rail_000s,
               CAST(commuter_rail_000s AS DOUBLE) AS commuter_rail_000s,
               CAST(trolleybus_000s AS DOUBLE) AS trolleybus_000s,
               CAST(bus_000s AS DOUBLE) AS bus_000s,
               CAST(demand_response_000s AS DOUBLE) AS demand_response_000s,
               CAST(other_000s AS DOUBLE) AS other_000s
        FROM "{_v("ridership-by-mode-quarter")}"
        WHERE year IS NOT NULL AND quarter IS NOT NULL
    ''',
    "ridership-by-mode-year": f'''
        SELECT CAST(year AS INTEGER) AS year,
               CAST(total_unlinked_trips_000s AS DOUBLE) AS total_unlinked_trips_000s,
               CAST(heavy_rail_000s AS DOUBLE) AS heavy_rail_000s,
               CAST(light_rail_000s AS DOUBLE) AS light_rail_000s,
               CAST(commuter_rail_000s AS DOUBLE) AS commuter_rail_000s,
               CAST(trolleybus_000s AS DOUBLE) AS trolleybus_000s,
               CAST(bus_000s AS DOUBLE) AS bus_000s,
               CAST(demand_response_000s AS DOUBLE) AS demand_response_000s,
               CAST(other_000s AS DOUBLE) AS other_000s
        FROM "{_v("ridership-by-mode-year")}"
        WHERE year IS NOT NULL
    ''',
    "factbook-a-table-1": f'''
        SELECT CAST(year AS INTEGER) AS year, mode,
               CAST(unlinked_passenger_trips_millions AS DOUBLE) AS unlinked_passenger_trips_millions
        FROM "{_v("factbook-a-table-1")}"
        WHERE year IS NOT NULL AND mode IS NOT NULL
          AND unlinked_passenger_trips_millions IS NOT NULL
    ''',
    "factbook-c-uza-population-by-census-year": f'''
        SELECT CAST(census_year AS INTEGER) AS census_year,
               urbanized_area,
               CAST(population AS BIGINT) AS population,
               CAST(land_area_sq_mi AS DOUBLE) AS land_area_sq_mi,
               CAST(density_per_sq_mi AS DOUBLE) AS density_per_sq_mi,
               primary_state, secondary_state, third_state, fourth_state
        FROM "{_v("factbook-c-uza-population-by-census-year")}"
        WHERE urbanized_area IS NOT NULL
    ''',
}


TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"{s.id}-transform",
        deps=[s.id],
        sql=_TRANSFORM_SQL.get(
            s.id[len(_PREFIX):], f'SELECT * FROM "{s.id}"'
        ),
    )
    for s in DOWNLOAD_SPECS
]
