"""DHS OHSS Yearbook of Immigration Statistics connector.

Source: Office of Homeland Security Statistics (OHSS), Yearbook of Immigration
Statistics. The yearbook publishes one consolidated multi-sheet xlsx workbook
per topic (Lawful Permanent Residents, Refugees, Naturalizations, Nonimmigrant
Admissions). Each workbook holds a TOC sheet plus one sheet per numbered table,
and every table has its own column layout, so each table is published as its own
Delta table.

Strategy (stateless full re-pull each run; corpus is tens of MB):
  * Workbook URLs embed the publication date (e.g. /system/files/2026-06/...) and
    change with every release, so we scrape the yearbook page for the current
    .xlsx links instead of hardcoding them.
  * The server returns 403 to default/library User-Agents, so every request
    carries an explicit browser User-Agent.
  * Each fetch downloads its topic workbook, extracts the entity's sheet, and
    melts the human-formatted sheet (title row, header row, section sub-headers,
    footnotes) into a uniform tidy long table: one row per (category, breakdown)
    cell with its numeric value. The transform SQL is then a thin typed
    projection.

Three details of the yearbook's human formatting carry meaning that a naive melt
would destroy, so the parser reads them explicitly:
  * Column A encodes a row hierarchy as cell INDENTATION, not as distinct labels.
    Table 30 repeats the same age labels under a 'Female' and a 'Male' leader
    row, so indentation is the only thing separating them; `parent_category`
    carries that leader. Elsewhere (Tables 7/8/9/25/26) the same mechanism marks
    subtotal-vs-component rows.
  * The NISuppTable sheets put a 'Description' column between the row label and
    the first data column. It is an attribute of the row, not an observation.
  * Footnote markers are superscript runs inside the shared string ('Australia'
    + superscript '1'), so they can be dropped exactly rather than guessed at by
    stripping trailing digits — which would destroy class codes like 'H2A'.
"""
import io
import re
import zipfile
import xml.etree.ElementTree as ET

import pyarrow as pa
import openpyxl
import requests

from subsets_utils import (
    NodeSpec,
    save_raw_parquet,
    transient_retry,
)

YEARBOOK_URL = "https://ohss.dhs.gov/topics/immigration/yearbook/2024"
BASE = "https://ohss.dhs.gov"
# ASCII-only User-Agent (required by httpx/urllib3); the OHSS server 403s the
# default library UA.
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0 Safari/537.36")
REQ_HEADERS = {
    "User-Agent": UA,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Connection": "keep-alive",
}
_SESSION = requests.Session()

# entity_id -> metadata. topic_slug matches the workbook filename key and the
# entity-id prefix; sheet is the exact worksheet name inside that workbook.
ENTITY_META = {
    "lawful-permanent-residents-lprsupptable-1": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "LPRSuppTable 1", "title": "Persons Obtaining Lawful Permanent Resident Status by State or Territory of Residence and Region and Country of Birth: Fiscal Year 2024"},
    "lawful-permanent-residents-lprsupptable-2": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "LPRSuppTable 2", "title": "Persons Obtaining Lawful Permanent Resident Status by Leading Core Based Statistical Areas (CBSA) of Residence and Region and Country of Birth: Fiscal Year 2024"},
    "lawful-permanent-residents-lprsupptable-3": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "LPRSuppTable 3", "title": "Persons Obtaining Lawful Permanent Resident Status by Region of Birth and Core Based Statistical Area (CBSA) of Residence: Fiscal Year 2024"},
    "lawful-permanent-residents-lprsupptable-4": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "LPRSuppTable 4", "title": "Immigrant Orphans Adopted by U.S. Citizens by Sex, Age, and State or Territory of Residence: Fiscal Year 2024"},
    "lawful-permanent-residents-table-1": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "Table 1", "title": "Persons Obtaining Lawful Permanent Resident Status: Fiscal Years 1820 to 2024"},
    "lawful-permanent-residents-table-10": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "Table 10", "title": "Persons Obtaining Lawful Permanent Resident Status by Broad Class of Admission and Region and Country of Birth: Fiscal Year 2024"},
    "lawful-permanent-residents-table-11": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "Table 11", "title": "Persons Obtaining Lawful Permanent Resident Status by Broad Class of Admission and Region and Country of Last Residence: Fiscal Year 2024"},
    "lawful-permanent-residents-table-12": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "Table 12", "title": "Immigrant Orphans Adopted by U.S. Citizens by Sex, Age, and Region and Country of Birth: Fiscal Year 2024"},
    "lawful-permanent-residents-table-2": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "Table 2", "title": "Persons Obtaining Lawful Permanent Resident Status by Region and Country of Last Residence: Fiscal Years 2015 to 2024"},
    "lawful-permanent-residents-table-3": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "Table 3", "title": "Persons Obtaining Lawful Permanent Resident Status by Region and Country of Birth: Fiscal Years 2015 to 2024"},
    "lawful-permanent-residents-table-4": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "Table 4", "title": "Persons Obtaining Lawful Permanent Resident Status by State or Territory of Residence: Fiscal Years 2015 to 2024"},
    "lawful-permanent-residents-table-5": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "Table 5", "title": "Persons Obtaining Lawful Permanent Resident Status by Core Based Statistical Area (CBSA) of Residence: Fiscal Years 2015 to 2024"},
    "lawful-permanent-residents-table-6": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "Table 6", "title": "Persons Obtaining Lawful Permanent Resident Status by Type and Major Class of Admission: Fiscal Years 2015 to 2024"},
    "lawful-permanent-residents-table-7": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "Table 7", "title": "Persons Obtaining Lawful Permanent Resident Status by Type and Detailed Class of Admission: Fiscal Year 2024"},
    "lawful-permanent-residents-table-8": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "Table 8", "title": "Persons Obtaining Lawful Permanent Resident Status by Sex, Age, Marital Status, and Occupation: Fiscal Year 2024"},
    "lawful-permanent-residents-table-9": {"topic_slug": "lawful-permanent-residents", "topic": "Lawful Permanent Residents", "sheet": "Table 9", "title": "Persons Obtaining Lawful Permanent Resident Status by Broad Class of Admission and Selected Demographic Characteristics: Fiscal Year 2024"},
    "naturalizations-natzsupptable1": {"topic_slug": "naturalizations", "topic": "Naturalizations", "sheet": "NATZSuppTable1", "title": "Persons Naturalized by State or Territory of Residence and Region and Country of Birth: Fiscal Year 2024"},
    "naturalizations-natzsupptable2": {"topic_slug": "naturalizations", "topic": "Naturalizations", "sheet": "NATZSuppTable2", "title": "Persons Naturalized by Leading Core Based Statistical Areas (CBSA) of Residence and Region and Country of Birth: Fiscal Year 2024"},
    "naturalizations-natzsupptable3": {"topic_slug": "naturalizations", "topic": "Naturalizations", "sheet": "NATZSuppTable3", "title": "Persons Naturalized by Region of Birth and Core Based Statistical Area (CBSA) of Residence: Fiscal Year 2024"},
    "naturalizations-table-21": {"topic_slug": "naturalizations", "topic": "Naturalizations", "sheet": "Table 21", "title": "Applications for Naturalization Filed, Persons Naturalized, and Applications for Naturalization Denied: Fiscal Years 1907 to 2024"},
    "naturalizations-table-22": {"topic_slug": "naturalizations", "topic": "Naturalizations", "sheet": "Table 22", "title": "Persons Naturalized by Region and Country of Birth: Fiscal Years 2015 to 2024"},
    "naturalizations-table-23": {"topic_slug": "naturalizations", "topic": "Naturalizations", "sheet": "Table 23", "title": "Persons Naturalized by State or Territory of Residence: Fiscal Years 2015 to 2024"},
    "naturalizations-table-24": {"topic_slug": "naturalizations", "topic": "Naturalizations", "sheet": "Table 24", "title": "Persons Naturalized by Core Based Statistical Area (CBSA) of Residence: Fiscal Years 2015 to 2024"},
    "naturalizations-table-25": {"topic_slug": "naturalizations", "topic": "Naturalizations", "sheet": "Table 25", "title": "Persons Naturalized by Sex, Age, Marital Status, and Occupation: Fiscal Year 2024"},
    "nonimmigrants-nisupptable1": {"topic_slug": "nonimmigrants", "topic": "Nonimmigrant Admissions", "sheet": "NISuppTable1", "title": "I-94 Nonimmigrant Admissions by Class of Admission and Country of Citizenship: Fiscal Year 2024"},
    "nonimmigrants-nisupptable2": {"topic_slug": "nonimmigrants", "topic": "Nonimmigrant Admissions", "sheet": "NISuppTable2", "title": "I-94 Nonimmigrant Admissions by Class of Admission and Country of Residence: Fiscal Year 2024"},
    "nonimmigrants-nisupptable3": {"topic_slug": "nonimmigrants", "topic": "Nonimmigrant Admissions", "sheet": "NISuppTable3", "title": "I-94 Nonimmigrant Admissions by Class of Admission and State or Territory of Destination: Fiscal Year 2024"},
    "nonimmigrants-table-26": {"topic_slug": "nonimmigrants", "topic": "Nonimmigrant Admissions", "sheet": "Table 26", "title": "Nonimmigrant Admissions by Class of Admission: Fiscal Years 2015 to 2024"},
    "nonimmigrants-table-27": {"topic_slug": "nonimmigrants", "topic": "Nonimmigrant Admissions", "sheet": "Table 27", "title": "I-94 Nonimmigrant Admissions by Region and Country of Citizenship: Fiscal Years 2015 to 2024"},
    "nonimmigrants-table-28": {"topic_slug": "nonimmigrants", "topic": "Nonimmigrant Admissions", "sheet": "Table 28", "title": "I-94 Nonimmigrant Admissions by Region and Country of Residence: Fiscal Years 2015 to 2024"},
    "nonimmigrants-table-29": {"topic_slug": "nonimmigrants", "topic": "Nonimmigrant Admissions", "sheet": "Table 29", "title": "I-94 Nonimmigrant Admissions by Selected Category of Admission and Region and Country of Citizenship: Fiscal Year 2024"},
    "nonimmigrants-table-30": {"topic_slug": "nonimmigrants", "topic": "Nonimmigrant Admissions", "sheet": "Table 30", "title": "I-94 Nonimmigrant Admissions by Selected Category of Admission, Age, and Sex: Fiscal Year 2024"},
    "nonimmigrants-table-31": {"topic_slug": "nonimmigrants", "topic": "Nonimmigrant Admissions", "sheet": "Table 31", "title": "I-94 Nonimmigrant Admissions by Selected Category of Admission and State or Territory of Destination: Fiscal Year 2024"},
    "nonimmigrants-table-32": {"topic_slug": "nonimmigrants", "topic": "Nonimmigrant Admissions", "sheet": "Table 32", "title": "I-94 Nonimmigrant Admissions by Selected Category of Admission and Month of Arrival: Fiscal Year 2024"},
    "nonimmigrants-table-33": {"topic_slug": "nonimmigrants", "topic": "Nonimmigrant Admissions", "sheet": "Table 33", "title": "I-94 Nonimmigrant Temporary Worker Admissions by Region and Country of Citizenship: Fiscal Year 2024"},
    "refugees-table-13": {"topic_slug": "refugees", "topic": "Refugees", "sheet": "Table 13", "title": "Refugee Arrivals: Fiscal Years 1980 to 2024"},
    "refugees-table-14": {"topic_slug": "refugees", "topic": "Refugees", "sheet": "Table 14", "title": "Refugee Arrivals by Region and Country of Nationality: Fiscal Years 2015 to 2024"},
    "refugees-table-15": {"topic_slug": "refugees", "topic": "Refugees", "sheet": "Table 15", "title": "Refugee Arrivals by Relationship to Principal Applicant and Sex, Age, and Marital Status: Fiscal Year 2024"},
}

SCHEMA = pa.schema([
    ("topic", pa.string()),
    ("table_label", pa.string()),
    ("title", pa.string()),
    ("section", pa.string()),
    ("parent_category", pa.string()),
    ("category", pa.string()),
    ("description", pa.string()),
    ("breakdown", pa.string()),
    ("value", pa.float64()),
    ("value_note", pa.string()),
])

# Non-numeric cell markers used in the yearbook (disclosure-suppressed, not
# applicable, rounds to zero, etc.).
_NULL_MARKERS = {"", "-", "--", "X", "(X)", "D", "(D)", "NA", "N/A", "*", "Z", "(Z)"}

# ---------------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------------


@transient_retry()
def _fetch(url: str) -> bytes:
    # DHS/Akamai currently rejects httpx HTTP/1.1 requests with 403 even when
    # browser headers are supplied. requests/urllib3 is accepted by the same
    # endpoint and still keeps the connector dependency-light.
    resp = _SESSION.get(url, headers=REQ_HEADERS, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.content


def _discover_workbooks() -> dict:
    """Scrape the yearbook page and return {topic_slug: xlsx_url} for every
    available topic workbook. The slug is the filename key (e.g.
    'lawful_permanent_residents' -> 'lawful-permanent-residents')."""
    html = _fetch(YEARBOOK_URL).decode("utf-8", errors="ignore")
    out = {}
    for href in re.findall(r'href="([^"]+\.xlsx)"', html):
        fn = href.rsplit("/", 1)[-1]
        m = re.search(r"ohss_yearbook_(.+?)_fy\d{4}", fn)
        if not m:
            continue
        slug = m.group(1).replace("_", "-")
        out.setdefault(slug, href if href.startswith("http") else BASE + href)
    return out


# ---------------------------------------------------------------------------
# Sheet parsing
# ---------------------------------------------------------------------------
_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"


def _footnote_map(content: bytes) -> dict:
    """{concatenated_text: text_without_superscript_runs} for every shared string
    whose footnote marker is a superscript run. openpyxl hands back the
    concatenation, so this maps it back to the clean label."""
    with zipfile.ZipFile(io.BytesIO(content)) as z:
        if "xl/sharedStrings.xml" not in z.namelist():
            return {}
        root = ET.fromstring(z.read("xl/sharedStrings.xml"))
    out = {}
    for si in root.findall(f"{_NS}si"):
        runs = si.findall(f"{_NS}r")
        if not runs:
            continue
        full, clean = [], []
        for r in runs:
            t = r.find(f"{_NS}t")
            text = t.text or "" if t is not None else ""
            full.append(text)
            props = r.find(f"{_NS}rPr")
            sup = props is not None and props.find(f"{_NS}vertAlign") is not None and \
                props.find(f"{_NS}vertAlign").get("val") == "superscript"
            if not sup:
                clean.append(text)
        f, c = "".join(full), "".join(clean)
        if c.strip() and c != f:
            out[f] = c
    return out


def _indent(row) -> float:
    """Indent level of a row's column-A cell. Blank cells carry no alignment."""
    if not row:
        return 0.0
    align = getattr(row[0], "alignment", None)
    return (align.indent or 0.0) if align is not None else 0.0


def _clean(value, footnotes: dict) -> str:
    """Cell value → display text: drop the superscript footnote marker, collapse
    the embedded newlines the yearbook uses to wrap long headers."""
    if value is None:
        return ""
    s = str(value)
    s = footnotes.get(s, s)
    return re.sub(r"\s+", " ", s).strip()


def _num_nonempty(row) -> int:
    return sum(1 for c in row if str(c if c is not None else "").strip() != "")


def _parse_num(raw: str):
    """Return (value, note). value is a float or None; note carries the raw
    string when it is a non-numeric marker."""
    t = raw.replace(",", "").replace("$", "").strip()
    if t in _NULL_MARKERS:
        return None, (raw.strip() or None)
    try:
        return float(t), None
    except ValueError:
        return None, raw.strip()


def _dedupe(headers):
    seen = {}
    out = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            out.append(f"{h} ({seen[h]})")
        else:
            seen[h] = 1
            out.append(h)
    return out


def _parse_sheet(rows, indents):
    """Melt a yearbook sheet into tidy long records:
    (section, parent_category, category, description, breakdown, value, note).

    `rows` are cleaned cell texts; `indents[i]` is the column-A indent level of
    row i, which is how the sheet encodes its row hierarchy."""
    # Header = first row with >= 2 non-empty cells (skips the 'Return to TOC'
    # link, the 'Table N.' marker, and the ALL-CAPS title which each occupy one
    # cell).
    header_idx = next((i for i, r in enumerate(rows) if _num_nonempty(r) >= 2), None)
    if header_idx is None:
        return []
    header = list(rows[header_idx])
    while header and header[-1] == "":
        header.pop()
    if not header:
        return []

    # The NISuppTables carry a 'Description' column between the row label and the
    # first observation: an attribute of the row, not a breakdown of it.
    has_desc = len(header) > 1 and header[1].lower() == "description"

    def split(cells):
        """row cells -> (label + observation cells, description)"""
        if not has_desc:
            return cells, ""
        desc = cells[1] if len(cells) > 1 else ""
        return [cells[0]] + list(cells[2:]), desc

    header, _ = split(header)

    # Repeated-block layout (e.g. 'Year','Number','Year','Number', ...): the
    # first header label reappears, marking the block width. These sheets are
    # flat year series — they carry no row hierarchy.
    period = None
    for j in range(1, len(header)):
        if header[j] == header[0] and header[0] != "":
            period = j
            break

    records = []
    data = range(header_idx + 1, len(rows))

    if period:
        base = _dedupe(header[:period])
        for i in data:
            r = rows[i]
            for c0 in range(0, len(r), period):
                chunk = r[c0:c0 + period]
                if _num_nonempty(chunk) == 0:
                    continue
                label = chunk[0]
                vals = chunk[1:len(base)]
                if not label or not any(v != "" for v in vals):
                    continue
                for k, colname in enumerate(base[1:]):
                    if k >= len(vals) or vals[k] == "":
                        continue
                    value, note = _parse_num(vals[k])
                    records.append((None, "", label, "", colname, value, note))
        return records

    base = _dedupe(header)
    section = None
    stack = []  # (indent, label) — the chain of ancestors of the current row
    for i in data:
        cells, desc = split(rows[i])
        label = cells[0]
        vals = cells[1:len(base)]
        has_vals = any(v != "" for v in vals)
        if not label and desc and has_vals:
            # The NISuppTables' grand-total row leaves the class code blank and
            # names itself in the Description column.
            label, desc = desc, ""
        if not label:
            continue
        if not has_vals:
            # A label with no values: a section header (REGION, AGE) or a
            # trailing footnote. Either way it opens a new hierarchy.
            section, stack = label, []
            continue
        indent = indents[i]
        while stack and stack[-1][0] >= indent:
            stack.pop()
        parent = stack[-1][1] if stack else ""
        stack.append((indent, label))
        for k, colname in enumerate(base[1:]):
            if k >= len(vals) or vals[k] == "":
                continue
            value, note = _parse_num(vals[k])
            records.append((section, parent, label, desc, colname, value, note))
    return records


def fetch_one(node_id: str) -> None:
    asset = node_id
    eid = node_id[len("dhs-"):]
    meta = ENTITY_META[eid]

    books = _discover_workbooks()
    url = books.get(meta["topic_slug"])
    if url is None:
        raise RuntimeError(
            f"workbook for topic '{meta['topic_slug']}' not found on {YEARBOOK_URL} "
            f"(available: {sorted(books)})"
        )

    content = _fetch(url)
    footnotes = _footnote_map(content)
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if meta["sheet"] not in wb.sheetnames:
        raise RuntimeError(
            f"sheet '{meta['sheet']}' not in workbook {url} (sheets: {wb.sheetnames})"
        )
    ws = wb[meta["sheet"]]
    grid, indents = [], []
    for row in ws.iter_rows():
        grid.append([_clean(c.value, footnotes) for c in row])
        indents.append(_indent(row))
    wb.close()

    records = _parse_sheet(grid, indents)
    if not records:
        raise RuntimeError(f"no data rows parsed from sheet '{meta['sheet']}' of {url}")

    n = len(records)
    table = pa.table(
        {
            "topic": [meta["topic"]] * n,
            "table_label": [meta["sheet"]] * n,
            "title": [meta["title"]] * n,
            "section": [r[0] for r in records],
            "parent_category": [r[1] for r in records],
            "category": [r[2] for r in records],
            "description": [r[3] for r in records],
            "breakdown": [r[4] for r in records],
            "value": [r[5] for r in records],
            "value_note": [r[6] for r in records],
        },
        schema=SCHEMA,
    )
    save_raw_parquet(table, asset)


DOWNLOAD_SPECS = [
    NodeSpec(id=f"dhs-{eid}", fn=fetch_one, kind="download")
    for eid in ENTITY_META
]
