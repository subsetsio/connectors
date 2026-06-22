import sys, io, zipfile, re
sys.path.insert(0, "src")
from subsets_utils import get
import openpyxl

def load_wb(content):
    try:
        return openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        zin = zipfile.ZipFile(io.BytesIO(content))
        out = io.BytesIO(); zout = zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED)
        for n in zin.namelist():
            if "drawing" in n.lower():
                continue
            data = zin.read(n)
            if n.endswith(".rels") or n.endswith(".xml"):
                data = re.sub(rb"<drawing[^>]*/>", b"", data)
                data = re.sub(rb"<Relationship[^>]*drawing[^>]*/>", b"", data)
            zout.writestr(n, data)
        zout.close(); out.seek(0)
        return openpyxl.load_workbook(out, data_only=True)

BASE = "https://live.laborstats.alaska.gov"

def dump(url, sheet=None, rows=9, cols=26, minr=1):
    print(f"\n\n##### {url}  sheet={sheet}")
    r = get(url, timeout=(10,120))
    print("status", r.status_code, "len", len(r.content))
    if r.status_code != 200: return
    wb = load_wb(r.content)
    print("sheets:", wb.sheetnames)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    print(f"dims rows={ws.max_row} cols={ws.max_column}")
    for ri, row in enumerate(ws.iter_rows(min_row=minr, max_row=minr+rows-1, max_col=cols, values_only=True), minr):
        cells = [f"{i+1}:{repr(v)}" for i, v in enumerate(row) if v is not None]
        print(f"R{ri}: " + " | ".join(cells))

dump(BASE + "/pop/estimates/data/AgeBySexBCA.xlsx", rows=9, cols=14)
dump(BASE + "/pop/estimates/data/AgeBySexPlace.xlsx", rows=9, cols=14)
dump(BASE + "/pop/estimates/data/TotalPopulationPlace.xlsx", rows=10, cols=12)
dump(BASE + "/pop/estimates/data/TotalPopulationTract.xlsx", rows=10, cols=12)
dump(BASE + "/pop/estimates/data/TotalPopulationANRC.xlsx", rows=10, cols=12)
dump(BASE + "/pop/estimates/data/TotalPopulationSchoolDistrict.xlsx", rows=10, cols=12)
dump(BASE + "/pop/estimates/data/AgeBySexByRaceAloneHispBCA.xlsx", sheet="2024", rows=9, cols=16)
dump(BASE + "/sites/default/files/2024-07/Statewide.xlsx", sheet="ProjComponentsOfChange", rows=10, cols=14)
dump(BASE + "/sites/default/files/2024-07/Alaska%20Native.xlsx", rows=6, cols=14)
dump(BASE + "/sites/default/files/2024-07/Boroughs%20and%20Census%20Areas.xlsx", rows=8, cols=14)
# QCEW sub-industry rows to see NAICS description indent
dump(BASE + "/sites/default/files/2026-05/Annual%20January%20to%20December%202025.xlsx", sheet="2025", minr=9, rows=6, cols=12)
