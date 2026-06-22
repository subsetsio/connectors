import sys, io, traceback, zipfile, re
sys.path.insert(0, "src")
import subsets_utils
from subsets_utils import get
import openpyxl

def load_wb(content):
    """Load xlsx, stripping broken drawing refs that crash openpyxl."""
    try:
        return openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        zin = zipfile.ZipFile(io.BytesIO(content))
        out = io.BytesIO(); zout = zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED)
        for n in zin.namelist():
            if "drawing" in n.lower():
                continue
            data = zin.read(n)
            if n.endswith(".rels") or n.endswith(".xml"):
                data = re.sub(rb"<drawing[^>]*/>", b"", data)
                data = re.sub(rb"<Relationship[^>]*drawing[^>]*/>", b"", data)
            zout.writestr(n, data)
        zout.close(); out.seek(0)
        return openpyxl.load_workbook(out, data_only=True)

BASE = "https://live.laborstats.alaska.gov"

def fetch(url):
    try:
        r = get(url, timeout=(10, 120))
        return r
    except Exception as e:
        print(f"  ERROR fetching: {e}")
        return None

def probe_csv(url, nlines=6):
    print(f"\n=== CSV: {url}")
    r = fetch(url)
    if r is None: return
    print(f"  status={r.status_code} ctype={r.headers.get('content-type')} len={len(r.content)}")
    if r.status_code != 200:
        return
    text = r.content.decode("utf-8", errors="replace")
    for i, line in enumerate(text.splitlines()[:nlines]):
        print(f"  L{i+1}: {line[:300]}")

def probe_xlsx(url, sheet_rows=10):
    print(f"\n=== XLSX: {url}")
    r = fetch(url)
    if r is None: return
    print(f"  status={r.status_code} ctype={r.headers.get('content-type')} len={len(r.content)}")
    if r.status_code != 200:
        return
    try:
        wb = load_wb(r.content)
    except Exception as e:
        print(f"  NOT a valid xlsx: {e}")
        print("  first bytes:", r.content[:120])
        return
    print(f"  sheets: {wb.sheetnames}")
    ws = wb[wb.sheetnames[0]]
    print(f"  primary sheet '{ws.title}' max_row={ws.max_row} max_col={ws.max_column}")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= sheet_rows: break
        cells = [("" if c is None else str(c))[:22] for c in row]
        print(f"  R{i+1}: {cells}")

def head(url):
    r = fetch(url)
    if r is None:
        print(f"  {url} -> ERR")
        return None
    print(f"  {url} -> {r.status_code} {r.headers.get('content-type')} {len(r.content)}b")
    return r

if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else ""
    args = sys.argv[2:]
    if cmd == "csv":
        for u in args: probe_csv(u)
    elif cmd == "xlsx":
        for u in args: probe_xlsx(u)
    elif cmd == "head":
        for u in args: head(u)
    else:
        print("usage: probe.py [csv|xlsx|head] URL...")
