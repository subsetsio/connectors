"""Alaska Department of Labor & Workforce Development (Research & Analysis).

The source has no API — it publishes statistical datasets as bulk CSV/XLSX files
at stable URLs under https://live.laborstats.alaska.gov. Each of the 14 entities
maps to one or more flat files that this module fetches, parses into a clean
long-format record list, and writes as ndjson. The XLSX files are
human-formatted (multi-row merged headers, spacer rows/columns, footnotes,
indented hierarchies, one-sheet-per-year, wide repeating blocks), so each parser
is tailored to its file family from observed layouts.

Strategy: stateless full re-pull every run (files are small, KB-low MB; no
incremental filter exists). Two entities (qcew, population-projections) discover
their current file URLs from the source's article pages because the per-release
URLs embed a publish-month folder and are not templatable; the rest use stable
known paths. A browser User-Agent is set per fetch because some HTML pages 403
non-browser agents (the file URLs themselves do not, but discovery pages do).

Coverage notes (documented, not silent): qcew publishes the two most recent
"Annual" statewide workbooks; population-projections covers statewide (low/middle/
high) and Alaska Native scenarios (the per-borough workbook is a 38-sheet book
skipped for now); nonfatal injuries and workplace fatalities publish only the
current survey year as BLS summary tables.
"""

import io
import re
import csv
import zipfile

import openpyxl
from subsets_utils import (
    NodeSpec,
    SqlNodeSpec,
    get,
    configure_http,
    transient_retry,
    save_raw_ndjson,
)

SLUG = "alaska-department-of-labor-and-workforce-development"
BASE = "https://live.laborstats.alaska.gov"
UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11,
    "december": 12,
}
MONTHYEAR = re.compile(
    r"^(January|February|March|April|May|June|July|August|September|October|"
    r"November|December)\s+\d{4}$"
)
_YEAR_RE = re.compile(r"(19|20)\d{2}")

# --------------------------------------------------------------------------- #
# Low-level helpers
# --------------------------------------------------------------------------- #


def _setup():
    """Set a browser User-Agent once per fetch process (some pages 403 else)."""
    configure_http(headers={"User-Agent": UA})


@transient_retry()
def _get_bytes(url: str) -> bytes:
    r = get(url, timeout=(15.0, 180.0))
    r.raise_for_status()
    return r.content


@transient_retry()
def _get_text(url: str) -> str:
    r = get(url, timeout=(15.0, 180.0))
    r.raise_for_status()
    return r.text


def load_wb(content: bytes):
    """Load an xlsx, stripping broken drawing refs that crash openpyxl on a
    few of the population workbooks."""
    try:
        return openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        zin = zipfile.ZipFile(io.BytesIO(content))
        out = io.BytesIO()
        zout = zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED)
        for n in zin.namelist():
            if "drawing" in n.lower():
                continue
            data = zin.read(n)
            if n.endswith(".rels") or n.endswith(".xml"):
                data = re.sub(rb"<drawing[^>]*/>", b"", data)
                data = re.sub(rb"<Relationship[^>]*drawing[^>]*/>", b"", data)
            zout.writestr(n, data)
        zout.close()
        out.seek(0)
        return openpyxl.load_workbook(out, data_only=True)


def _sheet_rows(wb, sheet) -> list:
    return list(wb[sheet].iter_rows(values_only=True))


def _C(rows, r, c):
    """1-based cell access; None if out of range."""
    if r < 1 or r > len(rows):
        return None
    row = rows[r - 1]
    if c < 1 or c > len(row):
        return None
    return row[c - 1]


_NULLS = {"", "-", "--", "*", "**", "n/a", "na", "(na)", "...", "(x)", "x", "(d)", "nd"}


def _num(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "").replace("$", "").replace("%", "")
    if s.lower() in _NULLS:
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return int(f) if f.is_integer() else f


def _txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _year(s):
    if not s:
        return None
    m = _YEAR_RE.search(str(s))
    return int(m.group(0)) if m else None


def _is_footnote(s):
    return bool(s) and s.lower().lstrip(" *").startswith(
        ("source", "note", "vintage", "table", "industry", "borough", "area name",
         "school district", "return", "**", "1 ", "2 ")
    )


def _csv_rows(text: str, header_match: str):
    """Return list[dict] from a CSV whose real header is the first row containing
    a cell equal to header_match (these files carry title/blank rows above it)."""
    reader = list(csv.reader(io.StringIO(text)))
    hdr_idx = None
    hm = header_match.strip().lower()
    for i, row in enumerate(reader):
        if any(c.strip().lower() == hm for c in row):
            hdr_idx = i
            break
    if hdr_idx is None:
        raise ValueError(f"CSV header '{header_match}' not found")
    header = [c.strip() for c in reader[hdr_idx]]
    out = []
    for row in reader[hdr_idx + 1:]:
        if not any(c.strip() for c in row):
            continue
        out.append({header[j]: (row[j] if j < len(row) else None) for j in range(len(header))})
    return out


# --------------------------------------------------------------------------- #
# CSV-based entities
# --------------------------------------------------------------------------- #


def fetch_labor_force_area(node_id: str) -> None:
    _setup()
    rows = _csv_rows(_get_text(BASE + "/labforce/csv/AKlaborforce.csv"), "Area Name")
    out = []
    for r in rows:
        out.append({
            "area_name": _txt(r.get("Area Name")),
            "area_type": _txt(r.get("Area Type")),
            "area_code": _txt(r.get("Area Code")),
            "year": _num(r.get("Year")),
            "month": _txt(r.get("month")),
            "period": _num(r.get("period")),
            "preliminary": _num(r.get("Preliminary if value is 1")),
            "labor_force": _num(r.get("Labor Force")),
            "employment": _num(r.get("Employment")),
            "unemployment": _num(r.get("Unemployment")),
            "unemployment_rate": _num(r.get("Unemployment Rate")),
        })
    save_raw_ndjson(out, node_id)


def fetch_wages_by_occupation(node_id: str) -> None:
    _setup()
    rows = _csv_rows(_get_text(BASE + "/wage/csv/Alaska_wages.csv"), "Occupation Title")
    out = []
    for r in rows:
        out.append({
            "soc": _txt(r.get("SOC")),
            "occupation_title": _txt(r.get("Occupation Title")),
            "employment": _num(r.get("Employment")),
            "mean_wage": _num(r.get("Mean")),
            "pct10": _num(r.get("10th")),
            "pct25": _num(r.get("25th")),
            "median_wage": _num(r.get("Median")),
            "pct75": _num(r.get("75th")),
            "pct90": _num(r.get("90th")),
            "mean_wage_rse": _num(r.get("Mean wage percent RSE")),
            "employment_rse": _num(r.get("Employment percent RSE")),
        })
    save_raw_ndjson(out, node_id)


def fetch_consumer_price_index(node_id: str) -> None:
    _setup()
    rows = _csv_rows(_get_text(BASE + "/cpi/csv/Alaska_CPI.csv"), "Period Num")
    out = []
    for r in rows:
        out.append({
            "area": _txt(r.get("Area")),
            "period": _txt(r.get("Period")),
            "period_num": _num(r.get("Period Num")),
            "year": _num(r.get("Year")),
            "cpi_index": _num(r.get("CPI-U Index Value")),
            "pct_change_12mo": _num(r.get("12-Month Percent Change")),
        })
    save_raw_ndjson(out, node_id)


def fetch_occupational_projections(node_id: str) -> None:
    _setup()
    rows = _csv_rows(_get_text(BASE + "/occupations/csv/alaska.csv"), "Occupation Code")
    out = []
    for r in rows:
        out.append({
            "occupation_code": _txt(r.get("Occupation Code")),
            "occupation_title": _txt(r.get("Occupation Title")),
            "base_year": _num(r.get("Base Year")),
            "projected_year": _num(r.get("Projected Year")),
            "base_employment": _num(r.get("Base Year Employment")),
            "projected_employment": _num(r.get("Projected Year Employment")),
            "numeric_change": _num(r.get("Numeric Change")),
            "percent_change": _num(r.get("Percentage Change")),
            "labor_force_exits": _num(r.get("Labor Force Exits")),
            "occupational_transfers": _num(r.get("Occupational Transfers")),
            "total_separations": _num(r.get("Total Occupational Separations")),
            "annual_openings": _num(r.get("Annual Average Openings")),
            "mean_hourly_wage": _num(r.get("hmean")),
        })
    save_raw_ndjson(out, node_id)


def fetch_ces_monthly_employment_by_industry(node_id: str) -> None:
    _setup()
    rows = _csv_rows(_get_text(BASE + "/labforce/csv/000000/01/ces.csv"), "areaname")
    out = []
    for r in rows:
        base = {
            "area_name": _txt(r.get("areaname")),
            "area_code": _txt(r.get("areacode")),
            "seasonally_adjusted": _num(r.get("adjusted")),
            "series_code": _txt(r.get("seriescode")),
            "industry": _txt(r.get("industry")),
            "year": _num(r.get("year")),
        }
        for mname, mnum in MONTHS.items():
            v = _num(r.get("emp_" + mname))
            if v is None:
                continue
            out.append({**base, "month": mnum, "employment": v})
    save_raw_ndjson(out, node_id)


# --------------------------------------------------------------------------- #
# Population estimates — XLSX families
# --------------------------------------------------------------------------- #

_TOTAL_FILES = [
    ("bca", "TotalPopulationBCA.xlsx"),
    ("place", "TotalPopulationPlace.xlsx"),
    ("tract", "TotalPopulationTract.xlsx"),
    ("anvsa", "TotalPopulationANVSA.xlsx"),
    ("anrc", "TotalPopulationANRC.xlsx"),
    ("subarea", "TotalPopulationBCASubarea.xlsx"),
    ("school_district", "TotalPopulationSchoolDistrict.xlsx"),
]


def _parse_total(rows, geo):
    # header row = first row carrying a month-year cell (e.g. "April 2020")
    hdr = None
    for i in range(1, len(rows) + 1):
        if any(isinstance(v, str) and MONTHYEAR.match(v.strip()) for v in rows[i - 1]):
            hdr = i
            break
    if hdr is None:
        return []
    hrow = rows[hdr - 1]
    area_col, fips_cols = None, []
    for c in range(1, len(hrow) + 1):
        v = hrow[c - 1]
        if isinstance(v, str):
            if "Name" in v:
                area_col = c
            if "FIPS" in v:
                fips_cols.append(c)
    if area_col is None:
        area_col = 1
    fips_col = max([c for c in fips_cols if c < area_col], default=None)
    # contiguous run of month-year columns after the area column (the population
    # estimate block; a blank-header gap separates the group-quarters block)
    period_cols, started = [], False
    for c in range(area_col + 1, len(hrow) + 1):
        v = hrow[c - 1]
        if isinstance(v, str) and MONTHYEAR.match(v.strip()):
            period_cols.append((c, v.strip()))
            started = True
        elif started:
            break
    out, last_area = [], None
    for r in range(hdr + 1, len(rows) + 1):
        a = _txt(_C(rows, r, area_col))
        if a and _is_footnote(a):
            break
        if a:
            last_area = a.lstrip()
        area = (a.lstrip() if a else last_area)
        if not area:
            continue
        fips = _txt(_C(rows, r, fips_col)) if fips_col else None
        for c, label in period_cols:
            pv = _num(_C(rows, r, c))
            if pv is None:
                continue
            out.append({
                "geo_level": geo, "area_name": area, "fips": fips,
                "period": label, "year": _year(label), "population": pv,
            })
    return out


def fetch_population_total(node_id: str) -> None:
    _setup()
    out = []
    for geo, fname in _TOTAL_FILES:
        try:
            wb = load_wb(_get_bytes(f"{BASE}/pop/estimates/data/{fname}"))
        except Exception:
            continue  # an optional geography file may be absent in a vintage
        out += _parse_total(_sheet_rows(wb, wb.sheetnames[0]), geo)
    save_raw_ndjson(out, node_id)


def _parse_coc_ak(rows):
    out = []
    for r in range(1, len(rows) + 1):
        period = _txt(_C(rows, r, 1))
        pop = _num(_C(rows, r, 2))
        if not period or not _YEAR_RE.search(period) or pop is None:
            continue
        out.append({
            "geo_level": "state", "area_name": "Alaska", "period": period,
            "population": pop, "population_change": _num(_C(rows, r, 3)),
            "growth_rate_pct": _num(_C(rows, r, 4)), "births": _num(_C(rows, r, 5)),
            "deaths": _num(_C(rows, r, 6)), "natural_increase": _num(_C(rows, r, 7)),
            "in_migration": _num(_C(rows, r, 8)), "out_migration": _num(_C(rows, r, 9)),
            "net_migration": _num(_C(rows, r, 10)),
        })
    return out


def _parse_coc_bca(rows):
    out, last_area = [], None
    for r in range(1, len(rows) + 1):
        area = _txt(_C(rows, r, 1))
        period = _txt(_C(rows, r, 2))
        if area and _is_footnote(area):
            break
        if area:
            last_area = area
        pop = _num(_C(rows, r, 3))
        if not period or pop is None:
            continue
        out.append({
            "geo_level": "region_bca", "area_name": last_area, "period": period,
            "population": pop, "population_change": _num(_C(rows, r, 4)),
            "growth_rate_pct": _num(_C(rows, r, 5)), "births": _num(_C(rows, r, 6)),
            "deaths": _num(_C(rows, r, 7)), "natural_increase": _num(_C(rows, r, 8)),
            "in_migration": None, "out_migration": None,
            "net_migration": _num(_C(rows, r, 9)),
        })
    return out


def fetch_population_components_of_change(node_id: str) -> None:
    _setup()
    out = []
    wb = load_wb(_get_bytes(BASE + "/pop/estimates/data/ComponentsOfChangeAK.xlsx"))
    out += _parse_coc_ak(_sheet_rows(wb, wb.sheetnames[0]))
    wb = load_wb(_get_bytes(BASE + "/pop/estimates/data/ComponentsOfChangeBCA.xlsx"))
    out += _parse_coc_bca(_sheet_rows(wb, wb.sheetnames[0]))
    save_raw_ndjson(out, node_id)


def _parse_agesex(rows, geo):
    sub = None
    for i in range(1, len(rows) + 1):
        vals = [str(v).strip().lower() for v in rows[i - 1] if v is not None]
        if "total" in vals and "male" in vals and "female" in vals:
            sub = i
            break
    if sub is None:
        return []
    srow = rows[sub - 1]
    age_col = area_col = bca_fips = place_fips = None
    blocks = []
    for c in range(1, len(srow) + 1):
        v = srow[c - 1]
        if not isinstance(v, str):
            continue
        s = v.strip()
        if s == "Age":
            age_col = c
        elif "Name" in s:
            area_col = c
        elif "FIPS" in s:
            if "lace" in s.lower():
                place_fips = c
            else:
                bca_fips = c
        elif s == "Total":
            blocks.append(c)
    if age_col is None or not blocks:
        return []
    out, last_area = [], None
    for r in range(sub + 1, len(rows) + 1):
        age = _txt(_C(rows, r, age_col))
        if age and _is_footnote(age):
            break
        if area_col:
            a = _txt(_C(rows, r, area_col))
            if a:
                last_area = a
        area = last_area if area_col else "Alaska"
        if not age:
            continue
        bf = _txt(_C(rows, r, bca_fips)) if bca_fips else None
        pf = _txt(_C(rows, r, place_fips)) if place_fips else None
        for bc in blocks:
            label = _txt(_C(rows, sub - 1, bc))
            for sex, off in (("Total", 0), ("Male", 1), ("Female", 2)):
                val = _num(_C(rows, r, bc + off))
                if val is None:
                    continue
                out.append({
                    "geo_level": geo, "area_name": area, "bca_fips": bf,
                    "place_fips": pf, "age": age, "period_label": label,
                    "year": _year(label), "sex": sex, "population": val,
                })
    return out


def fetch_population_age_sex(node_id: str) -> None:
    _setup()
    out = []
    for geo, fname in (("state", "AgeBySexAK.xlsx"), ("region_bca", "AgeBySexBCA.xlsx"),
                       ("place", "AgeBySexPlace.xlsx")):
        wb = load_wb(_get_bytes(f"{BASE}/pop/estimates/data/{fname}"))
        out += _parse_agesex(_sheet_rows(wb, wb.sheetnames[0]), geo)
    save_raw_ndjson(out, node_id)


def _parse_racehisp_ak(rows):
    hdr = None
    for i in range(1, len(rows) + 1):
        if any(isinstance(v, str) and MONTHYEAR.match(v.strip()) for v in rows[i - 1]):
            hdr = i
            break
    if hdr is None:
        return []
    hrow = rows[hdr - 1]
    period_cols = [(c, hrow[c - 1].strip()) for c in range(1, len(hrow) + 1)
                   if isinstance(hrow[c - 1], str) and MONTHYEAR.match(hrow[c - 1].strip())]
    out = []
    for r in range(hdr + 1, len(rows) + 1):
        race = _txt(_C(rows, r, 1))
        if not race:
            continue
        if _is_footnote(race):
            break
        for c, label in period_cols:
            v = _num(_C(rows, r, c))
            if v is None:
                continue
            out.append({
                "geo_level": "state", "area_name": "Alaska", "fips": None,
                "year": _year(label), "race": race, "population": v,
            })
    return out


_RACEHISP_BCA_COLS = {
    4: "Total", 6: "White", 7: "Alaska Native or American Indian",
    8: "Black or African American", 9: "Asian",
    10: "Native Hawaiian or Other Pacific Islander", 11: "Two or More Races",
    13: "Hispanic Origin (of any race)",
}


def _parse_racehisp_bca(wb):
    out = []
    for sheet in wb.sheetnames:
        if not sheet.strip().isdigit():
            continue
        year = int(sheet.strip())
        rows = _sheet_rows(wb, sheet)
        for r in range(1, len(rows) + 1):
            area = _txt(_C(rows, r, 2))
            if not area or _is_footnote(area):
                continue
            if _num(_C(rows, r, 4)) is None:
                continue
            fips = _txt(_C(rows, r, 1))
            for c, race in _RACEHISP_BCA_COLS.items():
                v = _num(_C(rows, r, c))
                if v is None:
                    continue
                out.append({
                    "geo_level": "region_bca", "area_name": area, "fips": fips,
                    "year": year, "race": race, "population": v,
                })
    return out


def fetch_population_race_hispanic(node_id: str) -> None:
    _setup()
    out = []
    wb = load_wb(_get_bytes(BASE + "/pop/estimates/data/RaceHispAK.xlsx"))
    out += _parse_racehisp_ak(_sheet_rows(wb, wb.sheetnames[0]))
    wb = load_wb(_get_bytes(BASE + "/pop/estimates/data/RaceHispBCA.xlsx"))
    out += _parse_racehisp_bca(wb)
    save_raw_ndjson(out, node_id)


def _parse_asrh_sheet(rows, geo, basis, year):
    sub = None
    for i in range(1, len(rows) + 1):
        vals = [str(v).strip().lower() for v in rows[i - 1] if v is not None]
        if "total" in vals and "male" in vals and "female" in vals:
            sub = i
            break
    if sub is None:
        return []
    srow = rows[sub - 1]
    age_col = area_col = fips_col = None
    blocks = []
    for c in range(1, len(srow) + 1):
        v = srow[c - 1]
        if not isinstance(v, str):
            continue
        s = v.strip()
        if s == "Age":
            age_col = c
        elif "Name" in s:
            area_col = c
        elif "FIPS" in s:
            fips_col = c
        elif s == "Total":
            blocks.append(c)
    if age_col is None or not blocks:
        return []
    spans = []
    for idx, bc in enumerate(blocks):
        end = blocks[idx + 1] if idx + 1 < len(blocks) else bc + 3
        spans.append((bc, end))

    def race_for(bc, end):
        parts = []
        for gr in (sub - 2, sub - 1):
            for c in range(bc, end):
                v = _C(rows, gr, c)
                if isinstance(v, str):
                    s = v.strip()
                    if s and s not in ("Total", "Male", "Female"):
                        parts.append(s)
        # dedupe duplicates (merged group cells repeat across the span) while
        # preserving order: "White White White" -> "White"
        parts = list(dict.fromkeys(parts))
        return " ".join(parts).strip() or "Total"

    out, last_area = [], None
    for r in range(sub + 1, len(rows) + 1):
        age = _txt(_C(rows, r, age_col))
        if age and _is_footnote(age):
            break
        if area_col:
            a = _txt(_C(rows, r, area_col))
            if a:
                last_area = a
        area = last_area if area_col else "Alaska"
        if not age:
            continue
        fips = _txt(_C(rows, r, fips_col)) if fips_col else None
        for bc, end in spans:
            race = race_for(bc, end)
            for sex, off in (("Total", 0), ("Male", 1), ("Female", 2)):
                val = _num(_C(rows, r, bc + off))
                if val is None:
                    continue
                out.append({
                    "geo_level": geo, "race_basis": basis, "area_name": area,
                    "fips": fips, "age": age, "year": year, "race": race,
                    "sex": sex, "population": val,
                })
    return out


_ASRH_FILES = [
    ("state", "alone", "AgeBySexByRaceAloneHispAK.xlsx"),
    ("state", "aic", "AgeBySexByRaceAICHispAK.xlsx"),
    ("region_bca", "alone", "AgeBySexByRaceAloneHispBCA.xlsx"),
    ("region_bca", "aic", "AgeBySexByRaceAICHispBCA.xlsx"),
]


def fetch_population_age_sex_race_hispanic(node_id: str) -> None:
    _setup()
    out = []
    for geo, basis, fname in _ASRH_FILES:
        wb = load_wb(_get_bytes(f"{BASE}/pop/estimates/data/{fname}"))
        for sheet in wb.sheetnames:
            if not sheet.strip().isdigit():
                continue
            out += _parse_asrh_sheet(_sheet_rows(wb, sheet), geo, basis, int(sheet.strip()))
    save_raw_ndjson(out, node_id)


# --------------------------------------------------------------------------- #
# Population projections — wide repeating blocks discovered from article page
# --------------------------------------------------------------------------- #


def _discover_xlsx(article_path):
    try:
        html = _get_text(BASE + article_path)
    except Exception:
        return []
    links = set(re.findall(r'href=["\']([^"\']+\.xlsx)', html, re.I))
    return [l if l.startswith("http") else BASE + l for l in links]


_AGE_BAND = re.compile(r"^\d{1,3}(-\d{1,3}|\+)?$")


def _is_age_band(age):
    """True for 5-year band / open-ended age labels (0-4, 85+, Total), False for
    period rows (2023-24) or stray numeric values bleeding in from the wide
    sheet's components-of-change block."""
    if age.strip().lower() == "total":
        return True
    if not _AGE_BAND.match(age.strip()):
        return False
    first = int(re.match(r"\d{1,3}", age.strip()).group(0))
    return first <= 120


def _parse_proj_sheet(rows, geo, scenario):
    hdr = None
    for i in range(1, len(rows) + 1):
        vals = [str(v).strip() for v in rows[i - 1] if v is not None]
        if "Age" in vals and "Total" in vals and "Male" in vals:
            hdr = i
            break
    if hdr is None:
        return []
    hrow = rows[hdr - 1]
    age_cols = [c for c in range(1, len(hrow) + 1)
                if isinstance(hrow[c - 1], str) and hrow[c - 1].strip() == "Age"]
    out = []
    for ac in age_cols:
        label = _txt(_C(rows, hdr - 1, ac))
        yr = _year(label)
        projected = bool(label and "proj" in label.lower())
        for r in range(hdr + 1, len(rows) + 1):
            age = _txt(_C(rows, r, ac))
            if not age or _is_footnote(age) or not _is_age_band(age):
                continue
            for sex, off in (("Total", 1), ("Male", 2), ("Female", 3)):
                v = _num(_C(rows, r, ac + off))
                if v is None:
                    continue
                out.append({
                    "geo_level": geo, "scenario": scenario, "period_label": label,
                    "year": yr, "is_projected": projected, "age": age,
                    "sex": sex, "population": v,
                })
    return out


def fetch_population_projections(node_id: str) -> None:
    _setup()
    links = _discover_xlsx("/article/alaska-population-projections")

    def pick(token, default):
        for l in links:
            if token.lower() in l.lower():
                return l
        return BASE + default

    out = []
    sw = pick("Statewide", "/sites/default/files/2024-07/Statewide.xlsx")
    wb = load_wb(_get_bytes(sw))
    for sheet, scn in (("AlaskaMiddle", "middle"), ("AlaskaLow", "low"),
                       ("AlaskaHigh", "high")):
        if sheet in wb.sheetnames:
            out += _parse_proj_sheet(_sheet_rows(wb, sheet), "statewide", scn)
    an = pick("Native", "/sites/default/files/2024-07/Alaska%20Native.xlsx")
    wb = load_wb(_get_bytes(an))
    out += _parse_proj_sheet(_sheet_rows(wb, wb.sheetnames[0]), "alaska_native", "middle")
    save_raw_ndjson(out, node_id)


# --------------------------------------------------------------------------- #
# QCEW — annual workbooks discovered from article page
# --------------------------------------------------------------------------- #


def _parse_qcew_sheet(rows, year):
    hdr = None
    for i in range(1, len(rows) + 1):
        vals = [str(v).strip().upper().replace("\n", " ") for v in rows[i - 1] if v is not None]
        if "AREANAME" in vals:
            hdr = i
            break
    if hdr is None:
        return []
    out = []
    for r in range(hdr + 1, len(rows) + 1):
        ac = _txt(_C(rows, r, 1))
        if not ac:
            continue
        avg_emp = _num(_C(rows, r, 24))
        est = _num(_C(rows, r, 11))
        wages = _num(_C(rows, r, 25))
        if avg_emp is None and est is None and wages is None:
            continue  # header / blank
        desc = None
        for c in range(4, 9):
            d = _txt(_C(rows, r, c))
            if d:
                desc = d
                break
        out.append({
            "area_code": ac, "area_name": _txt(_C(rows, r, 2)),
            "naics_code": _txt(_C(rows, r, 3)), "naics_description": desc,
            "year": year, "ownership": _txt(_C(rows, r, 10)),
            "establishments": est, "avg_employment": avg_emp,
            "total_wages": wages, "avg_monthly_wage": _num(_C(rows, r, 26)),
        })
    return out


def fetch_qcew(node_id: str) -> None:
    _setup()
    links = _discover_xlsx("/article/current-quarterly-census-employment-and-wages-qcew")
    annual = sorted({l for l in links if re.search(r"Annual.*\d{4}\.xlsx", l, re.I)})
    if not annual:
        annual = [BASE + "/sites/default/files/2026-05/"
                  "Annual%20January%20to%20December%202025.xlsx"]
    out = []
    for url in annual:
        wb = load_wb(_get_bytes(url))
        for sheet in wb.sheetnames:
            if re.fullmatch(r"\d{4}", sheet.strip()):
                out += _parse_qcew_sheet(_sheet_rows(wb, sheet), int(sheet.strip()))
    save_raw_ndjson(out, node_id)


# --------------------------------------------------------------------------- #
# BLS safety/health summary tables (current survey year)
# --------------------------------------------------------------------------- #

_INJURY_CASES = {
    3: "total_recordable", 4: "days_away_or_restricted_total",
    5: "days_away_from_work", 6: "job_transfer_or_restriction",
    8: "other_recordable",
}


def fetch_nonfatal_injuries_illnesses(node_id: str) -> None:
    _setup()
    wb = load_wb(_get_bytes(BASE + "/injill/xls/table01sum.xlsx"))
    rows = _sheet_rows(wb, wb.sheetnames[0])
    year = _year(_txt(_C(rows, 1, 1))) or _year(wb.sheetnames[0])
    out = []
    for r in range(1, len(rows) + 1):
        ind = _txt(_C(rows, r, 1))
        if not ind or _is_footnote(ind):
            continue
        vals = {c: _num(_C(rows, r, c)) for c in _INJURY_CASES}
        if all(v is None for v in vals.values()):
            continue
        naics = _txt(_C(rows, r, 2))
        for c, ct in _INJURY_CASES.items():
            if vals[c] is None:
                continue
            out.append({
                "industry": ind.strip(), "naics_code": naics, "year": year,
                "case_type": ct, "incidence_rate": vals[c],
            })
    save_raw_ndjson(out, node_id)


def fetch_workplace_fatalities(node_id: str) -> None:
    _setup()
    wb = load_wb(_get_bytes(BASE + "/fatal/xls/Table%20A-1.xlsx"))
    rows = _sheet_rows(wb, wb.sheetnames[0])
    year = _year(_txt(_C(rows, 1, 1))) or _year(wb.sheetnames[0])
    events = {3: "Total"}
    for c in range(4, 10):
        lab = _txt(_C(rows, 4, c))
        if lab:
            events[c] = re.sub(r"\d+$", "", lab).strip()
    out = []
    for r in range(1, len(rows) + 1):
        ind = _txt(_C(rows, r, 1))
        if not ind or _is_footnote(ind):
            continue
        vals = {c: _num(_C(rows, r, c)) for c in events}
        if all(v is None for v in vals.values()):
            continue
        naics = _txt(_C(rows, r, 2))
        for c, ev in events.items():
            if vals[c] is None:
                continue
            out.append({
                "industry": ind.strip(), "naics_code": naics, "year": year,
                "event_type": ev, "fatal_count": vals[c],
            })
    save_raw_ndjson(out, node_id)


# --------------------------------------------------------------------------- #
# Specs
# --------------------------------------------------------------------------- #

FETCH = {
    "labor-force-area": fetch_labor_force_area,
    "ces-monthly-employment-by-industry": fetch_ces_monthly_employment_by_industry,
    "wages-by-occupation": fetch_wages_by_occupation,
    "consumer-price-index": fetch_consumer_price_index,
    "occupational-projections": fetch_occupational_projections,
    "population-total": fetch_population_total,
    "population-components-of-change": fetch_population_components_of_change,
    "population-age-sex": fetch_population_age_sex,
    "population-race-hispanic": fetch_population_race_hispanic,
    "population-age-sex-race-hispanic": fetch_population_age_sex_race_hispanic,
    "population-projections": fetch_population_projections,
    "qcew": fetch_qcew,
    "nonfatal-injuries-illnesses": fetch_nonfatal_injuries_illnesses,
    "workplace-fatalities": fetch_workplace_fatalities,
}


def _did(entity: str) -> str:
    return f"{SLUG}-{entity}"


DOWNLOAD_SPECS = [
    NodeSpec(id=_did(e), fn=FETCH[e], kind="download") for e in FETCH
]

# Each transform reads the entity's raw ndjson view and publishes one table.
_SQL = {
    "labor-force-area": '''
        SELECT area_name, area_type, CAST(area_code AS VARCHAR) AS area_code,
               CAST(year AS INTEGER) AS year, month, CAST(period AS INTEGER) AS period,
               CAST(COALESCE(preliminary, 0) AS BOOLEAN) AS preliminary,
               CAST(labor_force AS BIGINT) AS labor_force,
               CAST(employment AS BIGINT) AS employment,
               CAST(unemployment AS BIGINT) AS unemployment,
               CAST(unemployment_rate AS DOUBLE) AS unemployment_rate
        FROM "{dep}"
        WHERE year IS NOT NULL
          AND (labor_force IS NOT NULL OR unemployment_rate IS NOT NULL)
    ''',
    "ces-monthly-employment-by-industry": '''
        SELECT area_name, CAST(area_code AS VARCHAR) AS area_code,
               CAST(COALESCE(seasonally_adjusted, 0) AS BOOLEAN) AS seasonally_adjusted,
               CAST(series_code AS VARCHAR) AS series_code, industry,
               CAST(year AS INTEGER) AS year, CAST(month AS INTEGER) AS month,
               CAST(employment AS BIGINT) AS employment
        FROM "{dep}"
        WHERE year IS NOT NULL AND month IS NOT NULL AND employment IS NOT NULL
    ''',
    "wages-by-occupation": '''
        SELECT CAST(soc AS VARCHAR) AS soc, occupation_title,
               CAST(employment AS BIGINT) AS employment,
               CAST(mean_wage AS DOUBLE) AS mean_wage,
               CAST(pct10 AS DOUBLE) AS pct10, CAST(pct25 AS DOUBLE) AS pct25,
               CAST(median_wage AS DOUBLE) AS median_wage,
               CAST(pct75 AS DOUBLE) AS pct75, CAST(pct90 AS DOUBLE) AS pct90,
               CAST(mean_wage_rse AS DOUBLE) AS mean_wage_rse,
               CAST(employment_rse AS DOUBLE) AS employment_rse
        FROM "{dep}"
        WHERE occupation_title IS NOT NULL
    ''',
    "consumer-price-index": '''
        SELECT area, period, CAST(period_num AS INTEGER) AS period_num,
               CAST(year AS INTEGER) AS year,
               CAST(cpi_index AS DOUBLE) AS cpi_index,
               CAST(pct_change_12mo AS DOUBLE) AS pct_change_12mo
        FROM "{dep}"
        WHERE year IS NOT NULL AND cpi_index IS NOT NULL
    ''',
    "occupational-projections": '''
        SELECT CAST(occupation_code AS VARCHAR) AS occupation_code, occupation_title,
               CAST(base_year AS INTEGER) AS base_year,
               CAST(projected_year AS INTEGER) AS projected_year,
               CAST(base_employment AS BIGINT) AS base_employment,
               CAST(projected_employment AS BIGINT) AS projected_employment,
               CAST(numeric_change AS BIGINT) AS numeric_change,
               CAST(percent_change AS DOUBLE) AS percent_change,
               CAST(labor_force_exits AS DOUBLE) AS labor_force_exits,
               CAST(occupational_transfers AS DOUBLE) AS occupational_transfers,
               CAST(total_separations AS DOUBLE) AS total_separations,
               CAST(annual_openings AS DOUBLE) AS annual_openings,
               CAST(mean_hourly_wage AS DOUBLE) AS mean_hourly_wage
        FROM "{dep}"
        WHERE occupation_title IS NOT NULL
    ''',
    "population-total": '''
        SELECT geo_level, area_name, CAST(fips AS VARCHAR) AS fips, period,
               CAST(year AS INTEGER) AS year, CAST(population AS BIGINT) AS population
        FROM "{dep}"
        WHERE area_name IS NOT NULL AND population IS NOT NULL
    ''',
    "population-components-of-change": '''
        SELECT geo_level, area_name, period,
               CAST(population AS BIGINT) AS population,
               CAST(population_change AS BIGINT) AS population_change,
               CAST(growth_rate_pct AS DOUBLE) AS growth_rate_pct,
               CAST(births AS BIGINT) AS births, CAST(deaths AS BIGINT) AS deaths,
               CAST(natural_increase AS BIGINT) AS natural_increase,
               CAST(in_migration AS BIGINT) AS in_migration,
               CAST(out_migration AS BIGINT) AS out_migration,
               CAST(net_migration AS BIGINT) AS net_migration
        FROM "{dep}"
        WHERE area_name IS NOT NULL AND population IS NOT NULL
    ''',
    "population-age-sex": '''
        SELECT geo_level, area_name, CAST(bca_fips AS VARCHAR) AS bca_fips,
               CAST(place_fips AS VARCHAR) AS place_fips, age, period_label,
               CAST(year AS INTEGER) AS year, sex,
               CAST(population AS BIGINT) AS population
        FROM "{dep}"
        WHERE area_name IS NOT NULL AND age IS NOT NULL AND population IS NOT NULL
    ''',
    "population-race-hispanic": '''
        SELECT geo_level, area_name, CAST(fips AS VARCHAR) AS fips,
               CAST(year AS INTEGER) AS year, race,
               CAST(population AS BIGINT) AS population
        FROM "{dep}"
        WHERE area_name IS NOT NULL AND race IS NOT NULL AND population IS NOT NULL
    ''',
    "population-age-sex-race-hispanic": '''
        SELECT geo_level, race_basis, area_name, CAST(fips AS VARCHAR) AS fips,
               age, CAST(year AS INTEGER) AS year, race, sex,
               CAST(population AS BIGINT) AS population
        FROM "{dep}"
        WHERE area_name IS NOT NULL AND age IS NOT NULL AND population IS NOT NULL
    ''',
    "population-projections": '''
        SELECT geo_level, scenario, period_label, CAST(year AS INTEGER) AS year,
               CAST(is_projected AS BOOLEAN) AS is_projected, age, sex,
               CAST(population AS BIGINT) AS population
        FROM "{dep}"
        WHERE year IS NOT NULL AND age IS NOT NULL AND population IS NOT NULL
    ''',
    "qcew": '''
        SELECT CAST(area_code AS VARCHAR) AS area_code, area_name,
               CAST(naics_code AS VARCHAR) AS naics_code, naics_description,
               CAST(year AS INTEGER) AS year, CAST(ownership AS VARCHAR) AS ownership,
               CAST(establishments AS BIGINT) AS establishments,
               CAST(avg_employment AS DOUBLE) AS avg_employment,
               CAST(total_wages AS BIGINT) AS total_wages,
               CAST(avg_monthly_wage AS DOUBLE) AS avg_monthly_wage
        FROM "{dep}"
        WHERE naics_code IS NOT NULL AND year IS NOT NULL
    ''',
    "nonfatal-injuries-illnesses": '''
        SELECT industry, CAST(naics_code AS VARCHAR) AS naics_code,
               CAST(year AS INTEGER) AS year, case_type,
               CAST(incidence_rate AS DOUBLE) AS incidence_rate
        FROM "{dep}"
        WHERE industry IS NOT NULL AND incidence_rate IS NOT NULL
    ''',
    "workplace-fatalities": '''
        SELECT industry, CAST(naics_code AS VARCHAR) AS naics_code,
               CAST(year AS INTEGER) AS year, event_type,
               CAST(fatal_count AS BIGINT) AS fatal_count
        FROM "{dep}"
        WHERE industry IS NOT NULL AND fatal_count IS NOT NULL
    ''',
}

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"{_did(e)}-transform",
        deps=[_did(e)],
        sql=_SQL[e].format(dep=_did(e)),
    )
    for e in FETCH
]
