import io
from subsets_utils import get, transient_retry
import openpyxl

@transient_retry(attempts=8, min_wait=5, max_wait=90)
def fetch(url):
    r = get(url, timeout=120)
    r.raise_for_status()
    return r.content

urls = {
 "MAIN":"https://cpds-data.org/wp-content/uploads/2026/05/cpds-1960-2023-update-2025.xlsx",
 "GOV":"https://cpds-data.org/wp-content/uploads/2025/06/government_composition_1960-2023_update_2025.xlsx",
}
for label,url in urls.items():
    content = fetch(url)
    print(f"\n=== {label} bytes={len(content)}")
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    print("sheets:", wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    print("ncols:", len(header))
    print("cols:", [str(c) for c in header])
    nrows=0
    for k,row in enumerate(it):
        nrows+=1
        if k<3: print(f"row{k}:", row[:8])
    print("datarows:", nrows)
