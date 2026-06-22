import io
from subsets_utils import get, transient_retry
import openpyxl
@transient_retry(attempts=8, min_wait=5, max_wait=90)
def fetch(url):
    r=get(url,timeout=120); r.raise_for_status(); return r.content
gov=fetch("https://cpds-data.org/wp-content/uploads/2025/06/government_composition_1960-2023_update_2025.xlsx")
wb=openpyxl.load_workbook(io.BytesIO(gov), read_only=True, data_only=True)
titles=set()
for cn in wb.sheetnames:
    rows=list(wb[cn].iter_rows(values_only=True))
    for i,r in enumerate(rows):
        if r and str(r[0]).strip()=="Year":
            titles.add(str(r[2]).strip()); break
print(sorted(titles))
