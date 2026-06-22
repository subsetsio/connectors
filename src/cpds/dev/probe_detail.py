import io
from subsets_utils import get, transient_retry
import openpyxl

@transient_retry(attempts=8, min_wait=5, max_wait=90)
def fetch(url):
    r = get(url, timeout=120); r.raise_for_status(); return r.content

# MAIN key check
main = fetch("https://cpds-data.org/wp-content/uploads/2026/05/cpds-1960-2023-update-2025.xlsx")
wb = openpyxl.load_workbook(io.BytesIO(main), read_only=True, data_only=True)
ws = wb["DATA"]
it = ws.iter_rows(values_only=True)
hdr = next(it)
rows=[r for r in it if r[0] is not None]
countries=set(r[1] for r in rows)
years=set(r[0] for r in rows)
pairs=set((r[1],r[0]) for r in rows)
print("MAIN non-null-year rows:", len(rows), "distinct countries:", len(countries), "year range:", min(years), max(years))
print("distinct (country,year) pairs:", len(pairs), "-> dup?", len(pairs)!=len(rows))
print("countries:", sorted(countries))

# GOV detail
gov = fetch("https://cpds-data.org/wp-content/uploads/2025/06/government_composition_1960-2023_update_2025.xlsx")
wb2 = openpyxl.load_workbook(io.BytesIO(gov), read_only=True, data_only=True)
ws2 = wb2["Australia"]
print("\n=== GOV Australia first 12 rows (cols 0-12):")
for i,row in enumerate(ws2.iter_rows(values_only=True)):
    if i>=12: break
    print(i, [str(c)[:14] if c is not None else "." for c in row[:13]])
