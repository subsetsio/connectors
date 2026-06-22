import re, io
from subsets_utils import get
import openpyxl

html = get("https://cpds-data.org/data/", timeout=60).text
links = re.findall(r'href="([^"]+\.xlsx)"', html)
print("XLSX LINKS FOUND:")
for l in sorted(set(links)):
    print(" ", l)

# inspect main cpds xlsx
main_url = next(l for l in links if "cpds" in l.lower() and "government" not in l.lower())
gov_url = next(l for l in links if "government" in l.lower())
print("\nMAIN:", main_url)
print("GOV :", gov_url)

for label, url in [("MAIN", main_url), ("GOV", gov_url)]:
    r = get(url, timeout=120)
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    print(f"\n=== {label} sheets:", wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    print(f"  ncols={len(header)} first 25 cols:", header[:25])
    r1 = next(rows); r2 = next(rows)
    print("  row1:", r1[:12])
    print("  row2:", r2[:12])
