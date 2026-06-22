"""CPDS — Comparative Political Data Set (Armingeon, Engler, Leemann, Weisstanner).

Two published subsets, each a single annually-updated .xlsx bulk download:

  * cpds-main                    — the main country-year panel (DATA sheet, ~315
                                   political/institutional/socio-economic columns).
  * cpds-government-composition  — the Government Composition supplement: one
                                   workbook sheet per country, a two-row merged
                                   header flattened into a long table.

Download URLs embed the WordPress upload month + annual 'update_YYYY' suffix and
change with each release, so we discover the current links by scraping the
canonical landing page rather than hard-coding a dated path. The host is
WordPress.com and intermittently answers 429 — handled by @transient_retry.
"""
from __future__ import annotations

import io
import re

import pyarrow as pa

from subsets_utils import NodeSpec, SqlNodeSpec, get, transient_retry, save_raw_parquet, save_raw_ndjson

DATA_PAGE = "https://cpds-data.org/data/"


@transient_retry(attempts=8, min_wait=5, max_wait=90)
def _fetch_bytes(url: str) -> bytes:
    r = get(url, timeout=180)
    r.raise_for_status()
    return r.content


def _xlsx_links() -> list[str]:
    """Absolute .xlsx links on the CPDS data landing page."""
    html = _fetch_bytes(DATA_PAGE).decode("utf-8", "replace")
    links = re.findall(r'href=["\']([^"\']+\.xlsx)["\']', html, flags=re.IGNORECASE)
    out: list[str] = []
    seen = set()
    for href in links:
        if href.startswith("//"):
            href = "https:" + href
        if href not in seen:
            seen.add(href)
            out.append(href)
    return out


def _pick(links: list[str], *, must: tuple[str, ...], forbid: tuple[str, ...]) -> str:
    cands = [
        l for l in links
        if all(m in l.lower() for m in must) and not any(f in l.lower() for f in forbid)
    ]
    if not cands:
        raise RuntimeError(
            f"no xlsx link on {DATA_PAGE} matching must={must} forbid={forbid}; "
            f"found {links}"
        )
    # Prefer the longest path (the dated upload URL) over any short alias.
    return sorted(cands, key=len, reverse=True)[0]


# ---- cpds-main -----------------------------------------------------------

def fetch_main(node_id: str) -> None:
    import pandas as pd

    links = _xlsx_links()
    url = _pick(links, must=("cpds", ".xlsx"),
                forbid=("government", "codebook", "by_government", "connection"))
    df = pd.read_excel(io.BytesIO(_fetch_bytes(url)), sheet_name="DATA", engine="openpyxl")

    # Drop the repeated section-separator key columns (year_01.., country_01..);
    # the canonical keys are the leading `year` and `country`.
    df = df.loc[:, ~df.columns.astype(str).str.match(r"^(year|country)_\d+$")]

    # Real data rows only — the sheet carries a long trailing run of blank rows.
    df = df[df["year"].notna()].copy()
    df["year"] = df["year"].astype("int64")

    # Object columns -> nullable pandas string so pyarrow gets a real string type
    # (never a null-typed column, which the Delta writer rejects).
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype("string")

    table = pa.Table.from_pandas(df, preserve_index=False)
    save_raw_parquet(table, node_id)


# ---- cpds-government-composition -----------------------------------------

_HEAD_TITLES = {"prime_minister", "president"}


def _slug(text: str) -> str:
    text = text.replace("%", "pct")
    text = re.sub(r"[^0-9a-zA-Z]+", "_", text).strip("_").lower()
    return text


def _gov_columns(top: tuple, sub: tuple) -> list[str | None]:
    """Flatten the two-row, group-merged header into one name per column.

    A group label on `top` forward-fills across its `sub` leaf columns
    (e.g. 'Number of Cabinet Posts' + {Right,Center,Left,Total}); standalone
    columns keep their single label; pure-spacer columns become None (dropped).
    """
    raw: list[str | None] = []
    cur: str | None = None
    width = max(len(top), len(sub))
    for j in range(width):
        t = top[j] if j < len(top) else None
        s = sub[j] if j < len(sub) else None
        if t is not None:
            cur = str(t).strip()
        if s is not None:
            raw.append(f"{cur} {str(s).strip()}")
        elif t is not None:
            raw.append(cur)
        else:
            raw.append(None)

    names: list[str | None] = []
    counts: dict[str, int] = {}
    for r in raw:
        if r is None:
            names.append(None)
            continue
        name = _slug(r)
        if name in _HEAD_TITLES:
            name = "head_of_government"
        if name in counts:
            counts[name] += 1
            name = f"{name}_{counts[name]}"
        else:
            counts[name] = 1
        names.append(name)
    return names


def _jsonable(v):
    if v is None:
        return None
    if hasattr(v, "isoformat"):  # datetime / date
        return v.isoformat()
    return v


def fetch_government_composition(node_id: str) -> None:
    import openpyxl

    links = _xlsx_links()
    url = _pick(links, must=("government_composition", ".xlsx"), forbid=("codebook",))
    wb = openpyxl.load_workbook(io.BytesIO(_fetch_bytes(url)), read_only=True, data_only=True)

    rows_out: list[dict] = []
    for country in wb.sheetnames:
        ws = wb[country]
        sheet_rows = list(ws.iter_rows(values_only=True))
        hidx = next(
            (i for i, r in enumerate(sheet_rows)
             if r and r[0] is not None and str(r[0]).strip() == "Year"),
            None,
        )
        if hidx is None:
            raise RuntimeError(f"{country}: no 'Year' header row found")
        cols = _gov_columns(sheet_rows[hidx], sheet_rows[hidx + 1])
        for r in sheet_rows[hidx + 2:]:
            if not r or r[0] is None:
                continue  # blank / spacer data row
            rec: dict = {"country": country}
            for j, name in enumerate(cols):
                if name is None:
                    continue
                rec[name] = _jsonable(r[j] if j < len(r) else None)
            rows_out.append(rec)

    save_raw_ndjson(rows_out, node_id)


DOWNLOAD_SPECS = [
    NodeSpec(id="cpds-main", fn=fetch_main, kind="download"),
    NodeSpec(id="cpds-government-composition", fn=fetch_government_composition, kind="download"),
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="cpds-main-transform",
        deps=("cpds-main",),
        sql='SELECT * FROM "cpds-main" ORDER BY country, year',
    ),
    SqlNodeSpec(
        id="cpds-government-composition-transform",
        deps=("cpds-government-composition",),
        sql='SELECT * FROM "cpds-government-composition" ORDER BY country, year',
    ),
]
