"""Banco Central de Bolivia — Boletin Estadistico connector.

Mechanism: `boletin_estadistico` (bulk per-table xlsx). The bulletin listing page
(static HTML) enumerates ~166 individual xlsx tables under a single monthly
date-stamped directory `/webdocs/publicacionesbcb/YYYY/MM/DD/<code>.xlsx`. The
numeric table codes (01_01, 02_01A, 15_01, ...) are the stable subset identity;
the date directory churns each month, so every fetch re-derives the current URL
from the listing page rather than hardcoding a path.

Shape: stateless full re-pull (decision shape 1). Each table is a small xlsx
(tens of KB, a few thousand populated cells); re-fetching the whole corpus every
refresh is cheap, and the bulletin is republished as a whole each month with
revisions, so there is no usable incremental delta filter anyway.

Raw format: these are human-formatted spreadsheets — scattered title/header
bands, merged cells, indented category hierarchies, pivot layouts — and every
one of the 163 tables has a different layout. No single semantic parser
generalizes across them, and a misparsing heuristic would silently emit garbage.
So each table is extracted to a faithful, uniform LONG CELL GRID: one row per
populated cell, carrying its (row, col) coordinate plus the numeric and/or text
value. This is type-stable (good for parquet), always non-empty, and lossless —
a consumer can reconstruct any table from its cells.
"""

import io

import pyarrow as pa
from subsets_utils import NodeSpec, SqlNodeSpec, get, save_raw_parquet, transient_retry
from constants import ENTITY_IDS

SLUG = "banco-central-de-bolivia"
LISTING_URL = "https://www.bcb.gob.bo/?q=pub_boletin-estadistico"
BASE = "https://www.bcb.gob.bo"

SCHEMA = pa.schema([
    ("table_code", pa.string()),
    ("sheet", pa.string()),
    ("row", pa.int32()),
    ("col", pa.int32()),
    ("value_num", pa.float64()),
    ("value_text", pa.string()),
])


def _norm(code: str) -> str:
    """Normalize a bulletin code (e.g. '02_01A') to the slug-suffix form used in
    spec ids ('02-01a')."""
    return code.lower().replace("_", "-")


@transient_retry()
def _get(url: str):
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp


def _resolve_url(suffix: str) -> tuple[str, str]:
    """Resolve the current xlsx URL for a table whose slug-suffix is `suffix`.

    Parses the static bulletin listing page for the current monthly directory.
    Returns (url, file_code). Raises KeyError if the code is absent from the
    listing (a real source change worth failing on).
    """
    import re

    html = _get(LISTING_URL).text
    pat = re.compile(r'href="(/webdocs/publicacionesbcb/\d{4}/\d{2}/\d{2}/([^"/]+?)\.xlsx)"')
    mapping = {}
    for href, code in pat.findall(html):
        mapping[_norm(code)] = (BASE + href, code)
    if suffix not in mapping:
        raise KeyError(
            f"table suffix {suffix!r} not found in bulletin listing "
            f"({len(mapping)} tables listed) — source layout may have changed"
        )
    return mapping[suffix]


def _coerce(v):
    """Map a cell value to (value_num, value_text). Numbers -> value_num;
    dates/strings -> value_text (with a conservative numeric parse attempt)."""
    import datetime

    if v is None:
        return None, None
    if isinstance(v, bool):
        return None, ("true" if v else "false")
    if isinstance(v, (int, float)):
        return float(v), None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return None, v.isoformat()
    s = str(v).strip()
    if not s:
        return None, None
    num = None
    try:
        num = float(s)
    except (ValueError, TypeError):
        num = None
    return num, s


def fetch_one(node_id: str) -> None:
    asset = node_id  # the runtime passes the spec id; it IS the asset name
    suffix = node_id[len(SLUG) + 1:]  # strip "banco-central-de-bolivia-"
    url, code = _resolve_url(suffix)

    import openpyxl

    content = _get(url).content
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    rows = []
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for r_idx, r in enumerate(ws.iter_rows(values_only=True)):
                for c_idx, v in enumerate(r):
                    if v is None:
                        continue
                    num, text = _coerce(v)
                    if num is None and text is None:
                        continue
                    rows.append({
                        "table_code": code,
                        "sheet": sheet,
                        "row": r_idx,
                        "col": c_idx,
                        "value_num": num,
                        "value_text": text,
                    })
    finally:
        wb.close()

    if not rows:
        raise AssertionError(f"{asset}: parsed 0 populated cells from {url}")

    table = pa.Table.from_pylist(rows, schema=SCHEMA)
    save_raw_parquet(table, asset)


DOWNLOAD_SPECS = [
    NodeSpec(
        id=f"{SLUG}-{_norm(eid)}",
        fn=fetch_one,
        kind="download",
    )
    for eid in ENTITY_IDS
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"{s.id}-transform",
        deps=[s.id],
        sql=f'''
            SELECT
                table_code,
                sheet,
                CAST(row AS INTEGER) AS row,
                CAST(col AS INTEGER) AS col,
                value_num,
                value_text
            FROM "{s.id}"
            WHERE value_num IS NOT NULL OR value_text IS NOT NULL
        ''',
    )
    for s in DOWNLOAD_SPECS
]
