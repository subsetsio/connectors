import io
import pyarrow as pa
import sys
sys.path.insert(0, "src")
sys.path.insert(0, "src/nodes")
import banco_central_de_bolivia as m

# Exercise resolve + parse for a couple of entities WITHOUT writing raw.
for suffix in ["01-01", "02-01a", "15-01"]:
    url, code = m._resolve_url(suffix)
    content = m._get(url).content
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for r_idx, r in enumerate(ws.iter_rows(values_only=True)):
            for c_idx, v in enumerate(r):
                if v is None:
                    continue
                num, text = m._coerce(v)
                if num is None and text is None:
                    continue
                rows.append({"table_code": code, "sheet": sheet, "row": r_idx,
                             "col": c_idx, "value_num": num, "value_text": text})
    wb.close()
    t = pa.Table.from_pylist(rows, schema=m.SCHEMA)
    import pyarrow.compute as pc
    nnum = pc.sum(pc.is_valid(t.column("value_num"))).as_py()
    print(f"{suffix} code={code} url=...{url[-30:]} cells={len(t)} numeric={nnum}")
