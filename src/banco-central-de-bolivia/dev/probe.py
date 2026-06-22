import io, re
from subsets_utils import get
import openpyxl

LISTING = "https://www.bcb.gob.bo/?q=pub_boletin-estadistico"
html = get(LISTING, timeout=(10,90)).text
# map code -> href
pat = re.compile(r'href="(/webdocs/publicacionesbcb/\d{4}/\d{2}/\d{2}/([^"/]+?)\.xlsx)"')
m = {code: href for href, code in pat.findall(html)}
print("num files in listing:", len(m))
print("sample codes:", list(m)[:5])

for code in ["01_01", "15_01", "11_05", "14_01", "09_01"]:
    href = m.get(code)
    if not href:
        print(f"\n### {code}: NOT IN LISTING"); continue
    url = "https://www.bcb.gob.bo" + href
    content = get(url, timeout=(10,120)).content
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    print(f"\n### {code}  url={url}")
    print("sheets:", wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    print("dims rows:", len(rows), "maxcol:", max((len(r) for r in rows), default=0))
    for i, r in enumerate(rows[:18]):
        # trim trailing None
        rr = list(r)
        while rr and rr[-1] is None: rr.pop()
        print(i, rr[:12])
    wb.close()
