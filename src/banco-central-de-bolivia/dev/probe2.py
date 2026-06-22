import io, re
from subsets_utils import get
import openpyxl

html = get("https://www.bcb.gob.bo/?q=pub_boletin-estadistico", timeout=(10,90)).text
pat = re.compile(r'href="(/webdocs/publicacionesbcb/\d{4}/\d{2}/\d{2}/([^"/]+?)\.xlsx)"')
m = {code: href for href, code in pat.findall(html)}

# Stress-test the wide-column file and a multi-sheet check; count populated cells.
for code in ["14_01", "01_01", "11_05"]:
    url = "https://www.bcb.gob.bo" + m[code]
    wb = openpyxl.load_workbook(io.BytesIO(get(url, timeout=(10,120)).content), data_only=True, read_only=True)
    total = 0
    maxc = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in ws.iter_rows(values_only=True):
            for ci, v in enumerate(r):
                if v is not None:
                    total += 1
                    if ci > maxc: maxc = ci
    print(code, "sheets", wb.sheetnames, "populated_cells", total, "max_populated_col", maxc)
    wb.close()
