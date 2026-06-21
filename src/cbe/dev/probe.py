import io, re
from subsets_utils import get, configure_http
import openpyxl

configure_http(headers={
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
})
BASE="https://www.cbe.org.eg"
CATS={"Inflation":"706A9057F8454F7284BE8143070D88C4","GDP":"DEF6421CA1354B128A1113D7A5BBFC66",
      "Interest Rates":"909707CDAD5C47529817D6146659E054","BOP":"232131B16F15454BB1E1933B2BFEB041"}

def cat_hrefs(guid):
    url=f"{BASE}/en/economic-research/time-series/downloadlist?category={guid}"
    h=get(url,timeout=(10,120)).text
    return re.findall(r'href="(/-/media/project/cbe/listing/time-series/[^"]+?\.xlsx)"',h)

def dump(url, label, maxr=22, maxc=10):
    r=get(BASE+url, timeout=(10,120))
    print(f"\n===== {label} =====\n{url}\nctype={r.headers.get('content-type')} bytes={len(r.content)}")
    wb=openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    print("sheets:", wb.sheetnames)
    ws=wb[wb.sheetnames[0]]
    for i,row in enumerate(ws.iter_rows(min_row=1,max_row=maxr,max_col=maxc,values_only=True)):
        cells=[("" if v is None else str(v)[:16]) for v in row]
        print(f"r{i+1:2d}|"+" | ".join(cells))

for cat,guid in CATS.items():
    hrefs=cat_hrefs(guid)
    # pick a monthly-ish recent file per dataset
    by_ds={}
    for h in hrefs:
        ds=h.split('/')[-2]
        by_ds.setdefault(ds,[]).append(h)
    print(f"\n######## CATEGORY {cat}: datasets={list(by_ds)}")
    for ds,files in by_ds.items():
        files.sort()
        dump(files[-1], f"{cat} / {ds} / {files[-1].split('/')[-1]}")
        break  # one dataset per category for now
