"""Generic CBE transposed-matrix XLSX parser (dev copy for iteration)."""
import io, re, datetime
import openpyxl

MONTHS={m:i for i,m in enumerate(["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"],1)}
ARAB=re.compile(r'[؀-ۿ]')
LATIN=re.compile(r'[A-Za-z]')
NUMRE=re.compile(r'^-?[\d,]+(\.\d+)?$')
YEAR=re.compile(r'(?:19|20)\d{2}')
QRE=re.compile(r'\bQ\s*([1-4])\b', re.I)

def month_in(tok):
    for w in re.split(r'[\s./,()]+', tok or ''):
        mn=MONTHS.get(w.strip().lower()[:3])
        if mn: return mn
    return None
def is_num(v):
    if isinstance(v,bool): return False
    if isinstance(v,(int,float)): return True
    return isinstance(v,str) and bool(NUMRE.match(v.strip()))
def to_float(v):
    if isinstance(v,bool): return None
    if isinstance(v,(int,float)): return float(v)
    if isinstance(v,str):
        try: return float(v.strip().replace(',',''))
        except: return None
    return None
def clean(v):
    if v is None: return None
    s=re.sub(r'\s+',' ',str(v)).strip()
    return s or None

def read_grid(ws,maxr=500,maxc=80):
    return [list(r) for r in ws.iter_rows(min_row=1,max_row=maxr,max_col=maxc,values_only=True)]

def _period_score(row):
    """Count cells that ARE a period token (short cell that is a month name or Qn).
    Long free-text footnote cells that merely contain a month word score 0."""
    s=0
    for c in row:
        cv=clean(c)
        if cv and len(cv)<=18 and (month_in(cv) or QRE.search(cv)):
            s+=1
    return s

def _year_only_score(row):
    s=0
    for c in row:
        cv=clean(c)
        if cv and len(cv)<=12 and YEAR.search(cv):
            s+=1
    return s

def _header_bottom(grid, scan=14):
    """Index of the last row of the period-header band (densest period-token row near
    the top; ties pick the lower row). Falls back to densest bare-year row, else None."""
    n=min(len(grid),scan)
    best_i,best=None,0
    for r in range(n):
        sc=_period_score(grid[r])
        if sc and sc>=best:
            best,best_i=sc,r
    if best_i is not None:
        return best_i
    best_i,best=None,0
    for r in range(n):
        sc=_year_only_score(grid[r])
        if sc and sc>=best:
            best,best_i=sc,r
    return best_i

def _ffill(grid, hdr_rows, c0, c1):
    filled={}
    for r in hdr_rows:
        row=grid[r]; last=None; o={}
        for c in range(c0,c1):
            v=clean(row[c]) if c<len(row) else None
            if v is not None: last=v
            o[c]=last
        filled[r]=o
    return filled

def _period(tokens, file_y0, file_y1, freq_tag):
    toks=[t for t in tokens if t]
    if not toks: return None
    joined=" | ".join(toks)
    years=[int(y) for y in YEAR.findall(joined)]
    mon=None; mon_year=None
    for t in toks:
        m=month_in(t)
        if m:
            mon=m
            ys=YEAR.findall(t)
            if ys: mon_year=int(ys[0])
            break
    qm=QRE.search(joined); quarter=int(qm.group(1)) if qm else None
    if not years and mon is None and quarter is None: return None
    dim_toks=[t for t in toks if not (YEAR.search(t) or month_in(t) or QRE.search(t))]
    dimension=" | ".join(dim_toks) or None
    year=None; date=None
    if mon is not None:
        year=mon_year or (years[0] if years else None)
        if year: date=datetime.date(year,mon,1)
    elif quarter is not None:
        y0=years[0] if years else file_y0
        if y0:
            mo,off={1:(7,0),2:(10,0),3:(1,1),4:(4,1)}[quarter]
            date=datetime.date(y0+off,mo,1); year=y0
    elif years:
        y0=years[0]
        fiscal = file_y1 and file_y0 and file_y1!=file_y0
        date=datetime.date(y0,7,1) if fiscal else datetime.date(y0,1,1)
        year=y0
    freq=freq_tag or ("monthly" if mon is not None else "quarterly" if quarter else "annual" if years else None)
    return {"period_label":joined,"dimension":dimension,"year":year,"date":date,"freq":freq}

def parse_sheet(grid, file_y0, file_y1, freq_tag):
    nrows=len(grid); ncols=max((len(r) for r in grid),default=0)
    if not ncols: return []
    period_bottom=_header_bottom(grid)
    if period_bottom is None: return []
    # Extend the header band past any sub-dimension rows (e.g. Public/Private/Total)
    # that sit below the period row: a real data row has >=2 numeric value cells.
    first_data=period_bottom+1
    for r in range(period_bottom+1, min(nrows, period_bottom+6)):
        ncount=sum(1 for c in range(ncols) if c<len(grid[r]) and is_num(grid[r][c]))
        if ncount>=2:
            first_data=r; break
    else:
        first_data=period_bottom+1
    hdr_rows=list(range(0,first_data))
    valcols=set()
    for r in range(first_data,nrows):
        for c in range(ncols):
            if c<len(grid[r]) and is_num(grid[r][c]): valcols.add(c)
    if not valcols: return []
    c0,c1=min(valcols),max(valcols)+1
    filled=_ffill(grid,hdr_rows,c0,c1)
    period_cols={}
    for c in range(c0,c1):
        p=_period([filled[r].get(c) for r in hdr_rows],file_y0,file_y1,freq_tag)
        if p: period_cols[c]=p
    if not period_cols: return []
    first_pc=min(period_cols)
    label_cols=[c for c in range(0,first_pc)] or [0]
    def score(c,rx): return sum(1 for r in range(first_data,nrows) if c<len(grid[r]) and clean(grid[r][c]) and rx.search(str(grid[r][c])))
    en_col=max(label_cols,key=lambda c:score(c,LATIN),default=None)
    ar_col=max(label_cols,key=lambda c:score(c,ARAB),default=None)
    out=[]
    for r in range(first_data,nrows):
        en=clean(grid[r][en_col]) if en_col is not None and en_col<len(grid[r]) else None
        ar=clean(grid[r][ar_col]) if ar_col is not None and ar_col<len(grid[r]) else None
        label=en or ar
        if not label: continue
        ll=(en or "").lower()
        if ll.startswith(("source",":")) or "title_en" in ll: continue
        for c,p in period_cols.items():
            if c>=len(grid[r]): continue
            val=to_float(grid[r][c])
            if val is None: continue
            out.append({"indicator_en":en,"indicator_ar":ar,"dimension":p["dimension"],
                        "period_label":p["period_label"],"year":p["year"],"date":p["date"],
                        "frequency":p["freq"],"value":val})
    return out

def parse_workbook(content, file_y0, file_y1, freq_tag):
    wb=openpyxl.load_workbook(io.BytesIO(content),read_only=True,data_only=True)
    rows=[]
    for sn in wb.sheetnames:
        try: rows+=parse_sheet(read_grid(wb[sn]),file_y0,file_y1,freq_tag)
        except Exception: pass
    return rows
