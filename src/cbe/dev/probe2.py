import io, re
from subsets_utils import get, configure_http
import openpyxl
configure_http(headers={"User-Agent":"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36","Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8","Accept-Language":"en-US,en;q=0.9"})
BASE="https://www.cbe.org.eg"
def hrefs(g):
    h=get(f"{BASE}/en/economic-research/time-series/downloadlist?category={g}",timeout=(10,120)).text
    return re.findall(r'href="(/-/media/project/cbe/listing/time-series/[^"]+?\.xlsx)"',h)
def dump(url,maxr=20,maxc=11):
    r=get(BASE+url,timeout=(10,120))
    wb=openpyxl.load_workbook(io.BytesIO(r.content),read_only=True,data_only=True)
    print(f"\n== {url.split('/')[-1]} sheets={wb.sheetnames}")
    ws=wb[wb.sheetnames[0]]
    for i,row in enumerate(ws.iter_rows(min_row=1,max_row=maxr,max_col=maxc,values_only=True)):
        print(f"r{i+1:2d}|"+" | ".join("" if v is None else str(v)[:14] for v in row))
# discount rate dataset - latest monthly file
dr=[h for h in hrefs("909707CDAD5C47529817D6146659E054") if h.split('/')[-2]=="the-discount-rate-and-interest-rates-on-deposits-and-loans-in-egyptian-pounds"]
dr.sort(); dump(dr[-1])
# gdp current quarterly 2022-2023
g=[h for h in hrefs("DEF6421CA1354B128A1113D7A5BBFC66") if h.split('/')[-2]=="gdp-by-expenditure-current" and "2022-2023" in h]
if g: dump(g[0])
