import io, re
from subsets_utils import get, configure_http
import openpyxl
configure_http(headers={"User-Agent":"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36","Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8","Accept-Language":"en-US,en;q=0.9"})
BASE="https://www.cbe.org.eg"
h=get(f"{BASE}/en/economic-research/time-series/downloadlist?category=706A9057F8454F7284BE8143070D88C4",timeout=(10,120)).text
files=sorted(x for x in re.findall(r'href="(/-/media/project/cbe/listing/time-series/[^"]+?\.xlsx)"',h) if x.split('/')[-2]=="wpi-and-ppi")
print(len(files),"files; latest:",files[-1])
r=get(BASE+files[-1],timeout=(10,120))
wb=openpyxl.load_workbook(io.BytesIO(r.content),read_only=True,data_only=True)
print("sheets",wb.sheetnames)
ws=wb[wb.sheetnames[0]]
for i,row in enumerate(ws.iter_rows(min_row=1,max_row=22,max_col=12,values_only=True)):
    print(f"r{i+1:2d}|"+" | ".join("" if v is None else str(v)[:13] for v in row))
