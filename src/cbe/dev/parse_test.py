import io, re, datetime
from subsets_utils import get, configure_http
import openpyxl

configure_http(headers={
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
})
BASE="https://www.cbe.org.eg"

MONTHS={m:i for i,m in enumerate(
    ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"],1)}
def month_num(tok):
    t=tok.strip().lower()[:3]
    return MONTHS.get(t)

ARAB=re.compile(r'[؀-ۿ]')
LATIN=re.compile(r'[A-Za-z]')
NUMRE=re.compile(r'^-?[\d,]+(\.\d+)?$')

def is_num(v):
    if isinstance(v,(int,float)) and not isinstance(v,bool): return True
    if isinstance(v,str) and NUMRE.match(v.strip()): return True
    return False
def to_float(v):
    if isinstance(v,(int,float)) and not isinstance(v,bool): return float(v)
    if isinstance(v,str):
        s=v.strip().replace(',','')
        try: return float(s)
        except: return None
    return None
def clean(v):
    if v is None: return None
    s=str(v).strip()
    s=re.sub(r'\s+',' ',s)
    return s or None

def read_grid(ws,maxr=400,maxc=60):
    grid=[]
    for row in ws.iter_rows(min_row=1,max_row=maxr,max_col=maxc,values_only=True):
        grid.append(list(row))
    # trim trailing empty cols
    return grid

def ffill_header(grid, hdr_rows, c0, c1):
    """forward-fill (left->right) merged header cells within header rows over value-col range"""
    filled={}
    for r in hdr_rows:
        row=grid[r]; last=None; out={}
        for c in range(c0,c1):
            v=clean(row[c]) if c<len(row) else None
            if v is not None: last=v
            out[c]=last  # propagate
        filled[r]=out
    return filled

def parse_period(tokens, file_y0, file_y1, freq_tag):
    """tokens: list of header strings for a column. return dict or None."""
    joined=" | ".join(t for t in tokens if t)
    low=joined.lower()
    # find years
    years=[int(y) for y in re.findall(r'(19|20)\d{2}', joined)]
    # month: look for a 'monthname year' or standalone month
    mon=None
    for tok in tokens:
        if not tok: continue
        for w in re.split(r'[\s./]+',tok):
            mn=month_num(w)
            if mn: mon=mn; break
        if mon: break
    qmatch=re.search(r'\bQ([1-4])\b', joined)
    quarter=int(qmatch.group(1)) if qmatch else None
    # Determine if this column is actually a period (must have year or month or quarter)
    if not years and mon is None and quarter is None:
        return None
    year=years[0] if years else None
    date=None; freq=None
    if mon is not None:
        # find a year near the month token
        y=None
        for tok in tokens:
            if tok and month_num_in(tok)==mon:
                ys=re.findall(r'(19|20)\d{2}',tok)
                if ys: y=int(ys[0]); break
        if y is None: y=year
        if y:
            date=datetime.date(y,mon,1); freq="monthly"; year=y
    elif quarter is not None and (years or file_y0):
        y0=years[0] if years else file_y0
        # egyptian fiscal year Y0/Y0+1 starting July
        qstart={1:(7,0),2:(10,0),3:(1,1),4:(4,1)}[quarter]
        date=datetime.date(y0+qstart[1],qstart[0],1); freq="quarterly"; year=y0
    elif years:
        y0=years[0]
        date=datetime.date(y0,7,1) if (file_y1 and file_y1!=file_y0) else datetime.date(y0,1,1)
        freq="annual"; year=y0
    return {"period_label":joined,"year":year,"quarter":quarter,"month":mon,"date":date,"freq":freq or (freq_tag or '').lower() or None}

def month_num_in(tok):
    for w in re.split(r'[\s./]+',tok or ''):
        mn=month_num(w)
        if mn: return mn
    return None

def parse_sheet(grid, file_y0, file_y1, freq_tag):
    nrows=len(grid); ncols=max((len(r) for r in grid),default=0)
    # find first data row: row with a text label in cols 0-2 AND >=2 numeric cells
    def numeric_cols_in_row(r):
        return [c for c in range(ncols) if c<len(grid[r]) and is_num(grid[r][c])]
    first_data=None
    for r in range(nrows):
        nc=numeric_cols_in_row(r)
        labs=[c for c in range(min(3,ncols)) if c<len(grid[r]) and clean(grid[r][c]) and not is_num(grid[r][c])]
        if len(nc)>=2 and labs:
            first_data=r; break
    if first_data is None: return []
    hdr_rows=list(range(0,first_data))
    if not hdr_rows: return []
    # value-column range = numeric columns observed across data rows
    valcols=set()
    for r in range(first_data,nrows):
        for c in numeric_cols_in_row(r): valcols.add(c)
    if not valcols: return []
    c0,c1=min(valcols),max(valcols)+1
    filled=ffill_header(grid,hdr_rows,c0,c1)
    # classify each candidate value col as a period col
    period_cols={}
    for c in range(c0,c1):
        toks=[filled[r].get(c) for r in hdr_rows]
        p=parse_period(toks,file_y0,file_y1,freq_tag)
        if p: period_cols[c]=p
    if not period_cols: return []
    # label columns: non-period cols left of first period col, pick latin & arabic
    first_pc=min(period_cols)
    label_cols=[c for c in range(0,first_pc)]
    # determine en/ar by content across data rows
    def colscore(c,rx):
        return sum(1 for r in range(first_data,nrows) if c<len(grid[r]) and clean(grid[r][c]) and rx.search(str(grid[r][c])))
    en_col=max(label_cols,key=lambda c:colscore(c,LATIN),default=None)
    ar_col=max(label_cols,key=lambda c:colscore(c,ARAB),default=None)
    out=[]
    for r in range(first_data,nrows):
        en=clean(grid[r][en_col]) if en_col is not None and en_col<len(grid[r]) else None
        ar=clean(grid[r][ar_col]) if ar_col is not None and ar_col<len(grid[r]) else None
        label=en or ar
        if not label: continue
        if label and (en or '').lower().startswith(('source','المصدر')): continue
        for c,p in period_cols.items():
            if c>=len(grid[r]): continue
            val=to_float(grid[r][c])
            if val is None: continue
            out.append({"indicator_en":en,"indicator_ar":ar,"period_label":p["period_label"],
                        "year":p["year"],"date":p["date"],"freq":p["freq"],"value":val})
    return out

def cat_hrefs(guid):
    url=f"{BASE}/en/economic-research/time-series/downloadlist?category={guid}"
    h=get(url,timeout=(10,120)).text
    return re.findall(r'href="(/-/media/project/cbe/listing/time-series/[^"]+?\.xlsx)"',h)

def parse_file(url, freq_tag, y0,y1):
    r=get(BASE+url,timeout=(10,120))
    wb=openpyxl.load_workbook(io.BytesIO(r.content),read_only=True,data_only=True)
    rows=[]
    for sn in wb.sheetnames:
        rows+=parse_sheet(read_grid(wb[sn]),y0,y1,freq_tag)
    return rows

TESTS=[
 ("706A9057F8454F7284BE8143070D88C4","cpi"),
 ("DEF6421CA1354B128A1113D7A5BBFC66","gdp-by-expenditure-current"),
 ("232131B16F15454BB1E1933B2BFEB041","egypts-balance-of-payments"),
 ("909707CDAD5C47529817D6146659E054","the-discount-rate-and-interest-rates-on-deposits-and-loans-in-egyptian-pounds"),
]
import collections
for guid,ds in TESTS:
    hrefs=[h for h in cat_hrefs(guid) if h.split('/')[-2]==ds]
    hrefs.sort()
    sample=hrefs[-3:] if len(hrefs)>=3 else hrefs
    print(f"\n#### {ds}  ({len(hrefs)} files) sampling {len(sample)}")
    for h in sample:
        fn=h.split('/')[-1]
        ys=[int(y) for y in re.findall(r'(19|20)\d{2}',fn)]
        y0,y1=(ys[0],ys[-1]) if ys else (None,None)
        ftag="monthly" if "month" in fn.lower() else "quarter" if "quart" in fn.lower() else "annual" if ("annual" in fn.lower() or "year" in fn.lower()) else None
        try:
            rows=parse_file(h,ftag,y0,y1)
        except Exception as e:
            print(f"  ERR {fn}: {type(e).__name__} {e}"); continue
        nd=sum(1 for x in rows if x['date'])
        inds=len(set(x['indicator_en'] for x in rows))
        print(f"  {fn[:55]:55s} rows={len(rows):4d} dated={nd:4d} ind_en={inds}")
        for x in rows[:2]: print("      ",{k:x[k] for k in ('indicator_en','period_label','date','freq','value')})
