"""Central Bank of Egypt (CBE) time-series connector.

CBE publishes its statistical time series as Excel (.xlsx) workbooks under a
stable media path, organised as Time-Series category pages -> datasets ->
per-(frequency x fiscal-year) workbook files. There is no JSON data API: the
file URLs are discovered by scraping each category's "download list" HTML page
(see research download_handoff). One Delta table is published per dataset (the
rank unit); a dataset's many workbook files are year/frequency partitions that
concatenate into one long-format table.

Fetch shape: stateless full re-pull (shape 1). Each workbook is small (KB), the
whole corpus is well under 100MB, and re-fetching every run picks up CBE's
in-place revisions and new fiscal-year files for free. No watermark/cursor.

WAF quirk: www.cbe.org.eg sits behind an Imperva WAF that rejects requests
lacking browser-like headers with a 269-byte "Request Rejected" HTML page
(HTTP 200, content-type text/html). A User-Agent alone is not enough -- we set
User-Agent + Accept + Accept-Language via configure_http, and verify every XLSX
response by content-type before parsing.

Parsing: CBE workbooks are *transposed matrices* -- indicators run down the
rows (col A Arabic, col B English label) and periods run across a multi-row
header band (year row + month/quarter row, sometimes a Public/Private/Total
sub-dimension row). Layouts vary across datasets and across years, so the parser
is generic: it locates the densest period-token header row, extends the header
band through any sub-dimension rows down to the first numeric data row,
forward-fills merged header cells, classifies each value column's period from
its forward-filled header chain, and melts every numeric data cell into one long
row -- (date, indicator_en, indicator_ar, dimension, period_label, frequency,
year, value). This degrades gracefully across the heterogeneous layouts instead
of hand-coding each one.
"""

import datetime
import html
import io
import re

import openpyxl
import pyarrow as pa

from subsets_utils import (
    NodeSpec,
    configure_http,
    get,
    save_raw_parquet,
    transient_retry,
)
from constants import CAT_SLUG_TO_GUID, ENTITY_IDS

BASE = "https://www.cbe.org.eg"
LIST_URL = BASE + "/en/economic-research/time-series/downloadlist?category={}"

# Browser-like headers required to pass the Imperva WAF (ASCII-only).
_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

_HREF_RE = re.compile(
    r'href="(/-/media/project/cbe/listing/time-series/[^"]+?\.xlsx)"', re.I
)

RAW_SCHEMA = pa.schema([
    ("indicator_en", pa.string()),
    ("indicator_ar", pa.string()),
    ("dimension", pa.string()),
    ("period_label", pa.string()),
    ("frequency", pa.string()),
    ("year", pa.int64()),
    ("date", pa.date32()),
    ("value", pa.float64()),
    ("source_file", pa.string()),
])


# --------------------------------------------------------------------------- #
# XLSX parsing (generic transposed-matrix melt)
# --------------------------------------------------------------------------- #
_MONTHS = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun",
     "jul", "aug", "sep", "oct", "nov", "dec"], 1)}
_ARAB = re.compile(r"[؀-ۿ]")
_LATIN = re.compile(r"[A-Za-z]")
_NUMRE = re.compile(r"^-?[\d,]+(\.\d+)?$")
_YEAR = re.compile(r"(?:19|20)\d{2}")
_QRE = re.compile(r"\bQ\s*([1-4])\b", re.I)


def _month_in(tok):
    for w in re.split(r"[\s./,()]+", tok or ""):
        mn = _MONTHS.get(w.strip().lower()[:3])
        if mn:
            return mn
    return None


def _is_num(v):
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    return isinstance(v, str) and bool(_NUMRE.match(v.strip()))


def _to_float(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip().replace(",", ""))
        except ValueError:
            return None
    return None


def _clean(v):
    if v is None:
        return None
    s = re.sub(r"\s+", " ", str(v)).strip()
    return s or None


def _read_grid(ws, maxr=600, maxc=80):
    return [list(r) for r in ws.iter_rows(
        min_row=1, max_row=maxr, max_col=maxc, values_only=True)]


def _period_score(row):
    """Cells that ARE a period token (short cell holding a month name or Qn).
    Long free-text footnotes that merely contain a month word score 0."""
    s = 0
    for c in row:
        cv = _clean(c)
        if cv and len(cv) <= 18 and (_month_in(cv) or _QRE.search(cv)):
            s += 1
    return s


def _year_only_score(row):
    s = 0
    for c in row:
        cv = _clean(c)
        if cv and len(cv) <= 12 and _YEAR.search(cv):
            s += 1
    return s


def _header_bottom(grid, scan=14):
    """Index of the densest period-token row near the top (ties -> lower row),
    falling back to the densest bare-year row, else None."""
    n = min(len(grid), scan)
    best_i, best = None, 0
    for r in range(n):
        sc = _period_score(grid[r])
        if sc and sc >= best:
            best, best_i = sc, r
    if best_i is not None:
        return best_i
    best_i, best = None, 0
    for r in range(n):
        sc = _year_only_score(grid[r])
        if sc and sc >= best:
            best, best_i = sc, r
    return best_i


def _ffill(grid, hdr_rows, c0, c1):
    """Forward-fill (left to right) merged header cells across the value-col range."""
    filled = {}
    for r in hdr_rows:
        row = grid[r]
        last = None
        o = {}
        for c in range(c0, c1):
            v = _clean(row[c]) if c < len(row) else None
            if v is not None:
                last = v
            o[c] = last
        filled[r] = o
    return filled


def _period(tokens, file_y0, file_y1, freq_tag):
    """Classify a value column's forward-filled header chain into a period."""
    toks = [t for t in tokens if t]
    if not toks:
        return None
    joined = " | ".join(toks)
    years = [int(y) for y in _YEAR.findall(joined)]
    mon = mon_year = None
    for t in toks:
        m = _month_in(t)
        if m:
            mon = m
            ys = _YEAR.findall(t)
            if ys:
                mon_year = int(ys[0])
            break
    qm = _QRE.search(joined)
    quarter = int(qm.group(1)) if qm else None
    if not years and mon is None and quarter is None:
        return None
    dim_toks = [t for t in toks
                if not (_YEAR.search(t) or _month_in(t) or _QRE.search(t))]
    dimension = " | ".join(dim_toks) or None
    year = date = None
    if mon is not None:
        year = mon_year or (years[0] if years else None)
        if year:
            date = datetime.date(year, mon, 1)
    elif quarter is not None:
        y0 = years[0] if years else file_y0
        if y0:
            # Egyptian fiscal year Y0/Y0+1 starts in July.
            mo, off = {1: (7, 0), 2: (10, 0), 3: (1, 1), 4: (4, 1)}[quarter]
            date = datetime.date(y0 + off, mo, 1)
            year = y0
    elif years:
        y0 = years[0]
        fiscal = bool(file_y1 and file_y0 and file_y1 != file_y0)
        date = datetime.date(y0, 7, 1) if fiscal else datetime.date(y0, 1, 1)
        year = y0
    freq = freq_tag or (
        "monthly" if mon is not None
        else "quarterly" if quarter
        else "annual" if years else None)
    return {"period_label": joined, "dimension": dimension,
            "year": year, "date": date, "freq": freq}


def _parse_sheet(grid, file_y0, file_y1, freq_tag):
    nrows = len(grid)
    ncols = max((len(r) for r in grid), default=0)
    if not ncols:
        return []
    period_bottom = _header_bottom(grid)
    if period_bottom is None:
        return []
    # Extend header band through sub-dimension rows (Public/Private/Total) down
    # to the first real data row (>=2 numeric value cells).
    first_data = period_bottom + 1
    for r in range(period_bottom + 1, min(nrows, period_bottom + 6)):
        ncount = sum(1 for c in range(ncols)
                     if c < len(grid[r]) and _is_num(grid[r][c]))
        if ncount >= 2:
            first_data = r
            break
    hdr_rows = list(range(0, first_data))
    valcols = set()
    for r in range(first_data, nrows):
        for c in range(ncols):
            if c < len(grid[r]) and _is_num(grid[r][c]):
                valcols.add(c)
    if not valcols:
        return []
    c0, c1 = min(valcols), max(valcols) + 1
    filled = _ffill(grid, hdr_rows, c0, c1)
    period_cols = {}
    for c in range(c0, c1):
        p = _period([filled[r].get(c) for r in hdr_rows],
                    file_y0, file_y1, freq_tag)
        if p:
            period_cols[c] = p
    if not period_cols:
        return []
    first_pc = min(period_cols)
    label_cols = list(range(0, first_pc)) or [0]

    def _score(c, rx):
        return sum(1 for r in range(first_data, nrows)
                   if c < len(grid[r]) and _clean(grid[r][c])
                   and rx.search(str(grid[r][c])))

    en_col = max(label_cols, key=lambda c: _score(c, _LATIN), default=None)
    ar_col = max(label_cols, key=lambda c: _score(c, _ARAB), default=None)
    out = []
    for r in range(first_data, nrows):
        en = _clean(grid[r][en_col]) if en_col is not None and en_col < len(grid[r]) else None
        ar = _clean(grid[r][ar_col]) if ar_col is not None and ar_col < len(grid[r]) else None
        label = en or ar
        if not label:
            continue
        ll = (en or "").lower()
        if ll.startswith(("source", ":")) or "title_en" in ll:
            continue
        for c, p in period_cols.items():
            if c >= len(grid[r]):
                continue
            val = _to_float(grid[r][c])
            if val is None:
                continue
            out.append({
                "indicator_en": en,
                "indicator_ar": ar,
                "dimension": p["dimension"],
                "period_label": p["period_label"],
                "frequency": p["freq"],
                "year": p["year"],
                "date": p["date"],
                "value": val,
            })
    return out


def _parse_workbook(content, file_y0, file_y1, freq_tag):
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows = []
    for sn in wb.sheetnames:
        rows += _parse_sheet(_read_grid(wb[sn]), file_y0, file_y1, freq_tag)
    wb.close()
    return rows


def _file_meta(fname):
    """Derive (frequency_tag, fiscal_year_start, fiscal_year_end) from a filename."""
    ys = [int(y) for y in _YEAR.findall(fname)]
    y0, y1 = (ys[0], ys[-1]) if ys else (None, None)
    low = fname.lower()
    if "month" in low:
        ft = "monthly"
    elif "quart" in low:
        ft = "quarterly"
    elif "annual" in low or "year" in low:
        ft = "annual"
    else:
        ft = None
    return ft, y0, y1


# --------------------------------------------------------------------------- #
# Fetch
# --------------------------------------------------------------------------- #
@transient_retry()
def _http_get(url):
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp


def _spec_id(entity_id):
    return f"cbe-{entity_id.lower().replace('_', '-')}"


# spec_id -> (category_slug, dataset_slug) recovered from the entity union.
_SPEC_LOOKUP = {}
for _eid in ENTITY_IDS:
    _cat, _ds = _eid.split("__", 1)
    _SPEC_LOOKUP[_spec_id(_eid)] = (_cat, _ds)


def _match_hrefs(page_html, ds_slug):
    """XLSX hrefs on a category page whose immediate parent folder is ds_slug.

    A dataset's files live in a folder named after the dataset slug. That folder
    usually nests under the category slug (.../time-series/<cat>/<ds>/file.xlsx),
    but for the bare "time-series" category it sits directly under time-series
    (.../time-series/<ds>/file.xlsx). Matching on the file's parent folder rather
    than a fixed category path handles both. Each href is html-unescaped because
    CBE encodes apostrophes etc. as entities (e.g. &#39;), which would otherwise
    produce a 400."""
    out = []
    for h in _HREF_RE.findall(page_html):
        href = html.unescape(h)
        parent = href.rsplit("/", 2)[-2] if href.count("/") >= 2 else ""
        if parent == ds_slug:
            out.append(href)
    return out


def _list_dataset_files(cat_slug, ds_slug):
    """Return the XLSX hrefs for one dataset by scraping its category page.

    Falls back to scanning every category page if the dataset's own page yields
    nothing (handles cross-listed datasets), and raises if still empty so a
    silent coverage loss becomes a loud failure."""
    guid = CAT_SLUG_TO_GUID.get(cat_slug)
    seen = []
    if guid:
        seen = _match_hrefs(_http_get(LIST_URL.format(guid)).text, ds_slug)
    if not seen:
        for g in set(CAT_SLUG_TO_GUID.values()):
            if g == guid:
                continue
            seen.extend(_match_hrefs(_http_get(LIST_URL.format(g)).text, ds_slug))
    files = sorted(set(seen))
    if not files:
        raise RuntimeError(
            f"no XLSX files found for {cat_slug}/{ds_slug} on any category page")
    return files


def fetch_one(node_id):
    """Fetch every workbook for one dataset, parse to long format, save parquet."""
    configure_http(headers=_HEADERS)  # browser headers for the WAF (per process)
    cat_slug, ds_slug = _SPEC_LOOKUP[node_id]
    files = _list_dataset_files(cat_slug, ds_slug)

    rows = []
    for href in files:
        fname = href.rsplit("/", 1)[-1]
        resp = _http_get(BASE + href)
        ctype = resp.headers.get("content-type", "")
        if "spreadsheet" not in ctype and "officedocument" not in ctype:
            raise RuntimeError(
                f"{fname}: expected an XLSX response, got content-type "
                f"{ctype!r} (WAF block or moved file)")
        ft, y0, y1 = _file_meta(fname)
        for rec in _parse_workbook(resp.content, y0, y1, ft):
            rec["source_file"] = fname
            rows.append(rec)

    if not rows:
        raise RuntimeError(f"{node_id}: parsed 0 rows from {len(files)} workbook(s)")

    table = pa.Table.from_pylist(rows, schema=RAW_SCHEMA)
    save_raw_parquet(table, node_id)


DOWNLOAD_SPECS = [
    NodeSpec(id=_spec_id(eid), fn=fetch_one, kind="download")
    for eid in ENTITY_IDS
]
