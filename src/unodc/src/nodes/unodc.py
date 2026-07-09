"""UNODC data portal connector.

Each entity is one UNODC workbook — 9 statistical collections plus the 3 region
classification workbooks — exposed as a single bulk .xlsx behind the ~65
https://data.unodc.org/datareport/<slug> pages. The dated file URL is
point-in-time (path embeds a YYYY-MM release directory) and the HTML links are
frequently stale/broken (a referenced .xlsx 404s while a sibling `_0.xlsx`
revision in the same directory is the live file). So we do NOT hardcode the URL:
each run scrapes the entity's datareport pages, collects every `YYYY-MM` release
directory mentioned anywhere on them, and probes `<basename>.xlsx` and
`<basename>_0.xlsx` in each — picking the newest directory that holds a valid
workbook (preferring the `_0` revision on a tie).

The corpus is small (a handful of .xlsx, low-GB total) and re-published wholesale
on each (roughly annual) release, with no incremental delta filter, so this is a
stateless full re-pull every run (TOOLS shape 1). Each workbook is parsed to
all-string NDJSON rows (VALUE columns carry censored entries like "<5", and the
four schema families — standard CTS, wildlife, SDG, region reference — differ);
casting is left to the compiled transform.
"""

import io
import re

import httpx
import openpyxl

from subsets_utils import (
    NodeSpec,
    get,
    save_raw_ndjson,
    transient_retry,
)

BASE = "https://data.unodc.org"

# entity_id -> (datareport pages to scrape for release dirs, base xlsx filename).
# Multiple pages per entity guard against a single page only surfacing a stale dir.
CONFIG = {
    "data-cts-access-and-functioning-of-justice": (
        ["cjs-arrested", "cjs-prosecuted", "cjs-convicted"],
        "data_cts_access_and_functioning_of_justice",
    ),
    "data-cts-corruption-and-economic-crime": (
        ["econ-bribery", "econ-corruption", "econ-environ"],
        "data_cts_corruption_and_economic_crime",
    ),
    "data-cts-intentional-homicide": (
        ["hom-victim", "hom-estimate", "hom-offenders"],
        "data_cts_intentional_homicide",
    ),
    "data-cts-prisons-and-prisoners": (
        ["prison-held", "prison-capacity", "prison-mortality"],
        "data_cts_prisons_and_prisoners",
    ),
    "data-cts-violent-and-sexual-crime": (
        ["violent-offences", "serious-assault", "sexual-violence"],
        "data_cts_violent_and_sexual_crime",
    ),
    "data-glotip": (
        ["tip-victims", "tip-offences", "tip-convicted"],
        "data_glotip",
    ),
    "data-iafq-firearms-trafficking": (
        ["firearm-seizures", "firearm-found", "firearm-tracing"],
        "data_iafq_firearms_trafficking",
    ),
    "data-wildlife-trafficking": (
        ["wildlife-seizures"],
        "data_wildlife_trafficking",
    ),
    "sdg-dataset": (
        ["sdg-16-1-1", "sdg-11-7-2", "sdg-15-7-1"],
        "sdg_dataset",
    ),
    # Region classification workbooks. Each is linked from the datareports of the
    # thematic area it applies to, so the pages that carry its release directory
    # are that area's — the drug datareports for WDR, the SDG ones for SDG.
    "data-portal-m49-regions": (
        ["hom-victim", "prison-held", "tip-victims"],
        "data_portal_m49_regions",
    ),
    "data-portal-sdg-regions": (
        ["sdg-16-1-1", "sdg-11-7-2", "sdg-15-7-1"],
        "data_portal_sdg_regions",
    ),
    "data-portal-wdr-regions": (
        ["drug-price", "drug-seizure", "druguse-prevalence"],
        "data_portal_wdr_regions",
    ),
}

# Anchor tokens identifying the real header row (a few title rows sit above it):
# `iso3_code` for the CTS family, `goal` for SDG, `iso_alpha3_code` for the region
# workbooks. Wildlife has neither and is matched on its own column pair below.
_HEADER_ANCHORS = ("iso3_code", "goal", "iso_alpha3_code")


@transient_retry(max_wait=60)
def _request(url: str) -> httpx.Response:
    """GET with transient-only retry. Returns the response without forcing 2xx —
    callers inspect status_code (404 is an expected outcome while probing URLs)."""
    resp = get(url, timeout=(10.0, 300.0))
    if resp.status_code == 429 or 500 <= resp.status_code < 600:
        resp.raise_for_status()  # surface as HTTPStatusError -> retried
    return resp


def _norm(cell) -> str:
    return re.sub(r"[^0-9a-z]+", "_", str(cell).strip().lower()).strip("_")


def _release_dirs(pages: list[str]) -> set[str]:
    """All YYYY-MM release directories referenced anywhere on the entity's pages."""
    dirs: set[str] = set()
    for page in pages:
        resp = _request(f"{BASE}/datareport/{page}")
        if resp.status_code != 200:
            continue  # a missing page is fine as long as another yields a dir
        dirs |= set(re.findall(r"/files/(\d{4}-\d{2})/", resp.text))
    return dirs


def _discover_and_download(pages: list[str], basename: str) -> bytes:
    """Find and download the current workbook for one collection.

    Newest release dir wins; within a dir the `_0` revision is preferred over the
    plain name. Returns the raw .xlsx bytes, or raises if nothing valid is found.
    """
    dirs = _release_dirs(pages)
    if not dirs:
        raise RuntimeError(f"unodc: no release directories found for {basename} via {pages}")

    candidates: list[tuple[str, bool, str]] = []
    for d in dirs:
        for fname, is_rev in ((f"{basename}_0.xlsx", True), (f"{basename}.xlsx", False)):
            candidates.append((d, is_rev, f"{BASE}/sites/dataportal.unodc.org/files/{d}/{fname}"))
    # newest dir first; on tie prefer the _0 revision
    candidates.sort(key=lambda c: (c[0], c[1]), reverse=True)

    for _d, _rev, url in candidates:
        resp = _request(url)
        if resp.status_code == 200 and resp.content[:2] == b"PK":
            return resp.content

    raise RuntimeError(
        f"unodc: no valid workbook for {basename} in release dirs {sorted(dirs)}"
    )


def _parse_workbook(content: bytes) -> list[dict]:
    """Parse the first sheet of a UNODC workbook into all-string row dicts.

    `reset_dimensions()` is mandatory: several workbooks ship a bogus <dimension>
    tag (e.g. A1:A1) that makes read-only mode yield zero data rows otherwise.
    Cells are stringified so the NDJSON column types stay stable (VALUE mixes
    numbers with censored strings like "<5"); the transform does the casting.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        ws.reset_dimensions()

        header: list[str] | None = None
        seen: dict[str, int] = {}
        rows: list[dict] = []
        for raw in ws.iter_rows(values_only=True):
            if header is None:
                cells = [_norm(c) if c is not None else "" for c in raw]
                is_header = any(a in cells for a in _HEADER_ANCHORS) or (
                    "geo" in cells and "taxonomic_group" in cells
                )
                if not is_header:
                    continue
                header = []
                for c in raw:
                    name = _norm(c) if c is not None else ""
                    if not name:
                        # keep positional alignment for trailing/blank header cells
                        name = f"col_{len(header)}"
                    if name == "txtvalue":
                        name = "value"
                    if name in seen:
                        seen[name] += 1
                        name = f"{name}_{seen[name]}"
                    else:
                        seen[name] = 0
                    header.append(name)
                continue

            row: dict[str, str] = {}
            has_value = False
            for idx, name in enumerate(header):
                cell = raw[idx] if idx < len(raw) else None
                if cell is None:
                    row[name] = None
                    continue
                text = str(cell).strip()
                row[name] = text if text != "" else None
                if row[name] is not None:
                    has_value = True
            if has_value:
                rows.append(row)
        return rows
    finally:
        wb.close()


def fetch_one(node_id: str) -> None:
    asset = node_id  # the runtime hands us the spec id; it IS the asset name
    entity = node_id[len("unodc-"):]
    pages, basename = CONFIG[entity]
    content = _discover_and_download(pages, basename)
    rows = _parse_workbook(content)
    if not rows:
        raise RuntimeError(f"unodc: parsed 0 data rows for {asset}")
    save_raw_ndjson(rows, asset)


DOWNLOAD_SPECS = [
    NodeSpec(id=f"unodc-{eid}", fn=fetch_one, kind="download")
    for eid in CONFIG
]
