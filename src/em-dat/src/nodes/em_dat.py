"""EM-DAT — CRED/UCLouvain Emergency Events Database Archive.

The entire public corpus is a single .xlsx workbook (~27k disaster events,
1900-present, 47 columns) published on the UCLouvain Dataverse. There is no
incremental query: CRED periodically re-snapshots the public portal into a new
Dataverse version, so this is a stateless full re-pull every refresh
(shape (1) in the implement guide) — small (~8MB) and cheap.

URL-stability caveat: the numeric datafile id changes with every Dataverse
version, but the dataset DOI is permanent. We resolve the current .xlsx file id
at fetch time via the native dataset API, then download it. No auth required
(public, CC-BY-NC-ND-4.0).
"""

import io

import httpx
import openpyxl
import pyarrow as pa

from subsets_utils import (
    NodeSpec,
    SqlNodeSpec,
    get,
    save_raw_parquet,
    transient_retry,
)

DATASET_DOI = "doi:10.14428/DVN/I0LTPH"
DATASET_API = "https://dataverse.uclouvain.be/api/datasets/:persistentId/"
ACCESS_BASE = "https://dataverse.uclouvain.be/api/access/datafile"

# The UCLouvain Dataverse is in Belgium; the TLS handshake from a US cloud
# runner can take well over 10s, so a tight connect timeout spuriously fails.
# Give connect (and read, for the ~8MB workbook) generous headroom; the retry
# decorator handles genuine transient stalls on top of this.
_TIMEOUT = httpx.Timeout(connect=60.0, read=600.0, write=120.0, pool=60.0)

# (source header, destination column, type)  — types: "str" | "int" | "float"
# Order mirrors the workbook; blanks ('' / None) become nulls.
COLUMNS = [
    ("DisNo.", "dis_no", "str"),
    ("Historic", "historic", "str"),
    ("Classification Key", "classification_key", "str"),
    ("Disaster Group", "disaster_group", "str"),
    ("Disaster Subgroup", "disaster_subgroup", "str"),
    ("Disaster Type", "disaster_type", "str"),
    ("Disaster Subtype", "disaster_subtype", "str"),
    ("External IDs", "external_ids", "str"),
    ("Event Name", "event_name", "str"),
    ("ISO", "iso", "str"),
    ("Country", "country", "str"),
    ("Subregion", "subregion", "str"),
    ("Region", "region", "str"),
    ("Location", "location", "str"),
    ("Origin", "origin", "str"),
    ("Associated Types", "associated_types", "str"),
    ("OFDA/BHA Response", "ofda_bha_response", "str"),
    ("Appeal", "appeal", "str"),
    ("Declaration", "declaration", "str"),
    ("AID Contribution ('000 US$)", "aid_contribution_000_usd", "float"),
    ("Magnitude", "magnitude", "float"),
    ("Magnitude Scale", "magnitude_scale", "str"),
    ("Latitude", "latitude", "float"),
    ("Longitude", "longitude", "float"),
    ("River Basin", "river_basin", "str"),
    ("Start Year", "start_year", "int"),
    ("Start Month", "start_month", "int"),
    ("Start Day", "start_day", "int"),
    ("End Year", "end_year", "int"),
    ("End Month", "end_month", "int"),
    ("End Day", "end_day", "int"),
    ("Total Deaths", "total_deaths", "int"),
    ("No. Injured", "no_injured", "int"),
    ("No. Affected", "no_affected", "int"),
    ("No. Homeless", "no_homeless", "int"),
    ("Total Affected", "total_affected", "int"),
    ("Reconstruction Costs ('000 US$)", "reconstruction_costs_000_usd", "float"),
    ("Reconstruction Costs, Adjusted ('000 US$)", "reconstruction_costs_adjusted_000_usd", "float"),
    ("Insured Damage ('000 US$)", "insured_damage_000_usd", "float"),
    ("Insured Damage, Adjusted ('000 US$)", "insured_damage_adjusted_000_usd", "float"),
    ("Total Damage ('000 US$)", "total_damage_000_usd", "float"),
    ("Total Damage, Adjusted ('000 US$)", "total_damage_adjusted_000_usd", "float"),
    ("CPI", "cpi", "float"),
    ("Admin Units", "admin_units", "str"),
    ("GADM Admin Units", "gadm_admin_units", "str"),
    ("Entry Date", "entry_date", "str"),
    ("Last Update", "last_update", "str"),
]

_ARROW_TYPE = {"str": pa.string(), "int": pa.int64(), "float": pa.float64()}
SCHEMA = pa.schema([(dest, _ARROW_TYPE[kind]) for _, dest, kind in COLUMNS])


@transient_retry()
def _fetch_dataset_metadata() -> dict:
    resp = get(DATASET_API, params={"persistentId": DATASET_DOI}, timeout=_TIMEOUT)
    resp.raise_for_status()
    return resp.json()


@transient_retry()
def _fetch_datafile(file_id: int) -> bytes:
    resp = get(f"{ACCESS_BASE}/{file_id}", timeout=_TIMEOUT)
    resp.raise_for_status()
    return resp.content


def _resolve_xlsx_file_id(metadata: dict) -> int:
    """Find the current *_emdat_archive.xlsx datafile id in the latest version."""
    files = metadata["data"]["latestVersion"]["files"]
    candidates = [
        f["dataFile"]
        for f in files
        if str(f["dataFile"].get("filename", "")).lower().endswith(".xlsx")
    ]
    if not candidates:
        raise RuntimeError(
            f"no .xlsx datafile found in Dataverse latestVersion for {DATASET_DOI}; "
            f"files present: {[f['dataFile'].get('filename') for f in files]}"
        )
    if len(candidates) > 1:
        raise RuntimeError(
            f"expected exactly one .xlsx in {DATASET_DOI}, found "
            f"{[c.get('filename') for c in candidates]}"
        )
    return int(candidates[0]["id"])


def _coerce(value, kind):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
    if kind == "str":
        return str(value)
    if kind == "int":
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None
    if kind == "float":
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    raise ValueError(f"unknown kind {kind!r}")


def fetch_one(node_id: str) -> None:
    asset = node_id  # the runtime passes the spec id; it IS the asset name

    metadata = _fetch_dataset_metadata()
    file_id = _resolve_xlsx_file_id(metadata)
    content = _fetch_datafile(file_id)

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    row_iter = ws.iter_rows(values_only=True)

    header = list(next(row_iter))
    expected = [src for src, _, _ in COLUMNS]
    if header != expected:
        raise RuntimeError(
            "EM-DAT workbook header drifted from expected schema.\n"
            f"  got:      {header}\n"
            f"  expected: {expected}"
        )

    # Accumulate column-major to build the typed Arrow table directly.
    columns = {dest: [] for _, dest, _ in COLUMNS}
    n_rows = 0
    for row in row_iter:
        if row is None or all(c is None for c in row):
            continue  # trailing blank rows that openpyxl sometimes emits
        for value, (_, dest, kind) in zip(row, COLUMNS):
            columns[dest].append(_coerce(value, kind))
        n_rows += 1

    wb.close()

    if n_rows == 0:
        raise RuntimeError(f"EM-DAT workbook (file id {file_id}) had no data rows")

    arrays = [
        pa.array(columns[dest], type=_ARROW_TYPE[kind])
        for _, dest, kind in COLUMNS
    ]
    table = pa.Table.from_arrays(arrays, schema=SCHEMA)
    save_raw_parquet(table, asset)


DOWNLOAD_SPECS = [
    NodeSpec(id="em-dat-events", fn=fetch_one, kind="download"),
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="em-dat-events-transform",
        deps=["em-dat-events"],
        sql='''
            SELECT * REPLACE (
                TRY_CAST(entry_date AS DATE)  AS entry_date,
                TRY_CAST(last_update AS DATE) AS last_update
            )
            FROM "em-dat-events"
            WHERE dis_no IS NOT NULL
        ''',
    ),
]
