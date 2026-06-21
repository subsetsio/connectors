"""Exercise the real parse logic + transform SQL end-to-end, without writing
to the production raw layer."""
import io
import duckdb
import pyarrow as pa
from nodes import em_dat
from subsets_utils import get

meta = em_dat._fetch_dataset_metadata()
fid = em_dat._resolve_xlsx_file_id(meta)
print("xlsx file id:", fid)
content = em_dat._fetch_datafile(fid)

import openpyxl
wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
it = ws.iter_rows(values_only=True)
header = list(next(it))
assert header == [s for s, _, _ in em_dat.COLUMNS], "header mismatch"
cols = {d: [] for _, d, _ in em_dat.COLUMNS}
n = 0
for row in it:
    if row is None or all(c is None for c in row):
        continue
    for v, (_, d, k) in zip(row, em_dat.COLUMNS):
        cols[d].append(em_dat._coerce(v, k))
    n += 1
wb.close()
arrays = [pa.array(cols[d], type=em_dat._ARROW_TYPE[k]) for _, d, k in em_dat.COLUMNS]
table = pa.Table.from_arrays(arrays, schema=em_dat.SCHEMA)
print("rows:", n, "cols:", table.num_columns)

con = duckdb.connect()
con.register("em-dat-events", table)
sql = em_dat.TRANSFORM_SPECS[0].sql
res = con.execute(sql).fetch_arrow_table()
print("transform rows:", res.num_rows, "cols:", res.num_columns)
print("entry_date type:", res.schema.field("entry_date").type)
print("last_update type:", res.schema.field("last_update").type)
# spot check key uniqueness
dn = res.column("dis_no").to_pylist()
print("unique dis_no:", len(set(dn)) == len(dn), "n=", len(dn))
print("sample:", res.slice(0, 1).to_pylist()[0])
