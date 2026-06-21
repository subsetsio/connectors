import io
import openpyxl
from subsets_utils import get

DOI = "doi:10.14428/DVN/I0LTPH"
DATASET_API = "https://dataverse.uclouvain.be/api/datasets/:persistentId/"

# 1. Resolve the current xlsx datafile id via the native dataset API
r = get(DATASET_API, params={"persistentId": DOI}, timeout=(10.0, 120.0))
r.raise_for_status()
meta = r.json()
lv = meta["data"]["latestVersion"]
print("versionNumber:", lv.get("versionNumber"), lv.get("versionMinorNumber"))
print("lastUpdateTime:", lv.get("lastUpdateTime"))
for f in lv["files"]:
    df = f["dataFile"]
    print(" file:", df["id"], df.get("filename"), df.get("contentType"), df.get("filesize"))

# find the archive xlsx
xlsx = next(
    f["dataFile"] for f in lv["files"]
    if f["dataFile"].get("filename", "").lower().endswith(".xlsx")
)
print("\nchosen xlsx id:", xlsx["id"], xlsx["filename"])

# 2. Download just enough to inspect the sheet header + a couple rows
url = f"https://dataverse.uclouvain.be/api/access/datafile/{xlsx['id']}"
resp = get(url, timeout=(10.0, 300.0))
resp.raise_for_status()
print("downloaded bytes:", len(resp.content))

wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
print("sheets:", wb.sheetnames)
ws = wb[wb.sheetnames[0]]
rows = ws.iter_rows(values_only=True)
header = next(rows)
print("ncols:", len(header))
print("header:", header)
for i, row in enumerate(rows):
    print("row", i, row)
    if i >= 2:
        break
# count total rows roughly
print("max_row:", ws.max_row)
