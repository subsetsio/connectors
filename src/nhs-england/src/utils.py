"""NHS England statistics: workbook discovery + generic long-format melt.

NHS England publishes each statistical work area as bespoke, human-formatted
Excel workbooks (banner blocks, multi-sheet, serial dates) reachable only by
scraping the work-area landing page (URLs are point-in-time — never stable).
There is no machine-readable catalog. Research's chosen strategy is
scrape_index -> bulk_files.

Because the workbooks are heterogeneous and there is no single tidy schema
across the ~50 areas, we melt every detected data block to a UNIFORM
long/tidy cell table: one row per (source_file, sheet, row_index, col_index)
with the detected column header, the row's leading label, and the value
(numeric or text). This is lossless, SQL-readable, and carries a clean grain;
the model stage derives periods/metrics from column_header/row_label.
"""
import datetime as dt
import io
import re
from urllib.parse import urljoin

import openpyxl
import pyarrow as pa
import xlrd

from subsets_utils import get

WORK_AREA_BASE = "https://www.england.nhs.uk/statistics/statistical-work-areas/"

# Uniform raw schema for every work-area table.
SCHEMA = pa.schema([
    ("source_file", pa.string()),
    ("sheet", pa.string()),
    ("row_index", pa.int32()),
    ("col_index", pa.int32()),
    ("column_header", pa.string()),
    ("row_label", pa.string()),
    ("value_num", pa.float64()),
    ("value_text", pa.string()),
])

# Safety ceilings — trip on source growth past expectation rather than
# silently truncate/hammer. They RAISE, never silently return.
MAX_SUBPAGES = 15
MAX_FILES_PER_AREA = 30
MAX_DISCOVERED = 400

# Fallback (non time-series) file names to drop — commentary, not data.
_NON_DATA_RE = re.compile(
    r"(technical[-_ ]?note|guidance|methodolog|specification|pre[-_ ]?release|"
    r"glossary|definition|read[-_ ]?me|faq|background|user[-_ ]?guide|"
    r"quality[-_ ]?statement|revisions[-_ ]?policy|annex)",
    re.I,
)
_TIME_SERIES_RE = re.compile(r"time[-_ ]?series", re.I)
_XLS_RE = re.compile(r"\.xlsx?($|\?)", re.I)
# Upload path date, e.g. /uploads/sites/2/2026/07/  -> (2026, 07) for recency.
_PATH_DATE_RE = re.compile(r"/uploads/sites/\d+/(\d{4})/(\d{2})/")


def _abs_links(page_url: str, html: str) -> list[str]:
    return [urljoin(page_url, h) for h in re.findall(r'href="([^"]+)"', html)]


def _path_date(url: str) -> tuple:
    m = _PATH_DATE_RE.search(url)
    return (int(m.group(1)), int(m.group(2))) if m else (0, 0)


def discover_files(slug: str) -> list[str]:
    """Return absolute .xls/.xlsx URLs for a work area: the landing page plus
    one level of same-section sub-pages (many areas hide their workbooks on
    per-topic sub-pages). PDFs are ignored."""
    landing = f"{WORK_AREA_BASE}{slug}/"
    resp = get(landing, timeout=(10, 60))
    resp.raise_for_status()
    links = _abs_links(landing, resp.text)

    files = {u for u in links if _XLS_RE.search(u)}

    sub_re = re.compile(
        rf"/statistical-work-areas/{re.escape(slug)}/[^\"?#]+/?$"
    )
    subpages = sorted({u for u in links if sub_re.search(u) and u.rstrip("/") != landing.rstrip("/")})
    if len(subpages) > MAX_SUBPAGES:
        # Deterministic slice; not silent — bounded by design for one refresh.
        subpages = subpages[:MAX_SUBPAGES]
    for sp in subpages:
        r = get(sp, timeout=(10, 60))
        if r.status_code != 200:
            continue
        for u in _abs_links(sp, r.text):
            if _XLS_RE.search(u):
                files.add(u)

    files = sorted(files)
    if len(files) > MAX_DISCOVERED:
        raise RuntimeError(
            f"{slug}: discovered {len(files)} workbook links (> {MAX_DISCOVERED}); "
            "source shape changed — review before bulk-fetching"
        )
    return files


def select_files(slug: str, files: list[str]) -> list[str]:
    """Pick the workbooks to melt. Prefer the full-history 'time series'
    workbooks (research: they contain the whole series, so a full re-fetch is
    the natural pattern). Otherwise fall back to the most recent data files."""
    ts = [u for u in files if _TIME_SERIES_RE.search(u.rsplit("/", 1)[-1])]
    if ts:
        chosen = ts
    else:
        data = [u for u in files if not _NON_DATA_RE.search(u.rsplit("/", 1)[-1])]
        # Most recent first by upload-path date.
        data.sort(key=_path_date, reverse=True)
        chosen = data[:MAX_FILES_PER_AREA]
    if len(chosen) > MAX_FILES_PER_AREA:
        chosen = sorted(chosen, key=_path_date, reverse=True)[:MAX_FILES_PER_AREA]
    return chosen


# ---- cell helpers -----------------------------------------------------------

def _is_num(v) -> bool:
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        return bool(re.fullmatch(r"-?\d{1,3}(,\d{3})*(\.\d+)?%?|-?\d+(\.\d+)?%?", v.strip()))
    return False


def _to_num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        pct = s.endswith("%")
        s = s.rstrip("%")
        try:
            f = float(s)
            return f / 100.0 if pct else f
        except ValueError:
            return None
    return None


def _celltext(v) -> str:
    if v is None:
        return ""
    if isinstance(v, dt.datetime):
        if v.hour == v.minute == v.second == 0:
            return v.date().isoformat()
        return v.isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def melt_sheet(grid: list, source_file: str, sheet: str) -> list[dict]:
    """Melt one sheet's grid (list of row-lists, dates as datetime) to long
    cell records. Skips the banner: the first data region is the first row
    with >=2 numeric cells; the header is the nearest text-rich row above it."""
    out = []
    n = len(grid)
    if n == 0:
        return out

    def numcount(r):
        return sum(1 for c in r if _is_num(c))

    first_data = next((i for i, r in enumerate(grid) if numcount(r) >= 2), None)
    if first_data is None:
        return out

    header = []
    for i in range(first_data - 1, -1, -1):
        if sum(1 for c in grid[i] if isinstance(c, str) and c.strip()) >= 2:
            header = grid[i]
            break

    for ri in range(first_data, n):
        r = grid[ri]
        if numcount(r) < 1:
            continue
        first_num_col = next((j for j, c in enumerate(r) if _is_num(c)), None)
        if first_num_col is None:
            continue
        label = " | ".join(t for t in (_celltext(c) for c in r[:first_num_col]) if t)
        for cj, v in enumerate(r):
            num = _to_num(v)
            text = _celltext(v)
            if num is None and not text:
                continue
            hdr = _celltext(header[cj]) if cj < len(header) else ""
            out.append({
                "source_file": source_file,
                "sheet": sheet,
                "row_index": ri,
                "col_index": cj,
                "column_header": hdr,
                "row_label": label,
                "value_num": num,
                "value_text": None if num is not None else text,
            })
    return out


def iter_sheets(url: str):
    """Yield (sheet_name, grid) for a workbook URL. Handles legacy .xls (BIFF,
    via xlrd) and .xlsx (via openpyxl); converts date cells to datetime."""
    resp = get(url, timeout=(10, 300))
    resp.raise_for_status()
    body = resp.content
    name = url.rsplit("/", 1)[-1].split("?")[0]
    if name.lower().endswith("xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
        try:
            for sn in wb.sheetnames:
                ws = wb[sn]
                yield name, sn, [list(row) for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
    else:
        wb = xlrd.open_workbook(file_contents=body)
        dm = wb.datemode
        for sn in wb.sheet_names():
            ws = wb.sheet_by_name(sn)
            grid = []
            for i in range(ws.nrows):
                row = []
                for j in range(ws.ncols):
                    cell = ws.cell(i, j)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            row.append(xlrd.xldate_as_datetime(cell.value, dm))
                        except Exception:
                            row.append(cell.value)
                    elif cell.ctype == xlrd.XL_CELL_EMPTY:
                        row.append(None)
                    else:
                        row.append(cell.value)
                grid.append(row)
            yield name, sn, grid
