"""CENACE (Mexican grid operator) wholesale-market price connector.

CENACE publishes Mercado Electrico Mayorista prices through three public REST
web services on ``ws01.cenace.gob.mx:8082`` that share one URL grammar::

    GET /SW{PML|PEND|PSC}/SIM/{sistema}/{proceso}/[{lista_nodos}/]{a}/{m}/{d}/{a}/{m}/{d}/JSON

``sistema`` in {SIN, BCA, BCS}; ``proceso`` in {MDA (day-ahead), MTR (real-time
ex-post)}; the per-request window is capped at 1-7 operation days. There is no
bulk dump, so a span is obtained by walking the window forward in 7-day chunks.

Shape: **stateless bounded re-pull** (one spec per price product). ``sistema``,
``proceso`` and the node/zone are *column values*, not separate specs — a single
``pml`` product covers every node across both markets and all three systems.
Each run, a product's fetch fn re-pulls a fixed trailing window
``today - BACKFILL_DAYS -> today+1`` in 7-day chunks and writes one raw parquet
batch per (system, market, window); the transform then ``overwrite()``s the
published table from this run's raw. This matches the harness storage model:
raw is **run-scoped** (``runs/<run_id>/raw/``) and transforms publish from the
current run's raw only, so we do NOT carry a cross-run watermark — that would
make a fresh run skip history it never re-fetched and shrink the table. Re-pull
is idempotent (deterministic batch ids overwrite in place) and picks up MDA/MTR
revisions for free. ``BACKFILL_DAYS`` is therefore the published trailing span,
sized so a full re-pull completes within the CI budget on this slow upstream;
widen it (and accept a longer run) for deeper coverage.

Per-product enumeration:
  * **PEND** / **PSC** — the service accepts a *node-less* path and returns the
    full system snapshot (all load zones / all reserve zones) in one call. No
    enumeration needed; one request per (system, market, window).
  * **PML** — node-level, so the path requires a comma-separated list of 1-20
    NodosP. The node universe (~2,600 NodosP, space-free CLAVE keys like
    "01PLO-115") is read from CENACE's published NodosP catalog XLSX (the latest
    version is discovered from the NodosP page); nodes are queried in chunks of
    20 and all chunks for a (system, market, window) accumulate into one batch.

Date bounds: the upper bound is discovered dynamically (today) so it never goes
stale; MDA is queryable through today+1, MTR only through today-7 (publication
lag), so each market's requested end is capped accordingly. Availability
semantics seen while probing: a window with no published data returns HTTP
400/404 ("No hay datos...") or 200 with status ZERO_RESULTS and an empty
Resultados list — all treated as "empty window", never an error. 5xx/429 are
transient (retried with bounded backoff); every request has a hard read timeout
so a single slow call can never stall the whole run.
"""

from datetime import date, datetime, timedelta, timezone
import io
import re
import time

import pyarrow as pa

from subsets_utils import (
    NodeSpec,
    SqlNodeSpec,
    get,
    transient_retry,
    save_raw_parquet,
)

SLUG = "cenace"
BASE = "https://ws01.cenace.gob.mx:8082"
CENACE_WWW = "https://www.cenace.gob.mx"

SYSTEMS = ("SIN", "BCA", "BCS")
MARKETS = ("MDA", "MTR")

WINDOW_DAYS = 7          # source cap: 1-7 operation days per request
PML_CHUNK = 20           # source cap: 1-20 NodosP per PML request
REQUEST_SPACING_S = 0.2  # politeness: a small gap between requests to one slow host

# Catalog discovery fallback (used only if the NodosP page can't be parsed).
CATALOG_FALLBACK_URL = (
    f"{CENACE_WWW}/Docs/MercadoOperacion/NodosP/"
    "Catálogo NodosP Sistema Eléctrico Nacional v2019 12 17.xlsx"
)
NODOSP_PAGE = f"{CENACE_WWW}/Paginas/SIM/NodosP.aspx"

# Per-product config. `group_key` is the response field naming each entity in the
# group; `num_fields` are coerced to float, `str_fields` kept as text;
# `node_listed` selects the PML node-list path vs the PEND/PSC node-less path;
# `backfill_days` is the trailing span re-pulled every run (see module docstring).
# PML carries ~2,600 nodes so its span is the tightest; PEND/PSC are one request
# per (system, market, window) so they can afford a wider span.
ENTITIES = {
    "pml": {
        "service": "SWPML",
        "group_key": "clv_nodo",
        "num_fields": ("pml", "pml_ene", "pml_per", "pml_cng"),
        "str_fields": (),
        "node_listed": True,
        "backfill_days": 21,
    },
    "pend": {
        "service": "SWPEND",
        "group_key": "zona_carga",
        "num_fields": ("pz", "pz_ene", "pz_per", "pz_cng"),
        "str_fields": (),
        "node_listed": False,
        "backfill_days": 120,
    },
    "psc": {
        "service": "SWPSC",
        "group_key": "clv_zona_reserva",
        "num_fields": ("pres",),
        "str_fields": ("tipo_res",),
        "node_listed": False,
        "backfill_days": 120,
    },
}


def _schema(cfg: dict) -> pa.Schema:
    fields = [
        ("sistema", pa.string()),
        ("proceso", pa.string()),
        (cfg["group_key"], pa.string()),
        ("fecha", pa.string()),
        ("hora", pa.int32()),
    ]
    fields += [(f, pa.float64()) for f in cfg["num_fields"]]
    fields += [(f, pa.string()) for f in cfg["str_fields"]]
    return pa.schema(fields)


# ---------------------------------------------------------------------------
# HTTP — one retried request. 5xx/429 are transient (retried); 400/404 mean the
# window has no published data (empty); any other 4xx is a real error.
# ---------------------------------------------------------------------------

@transient_retry(attempts=4, min_wait=2, max_wait=20)
def _fetch_json(url: str):
    # Hard read timeout (60s) + bounded retries: a single slow/hung call can
    # never stall the run — it fails fast and the window is retried or skipped.
    resp = get(url, timeout=(10.0, 60.0))
    code = resp.status_code
    if code >= 500 or code == 429:
        resp.raise_for_status()           # -> retried by the decorator
    if code in (400, 404):
        return None                        # window has no data (unpublished / out of range)
    resp.raise_for_status()                # any other 4xx is a genuine error
    return resp.json()


def _to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _parse(payload, sistema: str, proceso: str, cfg: dict) -> list[dict]:
    """Flatten one SW-* JSON payload into long-format rows."""
    if not payload:
        return []
    gk = cfg["group_key"]
    rows: list[dict] = []
    for group in payload.get("Resultados", []) or []:
        name = group.get(gk)
        for v in group.get("Valores", []) or []:
            row = {
                "sistema": sistema,
                "proceso": proceso,
                gk: name,
                "fecha": v.get("fecha"),
                "hora": _to_int(v.get("hora")),
            }
            for f in cfg["num_fields"]:
                row[f] = _to_float(v.get(f))
            for f in cfg["str_fields"]:
                row[f] = v.get(f)
            rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# NodosP catalog (PML only) — latest XLSX, CLAVE grouped by SISTEMA.
# ---------------------------------------------------------------------------

def _latest_catalog_url() -> str:
    try:
        html = get(NODOSP_PAGE, timeout=(10.0, 60.0)).text
        hrefs = re.findall(r'href="([^"]*NodosP[^"]*\.xlsx)"', html, re.I)
        if hrefs:
            def ver(h: str):
                m = re.search(r"(\d{4})[-\s]?(\d{2})[-\s]?(\d{2})", h)
                return tuple(int(x) for x in m.groups()) if m else (0, 0, 0)
            best = max(hrefs, key=ver)
            return best if best.startswith("http") else CENACE_WWW + best
    except Exception as e:
        print(f"[{SLUG}] catalog discovery failed ({type(e).__name__}: {e}); using fallback")
    return CATALOG_FALLBACK_URL


def _catalog_nodes_by_system() -> dict[str, list[str]]:
    import openpyxl
    url = _latest_catalog_url()
    content = get(url, timeout=(10.0, 180.0)).content
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: dict[str, list[str]] = {s: [] for s in SYSTEMS}
    rows = ws.iter_rows(values_only=True)
    next(rows, None)   # row 1: column grouping
    next(rows, None)   # row 2: real header
    for row in rows:
        if not row or len(row) < 4:
            continue
        sistema, clave = row[0], row[3]
        if sistema in out and clave:
            out[sistema].append(str(clave).strip())
    counts = {s: len(v) for s, v in out.items()}
    total = sum(counts.values())
    if total < 1000:
        raise AssertionError(f"NodosP catalog parsed only {total} nodes ({counts}); expected ~2,486")
    print(f"[{SLUG}] NodosP catalog: {total} nodes {counts} from {url}")
    return out


def _chunks(seq, n):
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


# ---------------------------------------------------------------------------
# Fetch — one generic stateless date-windowed re-pull for every product.
# ---------------------------------------------------------------------------

def _ymd(d: date) -> str:
    return f"{d.year:04d}/{d.month:02d}/{d.day:02d}"


def _log(msg: str) -> None:
    print(f"[{SLUG}] {msg}", flush=True)


def fetch_product(node_id: str) -> None:
    entity = node_id[len(SLUG) + 1:]
    cfg = ENTITIES[entity]
    schema = _schema(cfg)
    nodes_by_sys = _catalog_nodes_by_system() if cfg["node_listed"] else None

    today = datetime.now(timezone.utc).date()
    cur = today - timedelta(days=cfg["backfill_days"])
    end = today + timedelta(days=1)                  # MDA reaches today+1
    market_cap = {"MDA": today + timedelta(days=1), "MTR": today - timedelta(days=7)}
    window = timedelta(days=WINDOW_DAYS)
    _log(f"{entity}: re-pull {cur} .. {end} ({cfg['backfill_days']}d) in {WINDOW_DAYS}d windows")

    written = 0
    while cur <= end:
        win_end = min(cur + window - timedelta(days=1), end)
        for sistema in SYSTEMS:
            for proceso in MARKETS:
                req_end = min(win_end, market_cap[proceso])
                if cur > req_end:
                    continue                          # whole window beyond this market's availability
                date_path = f"{_ymd(cur)}/{_ymd(req_end)}"
                rows: list[dict] = []
                if cfg["node_listed"]:
                    for chunk in _chunks(nodes_by_sys[sistema], PML_CHUNK):
                        url = (f"{BASE}/{cfg['service']}/SIM/{sistema}/{proceso}/"
                               f"{','.join(chunk)}/{date_path}/JSON")
                        rows += _parse(_fetch_json(url), sistema, proceso, cfg)
                        time.sleep(REQUEST_SPACING_S)
                else:
                    url = f"{BASE}/{cfg['service']}/SIM/{sistema}/{proceso}/{date_path}/JSON"
                    rows += _parse(_fetch_json(url), sistema, proceso, cfg)
                    time.sleep(REQUEST_SPACING_S)
                if rows:
                    batch = f"{node_id}-{sistema.lower()}-{proceso.lower()}-{cur:%Y%m%d}"
                    save_raw_parquet(pa.Table.from_pylist(rows, schema=schema), batch)
                    written += 1
        _log(f"{entity}: window {cur:%Y-%m-%d} done ({written} batches so far)")
        cur = win_end + timedelta(days=1)
    _log(f"{entity}: complete, {written} batches written")


# ---------------------------------------------------------------------------
# Specs — one download + one thin publishing transform per product.
# ---------------------------------------------------------------------------

DOWNLOAD_SPECS = [
    NodeSpec(id=f"{SLUG}-{eid}", fn=fetch_product, kind="download")
    for eid in ENTITIES
]

# Each product's batches share one schema; DISTINCT is a cheap safety against any
# accidental duplicate row. The dep view glob-unions every
# `<id>-<system>-<market>-<window>` batch file automatically. fecha (text) is cast
# to a real DATE; the published table is long-format (system x market x node/zone x
# date x hour).
_TRANSFORM_SQL = {
    "pml": '''
        SELECT DISTINCT
            sistema, proceso, clv_nodo AS node,
            CAST(fecha AS DATE) AS date, hora AS hour,
            pml, pml_ene, pml_per, pml_cng
        FROM "{dep}"
        WHERE fecha IS NOT NULL AND hora IS NOT NULL
    ''',
    "pend": '''
        SELECT DISTINCT
            sistema, proceso, zona_carga AS load_zone,
            CAST(fecha AS DATE) AS date, hora AS hour,
            pz, pz_ene, pz_per, pz_cng
        FROM "{dep}"
        WHERE fecha IS NOT NULL AND hora IS NOT NULL
    ''',
    "psc": '''
        SELECT DISTINCT
            sistema, proceso, clv_zona_reserva AS reserve_zone,
            tipo_res AS reserve_type,
            CAST(fecha AS DATE) AS date, hora AS hour,
            pres
        FROM "{dep}"
        WHERE fecha IS NOT NULL AND hora IS NOT NULL
    ''',
}

# Per-product published grain (long-format one price per system x market x
# node/zone x date x hour); date is the observation period for freshness.
_TRANSFORM_KEY = {
    "pml": ("sistema", "proceso", "node", "date", "hour"),
    "pend": ("sistema", "proceso", "load_zone", "date", "hour"),
    "psc": ("sistema", "proceso", "reserve_zone", "reserve_type", "date", "hour"),
}

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"{s.id}-transform",
        deps=[s.id],
        sql=_TRANSFORM_SQL[s.id[len(SLUG) + 1:]].format(dep=s.id),
        key=_TRANSFORM_KEY[s.id[len(SLUG) + 1:]],
        temporal="date",
    )
    for s in DOWNLOAD_SPECS
]
