import io, zipfile
from subsets_utils import get

# 1) Catalog XLSX: structure + counts
url = "https://www.cenace.gob.mx/Docs/MercadoOperacion/NodosP/Cat%C3%A1logo%20NodosP%20Sistema%20El%C3%A9ctrico%20Nacional%20v2019%2012%2017.xlsx"
r = get(url, timeout=(10,120))
print("catalog xlsx:", r.status_code, len(r.content), "bytes")
import openpyxl
wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
print("sheets:", wb.sheetnames)
ws = wb[wb.sheetnames[0]]
rows = ws.iter_rows(values_only=True)
header = next(rows)
print("header:", header)
sample = [next(rows) for _ in range(4)]
for s in sample: print("  ", s)
cnt=0
sys_set=set(); zone_set=set()
for row in ws.iter_rows(values_only=True, min_row=2):
    cnt+=1
print("data rows:", cnt)
