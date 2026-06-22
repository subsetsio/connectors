import io, re
from collections import Counter
from subsets_utils import get
import openpyxl

# discover latest catalog link
for page in ["https://www.cenace.gob.mx/Paginas/SIM/NodosP.aspx"]:
    try:
        h = get(page, timeout=(10,60)).text
        links = re.findall(r'href="([^"]*NodosP[^"]*\.xlsx)"', h, re.I)
        print("links on", page, ":", links[:5])
    except Exception as e:
        print("page err", page, e)

url = "https://www.cenace.gob.mx/Docs/MercadoOperacion/NodosP/Cat%C3%A1logo%20NodosP%20Sistema%20El%C3%A9ctrico%20Nacional%20v2019%2012%2017.xlsx"
wb = openpyxl.load_workbook(io.BytesIO(get(url, timeout=(10,120)).content), read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
it = ws.iter_rows(values_only=True)
next(it); next(it)  # skip 2 header rows
sysc = Counter(); zones = {}; nperysys = Counter()
zoneset = set()
for row in it:
    if not row or not row[3]:
        continue
    sistema, ccr, zona, clave = row[0], row[1], row[2], row[3]
    sysc[sistema]+=1
    zoneset.add((sistema, zona))
print("nodes per system:", dict(sysc))
zc = Counter(s for s,z in zoneset)
print("zones per system:", dict(zc), "total zones:", len(zoneset))
print("sample zones:", sorted(z for s,z in zoneset)[:15])
