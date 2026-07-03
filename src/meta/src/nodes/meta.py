"""Meta connector — Data for Good / AI for Good at Meta, via the Humanitarian
Data Exchange (HDX) CKAN v3 portal (organization `meta`).

Mechanism (from research): CKAN at https://data.humdata.org/api/3/action.
Each accepted entity is one CKAN package; we publish one Delta table per
package. A package carries several resources (files); we ingest the tabular
ones (CSV/TSV, XLSX, and data files inside ZIPs), normalize them with DuckDB,
and write a single parquet per entity. The SQL transform then republishes it.

Strategy: stateless full re-pull. The corpus is a few GB of static files with
no incremental delta API, so every run re-fetches and overwrites. Each entity's
resources are unioned BY NAME (so heterogeneous geo-level / survey-wave files
with differing column sets coexist as one wide table), with a `_source_file`
column recording provenance. Per-file `read_csv_auto` lets DuckDB sniff each
file's delimiter independently (handles comma CSV and tab TSV in one package).
"""

import os
import re
import shutil
import tempfile
import zipfile

import duckdb
import httpx
import pyarrow as pa
from tenacity import (
    retry,
    retry_if_exception,
    stop_after_attempt,
    wait_exponential,
)

from subsets_utils import NodeSpec, SqlNodeSpec, get, raw_parquet_writer

CKAN = "https://data.humdata.org/api/3/action"

# The rank-accepted entity union (CKAN package slugs). One DOWNLOAD_SPEC each.
from constants import ENTITY_IDS

# Resources whose *name* matches this are documentation / metadata, not data.
_EXCLUDE_NAME = re.compile(
    r"(codebook|read[ _-]?me|licen[cs]e|dictionary|data[ _-]?description"
    r"|how to understand|\breadme\b)",
    re.I,
)

# Per-entity extra name exclusions (regex on resource name), to drop redundant
# or oversized variants when a cleaner equivalent already covers the data.
_ENTITY_EXCLUDE = {
    # RWI ships per-country files (filename = country) AND two pooled "93 Low and
    # Middle Income Countries" megafiles that duplicate them without a country
    # dimension — keep the per-country files only.
    "relative-wealth-index": re.compile(r"93 low and middle", re.I),
    # SCI's geoboundaries_adm1.csv is a ~465MB alternative-boundary restatement
    # of the gadm1 region pairs; gadm1.csv + country.csv already cover it.
    "social-connectedness-index": re.compile(r"geoboundaries", re.I),
}

# ---------------------------------------------------------------------------
# HTTP with honest retry classification
#
# HDX sits behind Cloudflare, which serves a non-JSON bot-challenge (observed as
# HTTP 202 with an HTML body) to the default datacenter User-Agent. A realistic
# browser UA clears it; any non-200 / non-JSON / HTML-challenge response is
# treated as transient and retried with backoff.
# ---------------------------------------------------------------------------

_BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json,text/plain,*/*",
    "Accept-Language": "en-US,en;q=0.9",
}

_TRANSIENT = (
    httpx.ConnectError, httpx.ConnectTimeout, httpx.ReadTimeout,
    httpx.WriteTimeout, httpx.PoolTimeout, httpx.RemoteProtocolError,
    httpx.ProxyError,
)


class _RetryableHTTP(Exception):
    """A 2xx-but-wrong / challenge response worth retrying."""


def _is_transient(exc: BaseException) -> bool:
    if isinstance(exc, (_TRANSIENT, _RetryableHTTP)):
        return True
    if isinstance(exc, httpx.HTTPStatusError):
        code = exc.response.status_code
        return code == 429 or 500 <= code < 600
    return False


def _looks_like_html(head: bytes) -> bool:
    sniff = head[:512].lstrip().lower()
    return sniff.startswith((b"<!doctype", b"<html", b"<head"))


@retry(
    retry=retry_if_exception(_is_transient),
    stop=stop_after_attempt(8),
    wait=wait_exponential(min=3, max=60),
    reraise=True,
)
def _get_json(url: str, **params) -> dict:
    resp = get(url, params=params, headers=_BROWSER_HEADERS, timeout=(10.0, 120.0))
    if resp.status_code != 200:
        raise _RetryableHTTP(f"{url} -> HTTP {resp.status_code}")
    try:
        return resp.json()
    except ValueError as e:
        raise _RetryableHTTP(f"{url} -> non-JSON body ({resp.text[:80]!r})") from e


@retry(
    retry=retry_if_exception(_is_transient),
    stop=stop_after_attempt(8),
    wait=wait_exponential(min=3, max=60),
    reraise=True,
)
def _download_to(url: str, dest: str) -> None:
    # Files range from KBs to ~250MB; stream to disk to bound memory. Sniff the
    # first chunk so a Cloudflare HTML challenge can't masquerade as data.
    with httpx.Client(follow_redirects=True, timeout=(10.0, 300.0)) as client:
        with client.stream("GET", url, headers=_BROWSER_HEADERS) as r:
            if r.status_code != 200:
                raise _RetryableHTTP(f"{url} -> HTTP {r.status_code}")
            first = True
            with open(dest, "wb") as f:
                for chunk in r.iter_bytes(8 * 1024 * 1024):
                    if first and _looks_like_html(chunk):
                        raise _RetryableHTTP(f"{url} -> HTML challenge body")
                    first = False
                    f.write(chunk)


# ---------------------------------------------------------------------------
# Resource selection + normalization to local CSV/TSV files
# ---------------------------------------------------------------------------

def _ext(url: str, name: str) -> str:
    path = url.split("?", 1)[0].lower()
    nm = (name or "").lower()
    for cand in (path, nm):
        for e in (".csv", ".tsv", ".xlsx", ".xls", ".zip"):
            if cand.endswith(e):
                return e
    return ""


def _select_resources(eid: str, resources: list) -> list:
    """Pick HDX-hosted tabular resources, dropping docs and external mirrors."""
    extra = _ENTITY_EXCLUDE.get(eid)
    csv_like, xlsx_like = [], []
    for r in resources:
        url = r.get("url") or ""
        name = r.get("name") or ""
        if "data.humdata.org" not in url:
            continue  # skip Google-Drive and other external mirrors
        if _EXCLUDE_NAME.search(name):
            continue
        if extra and extra.search(name):
            continue
        ext = _ext(url, name)
        if ext in (".csv", ".tsv", ".zip"):
            csv_like.append((r, ext))
        elif ext in (".xlsx", ".xls"):
            xlsx_like.append((r, ext))
    # Prefer real CSV/TSV/ZIP data; fall back to spreadsheets only when a
    # package ships nothing else (e.g. survey-on-gender-equality-at-home).
    return csv_like if csv_like else xlsx_like


def _safe(name: str) -> str:
    base = re.sub(r"\.(csv|tsv|xlsx?|zip|txt)$", "", name or "", flags=re.I)
    return re.sub(r"[^0-9A-Za-z]+", "_", base).strip("_")[:80] or "file"


def _xlsx_to_csv(src: str, dest: str) -> bool:
    """Convert the first worksheet of an .xlsx to a CSV. Returns False if empty."""
    import csv
    from openpyxl import load_workbook

    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb.active
    wrote = 0
    with open(dest, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            w.writerow(["" if v is None else v for v in row])
            wrote += 1
    wb.close()
    return wrote > 1  # header + at least one data row


def _materialize(eid: str, selected: list, workdir: str) -> list:
    """Download + normalize selected resources into local (path, label) CSVs."""
    out = []
    seen = {}

    def _label(name: str) -> str:
        lab = _safe(name)
        seen[lab] = seen.get(lab, 0) + 1
        return lab if seen[lab] == 1 else f"{lab}_{seen[lab]}"

    for idx, (r, ext) in enumerate(selected):
        url = r["url"]
        name = r.get("name") or f"resource_{idx}"
        raw = os.path.join(workdir, f"dl_{idx}{ext or '.bin'}")
        try:
            _download_to(url, raw)
        except Exception as e:  # permanent failure on one file: skip, keep going
            print(f"  ! skip {eid}/{name}: download failed ({type(e).__name__}: {e})")
            continue

        # NB: .xlsx is itself a zip container, so handle spreadsheets BEFORE the
        # zip magic-byte check below (which only exists to catch zips mislabeled
        # .txt, e.g. movement-range-maps).
        if ext in (".xlsx", ".xls"):
            target = os.path.join(workdir, f"x{idx}.csv")
            try:
                if _xlsx_to_csv(raw, target):
                    out.append((target, _label(name)))
            except Exception as e:
                print(f"  ! skip {eid}/{name}: xlsx parse failed ({type(e).__name__}: {e})")
            continue

        is_zip = ext == ".zip" or zipfile.is_zipfile(raw)
        if is_zip:
            try:
                with zipfile.ZipFile(raw) as zf:
                    for member in zf.namelist():
                        mlow = member.lower()
                        if mlow.endswith("/") or _EXCLUDE_NAME.search(member):
                            continue
                        if not mlow.endswith((".csv", ".tsv", ".txt")):
                            continue
                        target = os.path.join(workdir, f"z{idx}_{_safe(member)}.csv")
                        with zf.open(member) as srcf, open(target, "wb") as dstf:
                            shutil.copyfileobj(srcf, dstf)
                        out.append((target, _label(member)))
            except Exception as e:
                print(f"  ! skip {eid}/{name}: bad zip ({type(e).__name__}: {e})")
            continue

        out.append((raw, _label(name)))
    return out


def _q(s: str) -> str:
    return s.replace("'", "''")


def _union_sql(files: list, *, all_varchar: bool) -> str:
    opts = "header=true, null_padding=true, ignore_errors=true, sample_size=200000"
    if all_varchar:
        opts += ", all_varchar=true"
    parts = [
        f"SELECT *, '{_q(label)}' AS _source_file "
        f"FROM read_csv_auto('{_q(path)}', {opts})"
        for path, label in files
    ]
    return "\nUNION ALL BY NAME\n".join(parts)


# ---------------------------------------------------------------------------
# Download node — one generic fetch for every entity
# ---------------------------------------------------------------------------

def fetch_one(node_id: str) -> None:
    asset = node_id                       # the spec id IS the raw asset name
    eid = node_id[len("meta-"):]

    pkg = _get_json(f"{CKAN}/package_show", id=eid)
    resources = pkg["result"].get("resources", []) or []
    selected = _select_resources(eid, resources)
    if not selected:
        raise RuntimeError(f"{eid}: no ingestible tabular resources found")

    workdir = tempfile.mkdtemp(prefix=f"meta_{_safe(eid)}_")
    try:
        files = _materialize(eid, selected, workdir)
        if not files:
            raise RuntimeError(f"{eid}: all resource downloads/parses failed")

        con = duckdb.connect()
        for all_varchar in (False, True):
            sql = _union_sql(files, all_varchar=all_varchar)
            try:
                reader = con.sql(sql).fetch_record_batch()
                schema = reader.schema
                rows = 0
                with raw_parquet_writer(asset, schema) as w:
                    for batch in reader:
                        if batch.num_rows:
                            w.write_batch(batch)
                            rows += batch.num_rows
                if rows == 0:
                    raise RuntimeError(f"{eid}: union produced 0 rows")
                print(f"  -> {eid}: {rows:,} rows from {len(files)} file(s)"
                      f"{' (all_varchar fallback)' if all_varchar else ''}")
                return
            except (duckdb.Error, pa.ArrowInvalid) as e:
                if all_varchar:
                    raise
                print(f"  ~ {eid}: typed read failed ({type(e).__name__}: {e}); "
                      "retrying as all-varchar")
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


DOWNLOAD_SPECS = [
    NodeSpec(id=f"meta-{eid}", fn=fetch_one, kind="download")
    for eid in ENTITY_IDS
]

# One published Delta table per entity. The download already normalized and
# typed each package into a single parquet, so the transform is a thin
# republish (the 0-row guard in the runtime is the correctness gate).
#
# Per-entity grain declarations. Only entities with a genuinely unique key
# and/or a clear observation-period column are listed; the heterogeneous
# multi-level union-by-name tables (social-capital-atlas, relative-wealth-index,
# social-connectedness-index, movement-*, future-of-business, etc.) have no
# single stable grain and are left undeclared.
_GRAIN = {
    "commuting-zones": {"key": ("fbcz_id",), "temporal": "cz_gen_ds"},
    "cross-gender-ties": {"key": ("region_id",)},
    "facebook-business-activity-trends-during-covid19": {"temporal": "ds"},
    "facebook-business-activity-trends-during-crisis": {"temporal": "ds"},
    "international-migration-flows": {
        "key": ("country_from", "country_to", "migration_month"),
        "temporal": "migration_month",
    },
    "survey-on-gender-equality-at-home": {"temporal": "Year"},
    "uk-social-capital-atlas": {"temporal": "ds"},
}

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"{s.id}-transform",
        deps=[s.id],
        sql=f'SELECT * FROM "{s.id}"',
        **_GRAIN.get(s.id[len("meta-"):], {}),
    )
    for s in DOWNLOAD_SPECS
]
