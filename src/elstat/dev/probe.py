import re, io
from subsets_utils import get
import openpyxl

PUB = "https://www.statistics.gr/en/statistics/-/publication/{code}/-"
DOC_RE = re.compile(
    r'href="([^"]*?/statistics\?[^"]*?documents_WAR_publicationsportlet_INSTANCE_([A-Za-z0-9]+)[^"]*?documentID=(\d+)[^"]*?)"',
    re.IGNORECASE,
)

def find_docs(code):
    html = get(PUB.format(code=code), timeout=(10,60)).text
    docs = []
    seen = set()
    for m in DOC_RE.finditer(html):
        url = m.group(1).replace("&amp;", "&")
        inst, did = m.group(2), m.group(3)
        if did in seen: continue
        seen.add(did)
        docs.append((inst, did, url))
    return docs

def probe(code):
    print("="*70)
    print("CODE", code)
    docs = find_docs(code)
    print(f"  {len(docs)} document link(s)")
    ts_files = []
    for inst, did, url in docs:
        r = get(url, timeout=(10,90))
        cd = r.headers.get("content-disposition","")
        fn = re.search(r'filename="?([^"]+)"?', cd)
        fn = fn.group(1) if fn else "?"
        ct = r.headers.get("content-type","")
        print(f"    docID={did} inst={inst} bytes={len(r.content)} ct={ct.split(';')[0]} fn={fn}")
        if fn.lower().endswith(".xlsx") and "_TS_" in fn and ("_EN" in fn or "_P_EN" in fn or fn.endswith("EN.xlsx")):
            ts_files.append((fn, r.content))
    for fn, content in ts_files[:1]:
        print(f"  --- parsing TS file: {fn} ---")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        print("    sheets:", wb.sheetnames)
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 22: break
            cells = ["" if c is None else str(c)[:18] for c in row]
            print(f"    r{i}: {cells}")
        wb.close()

for code in ["SDT03","DKT87","SEL15","SPO03","SJO02"]:
    try:
        probe(code)
    except Exception as e:
        print("CODE", code, "ERROR", type(e).__name__, e)
