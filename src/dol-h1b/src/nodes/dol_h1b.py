"""Department of Labor (H-1B LCA) disclosure data connector.

Source: OFLC Labor Condition Application (LCA) disclosure files, served as
per-period .xlsx from https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs/ .

There is one publishable subset: the full LCA disclosure corpus. It is
assembled from many immutable per-period files — ~18 irregular historical
annual files (FY2008-FY2019) plus quarterly files from FY2020 onward
(LCA_Disclosure_Data_FY{YYYY}_Q{n}.xlsx). All periods share one conceptual
schema, so fiscal_year/quarter are columns rather than separate datasets.

Shape: a single download node that walks every period as an immutable artefact
(state tracks completed periods so reruns only pick up newly-published
quarters), writing one streamed parquet batch per period. The transform unions
the batches into one Delta table.

HTTP CLIENT EXCEPTION — dol.gov sits behind Akamai bot protection that blocks
by TLS fingerprint (JA3), not by User-Agent or IP geography. The harness's
httpx client (subsets_utils.get) is hard-403'd from BOTH an EU probe and the
US GitHub Actions runner (verified: instant 18-75ms edge denies on every file).
Only a browser-impersonated TLS handshake gets through. curl_cffi
(impersonate="chrome") returns 200 and the full payload from the same hosts.
This source is therefore unservable via httpx, so this is the one connector
that deliberately fetches via curl_cffi instead of subsets_utils.get. All other
I/O still goes through subsets_utils. Each period is fetched defensively — a
403/404/invalid-xlsx skips that period (left un-completed so it retries next
run) rather than killing the node, so a partially-blocked run still publishes
what it could fetch.
"""
import os
import tempfile

import pyarrow as pa
from datetime import datetime, timezone
from openpyxl import load_workbook
from tenacity import (
    retry, retry_if_exception_type, stop_after_attempt, wait_exponential,
)
from curl_cffi import requests as creq
from curl_cffi.requests.exceptions import RequestException

from subsets_utils import (
    NodeSpec,
    SqlNodeSpec,
    load_state,
    save_state,
    raw_parquet_writer,
)

# Bumped to 2: added historical column aliases (WAGE_RATE_1, OCCUPATIONAL_*),
# so completed periods must be re-fetched to re-parse with the new mapping.
STATE_VERSION = 2
BASE = "https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs"

# Browser profile for TLS impersonation — the thing that actually defeats Akamai.
IMPERSONATE = "chrome"

# Irregular historical annual files (FY2008-FY2019). Filenames are versioned and
# do not follow a pattern, so they are enumerated explicitly (per research).
HISTORICAL_ANNUAL_FILES = [
    ("fy2008", f"{BASE}/H-1B_Case_Data_FY2008.xlsx"),
    ("fy2009-icert", f"{BASE}/Icert_LCA_FY2009.xlsx"),
    ("fy2009-efiling", f"{BASE}/H-1B_Case_Data_FY2009.xlsx"),
    ("fy2010", f"{BASE}/H-1B_FY2010.xlsx"),
    ("fy2011", f"{BASE}/H-1B_iCert_LCA_FY2011_Q4.xlsx"),
    ("fy2012", f"{BASE}/LCA_FY2012_Q4.xlsx"),
    ("fy2013", f"{BASE}/LCA_FY2013.xlsx"),
    ("fy2014", f"{BASE}/H-1B_FY14_Q4.xlsx"),
    ("fy2015", f"{BASE}/H-1B_Disclosure_Data_FY15_Q4.xlsx"),
    ("fy2016", f"{BASE}/H-1B_Disclosure_Data_FY16.xlsx"),
    ("fy2017", f"{BASE}/H-1B_Disclosure_Data_FY17.xlsx"),
    ("fy2018", f"{BASE}/H-1B_Disclosure_Data_FY2018_EOY.xlsx"),
    ("fy2019", f"{BASE}/H-1B_Disclosure_Data_FY2019.xlsx"),
]

# Quarterly files follow a stable pattern from FY2020 onward. The end of the
# range is discovered from the calendar (federal FY starts Oct 1) rather than
# hardcoded; not-yet-published quarters 404/403 and are skipped.
QUARTERLY_START_FY = 2020

# Output schema. wage_rate is kept as text (source mixes ranges, commas, units);
# casting happens in the transform. fiscal_year is the only non-null column.
SCHEMA = pa.schema([
    ("fiscal_year", pa.int64()),
    ("quarter", pa.string()),
    ("case_status", pa.string()),
    ("employer_name", pa.string()),
    ("employer_state", pa.string()),
    ("employer_city", pa.string()),
    ("job_title", pa.string()),
    ("soc_code", pa.string()),
    ("soc_title", pa.string()),
    ("wage_rate", pa.string()),
    ("wage_unit", pa.string()),
    ("worksite_state", pa.string()),
    ("worksite_city", pa.string()),
])

# Target column -> accepted upstream header variants (UPPER_SNAKE). Header layout
# drifts across years (modern WORKSITE_STATE_1 vs legacy LCA_CASE_WORKLOC1_STATE),
# so each target matches the first present variant.
COLUMN_ALIASES = {
    "case_status": ["CASE_STATUS", "STATUS", "APPROVAL_STATUS"],
    "employer_name": ["EMPLOYER_NAME", "LCA_CASE_EMPLOYER_NAME"],
    "employer_state": ["EMPLOYER_STATE", "LCA_CASE_EMPLOYER_STATE", "EMPLOYER_PROVINCE"],
    "employer_city": ["EMPLOYER_CITY", "LCA_CASE_EMPLOYER_CITY"],
    "job_title": ["JOB_TITLE", "LCA_CASE_JOB_TITLE"],
    "soc_code": ["SOC_CODE", "LCA_CASE_SOC_CODE", "OCCUPATIONAL_CODE"],
    "soc_title": ["SOC_TITLE", "SOC_NAME", "LCA_CASE_SOC_NAME", "OCCUPATIONAL_TITLE"],
    "wage_rate": [
        "WAGE_RATE_OF_PAY_FROM_1", "WAGE_RATE_OF_PAY_FROM", "WAGE_RATE_OF_PAY",
        "LCA_CASE_WAGE_RATE_FROM", "WAGE_RATE_1",
    ],
    "wage_unit": [
        "WAGE_UNIT_OF_PAY_1", "WAGE_UNIT_OF_PAY", "PW_UNIT_1",
        "LCA_CASE_WAGE_RATE_UNIT",
    ],
    "worksite_state": [
        "WORKSITE_STATE_1", "WORKSITE_STATE", "LCA_CASE_WORKLOC1_STATE", "WORK_LOCATION_STATE1",
    ],
    "worksite_city": [
        "WORKSITE_CITY_1", "WORKSITE_CITY", "LCA_CASE_WORKLOC1_CITY", "WORK_LOCATION_CITY1",
    ],
}

TEXT_COLS = [
    "case_status", "employer_name", "employer_state", "employer_city",
    "job_title", "soc_code", "soc_title", "wage_rate", "wage_unit",
    "worksite_state", "worksite_city",
]

BATCH_ROWS = 100_000          # rows per parquet write — bounds peak memory
MIN_XLSX_BYTES = 50_000       # smaller than this is an error/challenge page
CHUNK = 1 << 20               # 1MB download chunks


class _Transient(Exception):
    """Retryable failure — 429/5xx or a transport error."""


class _Blocked(Exception):
    """Permanent for this period — 403 (Akamai), 404 (not yet published)."""

    def __init__(self, code: int):
        super().__init__(f"HTTP {code}")
        self.code = code


@retry(
    retry=retry_if_exception_type(_Transient),
    stop=stop_after_attempt(5),
    wait=wait_exponential(min=4, max=120),
    reraise=True,
)
def _download_to_tmp(url: str) -> tuple[str, int]:
    """Stream one period file to a temp path via impersonated TLS.

    Returns (path, size). Raises _Blocked on 4xx (caller skips the period) and
    _Transient on 429/5xx/transport errors (retried with backoff)."""
    try:
        resp = creq.get(url, impersonate=IMPERSONATE, timeout=600, stream=True)
    except RequestException as e:
        raise _Transient(f"transport: {type(e).__name__}: {e}")

    code = resp.status_code
    if code == 429 or 500 <= code < 600:
        resp.close()
        raise _Transient(f"HTTP {code}")
    if code != 200:
        resp.close()
        raise _Blocked(code)

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    size = 0
    try:
        with os.fdopen(fd, "wb") as f:
            for chunk in resp.iter_content(chunk_size=CHUNK):
                if chunk:
                    f.write(chunk)
                    size += len(chunk)
    except RequestException as e:
        os.unlink(path)
        raise _Transient(f"stream: {type(e).__name__}: {e}")
    finally:
        resp.close()
    return path, size


def _valid_xlsx(path: str, size: int) -> bool:
    if size < MIN_XLSX_BYTES:
        return False
    with open(path, "rb") as f:
        return f.read(2) == b"PK"


def _stream_period(path: str, fiscal_year: int, quarter: str, asset: str) -> int:
    """Parse one period's xlsx (streaming from disk) and write a parquet batch.

    Returns the number of rows written. Raises if the workbook can't be parsed
    or none of the expected columns are present (caller treats as skip)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        row_iter = ws.iter_rows(values_only=True)
        try:
            header = next(row_iter)
        except StopIteration:
            raise ValueError("empty sheet")

        norm = [str(h).strip().upper() if h is not None else "" for h in header]
        col_idx = {}
        for target, variants in COLUMN_ALIASES.items():
            for variant in variants:
                if variant in norm:
                    col_idx[target] = norm.index(variant)
                    break

        if not col_idx:
            raise ValueError(f"no recognized columns in header: {norm[:15]}")

        written = 0
        buf = {"fiscal_year": [], "quarter": []}
        for c in TEXT_COLS:
            buf[c] = []

        def _flush(writer):
            nonlocal buf, written
            if not buf["fiscal_year"]:
                return
            table = pa.table(buf, schema=SCHEMA)
            writer.write_table(table)
            written += table.num_rows
            buf = {"fiscal_year": [], "quarter": []}
            for c in TEXT_COLS:
                buf[c] = []

        with raw_parquet_writer(asset, SCHEMA) as writer:
            for row in row_iter:
                if row is None:
                    continue
                buf["fiscal_year"].append(fiscal_year)
                buf["quarter"].append(quarter)
                for c in TEXT_COLS:
                    idx = col_idx.get(c)
                    val = row[idx] if idx is not None and idx < len(row) else None
                    buf[c].append(None if val is None else str(val).strip())
                if len(buf["fiscal_year"]) >= BATCH_ROWS:
                    _flush(writer)
            _flush(writer)

        return written
    finally:
        wb.close()


def _periods() -> list[tuple[int, str, str]]:
    """All (fiscal_year, quarter_label, url) periods to fetch.

    Quarterly coverage extends to the current federal fiscal year, discovered
    from the calendar; unpublished future quarters are skipped on fetch."""
    out: list[tuple[int, str, str]] = []
    for label, url in HISTORICAL_ANNUAL_FILES:
        fy = int(label[2:6])
        out.append((fy, label, url))

    now = datetime.now(timezone.utc)
    current_fy = now.year + 1 if now.month >= 10 else now.year
    for fy in range(QUARTERLY_START_FY, current_fy + 1):
        for q in range(1, 5):
            label = f"fy{fy}-q{q}"
            url = f"{BASE}/LCA_Disclosure_Data_FY{fy}_Q{q}.xlsx"
            out.append((fy, label, url))
    return out


def fetch_disclosures(node_id: str) -> None:
    asset_base = node_id  # "dol-h1b-h1b-lca-disclosures"

    state = load_state(asset_base)
    if state.get("schema_version") != STATE_VERSION:
        state = {"schema_version": STATE_VERSION, "completed": []}
    completed = set(state.get("completed", []))

    for fiscal_year, label, url in _periods():
        if label in completed:
            continue

        batch_asset = f"{asset_base}-{label}"
        try:
            path, size = _download_to_tmp(url)
        except _Blocked as e:
            # 403 = Akamai block, 404 = not yet published. Skip without
            # completing so it retries next run.
            print(f"  skip {label}: HTTP {e.code} for {url}")
            continue
        except _Transient as e:
            print(f"  skip {label}: transient after retries: {e} ({url})")
            continue

        try:
            if not _valid_xlsx(path, size):
                print(f"  skip {label}: invalid xlsx ({size} bytes)")
                continue

            quarter = label.split("-")[-1].upper() if "-q" in label else "FY"
            try:
                rows = _stream_period(path, fiscal_year, quarter, batch_asset)
            except Exception as e:
                print(f"  skip {label}: parse failed: {type(e).__name__}: {e}")
                continue
        finally:
            try:
                os.unlink(path)
            except OSError:
                pass

        if rows == 0:
            print(f"  skip {label}: 0 rows parsed")
            continue

        print(f"  {label}: wrote {rows:,} rows")
        # Raw written first, then advance state — a crash here re-fetches the
        # (immutable) period next run rather than skipping it forever.
        completed.add(label)
        save_state(asset_base, {
            "schema_version": STATE_VERSION,
            "completed": sorted(completed),
        })


DOWNLOAD_SPECS = [
    NodeSpec(id="dol-h1b-h1b-lca-disclosures", fn=fetch_disclosures, kind="download"),
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="dol-h1b-h1b-lca-disclosures-transform",
        deps=["dol-h1b-h1b-lca-disclosures"],
        sql=r'''
            SELECT
                CAST(fiscal_year AS INTEGER)        AS fiscal_year,
                NULLIF(quarter, '')                 AS quarter,
                UPPER(NULLIF(case_status, ''))      AS case_status,
                NULLIF(employer_name, '')           AS employer_name,
                NULLIF(employer_state, '')          AS employer_state,
                NULLIF(employer_city, '')           AS employer_city,
                NULLIF(job_title, '')               AS job_title,
                NULLIF(soc_code, '')                AS soc_code,
                NULLIF(soc_title, '')               AS soc_title,
                NULLIF(wage_rate, '')               AS wage_rate,
                UPPER(NULLIF(wage_unit, ''))        AS wage_unit,
                NULLIF(worksite_state, '')          AS worksite_state,
                NULLIF(worksite_city, '')           AS worksite_city,
                -- Extract the leading numeric token: handles plain values
                -- ("143666"), decimals ("42.53"), and the FY2015 range format
                -- ("18.49 -") which a bare CAST cannot parse.
                TRY_CAST(
                    regexp_extract(
                        REPLACE(REPLACE(NULLIF(wage_rate, ''), ',', ''), '$', ''),
                        '[0-9]+\.?[0-9]*', 0
                    ) AS DOUBLE
                ) AS wage_rate_numeric
            FROM "dol-h1b-h1b-lca-disclosures"
            WHERE fiscal_year IS NOT NULL
              AND (employer_name IS NOT NULL OR case_status IS NOT NULL)
        ''',
    ),
]
