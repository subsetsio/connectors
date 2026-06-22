import io, json
from subsets_utils import get
import openpyxl

eu = json.load(open('/Users/nathansnellaert/Documents/hardened/data/sources/clio-infra/work/entity_union.json'))
ids = eu if isinstance(eu, list) else (eu.get('entities') or list(eu.keys()))
print("union type:", type(eu).__name__, "n=", len(ids), "sample:", ids[:3])
specids = {}
for e in ids:
    sid = f"clio-infra-{e.lower().replace('_','-')}"
    specids.setdefault(sid, []).append(e)
collisions = {k: v for k, v in specids.items() if len(v) > 1}
print("collisions:", collisions)

for stem in ["GDPperCapita", "ArmedConflicts(International)", "GoldStandard", "Polity2Index", "Long-TermGovernmentBondYield"]:
    url = f"https://clio-infra.eu/data/{stem}_Compact.xlsx"
    r = get(url, timeout=(10, 120))
    print("===", stem, r.status_code, r.headers.get("content-type"), len(r.content))
    if r.status_code != 200:
        continue
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    print("  sheets:", wb.sheetnames)
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == "data long format":
            ws = wb[name]
            break
    rows = list(ws.iter_rows(values_only=True))
    print("  header:", rows[0], "nrows:", len(rows))
    for rr in rows[1:4]:
        print("   ", rr, [type(x).__name__ for x in rr])
    vals = [rr[3] for rr in rows[1:300] if rr[3] is not None]
    print("  value types(sample):", set(type(v).__name__ for v in vals))
    ccodes = [rr[0] for rr in rows[1:300] if rr[0] is not None]
    print("  ccode types(sample):", set(type(v).__name__ for v in ccodes))
