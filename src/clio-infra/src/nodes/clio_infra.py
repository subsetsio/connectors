"""Clio Infra connector — historical socio-economic indicator time series.

Mechanism: bulk_xlsx. One stable XLSX per indicator at
``https://clio-infra.eu/data/<stem>_Compact.xlsx`` (modern-borders variant).
Each workbook has a tidy "Data Long Format" sheet with columns
(ccode, country.name, year, value) — the only sheet we read.

Shape: stateless full re-pull. Each file is the entire global time series for
one indicator (~0.1-2MB, well under any RAM concern), so every run fetches the
whole file and overwrites — no watermark, no cursor. Revisions are picked up
for free. One download spec per indicator (the entity union); one SQL transform
publishes one Delta table per indicator.
"""
import io

import pyarrow as pa
from subsets_utils import NodeSpec, get, transient_retry, save_raw_parquet
from constants import ENTITY_IDS

# spec id (lowercased, '_'->'-') -> original indicator file stem (the URL needs
# the exact upstream casing, e.g. "ArmedConflicts(International)").
_SPEC_TO_STEM = {f"clio-infra-{e.lower().replace('_', '-')}": e for e in ENTITY_IDS}

_LONG_SHEET = "data long format"

# Tidy panel schema, identical across every indicator. ccode = numeric country
# code (UN M49); value carries indicator-specific units (counts, ratios, dollars,
# index points) and may be negative (e.g. Polity2 democracy score).
SCHEMA = pa.schema([
    ("ccode", pa.int64()),
    ("country", pa.string()),
    ("year", pa.int32()),
    ("value", pa.float64()),
])


@transient_retry()  # 6 attempts, exponential backoff over transient net errors + 429/5xx
def _fetch_xlsx(url: str, referer: str) -> bytes:
    # The host is plain nginx and serves these files to a bare request from a
    # residential IP. It answered 403 to every request from the GitHub Actions
    # runners (run 20260709-164551, first request, no rate limiting). A Referer
    # covers the hotlink-protection case; if the deny is by IP the body below
    # carries nginx's reason into the run log.
    resp = get(url, timeout=(10.0, 120.0), headers={"Referer": referer})
    if resp.status_code == 403:
        raise AssertionError(
            f"403 fetching {url} (referer={referer}); "
            f"server={resp.headers.get('server')!r} body={resp.text[:300]!r}"
        )
    resp.raise_for_status()
    return resp.content


def _to_int(x):
    return None if x is None else int(round(float(x)))


def fetch_one(node_id: str) -> None:
    import openpyxl

    asset = node_id  # the runtime passes the spec id; it IS the asset name
    stem = _SPEC_TO_STEM[node_id]
    url = f"https://clio-infra.eu/data/{stem}_Compact.xlsx"
    content = _fetch_xlsx(url, referer=f"https://clio-infra.eu/Indicators/{stem}.html")

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = next((wb[n] for n in wb.sheetnames if n.strip().lower() == _LONG_SHEET), None)
    if ws is None:
        raise AssertionError(f"{asset}: no 'Data Long Format' sheet in {url}; sheets={wb.sheetnames}")

    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip().lower() if h is not None else "" for h in next(rows)]
    idx = {name: header.index(name) for name in ("ccode", "country.name", "year", "value") if name in header}
    missing = {"ccode", "country.name", "year", "value"} - set(idx)
    if missing:
        raise AssertionError(f"{asset}: long sheet missing columns {missing}; header={header}")

    ccode, country, year, value = [], [], [], []
    for r in rows:
        yv, vv = r[idx["year"]], r[idx["value"]]
        if yv is None or vv is None:
            continue  # long sheet is already observation-only; guard against stray blanks
        ccode.append(_to_int(r[idx["ccode"]]))
        cn = r[idx["country.name"]]
        country.append(None if cn is None else str(cn))
        year.append(_to_int(yv))
        value.append(float(vv))

    table = pa.table({"ccode": ccode, "country": country, "year": year, "value": value}, schema=SCHEMA)
    save_raw_parquet(table, asset)


DOWNLOAD_SPECS = [
    NodeSpec(id=f"clio-infra-{eid.lower().replace('_', '-')}", fn=fetch_one, kind="download")
    for eid in ENTITY_IDS
]
