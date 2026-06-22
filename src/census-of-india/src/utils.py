"""Shared helpers for the Census of India connector.

Two jobs:

1. TLS: censusindia.gov.in serves only its leaf certificate and omits the
   emSign intermediate, so the default certifi chain can't be built. We supply
   the missing intermediate (issued by the emSign root that IS in certifi) and
   install an httpx client that trusts the completed chain. This is NOT
   `verify=False` — verification is fully enforced, we just hand it the cert the
   server should have sent.

2. Parsing: ORGI publishes each census table as an Excel workbook (xlsx for the
   1991/2011 series, legacy .xls for 2001) with a multi-row, often merged header
   followed by a row of column numbers ("1","2","3",...) that marks where the
   data begins. `parse_census_excel` keys off that marker row, reconstructs
   column names from the header block, forward-fills the merged left-hand
   identifier columns, and infers per-column numeric vs text type.
"""

import io
import json
import re
import ssl

import certifi
import httpx

import subsets_utils.http_client as _hc

# emSign SSL CA - G1 (subject), issued by emSign Root CA - G1 (in certifi).
# Fetched from the leaf's AIA: http://repository.emsign.com/certs/emSignSSLCAG1.crt
_EMSIGN_INTERMEDIATE_PEM = """-----BEGIN CERTIFICATE-----
MIIEgjCCA2qgAwIBAgIKIXrVixxxPAAgkTANBgkqhkiG9w0BAQsFADBnMQswCQYD
VQQGEwJJTjETMBEGA1UECxMKZW1TaWduIFBLSTElMCMGA1UEChMcZU11ZGhyYSBU
ZWNobm9sb2dpZXMgTGltaXRlZDEcMBoGA1UEAxMTZW1TaWduIFJvb3QgQ0EgLSBH
MTAeFw0xODAyMTgxODMwMDBaFw0zMzAyMTgxODMwMDBaMGYxCzAJBgNVBAYTAklO
MRMwEQYDVQQLEwplbVNpZ24gUEtJMSUwIwYDVQQKExxlTXVkaHJhIFRlY2hub2xv
Z2llcyBMaW1pdGVkMRswGQYDVQQDExJlbVNpZ24gU1NMIENBIC0gRzEwggEiMA0G
CSqGSIb3DQEBAQUAA4IBDwAwggEKAoIBAQCU1fhvfUV6OJOhMHAzBZDvxxpa5bvT
S1x6S4rVQmP/+125Wj2gpxvtII2RyqXQFlZd3qAKMLgqgHGzeJcyjw6CXbzJHmri
liVGWmuLn/NUKjKJgP9zd6eGOHe6mT1WjB9ZZEFLsDYoXBthTwcHdLWK2quJrHgS
3hZiJnpkX+hmaY7DX89oUMI1uvCQaPljgTvtiR9vtmeg/GgyePX8K5EUMozX8ElR
DMWkzdFUYv0DVcSQcbN1R/IDhWW1vPHzU8kexAMO4B/E5sj6FGrAeMM36/uZ3AmF
4mt/0Ia2BKPsW/K2T3hkaNSTr2BKlUm6bRccONcNAyzyN258xcUcX+RJAgMBAAGj
ggEvMIIBKzAfBgNVHSMEGDAWgBT77w2GnrDj3am58SEXfz788HcrGjAdBgNVHQ4E
FgQUNNH3OTJFQEqZK32JaldprZWv4zcwDgYDVR0PAQH/BAQDAgEGMB0GA1UdJQQW
MBQGCCsGAQUFBwMBBggrBgEFBQcDAjA9BgNVHSAENjA0MDIGBFUdIAAwKjAoBggr
BgEFBQcCARYcaHR0cDovL3JlcG9zaXRvcnkuZW1zaWduLmNvbTASBgNVHRMBAf8E
CDAGAQH/AgEAMDIGCCsGAQUFBwEBBCYwJDAiBggrBgEFBQcwAYYWaHR0cDovL29j
c3AuZW1zaWduLmNvbTAzBgNVHR8ELDAqMCigJqAkhiJodHRwOi8vY3JsLmVtc2ln
bi5jb20/Um9vdENBRzEuY3JsMA0GCSqGSIb3DQEBCwUAA4IBAQAaBDZfBK+cP9Zk
lI7QN3mkpgD+mYfp/03P51cUNlfAFoYd1G/4lU468rg7JLTwqFXcDzcmrWt8lmdi
AMflxwLGeNObNS9RkpdiMDCdRItCHq00IMbbzj5rz+HSzAn6WsbLn9efn9WhO1MO
72d1SsEbVOTw/Z3sfPpWS8DSp91TRZuRKReVmD967QnsQGYNKUG6esTV73dOigHC
ndwglIXCUkaxTroFn7wT6Sqt9pklaqxBkEx/yzp0HxpZtC8uK6aOFx624S9yF8nk
6U7rbscn4kJYOF+0U9JshFkQ4+cx5kKd3cGNtmaTzemoZSGn+Aty6H6/oDPteLpE
cUPckzSa
-----END CERTIFICATE-----
"""

_CA_INSTALLED = False


def install_ca() -> None:
    """Install an httpx client into subsets_utils that trusts the completed
    emSign chain. Idempotent; call once at the top of each fetch fn (each spec
    runs in its own subprocess, so the module-level client starts fresh)."""
    global _CA_INSTALLED
    if _CA_INSTALLED and _hc._client is not None:
        return
    ctx = ssl.create_default_context(cafile=certifi.where())
    ctx.load_verify_locations(cadata=_EMSIGN_INTERMEDIATE_PEM)
    if _hc._client is not None:
        _hc._client.close()
    _hc._client = httpx.Client(
        timeout=_hc._client_config.get("timeout", 60),
        headers=_hc._client_config.get("headers", {"User-Agent": "DataIntegrations/1.0"}),
        follow_redirects=True,
        verify=ctx,
    )
    _CA_INSTALLED = True


# ----------------------------------------------------------------------------
# Excel parsing
# ----------------------------------------------------------------------------

_NIL_TOKENS = {"", "-", "--", "---", "----", "n", "na", "n.a.", "n.a", "nil",
               "neg", "@", "$", "&", "*", "..", "...", "x"}


def _clean_text(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _to_number(v):
    """Coerce a census cell to a number, or None for nil/footnote-only cells."""
    s = _clean_text(v)
    if s == "" or s.lower() in _NIL_TOKENS:
        return None
    # strip footnote markers and thousands separators, keep digits/sign/decimal
    s2 = re.sub(r"[,\s]", "", s)
    s2 = re.sub(r"[*@$&#%]+", "", s2)
    s2 = s2.replace("--", "")
    m = re.fullmatch(r"[-+]?\d*\.?\d+", s2)
    if not m:
        return None
    f = float(s2)
    return int(f) if f.is_integer() else f


def _snake(name, used):
    s = _clean_text(name).lower()
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    if not s:
        s = "col"
    base = s[:60]
    out = base
    i = 2
    while out in used:
        out = f"{base}_{i}"
        i += 1
    used.add(out)
    return out


def _read_matrix(blob: bytes, filename: str):
    """Return a dense list-of-lists of cell values with merged regions filled."""
    ext = filename.lower().split("?")[0].rsplit(".", 1)[-1]
    if ext == "xls":
        return _read_xls(blob)
    return _read_xlsx(blob)


def _read_xlsx(blob: bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        nrows, ncols = ws.max_row or 0, ws.max_column or 0
        grid = [[None] * ncols for _ in range(nrows)]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    grid[cell.row - 1][cell.column - 1] = cell.value
        # fill merged regions (top-left value spreads to the whole range)
        for rng in list(ws.merged_cells.ranges):
            val = grid[rng.min_row - 1][rng.min_col - 1]
            if val is None:
                continue
            for r in range(rng.min_row - 1, rng.max_row):
                for c in range(rng.min_col - 1, rng.max_col):
                    grid[r][c] = val
        return grid
    finally:
        wb.close()


def _read_xls(blob: bytes):
    import xlrd
    try:
        wb = xlrd.open_workbook(file_contents=blob, formatting_info=True)
        have_merge = True
    except Exception:
        wb = xlrd.open_workbook(file_contents=blob)
        have_merge = False
    sh = wb.sheet_by_index(0)
    grid = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    if have_merge:
        for rlo, rhi, clo, chi in sh.merged_cells:
            val = grid[rlo][clo]
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    grid[r][c] = val
    return grid


def _find_marker(grid):
    """Locate the column-number row that ORGI puts above the data: its non-blank
    cells begin 1,2,3,... and increment by one. We walk that contiguous prefix
    left-to-right and stop at the first break, so a stray integer elsewhere in
    the row (footnote, page number) doesn't disqualify it. The prefix columns
    are the real data columns. Returns (row_index, [cols]) or (None, None)."""
    for i, row in enumerate(grid[:40]):
        cols = []
        expect = 1
        for c, v in enumerate(row):
            s = _clean_text(v)
            if s == "":
                continue
            m = re.fullmatch(r"(\d+)(?:\.0+)?", s)
            if m and int(m.group(1)) == expect:
                cols.append(c)
                expect += 1
            else:
                break
        if len(cols) >= 5:
            return i, cols
    return None, None


def parse_census_excel(blob: bytes, filename: str):
    """Parse one ORGI census workbook into a list of row dicts.

    Returns [] if no marker row is found (caller treats that as a parse miss).
    """
    grid = _read_matrix(blob, filename)
    if not grid:
        return []
    marker_idx, cols = _find_marker(grid)
    if marker_idx is None:
        return []

    header_rows = grid[:marker_idx]
    data_rows = grid[marker_idx + 1:]

    # Build a name per real column from the header block, skipping caption/title
    # rows (a single filled cell, or one value spanning every column).
    parts = {c: [] for c in cols}
    for hrow in header_rows:
        filled = [(c, _clean_text(hrow[c])) for c in cols
                  if c < len(hrow) and _clean_text(hrow[c]) != ""]
        if len(filled) <= 1:
            continue
        if len(set(v for _, v in filled)) == 1:  # one value merged across all
            continue
        for c, v in filled:
            if not parts[c] or parts[c][-1] != v:
                parts[c].append(v)
    used = set()
    names = {c: _snake(" ".join(parts[c]) if parts[c] else f"col{idx + 1}", used)
             for idx, c in enumerate(cols)}

    # Drop fully-blank data rows.
    def blank(row):
        return all(c >= len(row) or _clean_text(row[c]) == "" for c in cols)

    data_rows = [r for r in data_rows if not blank(r)]
    if not data_rows:
        return []

    # Classify each real column as value (numeric) vs identifier, using a clean
    # HEAD sample of data rows — the contiguous block right under the marker is
    # real data, while footnotes/notes collect at the bottom and would otherwise
    # depress density. Nil tokens ("---", footnotes) count as numeric-compatible:
    # they are nulls within a numeric column, not evidence the column is text.
    sample = data_rows[:50]
    ns = len(sample)
    is_value = {}
    for c in cols:
        nonblank = [r[c] for r in sample if c < len(r) and _clean_text(r[c]) != ""]
        if not nonblank:
            is_value[c] = False
            continue
        dense = len(nonblank) >= 0.5 * ns
        numbers = sum(1 for v in nonblank if _to_number(v) is not None)
        nilish = sum(1 for v in nonblank if _clean_text(v).lower() in _NIL_TOKENS)
        numericish = numbers + nilish
        is_value[c] = (dense
                       and numericish >= 0.8 * len(nonblank)
                       and numbers >= 0.3 * len(nonblank))

    value_cols = [c for c in cols if is_value[c]]
    if not value_cols:
        return []

    # Drop junk rows: a real data row populates at least one value column.
    # Footnotes/notes (text only in the left margin) carry no value and go.
    data_rows = [r for r in data_rows
                 if any(c < len(r) and _clean_text(r[c]) != "" for c in value_cols)]
    if not data_rows:
        return []

    # Identifier block = leading real columns up to the first value column.
    id_block = []
    for c in cols:
        if is_value[c]:
            break
        id_block.append(c)

    # Forward-fill the identifier block (merged left-hand cells repeat downward).
    last = {c: None for c in id_block}
    out = []
    for r in data_rows:
        rec = {}
        for c in cols:
            cell = r[c] if c < len(r) else None
            if c in id_block:
                txt = _clean_text(cell)
                if txt == "":
                    txt = last[c]
                else:
                    last[c] = txt
                rec[names[c]] = txt if txt not in (None, "") else None
            elif is_value[c]:
                rec[names[c]] = _to_number(cell)
            else:
                txt = _clean_text(cell)
                rec[names[c]] = txt if txt != "" else None
        out.append(rec)
    return out


# ----------------------------------------------------------------------------
# Wide → tidy-long reshape
# ----------------------------------------------------------------------------
#
# Census tables span ~200 layouts with heterogeneous, often merged headers, so
# there is no single stable wide schema to publish — and within a multi-file
# entity (one .xls per state/UT) the geography column is even named differently
# from file to file (e.g. `india_state_union_territory` vs `state_district`).
# We melt every parsed workbook into a uniform long shape — (region, dimensions,
# measure, value) — so all five entities, .xls and .xlsx alike, land on ONE raw
# schema the SQL transform can publish from unchanged. Identifier columns are
# preserved verbatim in the `dimensions` JSON; only genuine numeric measures
# become value rows.

# A column whose NAME matches this is an identifier/dimension even when its
# cells are numeric: geographic/series codes, serial numbers, and the decadal
# `census_year` of an A-02 table (a time dimension, not a measured quantity).
_ID_NAME_RE = re.compile(
    r"(^|_)(code|sl_no|s_no|sr_no|serial|year|decade|t_r_u|tru|name|table_name)(_|$)"
)
_HAS_LETTER = re.compile(r"[A-Za-z]")

# Identifier columns whose NAME denotes a geographic place — preferred as the
# `region` label over generic codes (e.g. an A-11 table's constant `table_name`
# "A10ST1", which is a series code, not a place).
_GEO_NAME_RE = re.compile(
    r"(area_name|state|district|union_territory|territory|region|place|town|village|nation|ward|block)"
)


def _is_measure_col(col, rows):
    """A measure column is numeric across every non-null cell AND not named like
    an identifier. Anything else (text, or a numeric code/year) is a dimension."""
    if _ID_NAME_RE.search(col):
        return False
    seen_number = False
    for r in rows:
        v = r.get(col)
        if v is None:
            continue
        if isinstance(v, bool) or not isinstance(v, (int, float)):
            return False
        seen_number = True
    return seen_number


def excel_to_long(blob: bytes, filename: str):
    """Parse one census workbook and melt it to tidy-long rows.

    Returns a list of {region, dimensions, measure, value} dicts (value always a
    non-null float). Empty list if the workbook has no recognizable data block —
    the caller decides whether that's tolerable for the entity as a whole.
    """
    rows = parse_census_excel(blob, filename)
    if not rows:
        return []
    cols = list(rows[0].keys())
    measure_cols = [c for c in cols if _is_measure_col(c, rows)]
    id_cols = [c for c in cols if c not in measure_cols]

    # region: a human-readable geography label — the first identifier column
    # whose values are mostly alphabetic (the place name, not a numeric code).
    region_col = None
    for c in id_cols:
        sample = [str(r[c]) for r in rows[:30] if r.get(c) is not None]
        if sample and sum(1 for v in sample if _HAS_LETTER.search(v)) >= 0.6 * len(sample):
            region_col = c
            break
    if region_col is None and id_cols:
        region_col = id_cols[0]

    out = []
    for r in rows:
        dims = json.dumps({c: r.get(c) for c in id_cols}, ensure_ascii=False)
        region = r.get(region_col) if region_col else None
        region = str(region) if region is not None else None
        for m in measure_cols:
            v = r.get(m)
            if v is None:
                continue
            out.append({
                "region": region,
                "dimensions": dims,
                "measure": m,
                "value": float(v),
            })
    return out
