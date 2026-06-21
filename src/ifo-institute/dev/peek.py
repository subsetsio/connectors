import io, os, re, sys, time
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
from subsets_utils import get, configure_http
import openpyxl

configure_http(headers={
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
    "Referer": "https://www.ifo.de/en/ifo-time-series",
    "Accept-Language": "en-US,en;q=0.9",
})
BASE = "https://www.ifo.de"

FILES = {
    "gsk":      "/sites/default/files/secure/timeseries/gsk-e-202605.xlsx",
    "export":   "/sites/default/files/secure/timeseries/export-e-202605.xlsx",
    "empl":     "/sites/default/files/secure/timeseries/empl-e-202605.xlsx",
    "ostd":     "/sites/default/files/secure/timeseries/ostd-e-202605.xlsx",
    "sachsen":  "/sites/default/files/secure/timeseries/ku-sachsen-202604-LR-en.xlsx",
    "exklima":  "/sites/default/files/secure/timeseries/exklima-e-202604.xlsx",
    "imklima":  "/sites/default/files/secure/timeseries/imklima-e-202604.xlsx",
    "vintage":  "/sites/default/files/facts/vintage/Germany-ifo-vintage.xlsx",
}

want = sys.argv[1:] or list(FILES)

for key in want:
    path = FILES[key]
    r = get(BASE + path, timeout=(10.0, 120.0))
    ct = r.headers.get("content-type", "")
    print(f"\n######## {key}  {r.status_code} {ct} {len(r.content)}B")
    if not r.content[:2] == b"PK":
        print("  NOT a zip/xlsx (bot challenge?) -- first 200 bytes:")
        print("  ", r.content[:200])
        time.sleep(7); continue
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    print("  sheets:", wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"  --- sheet '{sn}'  dims={ws.calculate_dimension()}")
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(row)
            if i >= 14:
                break
        for i, row in enumerate(rows):
            # trim long rows
            disp = [("" if c is None else (str(c)[:18])) for c in row[:12]]
            print(f"    r{i}: {disp}")
    wb.close()
    time.sleep(7)
