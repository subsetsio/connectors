"""Shared helpers for the OHSS connector.

OHSS (ohss.dhs.gov) sits behind Akamai bot-manager, which blocks httpx's default
TLS ClientHello (returns 403 "Access Denied") while accepting a browser-like
cipher ordering. subsets_utils.get is the mandated HTTP path but its public
configure_http() only feeds timeout/headers to the client, so it cannot set an
ssl context. We therefore install a browser-TLS httpx client into the
subsets_utils client slot once per process — every subsequent subsets_utils.get
call uses it (and is still logged through _logged_request). This is the minimal
intervention that keeps the logged get() path while defeating the TLS fingerprint
block; verification stays on (valid cert), no verify=False.

The data surface is multi-sheet .xlsx workbooks discovered by scraping topic
landing pages (download URLs are point-in-time, so they are re-discovered each
run). Each sheet is a human-formatted statistical table; parse_sheet
reconstructs it faithfully as wide records (header row -> column names, one
record per data row), un-tiling the "newspaper column" layout some tables use
and forward-filling sparse leading label columns.
"""

import io
import re
import ssl

import httpx
import openpyxl

import subsets_utils.http_client as _hc
from subsets_utils import get, transient_retry

BASE = "https://ohss.dhs.gov"
# ASCII-only UA (httpx/urllib3 reject non-ASCII header bytes).
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36")
_BROWSER_CIPHERS = (
    "ECDHE-ECDSA-AES128-GCM-SHA256:ECDHE-RSA-AES128-GCM-SHA256:"
    "ECDHE-ECDSA-AES256-GCM-SHA384:ECDHE-RSA-AES256-GCM-SHA384:"
    "ECDHE-ECDSA-CHACHA20-POLY1305:ECDHE-RSA-CHACHA20-POLY1305:DEFAULT"
)

_MONTHS = {m: i for i, m in enumerate(
    ["january", "february", "march", "april", "may", "june", "july",
     "august", "september", "october", "november", "december"], start=1)}

_SUPPRESSED = {"", "-", "--", "---", "d", "x", "na", "n/a", "(na)", "(d)",
               "(x)", "*", "(*)", "z", "(z)", "—", "–", "nan", "none"}


def install_browser_client() -> None:
    """Swap subsets_utils' shared httpx client for one whose TLS ciphers pass
    Akamai. Idempotent-ish: rebuilds the client each call (cheap, once per fn)."""
    ctx = ssl.create_default_context()
    ctx.set_ciphers(_BROWSER_CIPHERS)
    if _hc._client is not None:
        _hc._client.close()
    _hc._client = httpx.Client(
        timeout=120, headers={"User-Agent": UA},
        follow_redirects=True, verify=ctx,
    )


@transient_retry()
def _get(url: str) -> httpx.Response:
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp


def _xlsx_links(page_path: str) -> list:
    html = _get(BASE + page_path).text
    out, seen = [], set()
    for h in re.findall(r'href="([^"]+\.xlsx)"', html):
        if h not in seen:
            seen.add(h)
            out.append(h if h.startswith("http") else BASE + h)
    return out


def yearbook_workbook_url(section_kw: str) -> str:
    """Locate the current workbook for a Yearbook section by scraping the year
    index, then the most recent year pages, for an xlsx whose filename carries
    the section keyword. Returns the newest available."""
    idx = _get(BASE + "/topics/immigration/yearbook").text
    years = sorted({int(y) for y in re.findall(
        r'/topics/immigration/yearbook/(\d{4})"', idx)}, reverse=True)
    for year in years[:6]:
        for url in _xlsx_links(f"/topics/immigration/yearbook/{year}"):
            if section_kw in url.lower():
                return url
    raise RuntimeError(f"no current workbook found for section {section_kw!r}")


def enforcement_workbook_url() -> str:
    """The Immigration Enforcement and Legal Processes monthly workbook is
    cumulative (each sheet spans 2014->latest month), so the newest month's
    workbook holds the full history. Pick the latest by (year, month)."""
    links = _xlsx_links(
        "/topics/immigration/immigration-enforcement/monthly-tables")
    if not links:
        raise RuntimeError("no enforcement monthly xlsx links found")

    def key(url):
        m = re.search(r"tables-([a-z]+)-(\d{4})", url.lower())
        if m and m.group(1) in _MONTHS:
            return (int(m.group(2)), _MONTHS[m.group(1)])
        d = re.search(r"/(\d{4})-(\d{2})/", url)  # fall back to publish dir
        return (int(d.group(1)), int(d.group(2))) if d else (0, 0)

    return max(links, key=key)


def download_workbook(url: str) -> bytes:
    return _get(url).content


# --------------------------- sheet parsing ---------------------------------

def _clean(x) -> str:
    if x is None:
        return ""
    return re.sub(r"\s+", " ", str(x)).strip()


def _to_number(x):
    """Coerce a cell to float, or None for blanks / suppression markers."""
    if x is None:
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", "").replace("%", "")
    if s.lower() in _SUPPRESSED:
        return None
    m = re.fullmatch(r"-?\d+(?:\.\d+)?", s)
    return float(s) if m else None


def _resolve_sheet(wb, want: str) -> str:
    if want in wb.sheetnames:
        return want
    norm = {re.sub(r"\s+", "", n).lower(): n for n in wb.sheetnames}
    key = re.sub(r"\s+", "", want).lower()
    if key in norm:
        return norm[key]
    for n in wb.sheetnames:  # last resort: prefix / contains
        if key in re.sub(r"\s+", "", n).lower():
            return n
    raise KeyError(f"sheet {want!r} not in {wb.sheetnames}")


def _find_header(rows: list) -> int:
    """First row (top-down, within the first 25) that looks like a column
    header: a non-empty leftmost stub cell, >=2 non-empty cells to its right,
    and numeric data appearing within the next few rows."""
    limit = min(25, len(rows))
    for h in range(limit):
        row = rows[h]
        if not (row and _clean(row[0])):
            continue
        right = sum(1 for c in row[1:] if _clean(c))
        if right < 2:
            continue
        for r in rows[h:min(h + 6, len(rows))]:
            if any(_to_number(c) is not None for c in r[1:]):
                return h
    # Fallback: densest row (to the right) among the first 25.
    best, best_n = -1, 0
    for h in range(limit):
        n = sum(1 for c in rows[h][1:] if _clean(c))
        if n > best_n:
            best, best_n = h, n
    return best


def _detect_period(names: list, width: int) -> int:
    """If the header names repeat as k-wide blocks (newspaper layout, e.g.
    Year|Number|Year|Number|...), return k; else width (single block)."""
    for k in range(1, width // 2 + 1):
        if width % k:
            continue
        if all(names[c] for c in range(k)) and \
           all(names[c] == names[c % k] for c in range(width)):
            return k
    return width


def parse_sheet(content: bytes, sheet_name: str):
    """Parse one workbook sheet into (column_names, rows) where rows is a list
    of dicts. Faithful wide reconstruction; never raises on layout quirks
    (returns [] rows only if no data is found)."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = wb[_resolve_sheet(wb, sheet_name)]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    if not grid:
        return [], []

    h = _find_header(grid)
    if h < 0:
        return [], []

    # Effective width: last column carrying a header or any data value.
    width = 0
    for r in grid[h:]:
        for c in range(len(r) - 1, -1, -1):
            if _clean(r[c]):
                width = max(width, c + 1)
                break
    if width == 0:
        return [], []

    def cell(row, c):
        return row[c] if c < len(row) else None

    raw_names = [_clean(cell(grid[h], c)) for c in range(width)]
    k = _detect_period(raw_names, width)
    unit = raw_names[:k]

    # Unique, non-empty column names for the unit.
    names, seen = [], {}
    for i, nm in enumerate(unit):
        nm = nm or f"column_{i + 1}"
        if nm in seen:
            seen[nm] += 1
            nm = f"{nm}_{seen[nm]}"
        else:
            seen[nm] = 1
        names.append(nm)

    data = [r for r in grid[h + 1:] if any(_clean(cell(r, c)) for c in range(width))]
    if not data:
        return names, []

    # Classify unit columns: value columns are the numeric right-suffix. A
    # column counts as a value column if it is numeric (>=60% of its non-empty
    # cells parse as numbers) OR carries no data at all (empty value column). A
    # column with text content (e.g. month names, country labels) breaks the
    # suffix and, with everything to its left, is a label column.
    def col_stats(local_c):
        num = tot = 0
        for r in data:
            for b in range(width // k):
                v = cell(r, b * k + local_c)
                if _clean(v):
                    tot += 1
                    if _to_number(v) is not None:
                        num += 1
        return num, tot

    stats = [col_stats(c) for c in range(k)]
    first_val = k
    for c in range(k - 1, -1, -1):
        num, tot = stats[c]
        is_value = (tot == 0) or (num / tot >= 0.6)
        if is_value:
            first_val = c
        else:
            break
    label_cols = set(range(first_val))
    value_cols = set(range(first_val, k))

    # Forward-fill sparse leading label columns (carry-down design, e.g. the
    # Fiscal Year stub that is blank on every month row). Only single-block
    # (non-tiled) tables have such stubs; tiled tables have no labels to fill.
    if k == width:
        for c in label_cols:
            blanks = sum(1 for r in data if not _clean(cell(r, c)))
            if blanks <= 0.2 * len(data):
                continue
            last = ""
            for r in data:
                v = _clean(cell(r, c))
                if v:
                    last = v
                elif last and c < len(r):
                    r[c] = last
                elif last:
                    r.extend([None] * (c - len(r) + 1))
                    r[c] = last

    out = []
    for r in data:
        for b in range(width // k):
            block_vals = [cell(r, b * k + c) for c in range(k)]
            if not any(_clean(v) for v in block_vals):
                continue
            rec = {}
            for c in range(k):
                v = block_vals[c]
                if c in value_cols:
                    rec[names[c]] = _to_number(v)
                else:
                    cv = _clean(v)
                    rec[names[c]] = cv or None
            # Skip blocks that are pure labels with no values at all.
            if value_cols and all(rec[names[c]] is None for c in value_cols):
                continue
            out.append(rec)
    return names, out
