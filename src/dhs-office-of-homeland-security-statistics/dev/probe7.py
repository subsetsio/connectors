import io, ssl, httpx, re
import subsets_utils.http_client as hc
from subsets_utils import get
import openpyxl

UA="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
BROWSER_CIPHERS=("ECDHE-ECDSA-AES128-GCM-SHA256:ECDHE-RSA-AES128-GCM-SHA256:"
 "ECDHE-ECDSA-AES256-GCM-SHA384:ECDHE-RSA-AES256-GCM-SHA384:"
 "ECDHE-ECDSA-CHACHA20-POLY1305:ECDHE-RSA-CHACHA20-POLY1305:DEFAULT")

def install():
    ctx=ssl.create_default_context(); ctx.set_ciphers(BROWSER_CIPHERS)
    if hc._client: hc._client.close()
    hc._client=httpx.Client(timeout=120, headers={"User-Agent":UA}, follow_redirects=True, verify=ctx)

install()
# scrape yearbook index -> latest LPR workbook
idx=get("https://ohss.dhs.gov/topics/immigration/yearbook").text
years=sorted({int(y) for y in re.findall(r'/topics/immigration/yearbook/(\d{4})"', idx)}, reverse=True)
print("years", years[:3])
lpr=None
for y in years[:3]:
    page=get(f"https://ohss.dhs.gov/topics/immigration/yearbook/{y}").text
    for h in re.findall(r'href="([^"]+\.xlsx)"', page):
        if "lawful_permanent_residents" in h.lower():
            lpr=h if h.startswith("http") else "https://ohss.dhs.gov"+h; break
    if lpr: print("LPR from year",y,lpr); break

content=get(lpr).content
print("xlsx bytes", len(content))
wb=openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
print("sheets", wb.sheetnames[:6])
ws=wb["Table 2"]
rows=[list(r) for r in ws.iter_rows(values_only=True)]
wb.close()
print("Table2 nrows", len(rows))
for i,r in enumerate(rows[:10]):
    print(i, "|".join(("" if c is None else str(c))[:18] for c in r[:7]))
