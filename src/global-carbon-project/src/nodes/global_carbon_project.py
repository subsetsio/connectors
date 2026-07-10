"""Global Carbon Project — Global Carbon Budget connector.

Mechanism (from research): ``bulk_xlsx`` — the official globalcarbonbudget.org
data hub. The budget is republished once a year as a small set of multi-sheet
XLSX workbooks reachable through opaque, version-specific numeric download ids
(``/download/<id>/``). Those ids change every release, so we DISCOVER the
current workbook urls at run time instead of hardcoding them: the stable hub
page (``/datahub/``) links to the "latest GCB data" page, which lists the three
workbooks. We then download each whole workbook, parse the relevant worksheet(s)
with openpyxl, and write a tidy long-format parquet per subset.

Fetch shape: **stateless full re-pull** (shape 1). The whole corpus is a few MB
across a handful of files and is republished annually, so every run re-fetches
and overwrites — revisions and back-corrections are picked up for free. No
watermark/cursor/state.

Each XLSX sheet is normalised to tidy long format in the fetch fn (SQL can only
read parquet/ndjson/csv, never xlsx):
  - global aggregate sheets  -> (year, series, value)
  - national country x year  -> (country, year, value)
  - national LUC (3 models)   -> (model, country, year, value)
  - regions membership        -> (region, country)
Transforms are NOT authored here: the model stage compiles one Delta table per
subset from the settled raw (thin cast/drop-null passes) as transforms/*.sql.
"""
import io
import re

import openpyxl
import pyarrow as pa

from subsets_utils import (
    NodeSpec,
    get,
    save_raw_parquet,
    transient_retry,
)

SLUG = "global-carbon-project"

# --- subset -> (workbook family, sheet(s), parse mode) ---------------------
# family is resolved to a live download url by _discover_workbooks().
GLOBAL = "global"            # (year, series, value)
NATIONAL = "national"        # (country, year, value)
NATIONAL_MODEL = "national_model"  # (model, country, year, value)
REGIONS = "regions"          # (region, country) — reference membership table

SUBSETS = {
    "global-carbon-budget":         ("gcb", "Global Carbon Budget", GLOBAL),
    "historical-budget":            ("gcb", "Historical Budget", GLOBAL),
    "fossil-emissions-by-category": ("gcb", "Fossil Emissions by Category", GLOBAL),
    "atmospheric-growth":           ("gcb", "Atmospheric Growth", GLOBAL),
    "ocean-sink":                   ("gcb", "Ocean Sink", GLOBAL),
    "terrestrial-sink":             ("gcb", "Terrestrial Sink", GLOBAL),
    "cement-carbonation-sink":      ("gcb", "Cement Carbonation Sink", GLOBAL),
    "national-fossil-territorial-emissions": ("national_fossil", "Territorial Emissions", NATIONAL),
    "national-fossil-consumption-emissions": ("national_fossil", "Consumption Emissions", NATIONAL),
    "national-fossil-emissions-transfers":   ("national_fossil", "Emissions Transfers", NATIONAL),
    "national-fossil-regions":               ("national_fossil", "Regions", REGIONS),
    "national-land-use-change-emissions":    ("national_luc", ["BLUE", "OSCAR", "LUCE"], NATIONAL_MODEL),
}

HUB_URL = "https://globalcarbonbudget.org/datahub/"

SCHEMA_GLOBAL = pa.schema([
    ("year", pa.int32()),
    ("series", pa.string()),
    ("value", pa.float64()),
])
SCHEMA_NATIONAL = pa.schema([
    ("country", pa.string()),
    ("year", pa.int32()),
    ("value", pa.float64()),
])
SCHEMA_NATIONAL_MODEL = pa.schema([
    ("model", pa.string()),
    ("country", pa.string()),
    ("year", pa.int32()),
    ("value", pa.float64()),
])
SCHEMA_REGIONS = pa.schema([
    ("region", pa.string()),
    ("country", pa.string()),
])

_DOWNLOAD_RE = re.compile(
    r'<a[^>]+href="(https://globalcarbonbudget\.org/download/\d+/[^"]*)"[^>]*>(.*?)</a>',
    re.S | re.I,
)
_LATEST_RE = re.compile(r'href="([^"]*latest-gcb-data[^"]*)"', re.I)

# ---------------------------------------------------------------------------
# HTTP with retry/backoff
# ---------------------------------------------------------------------------


@transient_retry()
def _get_text(url: str) -> str:
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.text


@transient_retry()
def _get_bytes(url: str) -> bytes:
    resp = get(url, timeout=(10.0, 300.0))
    resp.raise_for_status()
    return resp.content


# ---------------------------------------------------------------------------
# Discovery — map workbook family -> current download url
# ---------------------------------------------------------------------------
def _classify_family(label: str) -> str | None:
    t = label.lower()
    if "national" in t and "fossil" in t:
        return "national_fossil"
    if "national" in t and "land use change" in t:
        return "national_luc"
    if "global carbon budget" in t and "national" not in t:
        return "gcb"
    return None


def _discover_workbooks() -> dict[str, str]:
    """Resolve the three workbook families to their current /download/<id>/ urls.

    Hub page -> latest-data page -> the three workbook links. Raises if any of
    the three required families is missing (source layout changed) so the
    failure is loud rather than a silent partial.
    """
    hub = _get_text(HUB_URL)
    m = _LATEST_RE.search(hub)
    if not m:
        raise RuntimeError(f"could not find latest-gcb-data link on {HUB_URL}")
    latest_url = m.group(1)
    if latest_url.startswith("/"):
        latest_url = "https://globalcarbonbudget.org" + latest_url

    page = _get_text(latest_url)
    found: dict[str, str] = {}
    for url, raw_label in _DOWNLOAD_RE.findall(page):
        label = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", raw_label)).strip()
        fam = _classify_family(label)
        # first occurrence wins (latest page lists current version first)
        if fam and fam not in found:
            found[fam] = url

    missing = {"gcb", "national_fossil", "national_luc"} - set(found)
    if missing:
        raise RuntimeError(
            f"missing workbook families {sorted(missing)} on {latest_url}; "
            f"found: {sorted(found)}"
        )
    return found


def _load_workbook(url: str):
    content = _get_bytes(url)
    return openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)


# ---------------------------------------------------------------------------
# Cell coercion
# ---------------------------------------------------------------------------
def _as_year(v) -> int | None:
    if v is None:
        return None
    try:
        y = int(float(v))
    except (TypeError, ValueError):
        return None
    return y if 1700 <= y <= 2100 else None


def _as_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _sheet_matrix(ws) -> list[list]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


# ---------------------------------------------------------------------------
# Parsers — sheet -> tidy long rows
# ---------------------------------------------------------------------------
def _parse_global(ws) -> list[dict]:
    """Aggregate sheets: a 'Year' header row, then year-in-rows x series-in-cols.

    Detect the header row by its first cell == 'Year', map every non-empty
    column header to a series name, then unpivot each yearly data row.
    """
    rows = _sheet_matrix(ws)
    hdr_idx = None
    for i, row in enumerate(rows):
        if row and isinstance(row[0], str) and row[0].strip().lower() == "year":
            hdr_idx = i
            break
    if hdr_idx is None:
        raise RuntimeError(f"sheet {ws.title!r}: no 'Year' header row found")

    header = rows[hdr_idx]
    series_cols = {
        c: str(name).strip()
        for c, name in enumerate(header)
        if c >= 1 and name is not None and str(name).strip()
    }
    if not series_cols:
        raise RuntimeError(f"sheet {ws.title!r}: no series columns in header")

    out: list[dict] = []
    for row in rows[hdr_idx + 1:]:
        if not row:
            continue
        year = _as_year(row[0])
        if year is None:
            continue
        for c, series in series_cols.items():
            if c >= len(row):
                continue
            val = _as_float(row[c])
            if val is None:
                continue
            out.append({"year": year, "series": series, "value": val})
    return out


def _parse_national(ws, model: str | None = None) -> list[dict]:
    """Country sheets: country names across columns, year in rows.

    The country header is the row immediately above the first year row; col 0
    is the year column (sometimes carrying a unit label, which we skip).
    """
    rows = _sheet_matrix(ws)
    data_start = None
    for i, row in enumerate(rows):
        if row and _as_year(row[0]) is not None:
            data_start = i
            break
    if data_start is None or data_start == 0:
        raise RuntimeError(f"sheet {ws.title!r}: could not locate data/header rows")

    header = rows[data_start - 1]
    country_cols = {
        c: str(name).strip()
        for c, name in enumerate(header)
        if c >= 1 and name is not None and str(name).strip()
    }
    if not country_cols:
        raise RuntimeError(f"sheet {ws.title!r}: no country columns in header")

    out: list[dict] = []
    for row in rows[data_start:]:
        if not row:
            continue
        year = _as_year(row[0])
        if year is None:
            continue
        for c, country in country_cols.items():
            if c >= len(row):
                continue
            val = _as_float(row[c])
            if val is None:
                continue
            rec = {"country": country, "year": year, "value": val}
            if model is not None:
                rec = {"model": model, **rec}
            out.append(rec)
    return out


def _parse_regions(ws) -> list[dict]:
    """Regions sheet: one row per region, col 0 = region name, col 1 = a
    comma-separated list of member country names. Explode into a tidy
    (region, country) membership table — reference data joinable to the
    national emissions subsets via country.
    """
    out: list[dict] = []
    for row in _sheet_matrix(ws):
        if not row or len(row) < 2:
            continue
        region, members = row[0], row[1]
        if region is None or members is None:
            continue
        region = str(region).strip()
        if not region:
            continue
        for member in str(members).split(","):
            country = member.strip()
            if country:
                out.append({"region": region, "country": country})
    return out


# ---------------------------------------------------------------------------
# Fetch — one fn, dispatches on the subset id
# ---------------------------------------------------------------------------
def fetch_one(node_id: str) -> None:
    asset = node_id
    entity_id = node_id[len(SLUG) + 1:]  # strip "global-carbon-project-"
    family, sheet_spec, mode = SUBSETS[entity_id]

    url = _discover_workbooks()[family]
    wb = _load_workbook(url)

    if mode == GLOBAL:
        rows = _parse_global(wb[sheet_spec])
        schema = SCHEMA_GLOBAL
    elif mode == NATIONAL:
        rows = _parse_national(wb[sheet_spec])
        schema = SCHEMA_NATIONAL
    elif mode == NATIONAL_MODEL:
        rows = []
        for sheet in sheet_spec:  # one sheet per bookkeeping model
            rows.extend(_parse_national(wb[sheet], model=sheet))
        schema = SCHEMA_NATIONAL_MODEL
    elif mode == REGIONS:
        rows = _parse_regions(wb[sheet_spec])
        schema = SCHEMA_REGIONS
    else:  # pragma: no cover - guarded by SUBSETS
        raise RuntimeError(f"unknown parse mode {mode!r}")

    if not rows:
        raise RuntimeError(f"{asset}: parsed 0 rows from {url} sheet {sheet_spec!r}")

    table = pa.Table.from_pylist(rows, schema=schema)
    save_raw_parquet(table, asset)


# ---------------------------------------------------------------------------
# Specs
# ---------------------------------------------------------------------------
ENTITY_IDS = list(SUBSETS.keys())

DOWNLOAD_SPECS = [
    NodeSpec(id=f"{SLUG}-{eid}", fn=fetch_one, kind="download")
    for eid in ENTITY_IDS
]
