"""Shared HTTP + parsing helpers for the Fannie Mae connector.

The publishable surface is a small set of per-survey/index XLSX workbooks under
fanniemae.com/data-and-insights. Each workbook has its own bespoke sheet layout,
so each entity gets its own fetch fn (see the per-subset node modules) that
scrapes the dataset's data-and-insights page for the *current* download link
(the /media/document/xlsx/ URLs carry a release-date/quarter suffix that changes
every publication and must never be hardcoded), downloads the workbook, parses it
into clean long-format rows, and writes NDJSON raw.

This module holds only the code shared across those fetchers: the HTTP client,
the xlsx-link resolver, the workbook loader, and the cell-coercion helpers. It
contains no NodeSpec definitions.
"""

import datetime
import io
import re

import httpx
from openpyxl import load_workbook

from subsets_utils import get, transient_retry

# Browser User-Agent — the fanniemae.com host 403s non-browser fetchers. ASCII only.
_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
)
_HEADERS = {"User-Agent": _UA}
_HOST = "https://www.fanniemae.com"
_PAGE_BASE = "https://www.fanniemae.com/data-and-insights/surveys-indices/"


@transient_retry()
def _http_get(url: str) -> httpx.Response:
    resp = get(url, headers=_HEADERS, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp


def _resolve_xlsx_url(page_slug: str, prefix: str) -> str:
    """Scrape the dataset's data-and-insights page for the current xlsx link."""
    html = _http_get(_PAGE_BASE + page_slug).text
    for stem in re.findall(r"/media/document/xlsx/([A-Za-z0-9._-]+)", html):
        name = stem[:-5] if stem.endswith(".xlsx") else stem
        # exact-prefix form: <prefix>-<release>; or quarter-prefixed: <q#-YYYY>-<prefix>-...
        if name.startswith(prefix + "-") or prefix in name:
            return f"{_HOST}/media/document/xlsx/{stem}"
    raise AssertionError(
        f"no /media/document/xlsx/ link matching prefix {prefix!r} on page {page_slug!r} "
        "- page layout or download-link pattern changed"
    )


def _load_rows(node_id: str, page_slug: str, prefix: str):
    """Return (sheet_name, rows) where rows is a list of value tuples."""
    url = _resolve_xlsx_url(page_slug, prefix)
    content = _http_get(url).content
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = []
        empty_run = 0
        for row in ws.iter_rows(values_only=True):
            if all(c is None for c in row):
                empty_run += 1
                # read_only worksheets over-report dimensions with phantom blank
                # rows; stop after a long blank run past the real data.
                if empty_run > 60:
                    break
                rows.append(row)
                continue
            empty_run = 0
            rows.append(row)
        return wb.sheetnames[0], rows
    finally:
        wb.close()


def _num(v):
    """Coerce a cell to float, or None for blanks / 'n/a' / non-numeric."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.lower() in ("n/a", "na", "nan", "-", "--"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _isodate(v):
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    return None


def _ffill(header_row, ncols):
    """Forward-fill a header row across columns (for merged group labels)."""
    out = [None] * ncols
    last = None
    for i in range(ncols):
        v = header_row[i] if i < len(header_row) else None
        if v is not None and str(v).strip():
            last = str(v).strip()
        out[i] = last
    return out
