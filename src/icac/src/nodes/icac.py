"""ICAC (International Cotton Advisory Committee) — World Cotton Statistics.

The whole servable corpus is ONE multi-sheet xlsx workbook published on free
access at icac.org (research mechanism `bulk_xlsx`). Each of the four rank-active
subsets is carved out of a different family of sheets in that single workbook,
so every fetch fn downloads the workbook and extracts its slice into a clean,
long-format parquet that the SQL transform can read directly:

  * supply-and-use-balance  — Table 1 (World) + ~70 per-country balance sheets,
    all sharing one column layout. Long: country x season x metric.
  * extra-fine-cotton-supply — Table 3 (sections x country x year).
  * cotton-prices           — Tables 87a-87J, 88 (seasonal CIF / polyester) and
    89 (monthly NY futures). Long: source_table x quotation x period.
  * published-forecasts     — Tables 90-115 (one variable per sheet; horizon x
    publication-round columns x forecast-season rows).

Fetch shape: stateless full re-pull (shape 1). The workbook is ~8 MB and the
whole source is one file, so we simply re-download and overwrite every run; ICAC
releases a new snapshot ~twice a year and revisions are picked up for free.

HTTP note: icac.org sits behind Cloudflare bot management that rejects httpx by
TLS fingerprint (subsets_utils.get returns 403 on both the listing page and the
static file). curl_cffi impersonating Chrome passes the same check that the curl
CLI does, so this one connector fetches through curl_cffi instead of
subsets_utils.get. Everything else (parsing, raw I/O) is standard.
"""

import io
import re

import pyarrow as pa
from curl_cffi import requests as cffi_requests
from curl_cffi.requests.exceptions import (
    ConnectionError as CffiConnectionError,
    ProxyError as CffiProxyError,
    Timeout as CffiTimeout,
)
from openpyxl import load_workbook
from tenacity import (
    retry,
    retry_if_exception_type,
    stop_after_attempt,
    wait_exponential,
)

from subsets_utils import save_raw_parquet, NodeSpec, SqlNodeSpec

STATISTICS_PAGE = "https://icac.org/publications/statistics"
SITE_ROOT = "https://icac.org"
IMPERSONATE = "chrome"

_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6, "jul": 7,
    "jly": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "january": 1, "february": 2, "march": 3, "april": 4, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11,
    "december": 12,
}

_SEASON = re.compile(r"^\s*\d{4}/\d{2,4}\s*$")
_FC_SEASON = re.compile(r"^\d{2,4}/\d{2,4}$")
_MONTH_YEAR = re.compile(r"^[A-Z][a-z]{2} \d{2}$")
# Unit keywords for price-table header rows. Deliberately excludes "index"/
# "ratio" so that "A INDEX" / "B INDEX" label rows are NOT mistaken for units.
_UNIT_KW = re.compile(
    r"pound|kilogram|cents|dollar|yen|rupee|yuan|euro|nt\$|dm/|metric|tonne|hectare",
    re.I,
)


class _Transient(Exception):
    """Retryable HTTP status (429 / 5xx) from the Cloudflare-fronted host."""


@retry(
    retry=retry_if_exception_type(
        (_Transient, CffiTimeout, CffiConnectionError, CffiProxyError)
    ),
    wait=wait_exponential(multiplier=2, min=4, max=120),
    stop=stop_after_attempt(6),
    reraise=True,
)
def _http_get(url: str) -> cffi_requests.Response:
    resp = cffi_requests.get(url, impersonate=IMPERSONATE, timeout=120)
    if resp.status_code in (429, 500, 502, 503, 504):
        raise _Transient(f"{resp.status_code} for {url}")
    if resp.status_code >= 400:
        resp.raise_for_status()  # permanent 4xx -> HTTPError, not retried
    return resp


def _latest_workbook_url() -> str:
    """Scrape the statistics page for .xlsx links and pick the newest release.

    File naming is inconsistent across releases (e.g. 'World_Cotton_Statistics_
    Nov_2022_v2' vs 'world-cotton-statistics2021_november'), so we parse a
    (year, month) key out of each filename and take the maximum rather than
    guessing a path.
    """
    html = _http_get(STATISTICS_PAGE).text
    hrefs = re.findall(r'href="([^"]+\.xlsx[^"]*)"', html, re.I)
    if not hrefs:
        raise AssertionError(
            "no .xlsx links found on the ICAC statistics page; layout changed"
        )
    best, best_key = None, (-1, -1)
    for href in hrefs:
        fn = href.rsplit("/", 1)[-1].lower()
        ym = re.search(r"(?:19|20)\d{2}", fn)
        year = int(ym.group(0)) if ym else 0
        month = 0
        for tok, num in _MONTHS.items():
            if re.search(rf"[^a-z]{tok}[^a-z]|^{tok}|{tok}$", fn):
                month = num
        key = (year, month)
        if key > best_key:
            best_key, best = key, href
    return best if best.startswith("http") else SITE_ROOT + best


def _load_workbook():
    content = _http_get(_latest_workbook_url()).content
    return load_workbook(io.BytesIO(content), read_only=True, data_only=True)


# --- parsing helpers --------------------------------------------------------

def _s(x) -> str:
    return "" if x is None else str(x).strip()

def _norm(x) -> str:
    return " ".join(_s(x).split())

def _ffill(row) -> list:
    out, last = [], ""
    for x in row:
        v = _s(x)
        if v:
            last = v
        out.append(last)
    return out

def _rows(ws) -> list:
    return [list(r) for r in ws.iter_rows(values_only=True)]


# Canonical metric names, keyed by the header text with all whitespace stripped
# (one country sheet carries a literal line break inside "Consumption").
_BAL_CANON = {
    "area": "Area",
    "yield": "Yield",
    "production": "Production",
    "beginningstocks": "Beginning Stocks",
    "imports": "Imports",
    "consumption": "Consumption",
    "exports": "Exports",
    "endingstocks": "Ending Stocks",
    "s/u*": "S/U *",
    "s/u": "S/U *",
}

_BAL_UNITS = {
    "area": "'000 ha",
    "yield": "kg/ha",
    "s/u *": "ratio",
}


def _canon_metric(cell) -> str:
    key = re.sub(r"\s+", "", _s(cell)).lower()
    return _BAL_CANON.get(key, _norm(cell))


def parse_balance(wb) -> list:
    """Table 1 (World) + per-country balance sheets -> long supply/use rows.

    A balance sheet is identified structurally: header row (index 4) has 'Area'
    in column index 2. Metric columns run from index 2; season is in column 1.
    """
    out = []
    for name in wb.sheetnames:
        data = _rows(wb[name])
        if len(data) < 8:
            continue
        hdr = [_norm(x) for x in data[4]]
        if len(hdr) < 3 or hdr[2] != "Area":
            continue
        country = _s(data[1][1]) if len(data[1]) > 1 else ""
        if not country:
            country = name
        metrics = {ci: _canon_metric(data[4][ci])
                   for ci in range(2, len(hdr)) if hdr[ci]}
        for r in data[7:]:
            season = _s(r[1]) if len(r) > 1 else ""
            if not _SEASON.match(season):
                continue
            year_begin = int(season.split("/")[0])
            for ci, metric in metrics.items():
                if ci >= len(r) or r[ci] is None or _s(r[ci]) == "":
                    continue
                try:
                    value = float(r[ci])
                except (TypeError, ValueError):
                    continue
                out.append({
                    "country": country,
                    "season": season,
                    "year_begin": year_begin,
                    "metric": metric,
                    "unit": _BAL_UNITS.get(metric.lower(), "'000 metric tonnes"),
                    "value": value,
                })
    return out


def parse_extra_fine(wb) -> list:
    """Table 3 — sections (BEGINNING STOCKS, PRODUCTION, ...) x country x year."""
    data = _rows(wb["3"])
    hi = next(
        i for i, r in enumerate(data)
        if _s(r[0]).lower().startswith("years beginning")
    )
    years = {
        ci: int(data[hi][ci])
        for ci in range(1, len(data[hi]))
        if _s(data[hi][ci]).isdigit()
    }
    out, section = [], None
    for r in data[hi + 1:]:
        raw = r[0]
        label = _s(raw)
        if not label:
            continue
        indented = isinstance(raw, str) and raw[:1] == " "
        if not indented:
            # Section header (e.g. 'PRODUCTION'); skip footnotes/aggregate rows.
            if label[0].isdigit() or label.startswith(("*", "1/", "2/")):
                continue
            section = label.rstrip("* ").strip()
            continue
        if section is None:
            continue
        country = label
        for ci, yr in years.items():
            if ci >= len(r) or r[ci] is None or _s(r[ci]) == "":
                continue
            try:
                value = float(r[ci])
            except (TypeError, ValueError):
                continue
            out.append({
                "item": section,
                "country": country,
                "year_begin": yr,
                "season": f"{yr}/{str(yr + 1)[-2:]}",
                "value": value,
            })
    return out


def _parse_price_seasonal(wb, name) -> list:
    """Generic parser for the season-keyed price sheets (87*, 88).

    Multi-row headers: the first non-unit header row is the grouping row
    (forward-filled), remaining non-unit rows are detail, and rows containing a
    unit keyword become the column unit (a single centered unit applies to all).
    """
    data = _rows(wb[name])
    ds = next((i for i, r in enumerate(data) if _SEASON.match(_s(r[0]))), None)
    if ds is None:
        return []
    title = _norm(data[1][0]) or _norm(data[0][0])
    # Header block = rows between the title/courtesy lines and the first data
    # row. The grouping row (country or CURRENT/FORWARD) sits at varying depth
    # across sheets, so anchor on the last title line rather than a fixed index.
    title_end = max(
        [i for i, r in enumerate(data[:ds])
         if "COURTESY" in _s(r[0]).upper() or "PRICES OF" in _s(r[0]).upper()],
        default=2,
    )
    header = data[title_end + 1:ds]
    unit_rows = [h for h in header if any(_UNIT_KW.search(_s(c)) for c in h[1:])]
    label_rows = [h for h in header
                  if h not in unit_rows and any(_s(c) for c in h[1:])]
    group = _ffill(label_rows[0]) if label_rows else []
    ncol = max(len(r) for r in data)

    def label(ci):
        parts = []
        if ci < len(group) and group[ci]:
            parts.append(group[ci])
        for h in label_rows[1:]:
            frag = _s(h[ci]) if ci < len(h) else ""
            if frag and not re.match(r"^\d+/$", frag):  # drop footnote refs
                parts.append(frag)
        return _norm(" ".join(parts))

    def unit(ci):
        parts = []
        for h in unit_rows:
            nz = [_s(c) for c in h[1:] if _s(c)]
            if len(nz) == 1:
                parts.append(nz[0])
            elif ci < len(h) and _s(h[ci]):
                parts.append(_s(h[ci]))
        return _norm(" ".join(parts))

    out = []
    started = False
    for r in data[ds:]:
        period = _s(r[0])
        if not _SEASON.match(period):
            if started:
                break  # end of the absolute block; an indexed copy follows
            continue
        started = True
        for ci in range(1, ncol):
            if ci >= len(r) or r[ci] is None or _s(r[ci]) == "":
                continue
            try:
                value = float(r[ci])
            except (TypeError, ValueError):
                continue
            q = label(ci) or f"col{ci}"
            u = unit(ci)
            out.append({
                "source_table": name,
                "table_title": title,
                "quotation": f"{q} ({u})" if u else q,
                "period": period,
                "period_type": "season",
                "value": value,
            })
    return out


def _parse_price_futures(wb) -> list:
    """Table 89 — monthly NY cotton futures by contract expiry month/year."""
    if "89" not in wb.sheetnames:
        return []
    data = _rows(wb["89"])
    title = _norm(data[0][0])
    group = _ffill(data[4])
    years = data[5]
    ds = next(
        (i for i, r in enumerate(data) if _MONTH_YEAR.match(_s(r[0]))), None
    )
    if ds is None:
        return []
    ncol = max(len(r) for r in data)
    out = []
    started = False
    for r in data[ds:]:
        period = _s(r[0])
        if not _MONTH_YEAR.match(period):
            if started:
                break  # end of the absolute block; an indexed copy follows
            continue
        started = True
        for ci in range(1, ncol):
            if ci >= len(r) or r[ci] is None or _s(r[ci]) == "":
                continue
            try:
                value = float(r[ci])
            except (TypeError, ValueError):
                continue
            g = group[ci] if ci < len(group) else ""
            y = _s(years[ci]) if ci < len(years) else ""
            out.append({
                "source_table": "89",
                "table_title": title,
                "quotation": _norm(f"{g} {y}") or f"col{ci}",
                "period": period,
                "period_type": "month",
                "value": value,
            })
    return out


def parse_prices(wb) -> list:
    out = []
    for name in wb.sheetnames:
        if re.match(r"^87[a-zA-Z]$", name) or name == "88":
            out += _parse_price_seasonal(wb, name)
    out += _parse_price_futures(wb)
    return out


def parse_forecasts(wb) -> list:
    """Tables 90-115 — published forecasts, one variable per sheet.

    Columns are (horizon, publication_round) pairs; rows are forecast seasons.
    """
    out = []
    for name in wb.sheetnames:
        data = _rows(wb[name])
        if not data or not _s(data[0][0]).lower().startswith("forecasts"):
            continue
        variable = re.sub(r"(?i)^forecasts of (the )?", "", _norm(data[0][0]))
        unit = _norm(data[8][0]) if len(data) > 8 else ""
        rounds = data[5] if len(data) > 5 else []
        horizons = _ffill(data[6]) if len(data) > 6 else []
        ds = next(
            (i for i, r in enumerate(data) if _FC_SEASON.match(_s(r[0]))), None
        )
        if ds is None:
            continue
        ncol = max(len(r) for r in data)
        started = False
        for r in data[ds:]:
            fs = _s(r[0])
            if not _FC_SEASON.match(fs):
                if started:
                    break  # end of the absolute block; an indexed copy follows
                continue
            started = True
            for ci in range(1, ncol):
                if ci >= len(r) or r[ci] is None or _s(r[ci]) == "":
                    continue
                rnd = _norm(rounds[ci]) if ci < len(rounds) else ""
                if not rnd:
                    continue
                try:
                    value = float(r[ci])
                except (TypeError, ValueError):
                    continue
                out.append({
                    "variable": variable,
                    "forecast_season": fs,
                    "horizon": horizons[ci] if ci < len(horizons) else "",
                    "publication_round": rnd,
                    "unit": unit,
                    "value": value,
                })
    return out


# --- download nodes ---------------------------------------------------------

_BALANCE_SCHEMA = pa.schema([
    ("country", pa.string()),
    ("season", pa.string()),
    ("year_begin", pa.int64()),
    ("metric", pa.string()),
    ("unit", pa.string()),
    ("value", pa.float64()),
])

_EXTRA_FINE_SCHEMA = pa.schema([
    ("item", pa.string()),
    ("country", pa.string()),
    ("year_begin", pa.int64()),
    ("season", pa.string()),
    ("value", pa.float64()),
])

_PRICES_SCHEMA = pa.schema([
    ("source_table", pa.string()),
    ("table_title", pa.string()),
    ("quotation", pa.string()),
    ("period", pa.string()),
    ("period_type", pa.string()),
    ("value", pa.float64()),
])

_FORECASTS_SCHEMA = pa.schema([
    ("variable", pa.string()),
    ("forecast_season", pa.string()),
    ("horizon", pa.string()),
    ("publication_round", pa.string()),
    ("unit", pa.string()),
    ("value", pa.float64()),
])


def fetch_balance(node_id: str) -> None:
    rows = parse_balance(_load_workbook())
    if not rows:
        raise AssertionError(f"{node_id}: parsed 0 balance rows")
    save_raw_parquet(pa.Table.from_pylist(rows, schema=_BALANCE_SCHEMA), node_id)


def fetch_extra_fine(node_id: str) -> None:
    rows = parse_extra_fine(_load_workbook())
    if not rows:
        raise AssertionError(f"{node_id}: parsed 0 extra-fine rows")
    save_raw_parquet(pa.Table.from_pylist(rows, schema=_EXTRA_FINE_SCHEMA), node_id)


def fetch_prices(node_id: str) -> None:
    rows = parse_prices(_load_workbook())
    if not rows:
        raise AssertionError(f"{node_id}: parsed 0 price rows")
    save_raw_parquet(pa.Table.from_pylist(rows, schema=_PRICES_SCHEMA), node_id)


def fetch_forecasts(node_id: str) -> None:
    rows = parse_forecasts(_load_workbook())
    if not rows:
        raise AssertionError(f"{node_id}: parsed 0 forecast rows")
    save_raw_parquet(pa.Table.from_pylist(rows, schema=_FORECASTS_SCHEMA), node_id)


DOWNLOAD_SPECS = [
    NodeSpec(id="icac-supply-and-use-balance", fn=fetch_balance, kind="download"),
    NodeSpec(id="icac-extra-fine-cotton-supply", fn=fetch_extra_fine, kind="download"),
    NodeSpec(id="icac-cotton-prices", fn=fetch_prices, kind="download"),
    NodeSpec(id="icac-published-forecasts", fn=fetch_forecasts, kind="download"),
]


# --- transform nodes (one published Delta table per subset) -----------------

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="icac-supply-and-use-balance-transform",
        deps=["icac-supply-and-use-balance"],
        sql='''
            SELECT
                country,
                season,
                CAST(year_begin AS INTEGER) AS year_begin,
                metric,
                unit,
                CAST(value AS DOUBLE) AS value
            FROM "icac-supply-and-use-balance"
            WHERE value IS NOT NULL
        ''',
    ),
    SqlNodeSpec(
        id="icac-extra-fine-cotton-supply-transform",
        deps=["icac-extra-fine-cotton-supply"],
        sql='''
            SELECT
                item,
                country,
                CAST(year_begin AS INTEGER) AS year_begin,
                season,
                CAST(value AS DOUBLE) AS value
            FROM "icac-extra-fine-cotton-supply"
            WHERE value IS NOT NULL
        ''',
    ),
    SqlNodeSpec(
        id="icac-cotton-prices-transform",
        deps=["icac-cotton-prices"],
        sql='''
            SELECT
                source_table,
                table_title,
                quotation,
                period,
                period_type,
                CAST(value AS DOUBLE) AS value
            FROM "icac-cotton-prices"
            WHERE value IS NOT NULL
        ''',
    ),
    SqlNodeSpec(
        id="icac-published-forecasts-transform",
        deps=["icac-published-forecasts"],
        sql='''
            SELECT
                variable,
                forecast_season,
                horizon,
                publication_round,
                unit,
                CAST(value AS DOUBLE) AS value
            FROM "icac-published-forecasts"
            WHERE value IS NOT NULL
        ''',
    ),
]
