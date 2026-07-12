import csv
import io
import re
from html.parser import HTMLParser
from urllib.parse import unquote, urljoin

import openpyxl
import pyarrow as pa
import requests
from openpyxl.utils import get_column_letter
from subsets_utils import NodeSpec, save_raw_parquet

SLUG = "u-s-citizenship-and-immigration-services"
PREFIX = f"{SLUG}-"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}
MAIN_PAGE = (
    "https://www.uscis.gov/tools/reports-and-studies/"
    "immigration-and-citizenship-data"
)

MONTHS = (
    "january|february|march|april|may|june|july|august|september|"
    "october|november|december"
)

ENTITY_IDS = [
    "active_daca_recipients",
    "appropriation_requirement",
    "daca_performance_data",
    "eb_i140_i360_i526_performancedata",
    "h1b-employer-data-hub",
    "h2a-employer-data-hub",
    "h2b-employer-data-hub",
    "i129_quarterly_request_for_evidence",
    "i130_awaiting_a_visa_availability",
    "i130_performance_data",
    "i140",
    "i140_rec_by_class_country",
    "i360_sij_congressional",
    "i360_sij_performance_data",
    "i360_vawa_performance_data",
    "i485_performance_data",
    "i526_i526e_r_cob_pref",
    "i765_application_for_employment",
    "i765_p_allcat_c08",
    "i821_radp",
    "i907_premium_processing",
    "i914t_visastatistics",
    "i918u_visastatistics",
    "n400_performance_data",
    "net_backlog_frontlog",
    "quarterly_all_forms",
    "waivers",
]

HUB_PAGES = {
    "h1b-employer-data-hub": "https://www.uscis.gov/archive/h-1b-employer-data-hub-files",
    "h2a-employer-data-hub": "https://www.uscis.gov/archive/h-2a-employer-data-hub-files",
    "h2b-employer-data-hub": "https://www.uscis.gov/archive/h-2b-employer-data-hub-files",
}

SCHEMA = pa.schema(
    [
        ("entity_id", pa.string()),
        ("source_url", pa.string()),
        ("source_file", pa.string()),
        ("source_format", pa.string()),
        ("source_period", pa.string()),
        ("source_year", pa.int64()),
        ("sheet_name", pa.string()),
        ("row_index", pa.int64()),
        ("column_index", pa.int64()),
        ("column_label", pa.string()),
        ("value_text", pa.string()),
    ]
)


class LinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.hrefs: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        for key, value in attrs:
            if key.lower() == "href" and value:
                self.hrefs.append(value)


def _html(url: str) -> str:
    response = requests.get(url, headers=HEADERS, timeout=60)
    response.raise_for_status()
    return response.text


def _hrefs(url: str) -> list[str]:
    parser = LinkParser()
    parser.feed(_html(url))
    return [urljoin(url, href) for href in parser.hrefs]


def _filename(url: str) -> str:
    return unquote(url.rsplit("/", 1)[-1])


def _family(filename: str) -> str:
    name = filename.lower()
    name = re.sub(r"\.(xlsx|csv|xls)$", "", name)
    name = re.sub(r"_fy\d{4}.*$", "", name)
    name = re.sub(rf"_({MONTHS})_\d{{4}}(_v\d+)?$", "", name)
    name = re.sub(r"_v\d+$", "", name)
    name = re.sub(r"_\d{4}$", "", name)
    return name.strip("_")


def _period(filename: str) -> str | None:
    name = filename.lower()
    fiscal = re.search(r"fy\d{4}(?:_q\d)?", name)
    if fiscal:
        return fiscal.group(0)
    monthly = re.search(rf"({MONTHS})_\d{{4}}", name)
    if monthly:
        return monthly.group(0)
    year = re.search(r"(20\d{2})", name)
    if year:
        return year.group(1)
    return None


def _year(filename: str) -> int | None:
    period = _period(filename) or ""
    match = re.search(r"(20\d{2})", period)
    return int(match.group(1)) if match else None


def _spec_id(entity_id: str) -> str:
    return f"{SLUG}-{entity_id.lower().replace('_', '-')}"


SPEC_TO_ENTITY = {_spec_id(entity_id): entity_id for entity_id in ENTITY_IDS}


def _report_urls(entity_id: str) -> list[str]:
    urls = []
    for href in _hrefs(MAIN_PAGE):
        filename = _filename(href)
        if not re.search(r"\.(xlsx|csv|xls)$", filename, re.I):
            continue
        if "/document/reports/" not in href and "/document/data/" not in href:
            continue
        if _family(filename) == entity_id:
            urls.append(href)
    return sorted(set(urls), key=lambda url: (_year(_filename(url)) or 0, _filename(url)))


def _hub_urls(entity_id: str) -> list[str]:
    page = HUB_PAGES[entity_id]
    marker = entity_id.split("-", 1)[0].upper()
    urls = []
    for href in _hrefs(page):
        filename = _filename(href)
        if not filename.lower().endswith(".csv"):
            continue
        normalized = re.sub(r"[^A-Z0-9]", "", filename.upper())
        if marker in normalized:
            urls.append(href)
    return sorted(set(urls), key=lambda url: (_year(_filename(url)) or 0, _filename(url)))


def _urls_for_entity(entity_id: str) -> list[str]:
    if entity_id in HUB_PAGES:
        urls = _hub_urls(entity_id)
    else:
        urls = _report_urls(entity_id)
    if not urls:
        raise RuntimeError(f"no downloadable files discovered for {entity_id}")
    return urls


def _download(url: str) -> bytes:
    response = requests.get(url, headers=HEADERS, timeout=120)
    response.raise_for_status()
    return response.content


def _csv_rows(entity_id: str, url: str, content: bytes) -> list[dict]:
    filename = _filename(url)
    period = _period(filename)
    year = _year(filename)
    text = content.decode("utf-8-sig", "replace")
    reader = csv.reader(io.StringIO(text))
    headers: list[str] = []
    rows = []
    for row_index, row in enumerate(reader, start=1):
        if row_index == 1:
            headers = [cell.strip() for cell in row]
        for column_index, value in enumerate(row, start=1):
            value_text = str(value).strip()
            if not value_text:
                continue
            rows.append(
                {
                    "entity_id": entity_id,
                    "source_url": url,
                    "source_file": filename,
                    "source_format": "csv",
                    "source_period": period,
                    "source_year": year,
                    "sheet_name": None,
                    "row_index": row_index,
                    "column_index": column_index,
                    "column_label": headers[column_index - 1] if column_index <= len(headers) else None,
                    "value_text": value_text,
                }
            )
    return rows


def _xlsx_rows(entity_id: str, url: str, content: bytes) -> list[dict]:
    filename = _filename(url)
    period = _period(filename)
    year = _year(filename)
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows = []
    for sheet in workbook.worksheets:
        for row_index, cells in enumerate(sheet.iter_rows(), start=1):
            for column_index, cell in enumerate(cells, start=1):
                if cell.value is None:
                    continue
                value_text = str(cell.value).strip()
                if not value_text:
                    continue
                rows.append(
                    {
                        "entity_id": entity_id,
                        "source_url": url,
                        "source_file": filename,
                        "source_format": "xlsx",
                        "source_period": period,
                        "source_year": year,
                        "sheet_name": sheet.title,
                        "row_index": row_index,
                        "column_index": column_index,
                        "column_label": get_column_letter(column_index),
                        "value_text": value_text,
                    }
                )
    return rows


def fetch_one(node_id: str) -> None:
    entity_id = SPEC_TO_ENTITY[node_id]
    rows = []
    for url in _urls_for_entity(entity_id):
        content = _download(url)
        filename = _filename(url)
        if filename.lower().endswith(".csv"):
            rows.extend(_csv_rows(entity_id, url, content))
        elif filename.lower().endswith((".xlsx", ".xls")):
            rows.extend(_xlsx_rows(entity_id, url, content))
        else:
            raise RuntimeError(f"unsupported file format for {url}")
    if not rows:
        raise RuntimeError(f"no non-empty cells extracted for {entity_id}")
    save_raw_parquet(pa.Table.from_pylist(rows, schema=SCHEMA), node_id)


DOWNLOAD_SPECS = [NodeSpec(id=_spec_id(entity_id), fn=fetch_one) for entity_id in ENTITY_IDS]
