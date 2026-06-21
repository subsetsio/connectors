import sys, os; sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
import sys, json, re, io
from subsets_utils import get
import openpyxl

def page_data(slug, lang="en"):
    url = f"https://statistique.quebec.ca/{lang}/produit/tableau/{slug}"
    r = get(url, timeout=(10,60))
    if r.status_code != 200:
        return None, r.status_code
    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', r.text, re.S)
    if not m: return None, "no __NEXT_DATA__"
    d = json.loads(m.group(1))
    return d["props"]["pageProps"]["data"], 200

def resolve_xlsx_url(data, lang="en"):
    t = data.get("type")
    if t == "dynamique":
        return f"https://statistique.quebec.ca/docs-ken/multimedia/Fichier_complet_{data['no']}.xlsx"
    elif t == "statique":
        return f"https://statistique.quebec.ca/{lang}/fichier/{data['excel']}"
    return None

def dump(slug, lang="en"):
    print("="*90); print("SLUG:", slug, "lang:", lang)
    data, st = page_data(slug, lang)
    if data is None:
        print("  page failed:", st); return
    print("  type:", data.get("type"), "| no:", data.get("no"), "| excel:", data.get("excel"),
          "| nom:", (data.get("nom") or "")[:70])
    url = resolve_xlsx_url(data, lang)
    print("  xlsx url:", url)
    r = get(url, timeout=(10,120))
    print("  xlsx http:", r.status_code, "bytes:", len(r.content))
    if r.status_code != 200: return
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    print("  sheets:", wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    print(f"  dims: {ws.max_row} rows x {ws.max_column} cols")
    rows = list(ws.iter_rows(min_row=1, max_row=min(18, ws.max_row), values_only=True))
    for i, row in enumerate(rows, 1):
        cells = [("" if c is None else str(c))[:18] for c in row[:8]]
        print(f"   r{i:>2}: {cells}")

for slug, lang in [
    ("population-composantes-accroissement-demographique-trimestre-quebec","fr"),
    ("person-years-wage-bill-and-paid-hours-for-core-drilling-quebec","en"),
    ("gross-domestic-product-expenditure-quebec","en"),
]:
    try: dump(slug, lang)
    except Exception as e:
        import traceback; traceback.print_exc()
