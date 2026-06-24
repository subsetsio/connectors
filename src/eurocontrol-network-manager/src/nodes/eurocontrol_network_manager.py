"""Eurocontrol Network Manager — EUROCONTROL Performance Review bulk datasets.

The public statistical surface of the Network Manager's operational data
(Aviation Intelligence Portal / ansperformance.eu). Each subset is one bulk
file served at a stable URL of the form
``https://www.eurocontrol.int/performance/data/download/{xls|parquet}/{name}.{ext}``.

Fetch shape: **stateless full re-pull** (shape 1). Every file is a full
historical snapshot with no incremental/`since` filter, so each run re-fetches
the whole corpus and overwrites. No state, no watermark.

Per-dataset extraction (the files are not uniform):
  * Most datasets ship as a multi-sheet ``.xlsx`` whose flat data lives on a
    single ``DATA`` sheet — read it with DuckDB's excel reader as strings
    (robust: some numeric columns carry sentinels like ``NO DATA RECEIVED``),
    then coerce each column to BIGINT/DOUBLE where it cleanly casts.
  * ``g2g_emissions`` exposes a native parquet endpoint — fetch it directly.
  * ``ANSP_Financial_Data`` spreads its data across one sheet per year
    (``2024``..``2017``) with an identical header — union them.
  * ``ACE_Yearly_Operational_Data`` keeps its flat table on the
    ``Operational Data`` sheet behind a 3-row pivot header — read from the data
    rows under the row-1 header.
"""
import io
import os
import re
import tempfile

import duckdb
import openpyxl
import pyarrow.parquet as pq

from subsets_utils import (
    NodeSpec,
    SqlNodeSpec,
    get,
    transient_retry,
    save_raw_parquet,
    save_raw_ndjson,
)
from constants import ENTITY_IDS

SLUG = "eurocontrol-network-manager"
BASE = "https://www.eurocontrol.int/performance/data/download"


def _spec_id(entity_id: str) -> str:
    return f"{SLUG}-{entity_id.lower().replace('_', '-')}"


# Recover the original entity id (and thus the file name) from a spec id.
SLUG_TO_ENTITY = {_spec_id(e): e for e in ENTITY_IDS}

# Datasets that need something other than the generic flat-"DATA"-sheet read.
_PARQUET_DATASETS = {"g2g_emissions"}


@transient_retry()
def _download(url: str) -> bytes:
    resp = get(url, timeout=(10.0, 300.0))
    resp.raise_for_status()
    return resp.content


def _fetch_parquet(asset: str, entity_id: str) -> None:
    """g2g_emissions: a native parquet endpoint, already typed and clean."""
    content = _download(f"{BASE}/parquet/{entity_id}.parquet")
    table = pq.read_table(io.BytesIO(content))
    save_raw_parquet(table, asset)


def _fetch_data_sheet(asset: str, entity_id: str) -> None:
    """Read the flat 'DATA' sheet of a multi-sheet workbook.

    DuckDB's excel reader is used with ``all_varchar`` (robust against sentinel
    strings in numeric columns), then every column that cleanly casts is
    promoted to BIGINT/DOUBLE so published types are meaningful.
    """
    content = _download(f"{BASE}/xls/{entity_id}.xlsx")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as fh:
        fh.write(content)
        path = fh.name
    try:
        con = duckdb.connect()
        con.execute("INSTALL excel; LOAD excel;")
        con.execute(
            f"CREATE TABLE raw AS "
            f"SELECT * FROM read_xlsx('{path}', sheet='DATA', all_varchar=true)"
        )
        cols = [r[0] for r in con.execute("DESCRIBE raw").fetchall()]
        select_parts = []
        for c in cols:
            q = c.replace('"', '""')
            nonempty, as_int, as_dbl = con.execute(
                f'SELECT '
                f'count(*) FILTER (WHERE "{q}" IS NOT NULL AND "{q}" <> \'\'), '
                f'count(*) FILTER (WHERE "{q}" <> \'\' AND TRY_CAST("{q}" AS BIGINT) IS NOT NULL), '
                f'count(*) FILTER (WHERE "{q}" <> \'\' AND TRY_CAST("{q}" AS DOUBLE) IS NOT NULL) '
                f'FROM raw'
            ).fetchone()
            if nonempty and as_int / nonempty >= 0.95:
                select_parts.append(f'TRY_CAST("{q}" AS BIGINT) AS "{q}"')
            elif nonempty and as_dbl / nonempty >= 0.95:
                select_parts.append(f'TRY_CAST("{q}" AS DOUBLE) AS "{q}"')
            else:
                select_parts.append(f'NULLIF("{q}", \'\') AS "{q}"')
        table = con.execute(
            f'SELECT {", ".join(select_parts)} FROM raw'
        ).to_arrow_table()
        save_raw_parquet(table, asset)
    finally:
        os.unlink(path)


def _fetch_ansp_financial(asset: str, entity_id: str) -> None:
    """ANSP financial data: one sheet per year, identical header. Union them."""
    content = _download(f"{BASE}/xls/{entity_id}.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        rows = []
        for sheet_name in wb.sheetnames:
            if not re.fullmatch(r"\d{4}", sheet_name):
                continue
            it = wb[sheet_name].iter_rows(values_only=True)
            header = next(it)
            header = [
                str(c).strip() if c is not None else f"col{i}"
                for i, c in enumerate(header)
            ]
            for r in it:
                if not r or r[0] is None:
                    continue
                rows.append(
                    {h: (r[i] if i < len(r) else None) for i, h in enumerate(header)}
                )
    finally:
        wb.close()
    if not rows:
        raise ValueError(f"{asset}: no year-sheet rows extracted")
    save_raw_ndjson(rows, asset)


def _fetch_ace_yearly(asset: str, entity_id: str) -> None:
    """ACE yearly operational data: flat table on the 'Operational Data' sheet,
    header on row 1 (1-based), data from row 4. Multi-row merged-header helper
    columns (no header text) are dropped."""
    content = _download(f"{BASE}/xls/{entity_id}.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        all_rows = list(wb["Operational Data"].iter_rows(values_only=True))
    finally:
        wb.close()
    header = [
        (str(c).strip().replace("\n", " ") if c is not None else None)
        for c in all_rows[1]
    ]
    rows = []
    for r in all_rows[4:]:
        if len(r) < 2 or r[1] is None:
            continue
        rec = {}
        for i, h in enumerate(header):
            if not h:
                continue
            v = r[i] if i < len(r) else None
            rec[h] = None if v == "" else v
        rows.append(rec)
    if not rows:
        raise ValueError(f"{asset}: no rows extracted from 'Operational Data'")
    save_raw_ndjson(rows, asset)


def fetch_one(node_id: str) -> None:
    asset = node_id
    entity_id = SLUG_TO_ENTITY[node_id]
    if entity_id in _PARQUET_DATASETS:
        _fetch_parquet(asset, entity_id)
    elif entity_id == "ANSP_Financial_Data":
        _fetch_ansp_financial(asset, entity_id)
    elif entity_id == "ACE_Yearly_Operational_Data":
        _fetch_ace_yearly(asset, entity_id)
    else:
        _fetch_data_sheet(asset, entity_id)


DOWNLOAD_SPECS = [
    NodeSpec(id=_spec_id(e), fn=fetch_one, kind="download") for e in ENTITY_IDS
]

# One published table per subset. Each file is already a clean single-table
# snapshot, so the transform is a thin pass-through over the raw view.
TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"{s.id}-transform",
        deps=[s.id],
        sql=f'SELECT * FROM "{s.id}"',
    )
    for s in DOWNLOAD_SPECS
]
