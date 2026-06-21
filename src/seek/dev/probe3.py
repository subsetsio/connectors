import io, re
from subsets_utils import get, configure_http
import openpyxl
configure_http(headers={"User-Agent": "Mozilla/5.0"})

EMP="https://ap-southeast-2-seek-apac.graphassets.com/AEzBCRO50TYyqbV6XzRDQz/cmqingfsddr7u072s5xl19vkr"
ASI="https://ap-southeast-2-seek-apac.graphassets.com/AEzBCRO50TYyqbV6XzRDQz/cmqhn27z0x15f0643ue1gvan4"

def rows_from(content, sheet):
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet]
    ws.reset_dimensions()  # ignore broken <dimension>; recompute from actual cells
    out=[]
    for r in ws.iter_rows(values_only=True):
        out.append(r)
    wb.close()
    return out

emp = get(EMP, timeout=(10,180)).content
asi = get(ASI, timeout=(10,180)).content
print("emp bytes", len(emp), "asi bytes", len(asi))

wb=openpyxl.load_workbook(io.BytesIO(emp), read_only=True)
print("emp sheets", wb.sheetnames); wb.close()
wb=openpyxl.load_workbook(io.BytesIO(asi), read_only=True)
print("asi sheets", wb.sheetnames); wb.close()

for name,(c,sheet) in {"jobad":(emp,"SEEK Job Ad Index"),"appsperad":(emp,"SEEK Applications per Ad Index"),"asi":(asi,"Sheet 1")}.items():
    rs = rows_from(c, sheet)
    print(f"=== {name}: {len(rs)} rows incl header")
    print("   header:", rs[0])
    print("   row1:", rs[1])
    print("   last:", rs[-1])
