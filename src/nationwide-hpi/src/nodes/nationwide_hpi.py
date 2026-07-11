"""Nationwide House Price Index connector.

The Nationwide Building Society publishes its House Price Index as ~24 Excel
spreadsheets under stable URLs (https://www.nationwide.co.uk/media/hpi/download/<slug>).
Each file is a distinct House Price Index time series. There is no machine-readable
API and no incremental filter: the whole corpus is tiny (tens of KB per file) so we
re-fetch every file in full on each run (stateless full re-pull) and overwrite.

The spreadsheets use several different layouts (tidy single-region UK series, wide
region tables with paired GBP/INDEX columns, affordability tables keyed by region or
occupation, multi-block historical series). Rather than push that heterogeneity into
SQL, each download fn parses its file in Python into ONE uniform long-format table:

    date (DATE) | period_label (str) | category (str) | measure (str) | value (DOUBLE)

`category` is the breakdown dimension (UK region, occupation group, property class, or
just "UK"); `measure` names the metric (price_gbp, index, annual_pct_change, hper,
mtg_pct_take_home, ...). Each transform is then a thin type-and-filter pass over its
own asset, publishing one Delta table per subset.
"""
import datetime as _dt
import io
import re

import openpyxl
import pyarrow as pa

from subsets_utils import (
    NodeSpec,
    get,
    save_raw_parquet,
    transient_retry,
)

SLUG = "nationwide-hpi"
BASE = "https://www.nationwide.co.uk/media/hpi/download"

ENTITY_IDS = [
    "all-buyers-hper-by-region",
    "all-prop",
    "annual-percentage-change-in-regional-house-prices",
    "chart-data-download-annual-percentage-change-in-uk-house-prices",
    "detached-post-1991",
    "first-by",
    "flats-post-1991",
    "fowner",
    "ftb-hper-by-broad-occupation",
    "ftb-hper-by-region",
    "ftb-mtg-payments-by-broad-occupation",
    "ftb-mtg-payments-by-detailed-occupation",
    "ftb-mtg-payments-by-region",
    "monthly",
    "new-prop",
    "not-new-prop",
    "quarterly",
    "seasonal-regional",
    "semi-detached-post-1991",
    "terraced-post-1991",
    "uk-house-price-since-1952",
    "uk-house-prices-adjusted-for-inflation",
]

SCHEMA = pa.schema(
    [
        ("date", pa.date32()),
        ("period_label", pa.string()),
        ("category", pa.string()),
        ("measure", pa.string()),
        ("value", pa.float64()),
    ]
)

# ----------------------------------------------------------------------------
# HTTP
# ----------------------------------------------------------------------------


@transient_retry()
def _download_xlsx(slug: str) -> bytes:
    resp = get(f"{BASE}/{slug}", timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.content


def _rows(content: bytes, sheet: str | None = None) -> list[tuple]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    return list(ws.iter_rows(values_only=True))


# ----------------------------------------------------------------------------
# Value / period helpers
# ----------------------------------------------------------------------------
_QUARTER_MONTH = {1: 1, 2: 4, 3: 7, 4: 10}


def _num(v):
    """Return a float if the cell is a real number, else None (skip N/A, '', ' ')."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return f if f == f else None  # drop NaN
    return None


def _clean(s) -> str:
    return re.sub(r"\s+", " ", str(s)).strip()


def _parse_period(v):
    """Normalize a period cell to (date, label) or (None, None) if not a period."""
    if v is None:
        return None, None
    if isinstance(v, _dt.datetime):
        return _dt.date(v.year, v.month, 1), v.date().isoformat()
    if isinstance(v, _dt.date):
        return _dt.date(v.year, v.month, 1), v.isoformat()
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        year = int(v)
        if 1900 <= year <= 2100 and float(v) == year:
            return _dt.date(year, 1, 1), str(year)
        return None, None
    label = _clean(v)
    m = re.search(r"Q([1-4])", label, re.I)
    y = re.search(r"(19|20)\d{2}", label)
    if m and y:
        year = int(y.group(0))
        month = _QUARTER_MONTH[int(m.group(1))]
        return _dt.date(year, month, 1), label
    if y and "=" not in label:  # a bare year string like "1952"
        return _dt.date(int(y.group(0)), 1, 1), label
    return None, None


# ----------------------------------------------------------------------------
# Per-family parsers — each returns list[dict] in the uniform long schema
# ----------------------------------------------------------------------------
def _emit(out, date, label, category, measure, value):
    val = _num(value)
    if val is None:
        return
    out.append(
        {
            "date": date,
            "period_label": label,
            "category": category,
            "measure": _clean(measure),
            "value": val,
        }
    )


def _parse_tidy_uk(rows, header_row, data_start, measures=None):
    """Col A = period; each named column (from header_row) is a measure. category=UK."""
    out = []
    if measures is None:
        header = rows[header_row]
        measures = [(_clean(h), c) for c, h in enumerate(header) if c >= 1 and h not in (None, "")]
    for row in rows[data_start:]:
        date, label = _parse_period(row[0])
        if date is None:
            continue
        for name, c in measures:
            if c < len(row):
                _emit(out, date, label, "UK", name, row[c])
    return out


def _parse_wide_pairs(rows, region_row=0, unit_row=2, data_start=3):
    """Wide region table: paired GBP / INDEX columns under each region name."""
    out = []
    regions = rows[region_row]
    units = rows[unit_row]
    cols = []  # (category, price_col, index_col)
    for c, u in enumerate(units):
        if _clean(u) == "£" and c < len(regions) and regions[c] not in (None, ""):
            idx_col = c + 1 if (c + 1 < len(units) and _clean(units[c + 1]).upper() == "INDEX") else None
            cols.append((_clean(regions[c]), c, idx_col))
    for row in rows[data_start:]:
        date, label = _parse_period(row[0])
        if date is None:
            continue
        for category, pc, ic in cols:
            if pc < len(row):
                _emit(out, date, label, category, "price_gbp", row[pc])
            if ic is not None and ic < len(row):
                _emit(out, date, label, category, "index", row[ic])
    return out


def _parse_regional_pct(rows, header_row=4, data_start=5):
    """Region pairs of (GBP price, annual % change)."""
    out = []
    header = rows[header_row]
    cols = []  # (category, price_col, pct_col)
    for c, h in enumerate(header):
        if c >= 1 and h not in (None, "") and _clean(h) != "%":
            pct_col = c + 1 if (c + 1 < len(header) and _clean(header[c + 1]) == "%") else None
            cols.append((_clean(h), c, pct_col))
    for row in rows[data_start:]:
        date, label = _parse_period(row[0])
        if date is None:
            continue
        for category, pc, qc in cols:
            if pc < len(row):
                _emit(out, date, label, category, "price_gbp", row[pc])
            if qc is not None and qc < len(row):
                _emit(out, date, label, category, "annual_pct_change", row[qc])
    return out


def _parse_by_label(rows, header_row, data_start, measure):
    """One value column per labelled category (region / occupation). Col A = period."""
    out = []
    header = rows[header_row]
    cols = [(c, _clean(h)) for c, h in enumerate(header) if c >= 1 and h not in (None, "")]
    for row in rows[data_start:]:
        date, label = _parse_period(row[0])
        if date is None:
            continue
        for c, category in cols:
            if c < len(row):
                _emit(out, date, label, category, measure, row[c])
    return out


def _parse_seasonal_regional(rows, header_row=2, data_start=3):
    """Seasonal regional sheet: adjusted index block followed by QoQ change block."""
    out = []
    header = rows[header_row]
    blocks = [
        ("seasonally_adjusted_index", range(1, 15)),
        ("quarter_on_quarter_change_seasonally_adjusted", range(15, 29)),
    ]
    cols = []
    for measure, col_range in blocks:
        for c in col_range:
            if c < len(header) and header[c] not in (None, ""):
                cols.append((c, _clean(header[c]), measure))
    for row in rows[data_start:]:
        date, label = _parse_period(row[0])
        if date is None:
            continue
        for c, category, measure in cols:
            if c < len(row):
                _emit(out, date, label, category, measure, row[c])
    return out


def _parse_since1952(rows, block_row=3, data_start=6):
    """Multi-block historical series: each property-class block has Index/Price/Annual Change."""
    out = []
    blocks = [(c, _clean(h)) for c, h in enumerate(rows[block_row]) if c >= 1 and h not in (None, "")]
    sub = ["index", "price_gbp", "annual_change_pct"]
    for row in rows[data_start:]:
        date, label = _parse_period(row[0])
        if date is None:
            continue
        for bc, category in blocks:
            for offset, measure in enumerate(sub):
                c = bc + offset
                if c < len(row):
                    _emit(out, date, label, category, measure, row[c])
    return out


# ----------------------------------------------------------------------------
# Slug -> parser dispatch
# ----------------------------------------------------------------------------
def _parse(slug: str, content: bytes) -> list[dict]:
    wide = {
        "all-prop",
        "detached-post-1991",
        "first-by",
        "flats-post-1991",
        "fowner",
        "new-prop",
        "semi-detached-post-1991",
        "terraced-post-1991",
    }
    if slug in wide:
        return _parse_wide_pairs(_rows(content))
    if slug == "monthly":
        return _parse_tidy_uk(_rows(content), header_row=0, data_start=1)
    if slug == "quarterly":
        return _parse_tidy_uk(_rows(content), header_row=0, data_start=1)
    if slug == "uk-house-prices-adjusted-for-inflation":
        return _parse_tidy_uk(_rows(content), header_row=2, data_start=3)
    if slug == "not-new-prop":
        # UK-only single price column (data begins ~2005); no usable header row.
        return _parse_tidy_uk(
            _rows(content), header_row=None, data_start=3, measures=[("price_gbp", 1)]
        )
    if slug == "annual-percentage-change-in-regional-house-prices":
        return _parse_regional_pct(_rows(content))
    if slug == "seasonal-regional":
        return _parse_seasonal_regional(_rows(content))
    if slug == "all-buyers-hper-by-region":
        return _parse_by_label(_rows(content), header_row=4, data_start=5, measure="hper")
    if slug == "ftb-hper-by-region":
        return _parse_by_label(_rows(content), header_row=4, data_start=5, measure="hper")
    if slug == "ftb-mtg-payments-by-region":
        return _parse_by_label(
            _rows(content), header_row=4, data_start=5, measure="mtg_pct_take_home"
        )
    if slug == "ftb-hper-by-broad-occupation":
        return _parse_by_label(_rows(content), header_row=4, data_start=5, measure="hper")
    if slug == "ftb-mtg-payments-by-broad-occupation":
        return _parse_by_label(
            _rows(content), header_row=4, data_start=5, measure="mtg_pct_take_home"
        )
    if slug == "ftb-mtg-payments-by-detailed-occupation":
        # The 'Data_sheet' tab is the tidy form: occupation names header, year-keyed rows.
        return _parse_by_label(
            _rows(content, sheet="Data_sheet"),
            header_row=2,
            data_start=3,
            measure="mtg_pct_take_home",
        )
    if slug == "uk-house-price-since-1952":
        return _parse_since1952(_rows(content))
    if slug == "chart-data-download-annual-percentage-change-in-uk-house-prices":
        # Monthly-frequency annual % change for the UK headline series.
        return _parse_tidy_uk(
            _rows(content, sheet="UK_monthly"),
            header_row=None,
            data_start=6,
            measures=[("annual_pct_change", 1)],
        )
    raise ValueError(f"no parser registered for slug {slug!r}")


def fetch_one(node_id: str) -> None:
    asset = node_id  # the spec id IS the asset name
    slug = node_id[len(SLUG) + 1 :]  # strip "nationwide-hpi-"
    content = _download_xlsx(slug)
    rows = _parse(slug, content)
    if not rows:
        raise AssertionError(f"{slug}: parsed 0 rows from spreadsheet (layout may have changed)")
    table = pa.Table.from_pylist(rows, schema=SCHEMA)
    save_raw_parquet(table, asset)


DOWNLOAD_SPECS = [
    NodeSpec(id=f"{SLUG}-{eid}", fn=fetch_one, kind="download") for eid in ENTITY_IDS
]
