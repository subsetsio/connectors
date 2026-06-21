"""Global Energy Monitor connector — one full Excel workbook per tracker slug.

GEM distributes each tracker as a single full-corpus Excel (.xlsx) workbook
behind the Supabase "data release hub" download form (research mechanism
`release_hub_presign`). There is no stable direct file URL: you mint a
short-lived capability token, exchange it for a presigned DigitalOcean Spaces
URL, and GET that. The flow is unauthenticated (the publishable key ships in
GEM's public web bundle) but the form "requires" a contact email — we send a
real research address.

Each accepted entity is one slug → one workbook → one published Delta table.
Strategy is stateless full re-pull: the workbook is the entire corpus and
releases are only quarterly/bi-annual, so we re-fetch the whole file each run
and overwrite. There is no incremental query surface.

Parsing: GEM workbooks bundle an "About"/"Metadata"/dictionary sheet, a primary
data sheet, and (sometimes) supplemental/pivot sheets. The primary registry is
consistently the FIRST non-junk sheet carrying a wide header row, so we select
that, detect its header row (workbooks occasionally prepend a title row),
normalize column names to snake_case, coerce each column to a single type
(int/float/str, with GEM's sentinel strings -> null), and write NDJSON. The
SQL transform is a thin typed passthrough that publishes the table.

Multi-table workbooks (e.g. the energy-ownership, methane, iron/steel-unit and
latin-america files carry several distinct sheets) publish only their primary
sheet under the one-table-per-slug contract; the supplemental sheets are not
materialized.
"""

import io
import json
import re

import openpyxl

from subsets_utils import (
    NodeSpec,
    SqlNodeSpec,
    get,
    post,
    save_raw_ndjson,
    transient_retry,
)

SLUG = "gem-global-energy-monitor"

# Entity union (rank-accepted slugs) — copied from work/entity_union.json.
from constants import ENTITY_IDS

# --- Release-hub flow constants ----------------------------------------------
_SUPABASE = "https://auxunjnrktkmeqyoyngm.supabase.co"
_MINT_URL = f"{_SUPABASE}/rest/v1/rpc/mint_submission"
_PRESIGN_URL = f"{_SUPABASE}/functions/v1/presign"
# Publishable key embedded in GEM's public gem-download-form.bundle.js.
_PUBLISHABLE_KEY = "sb_publishable_8mQAV8B2HhveNc5T8VGqPQ_1lgsFAvz"
_CONTACT = {
    "name": "Nathan Snellaert",
    "email": "nathansnellaert@gmail.com",
    "organization": "subsets.io",
    "sector": "Academic / Research",
    "country": "United States",
    "use_case": (
        "Integrating Global Energy Monitor tracker data into the subsets.io "
        "open data catalog for research and statistical analysis of the global "
        "energy transition across all sectors covered by GEM."
    ),
}

# --- Transport / retry --------------------------------------------------------


@transient_retry()
def _mint_token(slug: str) -> str:
    body = dict(_CONTACT)
    body.update({
        "license_text": "CC-BY-4.0",
        "email_optin": False,
        "request_mode": "slugs",
        "requested_slugs": [slug],
    })
    resp = post(
        _MINT_URL,
        headers={
            "apikey": _PUBLISHABLE_KEY,
            "authorization": f"Bearer {_PUBLISHABLE_KEY}",
            "content-type": "application/json",
        },
        json=body,
        timeout=(10.0, 60.0),
    )
    resp.raise_for_status()
    token = resp.json().get("capability_token")
    if not token:
        raise RuntimeError(f"mint_submission returned no capability_token for {slug}")
    return token


@transient_retry()
def _presign(token: str, slug: str) -> str:
    resp = post(
        _PRESIGN_URL,
        headers={
            "authorization": f"Bearer {token}",
            "content-type": "application/json",
        },
        timeout=(10.0, 60.0),
    )
    resp.raise_for_status()
    urls = resp.json().get("urls") or []
    if not urls:
        raise RuntimeError(f"presign returned no urls for {slug}")
    # Single-slug requests resolve to one file; prefer the entry whose slug
    # matches, else take the first.
    for u in urls:
        if u.get("slug") == slug and u.get("url"):
            return u["url"]
    if urls[0].get("url"):
        return urls[0]["url"]
    raise RuntimeError(f"presign url missing for {slug}")


@transient_retry()
def _download_xlsx(url: str) -> bytes:
    resp = get(url, timeout=(10.0, 600.0))
    resp.raise_for_status()
    return resp.content


# --- Workbook parsing ---------------------------------------------------------
# Sheet names that never hold the primary registry.
_JUNK_RE = re.compile(
    r"(about|metadata|dictionary|column key|read\s?me|readme|cover|glossary|"
    r"acronym|copyright|introduction|terminolog|notes?|legend|sources?|"
    r"contents|instructions)", re.I,
)
# Lowercased values GEM uses for "missing" — coerced to null.
_NULLS = {
    "", "n/a", "na", "unknown", "tbd", "tbc", "not found", "--", "-", "none",
    "null", "not applicable", "not available", "no data", "?",
}
_MIN_HEADER_WIDTH = 8  # primary registries have >=12 cols; supplements have <=3.


def _norm_col(value, idx: int) -> str:
    if value is None or str(value).strip() == "":
        return f"col_{idx}"
    s = re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")
    if not s:
        return f"col_{idx}"
    if s[0].isdigit():
        s = "n_" + s
    return s


def _detect_header(head_rows) -> tuple[int, int]:
    """Header = the row (within the first 15) with the most non-empty cells."""
    best_i, best_n = 0, -1
    for i, row in enumerate(head_rows[:15]):
        n = sum(1 for c in row if c is not None and str(c).strip() != "")
        if n > best_n:
            best_n, best_i = n, i
    return best_i, best_n


def _pick_sheet(wb):
    """First non-junk sheet whose header row is wide — GEM's primary registry."""
    for idx, ws in enumerate(wb.worksheets):
        head = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            head.append(row)
            if i >= 15:
                break
        if _JUNK_RE.search(ws.title or "") or (ws.title or "").strip().startswith("#"):
            continue
        hi, hn = _detect_header(head)
        if hn < _MIN_HEADER_WIDTH:
            continue
        return ws.title, hi
    raise RuntimeError("no primary data sheet found in workbook")


def _as_number(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "")
    if s == "":
        return None
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        return None


def _column_caster(values):
    """Return a (cast_fn) for a column based on whether every non-null value is
    numeric. Ints stay ints; mixed/decimal -> float; anything else -> str."""
    nums = []
    all_numeric = True
    for v in values:
        if v is None:
            continue
        n = _as_number(v)
        if n is None:
            all_numeric = False
            break
        nums.append(n)
    if all_numeric and nums:
        if all(isinstance(n, int) or float(n).is_integer() for n in nums):
            return lambda v: (int(_as_number(v)) if v is not None else None)
        return lambda v: (float(_as_number(v)) if v is not None else None)
    return lambda v: (str(v) if v is not None else None)


def _parse_workbook(blob: bytes):
    """Parse the primary sheet into (columns, list-of-tuples rows, casters)."""
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    try:
        title, header_idx = _pick_sheet(wb)
        ws = wb[title]
        columns = None
        rows = []
        for i, raw in enumerate(ws.iter_rows(values_only=True)):
            if i < header_idx:
                continue
            if i == header_idx:
                seen = {}
                columns = []
                for j, cell in enumerate(raw):
                    name = _norm_col(cell, j)
                    if name in seen:
                        seen[name] += 1
                        name = f"{name}_{seen[name]}"
                    else:
                        seen[name] = 1
                    columns.append(name)
                continue
            if all(c is None or str(c).strip() == "" for c in raw):
                continue
            cleaned = []
            for j in range(len(columns)):
                v = raw[j] if j < len(raw) else None
                if isinstance(v, str):
                    v = v.strip()
                    if v.lower() in _NULLS:
                        v = None
                elif v is not None and str(v).strip().lower() in _NULLS:
                    v = None
                cleaned.append(v)
            rows.append(tuple(cleaned))
    finally:
        wb.close()
    if not columns:
        raise RuntimeError(f"sheet '{title}' had no header")
    casters = [_column_caster([r[k] for r in rows]) for k in range(len(columns))]
    return columns, rows, casters


def fetch_one(node_id: str) -> None:
    asset = node_id  # the spec id IS the asset name
    slug = node_id[len(SLUG) + 1:]  # strip "gem-global-energy-monitor-"

    token = _mint_token(slug)
    url = _presign(token, slug)
    blob = _download_xlsx(url)
    print(f"  {slug}: downloaded {len(blob):,} bytes")

    columns, rows, casters = _parse_workbook(blob)
    if not rows:
        raise RuntimeError(f"{slug}: primary sheet produced 0 data rows")
    print(f"  {slug}: {len(rows):,} rows x {len(columns)} cols")

    def _records():
        for r in rows:
            yield {columns[k]: casters[k](r[k]) for k in range(len(columns))}

    # Stream-serialize; only the compressed buffer is held in memory.
    save_raw_ndjson(_records(), asset, compression="zstd")


DOWNLOAD_SPECS = [
    NodeSpec(
        id=f"{SLUG}-{eid.lower().replace('_', '-')}",
        fn=fetch_one,
        kind="download",
    )
    for eid in ENTITY_IDS
]

# One published Delta table per slug: a thin typed passthrough of the parsed
# registry (NDJSON columns are already coerced to a single type each).
TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"{s.id}-transform",
        deps=[s.id],
        sql=f'SELECT * FROM "{s.id}"',
    )
    for s in DOWNLOAD_SPECS
]
