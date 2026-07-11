"""SAPS Crime Statistics workbook downloads.

SAPS publishes detailed crime statistics as human-formatted XLSX workbooks on
two static HTML index pages. There is no API and no single all-history export,
so this node re-discovers the XLSX links on every run, downloads the workbooks,
and normalizes the `RAW Data` sheet into SQL-readable parquet fragments.

The workbook measure columns vary by release. Raw keeps a stable generic schema:
dimension columns plus numbered `measure_NN_label` / `measure_NN_value` pairs.
Transforms can unpivot those pairs after profiling the real raw output.
"""

from __future__ import annotations

import io
import re
import urllib.parse
from html.parser import HTMLParser
from time import sleep

import httpx
import openpyxl
import pyarrow as pa

from subsets_utils import NodeSpec, save_raw_parquet


PREFIX = "saps-crime-stats-"
CRIME_STATS_ASSET_ID = "saps-crime-stats-crime-statistics"
RELEASES_ASSET_ID = "saps-crime-stats-releases"
INDEX_URLS = (
    "https://www.saps.gov.za/services/crimestats.php",
    "https://www.saps.gov.za/services/older_crimestats.php",
)
MAX_MEASURES = 80


BASE_FIELDS = [
    ("source_workbook", pa.string()),
    ("source_page", pa.string()),
    ("source_url", pa.string()),
    ("release_year_start", pa.int16()),
    ("release_year_end", pa.int16()),
    ("sheet", pa.string()),
    ("excel_row_number", pa.int32()),
    ("release_title", pa.string()),
    ("crime_category_national_contribution_placement", pa.string()),
    ("crime_category_provincial_contribution_placement", pa.string()),
    ("component_level", pa.string()),
    ("station_crime_category", pa.string()),
    ("station", pa.string()),
    ("district", pa.string()),
    ("province", pa.string()),
    ("crime_category", pa.string()),
    ("crime_code", pa.string()),
]
MEASURE_FIELDS = [
    field
    for idx in range(1, MAX_MEASURES + 1)
    for field in (
        (f"measure_{idx:02d}_label", pa.string()),
        (f"measure_{idx:02d}_value", pa.string()),
    )
]
SCHEMA = pa.schema(BASE_FIELDS + MEASURE_FIELDS)
RELEASE_SCHEMA = pa.schema(
    [
        ("source_page", pa.string()),
        ("source_workbook", pa.string()),
        ("source_url", pa.string()),
        ("resolved_url", pa.string()),
        ("release_year_start", pa.int16()),
        ("release_year_end", pa.int16()),
        ("fragment", pa.string()),
    ]
)


class _LinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.hrefs: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        attrs_dict = dict(attrs)
        href = attrs_dict.get("href") or ""
        if ".xlsx" in href.lower():
            self.hrefs.append(href)


def _client() -> httpx.Client:
    # SAPS serves an incomplete certificate chain to Python's trust store.
    return httpx.Client(
        verify=False,
        follow_redirects=True,
        timeout=httpx.Timeout(300.0, connect=20.0, read=120.0),
        headers={"User-Agent": "Mozilla/5.0 (compatible; subsets-saps-crime-stats/1.0)"},
    )


def _get_text(client: httpx.Client, url: str) -> str:
    resp = client.get(url)
    resp.raise_for_status()
    return resp.text


def _get_bytes(client: httpx.Client, url: str) -> bytes:
    last_error: Exception | None = None
    for attempt in range(3):
        try:
            resp = client.get(url)
            resp.raise_for_status()
            content = resp.content
            if content[:4] != b"PK\x03\x04":
                raise RuntimeError(f"{url}: expected XLSX bytes, got {content[:16]!r}")
            expected = resp.headers.get("content-length")
            if expected and len(content) < int(expected):
                raise RuntimeError(
                    f"{url}: truncated response {len(content)} < {expected} bytes"
                )
            return content
        except httpx.HTTPStatusError as exc:
            last_error = exc
            status = exc.response.status_code
            if 400 <= status < 500:
                raise
            if attempt < 2:
                sleep(2 ** (attempt + 1))
        except Exception as exc:
            last_error = exc
            if attempt < 2:
                sleep(2 ** (attempt + 1))
    assert last_error is not None
    raise last_error


def _discover_workbooks(client: httpx.Client) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    seen: set[str] = set()
    for page in INDEX_URLS:
        parser = _LinkParser()
        parser.feed(_get_text(client, page))
        for href in parser.hrefs:
            url = urllib.parse.urljoin(page, href)
            key = urllib.parse.unquote(url).lower()
            if key in seen:
                continue
            seen.add(key)
            out.append(
                {
                    "url": url,
                    "source_page": page,
                    "filename": urllib.parse.unquote(url.rsplit("/", 1)[-1]),
                }
            )
    return out


def _candidate_urls(meta: dict[str, str]) -> list[str]:
    url = meta["url"]
    filename = url.rsplit("/", 1)[-1]
    release_start, release_end = _release_years(meta["filename"])
    urls: list[str] = []
    if "/services/downloads/" in url:
        base = "https://www.saps.gov.za/services/downloads/"
        path_after_downloads = url.split("/services/downloads/", 1)[1]
        if "/" in path_after_downloads:
            urls.append(url)
        for year in (release_end, release_start):
            if year is not None:
                urls.append(f"{base}{year}/{filename}")
        urls.append(url)
    else:
        urls.append(url)
        urls.append(f"https://www.saps.gov.za/services/downloads/{filename}")

    seen: set[str] = set()
    out: list[str] = []
    for candidate in urls:
        key = urllib.parse.unquote(candidate).lower()
        if key not in seen:
            seen.add(key)
            out.append(candidate)
    return out


def _clean(value) -> str | None:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    text = " ".join(str(value).split())
    return text if text else None


def _header_key(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _fragment_name(filename: str) -> str:
    base = filename.rsplit(".", 1)[0].lower()
    base = re.sub(r"[^a-z0-9]+", "-", base).strip("-")
    return base[:120] or "workbook"


def _release_years(filename: str) -> tuple[int | None, int | None]:
    years = [int(y) for y in re.findall(r"(20\d{2})", filename)]
    if len(years) >= 2:
        return min(years[:2]), max(years[:2])
    if len(years) == 1:
        return years[0], None
    return None, None


def _find_header(rows: list[tuple]) -> int:
    required = {"station", "district", "province", "crimecategory", "code"}
    for idx, row in enumerate(rows):
        normalized = {_header_key(v) for v in row if v is not None}
        if required.issubset(normalized):
            return idx
    raise RuntimeError("detailed data header row not found")


def _header_positions(headers: list[str | None]) -> dict[str, int]:
    aliases = {
        "crime_category_national_contribution_placement": {
            "crimecategorynationalcontributionplacement",
            "nationalcontributionplacement",
        },
        "crime_category_provincial_contribution_placement": {
            "crimecategoryprovincialcontributionplacement",
            "provincialcontributionplacement",
        },
        "component_level": {"componentlevel"},
        "station_crime_category": {"stationcrimecategory"},
        "station": {"station"},
        "district": {"district"},
        "province": {"province"},
        "crime_category": {"crimecategory"},
        "crime_code": {"code", "crimecode"},
    }
    keys = [_header_key(header) for header in headers]
    positions: dict[str, int] = {}
    for field, field_aliases in aliases.items():
        for idx, key in enumerate(keys):
            if key in field_aliases:
                positions[field] = idx
                break
    missing = {"station", "district", "province", "crime_category", "crime_code"} - set(
        positions
    )
    if missing:
        raise RuntimeError(f"detailed data header missing columns: {sorted(missing)}")
    return positions


def _release_title(rows: list[tuple], header_idx: int) -> str | None:
    titles: list[str] = []
    for row in rows[:header_idx]:
        for value in row:
            text = _clean(value)
            if text and "comparison" in text.lower():
                titles.append(text)
    return " | ".join(dict.fromkeys(titles)) or None


def _read_detailed_sheet(content: bytes, filename: str) -> tuple[str, list[tuple], int]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        failures: list[str] = []
        preferred = [name for name in wb.sheetnames if name == "RAW Data"]
        fallback = [name for name in wb.sheetnames if name not in preferred]
        for sheet_name in preferred + fallback:
            ws = wb[sheet_name]
            ws.reset_dimensions()
            rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
            try:
                return sheet_name, rows, _find_header(rows)
            except Exception as exc:
                failures.append(f"{sheet_name}: {exc}")
    finally:
        wb.close()

    raise RuntimeError(
        f"{filename}: detailed data sheet not found; checked {len(failures)} sheets"
    )


def _parse_workbook(content: bytes, meta: dict[str, str]) -> pa.Table:
    sheet_name, rows, header_idx = _read_detailed_sheet(content, meta["filename"])
    headers = [_clean(v) for v in rows[header_idx]]
    positions = _header_positions(headers)
    measure_headers = headers[9 : 9 + MAX_MEASURES]
    release_start, release_end = _release_years(meta["filename"])
    title = _release_title(rows, header_idx)

    columns: dict[str, list] = {name: [] for name in SCHEMA.names}

    def row_value(row: tuple, field: str) -> str | None:
        idx = positions.get(field)
        if idx is None or len(row) <= idx:
            return None
        return _clean(row[idx])

    def append_base(row: tuple, excel_row_number: int) -> None:
        values = {
            "source_workbook": meta["filename"],
            "source_page": meta["source_page"],
            "source_url": meta["url"],
            "release_year_start": release_start,
            "release_year_end": release_end,
            "sheet": sheet_name,
            "excel_row_number": excel_row_number,
            "release_title": title,
            "crime_category_national_contribution_placement": row_value(
                row, "crime_category_national_contribution_placement"
            ),
            "crime_category_provincial_contribution_placement": row_value(
                row, "crime_category_provincial_contribution_placement"
            ),
            "component_level": row_value(row, "component_level"),
            "station_crime_category": row_value(row, "station_crime_category"),
            "station": row_value(row, "station"),
            "district": row_value(row, "district"),
            "province": row_value(row, "province"),
            "crime_category": row_value(row, "crime_category"),
            "crime_code": row_value(row, "crime_code"),
        }
        for key, value in values.items():
            columns[key].append(value)

    row_count = 0
    for excel_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not any(row[:9]):
            continue
        if _clean(row[4] if len(row) > 4 else None) is None:
            continue
        append_base(row, excel_idx)
        for idx in range(MAX_MEASURES):
            label_col = f"measure_{idx + 1:02d}_label"
            value_col = f"measure_{idx + 1:02d}_value"
            label = measure_headers[idx] if idx < len(measure_headers) else None
            value = row[idx + 9] if len(row) > idx + 9 else None
            columns[label_col].append(label)
            columns[value_col].append(_clean(value))
        row_count += 1

    if row_count == 0:
        raise RuntimeError(f"{meta['filename']}: parsed 0 RAW Data rows")
    return pa.table(columns, schema=SCHEMA)


def fetch_crime_statistics(node_id: str) -> None:
    if node_id != CRIME_STATS_ASSET_ID:
        raise ValueError(f"unexpected node id {node_id!r}")

    failures: list[str] = []
    written = 0
    with _client() as client:
        for workbook in _discover_workbooks(client):
            try:
                content = None
                errors: list[str] = []
                for candidate in _candidate_urls(workbook):
                    try:
                        workbook["resolved_url"] = candidate
                        content = _get_bytes(client, candidate)
                        break
                    except Exception as exc:
                        errors.append(f"{candidate}: {exc}")
                if content is None:
                    raise RuntimeError("; ".join(errors[:3]))
                table = _parse_workbook(content, workbook)
            except Exception as exc:
                failures.append(f"{workbook['filename']}: {exc}")
                continue
            save_raw_parquet(table, node_id, fragment=_fragment_name(workbook["filename"]))
            written += 1

    if written < 20:
        detail = "; ".join(failures[:5])
        raise RuntimeError(
            f"{node_id}: wrote only {written} workbook fragments; expected at least 20. "
            f"First failures: {detail}"
        )


def fetch_releases(node_id: str) -> None:
    if node_id != RELEASES_ASSET_ID:
        raise ValueError(f"unexpected node id {node_id!r}")

    rows: list[dict[str, object]] = []
    with _client() as client:
        for workbook in _discover_workbooks(client):
            release_start, release_end = _release_years(workbook["filename"])
            resolved_url = None
            for candidate in _candidate_urls(workbook):
                try:
                    resp = client.head(candidate)
                    if resp.status_code == 200:
                        resolved_url = candidate
                        break
                except Exception:
                    continue
            rows.append(
                {
                    "source_page": workbook["source_page"],
                    "source_workbook": workbook["filename"],
                    "source_url": workbook["url"],
                    "resolved_url": resolved_url,
                    "release_year_start": release_start,
                    "release_year_end": release_end,
                    "fragment": _fragment_name(workbook["filename"]),
                }
            )

    if len(rows) < 20:
        raise RuntimeError(f"{node_id}: discovered only {len(rows)} workbook releases")
    save_raw_parquet(pa.Table.from_pylist(rows, schema=RELEASE_SCHEMA), node_id)


DOWNLOAD_SPECS = [
    NodeSpec(id=CRIME_STATS_ASSET_ID, fn=fetch_crime_statistics, kind="download"),
    NodeSpec(id=RELEASES_ASSET_ID, fn=fetch_releases, kind="download"),
]
