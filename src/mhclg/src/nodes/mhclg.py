"""MHCLG connector — gov.uk statistical publications.

Mechanism (research-chosen): `gov_uk` REST. Each rank-accepted entity is a gov.uk
publication (a `statistical_data_set` "live table" or a designated statistics
release). Per entity we resolve its attachments from the Content API
(`https://www.gov.uk/api/content/<base_path>`) and download every tabular
attachment (CSV / XLS / XLSX / ODS).

The attachments are heterogeneous, multi-sheet spreadsheets with irregular
title/merged-header layouts that differ table-to-table and year-to-year, so no
single rectangular schema fits a publication — and some publications ship
record-level microdata in 30 MB+ ODS files. We therefore publish a faithful,
uniform **row-level long extraction**: one row per non-empty source row,
carrying its source attachment, table title, sheet, 0-based row index, and the
row's cell values as a JSON-encoded array of strings. A consumer reconstructs
any underlying table by filtering on attachment_title / sheet_name and parsing
`cells`.

Every reader streams (chunked CSV, openpyxl read-only, lxml row-level iterparse
for ODS) and rows are flushed in bounded parquet batches, so memory stays flat
even on the 200 MB fire-statistics publication.

Fetch shape: stateless full re-pull (shape 1). Asset URLs are content-hash paths
that change on re-upload, so we re-resolve them from the Content API every run
and never store a watermark.
"""

import io
import json
import zipfile

import pyarrow as pa

from subsets_utils import (
    NodeSpec,
    SqlNodeSpec,
    get,
    transient_retry,
    raw_parquet_writer,
    raw_asset_exists,
)
from constants import ENTITY_IDS, BASE_PATHS

CONTENT_API = "https://www.gov.uk/api/content/"

_CSV_TYPES = {"text/csv", "text/plain"}
_XLSX_TYPES = {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
_XLS_TYPES = {"application/vnd.ms-excel"}
_ODS_TYPES = {"application/vnd.oasis.opendocument.spreadsheet"}
_TABULAR_TYPES = _CSV_TYPES | _XLSX_TYPES | _XLS_TYPES | _ODS_TYPES

SCHEMA = pa.schema([
    ("attachment_filename", pa.string()),
    ("attachment_title", pa.string()),
    ("content_type", pa.string()),
    ("sheet_name", pa.string()),
    ("row_index", pa.int64()),
    ("n_cols", pa.int64()),
    ("cells", pa.string()),  # JSON array of stringified cell values
])

_BATCH_ROWS = 20_000  # source rows per parquet row group flush

# ODS XML namespaces
_NS_T = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
_NS_O = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"
_NS_TX = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"


def _qn(ns, tag):
    return "{%s}%s" % (ns, tag)


@transient_retry()
def _get_json(url):
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.json()


@transient_retry()
def _get_bytes(url):
    resp = get(url, timeout=(10.0, 300.0))
    resp.raise_for_status()
    return resp.content


def _clean(cells):
    """Trim trailing empties; return list or None if the row is all-empty."""
    out = ["" if c is None else str(c).strip() for c in cells]
    while out and out[-1] == "":
        out.pop()
    if not out or all(c == "" for c in out):
        return None
    return out


def _rows_csv(raw):
    import pandas as pd

    for chunk in pd.read_csv(
        io.BytesIO(raw), header=None, dtype=str, keep_default_na=False,
        na_values=[], chunksize=10_000, low_memory=False, encoding_errors="replace",
    ):
        base = chunk.index[0]
        for off, row in enumerate(chunk.itertuples(index=False, name=None)):
            cells = _clean(row)
            if cells is not None:
                yield "csv", base + off, cells


def _rows_xlsx(raw):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            for ri, row in enumerate(ws.iter_rows(values_only=True)):
                cells = _clean(row)
                if cells is not None:
                    yield ws.title, ri, cells
    finally:
        wb.close()


def _rows_xls(raw):
    import xlrd

    book = xlrd.open_workbook(file_contents=raw)
    for sheet in book.sheets():
        for ri in range(sheet.nrows):
            cells = _clean(sheet.row_values(ri))
            if cells is not None:
                yield sheet.name, ri, cells


_MAX_COLS = 4096  # cap a repeated-cell run so full-width padding can't explode a row


def _rows_ods(raw):
    """Stream ODS rows via lxml row-level iterparse, freeing each row's subtree
    so a record-level ODS stays flat in memory.

    ODS encodes blank padding with `number-columns-repeated` (up to 16384, the
    sheet width) and `number-rows-repeated` (up to ~1.05M, the sheet height).
    Naively expanding every repeated empty cell turns an 87 MB file into billions
    of throwaway list entries (O(rows × 16384)). Instead we defer empty cells into
    a running counter and only materialise them when a real value follows — so a
    trailing full-width pad is never built (it's dropped, exactly as `_clean`
    would), interior gaps stay bounded by `_MAX_COLS`, and only true data cells
    cost work. Blank rows still advance the index without emitting."""
    from lxml import etree

    row_tag = _qn(_NS_T, "table-row")
    cell_tag = _qn(_NS_T, "table-cell")
    table_tag = _qn(_NS_T, "table")
    name_attr = _qn(_NS_T, "name")
    rep_rows_attr = _qn(_NS_T, "number-rows-repeated")
    rep_cols_attr = _qn(_NS_T, "number-columns-repeated")
    value_attr = _qn(_NS_O, "value")
    p_tag = _qn(_NS_TX, "p")

    zf = zipfile.ZipFile(io.BytesIO(raw))
    with zf.open("content.xml") as f:
        # only fire end events for rows and tables — skipping per-cell/per-text
        # callbacks is the difference between minutes and seconds on big ODS.
        ctx = etree.iterparse(f, events=("end",), tag=(row_tag, table_tag))
        table_rows = {}  # id(table) -> running row index
        for _event, el in ctx:
            if el.tag == table_tag:
                el.clear()
                continue
            parent = el.getparent()
            cur_table = (parent.get(name_attr) or "table") if parent is not None else "table"
            row_idx = table_rows.get(id(parent), 0)
            rep = int(el.get(rep_rows_attr) or 1)
            cells = []
            pending_empty = 0  # deferred empties — flushed only before a real value
            for cell in el.iterfind(cell_tag):
                crep = int(cell.get(rep_cols_attr) or 1)
                if crep > _MAX_COLS:
                    crep = _MAX_COLS
                val = cell.get(value_attr)
                if val is None:
                    ps = cell.findall(p_tag)
                    val = "".join("".join(p.itertext()) for p in ps) if ps else ""
                if val == "":
                    pending_empty += crep
                else:
                    if pending_empty:
                        cells.extend([""] * pending_empty)
                        pending_empty = 0
                    cells.extend([val] * crep)
            # trailing pending_empty is intentionally dropped (== _clean's trim)
            cleaned = _clean(cells)
            if cleaned is None:
                row_idx += rep  # blank run — advance index without emitting
            else:
                for _ in range(rep):
                    yield cur_table, row_idx, cleaned
                    row_idx += 1
            table_rows[id(parent)] = row_idx
            # free this row and its already-processed previous siblings
            el.clear()
            prev = el.getprevious()
            while prev is not None:
                parent.remove(prev)
                prev = el.getprevious()


def _rows_for(content_type, raw):
    if content_type in _CSV_TYPES:
        return _rows_csv(raw)
    if content_type in _XLSX_TYPES:
        return _rows_xlsx(raw)
    if content_type in _XLS_TYPES:
        return _rows_xls(raw)
    if content_type in _ODS_TYPES:
        return _rows_ods(raw)
    return iter(())


def fetch_one(node_id: str) -> None:
    asset = node_id  # the spec id IS the asset name
    # Resume contract: continuations re-invoke every pending download, so a node
    # already materialised in a prior invocation must be a no-op. Without this,
    # each continuation re-pulls all completed assets and starves the budget for
    # the not-yet-done heavy nodes — the DAG never converges.
    if raw_asset_exists(asset):
        print(f"  -> skip {asset} (raw already present)")
        return
    entity_id = node_id[len("mhclg-"):]
    base_path = BASE_PATHS[entity_id]

    doc = _get_json(CONTENT_API + base_path)
    attachments = (doc.get("details") or {}).get("attachments") or []

    buf = {k: [] for k in SCHEMA.names}
    wrote_any = False
    with raw_parquet_writer(asset, SCHEMA) as writer:
        def flush():
            nonlocal wrote_any
            if not buf["row_index"]:
                return
            writer.write_table(pa.table(buf, schema=SCHEMA))
            for v in buf.values():
                v.clear()
            wrote_any = True

        for att in attachments:
            ct = att.get("content_type")
            url = att.get("url")
            if ct not in _TABULAR_TYPES or not url:
                continue
            filename = att.get("filename") or url.rsplit("/", 1)[-1]
            title = att.get("title") or filename
            raw = _get_bytes(url)
            try:
                for sheet_name, row_index, cells in _rows_for(ct, raw):
                    buf["attachment_filename"].append(filename)
                    buf["attachment_title"].append(title)
                    buf["content_type"].append(ct)
                    buf["sheet_name"].append(str(sheet_name))
                    buf["row_index"].append(int(row_index))
                    buf["n_cols"].append(len(cells))
                    buf["cells"].append(json.dumps(cells, ensure_ascii=False))
                    if len(buf["row_index"]) >= _BATCH_ROWS:
                        flush()
            except Exception as exc:  # noqa: BLE001 - one bad file shouldn't sink the asset
                print(f"  !! parse failed {asset} {filename} ({ct}): {type(exc).__name__}: {exc}")
                continue
        flush()

        if not wrote_any:
            # No parseable rows — write a schema-only table so the asset exists;
            # the transform's 0-row guard surfaces the problem.
            writer.write_table(SCHEMA.empty_table())


DOWNLOAD_SPECS = [
    NodeSpec(
        id=f"mhclg-{eid.lower().replace('_', '-')}",
        fn=fetch_one,
        kind="download",
    )
    for eid in ENTITY_IDS
]

# One published Delta table per publication: a thin pass over the row extraction,
# dropping any all-empty rows. Uniform across all subsets (uniform raw schema).
TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"{s.id}-transform",
        deps=[s.id],
        sql=f'''
            SELECT
                attachment_filename,
                attachment_title,
                content_type,
                sheet_name,
                CAST(row_index AS BIGINT) AS row_index,
                CAST(n_cols AS BIGINT) AS n_cols,
                cells
            FROM "{s.id}"
            WHERE n_cols > 0
        ''',
    )
    for s in DOWNLOAD_SPECS
]
