"""National Bank of Rwanda (BNR) download nodes."""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from urllib.parse import urljoin

import pyarrow as pa
import xlrd
from openpyxl import load_workbook

from constants import CURRENCIES, DOCUMENT_ENDPOINTS, DOCUMENT_ENTITY_IDS
from subsets_utils import NodeSpec, get, save_raw_parquet

_FX_BASE = "https://fxrates.bnr.rw/currency_history/"
_BNR_BASE = "https://www.bnr.rw"
_START_DATE = "01/01/2000"
_PREFIX = "national-bank-of-rwanda-"

_UA_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120 Safari/537.36"
    )
}

_FX_SCHEMA = pa.schema([
    ("currency_name", pa.string()),
    ("buying_rate", pa.string()),
    ("average_rate", pa.string()),
    ("selling_rate", pa.string()),
    ("post_date", pa.string()),
])

_WORKBOOK_CELL_SCHEMA = pa.schema([
    ("source_entity_id", pa.string()),
    ("source_endpoint", pa.string()),
    ("source_url", pa.string()),
    ("source_file", pa.string()),
    ("release_name", pa.string()),
    ("source_modified_date", pa.date32()),
    ("sheet_name", pa.string()),
    ("row_number", pa.int32()),
    ("column_number", pa.int32()),
    ("cell_ref", pa.string()),
    ("cell_value", pa.string()),
])


def fetch_exchange_rates(node_id: str) -> None:
    end_date = date.today().strftime("%d/%m/%Y")
    rows: list[dict] = []
    for currency in CURRENCIES:
        resp = get(
            _FX_BASE,
            params={
                "currency_name": currency,
                "start_date": _START_DATE,
                "end_date": end_date,
            },
            timeout=(10.0, 120.0),
        )
        resp.raise_for_status()
        for record in resp.json():
            rows.append({
                "currency_name": _s(record.get("currency_name") or currency),
                "buying_rate": _s(record.get("buying_rate")),
                "average_rate": _s(record.get("average_rate")),
                "selling_rate": _s(record.get("selling_rate")),
                "post_date": _s(record.get("post_date")),
            })

    if not rows:
        raise AssertionError(f"{node_id}: fetched 0 FX rows")

    save_raw_parquet(pa.Table.from_pylist(rows, schema=_FX_SCHEMA), node_id)


def fetch_document_workbook(node_id: str) -> None:
    entity_id = _entity_id_from_node(node_id)
    match = _find_latest_document(entity_id)
    file_path = match.get("file") or ""
    source_url = urljoin(_BNR_BASE, file_path)

    resp = get(source_url, headers=_UA_HEADERS, timeout=(10.0, 180.0))
    resp.raise_for_status()
    rows = _workbook_cells(entity_id, match, source_url, resp.content)
    if not rows:
        raise AssertionError(f"{node_id}: workbook contained 0 non-empty cells from {source_url}")

    save_raw_parquet(pa.Table.from_pylist(rows, schema=_WORKBOOK_CELL_SCHEMA), node_id)


def _find_latest_document(entity_id: str) -> dict:
    best: tuple[str, str, dict] | None = None
    for endpoint in DOCUMENT_ENDPOINTS:
        resp = get(f"{_BNR_BASE}/{endpoint}", headers=_UA_HEADERS, timeout=(10.0, 60.0))
        resp.raise_for_status()
        records = resp.json()
        if not isinstance(records, list):
            continue
        for record in records:
            file_path = (record.get("file") or "").strip()
            if not file_path.lower().endswith((".xls", ".xlsx")):
                continue
            name = (record.get("name") or record.get("title") or "").strip()
            if _logical_id(name) != entity_id:
                continue
            modified = record.get("date_last_modified") or ""
            if best is None or modified > best[0]:
                best = (modified, endpoint, record)

    if best is None:
        raise AssertionError(f"no current Excel document found for entity_id={entity_id!r}")

    record = dict(best[2])
    record["_endpoint"] = best[1]
    return record


def _workbook_cells(entity_id: str, record: dict, source_url: str, content: bytes) -> list[dict]:
    source_file = (record.get("file") or "").strip()
    release_name = (record.get("name") or record.get("title") or "").strip()
    modified_date = _parse_modified_date(record.get("date_last_modified"))
    rows: list[dict] = []

    if content.startswith(b"PK\x03\x04"):
        return _xlsx_cells(entity_id, record, source_url, content)
    if content.startswith(b"\xd0\xcf\x11\xe0"):
        return _xls_cells(entity_id, record, source_url, content)

    raise AssertionError(f"{source_url} is not an Excel workbook")


def _xlsx_cells(entity_id: str, record: dict, source_url: str, content: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    source_file = (record.get("file") or "").strip()
    release_name = (record.get("name") or record.get("title") or "").strip()
    modified_date = _parse_modified_date(record.get("date_last_modified"))
    rows: list[dict] = []

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = _cell_value(cell.value)
                if value is None:
                    continue
                rows.append({
                    "source_entity_id": entity_id,
                    "source_endpoint": record.get("_endpoint"),
                    "source_url": source_url,
                    "source_file": source_file,
                    "release_name": release_name,
                    "source_modified_date": modified_date,
                    "sheet_name": sheet.title,
                    "row_number": cell.row,
                    "column_number": cell.column,
                    "cell_ref": cell.coordinate,
                    "cell_value": value,
                })

    return rows


def _xls_cells(entity_id: str, record: dict, source_url: str, content: bytes) -> list[dict]:
    workbook = xlrd.open_workbook(file_contents=content)
    source_file = (record.get("file") or "").strip()
    release_name = (record.get("name") or record.get("title") or "").strip()
    modified_date = _parse_modified_date(record.get("date_last_modified"))
    rows: list[dict] = []

    for sheet in workbook.sheets():
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                value = _xlrd_cell_value(cell, workbook.datemode)
                if value is None:
                    continue
                row_number = row_idx + 1
                column_number = col_idx + 1
                rows.append({
                    "source_entity_id": entity_id,
                    "source_endpoint": record.get("_endpoint"),
                    "source_url": source_url,
                    "source_file": source_file,
                    "release_name": release_name,
                    "source_modified_date": modified_date,
                    "sheet_name": sheet.name,
                    "row_number": row_number,
                    "column_number": column_number,
                    "cell_ref": f"{_excel_col(column_number)}{row_number}",
                    "cell_value": value,
                })

    return rows


def _logical_id(name: str) -> str:
    s = name.strip()
    s = re.sub(r"\.(xlsx?|pdf|docx)$", "", s, flags=re.I)
    s = re.sub(r"[_\.]+", " ", s)
    s = re.split(r"\b(?:up\s*to|as\s*of|on|issued\s+from|from)\b", s, flags=re.I)[0]
    s = re.sub(r"\bQ[1-4]\b", " ", s, flags=re.I)
    s = re.sub(r"\b(?:from|to)\b", " ", s, flags=re.I)
    s = re.sub(
        r"\b(jan(uary)?|feb(ruary)?|mar(ch)?|apr(il)?|may|jun(e)?|jul(y)?|"
        r"aug(ust)?|sep(t)?(ember)?|oct(ober)?|nov(ember)?|dec(ember)?)\b",
        " ",
        s,
        flags=re.I,
    )
    s = re.sub(r"\bFY\b", " ", s, flags=re.I)
    s = re.sub(r"\b\d{4}\b", " ", s)
    s = re.sub(r"\b\d{1,2}(st|nd|rd|th)\b", " ", s, flags=re.I)
    s = re.sub(r"\b\d{1,2}\b", " ", s)
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return re.sub(r"-+", "-", s) or "document"


def _entity_id_from_node(node_id: str) -> str:
    if not node_id.startswith(_PREFIX):
        raise AssertionError(f"unexpected node id {node_id!r}")
    return node_id.removeprefix(_PREFIX)


def _parse_modified_date(value) -> date | None:
    if not value:
        return None
    text = str(value).replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _cell_value(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _xlrd_cell_value(cell, datemode: int) -> str | None:
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(cell.value, datemode).isoformat()
        except (OverflowError, ValueError, xlrd.XLDateError):
            return str(cell.value)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return "true" if cell.value else "false"
    if cell.ctype == xlrd.XL_CELL_NUMBER and float(cell.value).is_integer():
        return str(int(cell.value))
    text = str(cell.value).strip()
    return text or None


def _excel_col(number: int) -> str:
    label = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        label = chr(65 + remainder) + label
    return label


def _s(value) -> str | None:
    return None if value is None else str(value)


DOWNLOAD_SPECS = [
    NodeSpec(
        id="national-bank-of-rwanda-exchange-rates",
        fn=fetch_exchange_rates,
        kind="download",
    ),
    *[
        NodeSpec(
            id=f"national-bank-of-rwanda-{entity_id}",
            fn=fetch_document_workbook,
            kind="download",
        )
        for entity_id in DOCUMENT_ENTITY_IDS
    ],
]
