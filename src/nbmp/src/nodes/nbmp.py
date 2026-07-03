"""National Bat Monitoring Programme (NBMP) — official population-trend statistics.

Source: JNCC/BCT 'Population trends for bat species in the UK', a single annual
.xlsx workbook published under Open Government Licence v3.0. There is no API; the
workbook is the bulk export for the whole programme.

Fetch shape: stateless full re-pull (shape 1). The file is ~185 KB and published
once a year — re-fetch the whole thing each run and overwrite. The download link
on the entry page points at an opaque CMS media slug that changes every annual
release, so we resolve the current .xlsx href from the entry page rather than
hardcoding it.

Two publishable subsets, one per data sheet (each a distinct schema):
  - nbmp-trend-indices     <- sheet '1_Trend_indices_and_CIs' (annual index series)
  - nbmp-population-trends <- sheet '2_Population_trends'      (short/long-term trends)

Raw is stored as parquet with an all-string schema: the workbook mixes numeric
cells with literal 'Na' markers in numeric columns, so we keep every value as
text and TRY_CAST to numeric types in the transform SQL.
"""

import io
import re
from urllib.parse import urljoin

import pyarrow as pa

from subsets_utils import (
    NodeSpec,
    SqlNodeSpec,
    get,
    save_raw_parquet,
    transient_retry,
)

ENTRY_URL = "https://jncc.gov.uk/our-work/bat-monitoring-official-statistics/"

SHEET1 = "1_Trend_indices_and_CIs"
SHEET2 = "2_Population_trends"

SHEET1_COLS = [
    "geographical_scale",
    "species",
    "survey",
    "official_statistic",
    "year",
    "smoothed_index",
    "se",
    "lower_95_ci",
    "upper_95_ci",
]
SHEET2_COLS = [
    "country",
    "species",
    "survey_type",
    "term",
    "average_number_of_sites",
    "trend_pct_change",
    "lower_confidence_limit",
    "upper_confidence_limit",
    "significance_of_change",
]


# --- fetch helpers -----------------------------------------------------------

@transient_retry()
def _http_get(url: str):
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp


def _resolve_xlsx_url() -> str:
    """Scrape the entry page for the current 'bats-nbmp-dataset...*.xlsx' link."""
    html = _http_get(ENTRY_URL).text
    m = re.search(r'href=["\']([^"\']*bats-nbmp-dataset[^"\']*\.xlsx)["\']',
                  html, re.IGNORECASE)
    if not m:
        raise AssertionError(
            f"could not find a 'bats-nbmp-dataset*.xlsx' link on {ENTRY_URL}"
        )
    return urljoin(ENTRY_URL, m.group(1))


def _load_workbook():
    import openpyxl
    url = _resolve_xlsx_url()
    content = _http_get(url).content
    return openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)


def _norm(v):
    if v is None:
        return None
    s = str(v).replace("\n", " ").replace("\r", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s or None


def _clean_country(c):
    """The workbook has a corrupt cell ('UK+H20...') — keep the leading name."""
    if not c:
        return c
    m = re.match(r"^[A-Za-z ]+", c)
    return m.group(0).strip() if m else c


def _save_strings(rows, columns, asset):
    schema = pa.schema([(c, pa.string()) for c in columns])
    cols = {c: [r.get(c) for r in rows] for c in columns}
    table = pa.table(cols, schema=schema)
    save_raw_parquet(table, asset)


# --- download fetch fns ------------------------------------------------------

def fetch_trend_indices(node_id: str) -> None:
    asset = node_id  # nbmp-trend-indices
    wb = _load_workbook()
    ws = wb[SHEET1]
    grid = [[_norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
    hdr_i = next(
        i for i, r in enumerate(grid)
        if r[:2] == ["Geographical scale", "Species"]
    )
    rows = []
    for r in grid[hdr_i + 1:]:
        if not any(r):
            continue
        rows.append(dict(zip(SHEET1_COLS, r[:len(SHEET1_COLS)])))
    if not rows:
        raise AssertionError(f"no data rows parsed from {SHEET1}")
    _save_strings(rows, SHEET1_COLS, asset)


def fetch_population_trends(node_id: str) -> None:
    asset = node_id  # nbmp-population-trends
    wb = _load_workbook()
    ws = wb[SHEET2]
    grid = [[_norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
    term = None
    rows = []
    for r in grid:
        first = r[0] or ""
        if first.startswith("Table"):
            if "Short-term" in first:
                term = "short_term"
            elif "Long-term" in first:
                term = "long_term"
            else:
                term = None
            continue
        if r[:2] == ["Country", "Species"]:
            continue
        if term and r[0] and r[1]:
            rows.append({
                "country": _clean_country(r[0]),
                "species": r[1],
                "survey_type": r[2],
                "term": term,
                "average_number_of_sites": r[3],
                "trend_pct_change": r[4],
                "lower_confidence_limit": r[5],
                "upper_confidence_limit": r[6],
                "significance_of_change": r[7],
            })
    if not rows:
        raise AssertionError(f"no data rows parsed from {SHEET2}")
    _save_strings(rows, SHEET2_COLS, asset)


DOWNLOAD_SPECS = [
    NodeSpec(id="nbmp-trend-indices", fn=fetch_trend_indices, kind="download"),
    NodeSpec(id="nbmp-population-trends", fn=fetch_population_trends, kind="download"),
]


# --- transforms — one published Delta table per subset -----------------------

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="nbmp-trend-indices-transform",
        deps=["nbmp-trend-indices"],
        key=("geographical_scale", "species", "survey", "year"),
        temporal="year",
        sql='''
            SELECT
                geographical_scale,
                species,
                survey,
                official_statistic,
                CAST(TRY_CAST(year AS DOUBLE) AS INTEGER) AS year,
                TRY_CAST(smoothed_index AS DOUBLE)        AS smoothed_index,
                TRY_CAST(se AS DOUBLE)                    AS standard_error,
                TRY_CAST(lower_95_ci AS DOUBLE)           AS lower_95_ci,
                TRY_CAST(upper_95_ci AS DOUBLE)           AS upper_95_ci
            FROM "nbmp-trend-indices"
            WHERE TRY_CAST(year AS INTEGER) IS NOT NULL
              AND TRY_CAST(smoothed_index AS DOUBLE) IS NOT NULL
              AND species IS NOT NULL
        ''',
    ),
    SqlNodeSpec(
        id="nbmp-population-trends-transform",
        deps=["nbmp-population-trends"],
        key=("country", "species", "survey_type", "term"),
        sql='''
            SELECT
                country,
                species,
                survey_type,
                term,
                TRY_CAST(average_number_of_sites AS INTEGER) AS average_number_of_sites,
                TRY_CAST(trend_pct_change AS DOUBLE)         AS trend_pct_change,
                TRY_CAST(lower_confidence_limit AS DOUBLE)   AS lower_confidence_limit,
                TRY_CAST(upper_confidence_limit AS DOUBLE)   AS upper_confidence_limit,
                significance_of_change
            FROM "nbmp-population-trends"
            WHERE term IS NOT NULL
              AND species IS NOT NULL
              AND country IS NOT NULL
        ''',
    ),
]
