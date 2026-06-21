import io
import openpyxl
from subsets_utils import get

BASE = "https://www.philadelphiafed.org/-/media/FRBP/Assets/Surveys-And-Data"
URLS = {
    "meanLevel": f"{BASE}/survey-of-professional-forecasters/historical-data/meanLevel.xlsx",
    "routput_fst": f"{BASE}/real-time-data/data-files/xlsx/routput_first_second_third.xlsx",
    "dispersion1": f"{BASE}/survey-of-professional-forecasters/historical-data/Dispersion_1.xlsx",
}

for name, url in URLS.items():
    print(f"\n{'='*70}\n{name}: {url}")
    content = get(url, timeout=120).content
    # read_only avoids the eager date-conversion TypeError
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    print(f"  sheets({len(wb.sheetnames)}): {wb.sheetnames[:15]}")
    for sn in wb.sheetnames[:2]:
        ws = wb[sn]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(row)
            if i >= 6:
                break
        print(f"  -- '{sn}' first rows:")
        for r in rows:
            print("    ", r[:12])
    wb.close()
