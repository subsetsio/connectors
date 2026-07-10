"""EDGAR Community GHG inventory (European Commission JRC).

Source: https://edgar.jrc.ec.europa.eu/dataset_ghg2025
Release: EDGAR_2025_GHG (2025 release).

Publishes one long table — emissions by country x IPCC-2006 sector x gas x
year — built by unpivoting the per-substance XLSX workbooks served as ZIPs
from the JRC open-data FTP.

License: we ingest ONLY the CC BY 4.0 substances (EDGAR CH4, EDGAR N2O,
EDGAR F-gases, EDGAR CO2bio). We deliberately DO NOT fetch IEA-EDGAR CO2 or
the EDGAR AR5 GHG total — both embed IEA fossil-CO2 data under CC BY-NC-ND
4.0 (non-commercial / no-derivatives), which we may not redistribute.

Each workbook's "IPCC 2006" sheet is a wide table: a 9-row metadata
preamble, then a header row (index 9) of 8 identifier columns
(IPCC_annex, C_group_IM24_sh, Country_code_A3, Name,
ipcc_code_2006_for_standard_report, ...name, Substance, fossil_bio) followed
by Y_1970..Y_2024 year columns (F-gases start Y_1990). Values are in Gg
(gigagrams = kilotonnes) of the named substance. The F-gases workbook
resolves into ~25 individual species (HFC-125, SF6, NF3, ...) via the
Substance column, which we carry through verbatim as `gas`.

Re-pull strategy: stateless full re-pull. The whole CC-BY tabular set is a
few tens of MB and parses in seconds, so we re-download and overwrite every
run; revisions to historical years are picked up for free.
"""
import io
import os
import zipfile

import openpyxl
import pyarrow as pa

from subsets_utils import (
    MaintainSpec,
    NodeSpec,
    get_client,
    list_raw_fragments,
    save_raw_parquet,
    transient_retry,
)

BASE_URL = (
    "https://jeodpp.jrc.ec.europa.eu/ftp/jrc-opendata/EDGAR/datasets/EDGAR_2025_GHG"
)

# CC BY 4.0 substance packages only. (gas-group label, zip filename) — the
# real per-row gas/species comes from the workbook's Substance column.
GAS_PACKAGES = [
    ("ch4", "EDGAR_CH4_1970_2024.zip"),
    ("n2o", "EDGAR_N2O_1970_2024.zip"),
    ("f-gases", "EDGAR_F-gases_1990_2024.zip"),
    ("co2bio", "EDGAR_CO2bio_1970_2024.zip"),
]

SHEET = "IPCC 2006"
HEADER_ROW = 9  # 0-indexed: 9-row metadata preamble precedes the column header

SCHEMA = pa.schema([
    ("country_code", pa.string()),
    ("country_name", pa.string()),
    ("ipcc_annex", pa.string()),
    ("c_group", pa.string()),
    ("ipcc_sector_code", pa.string()),
    ("ipcc_sector_name", pa.string()),
    ("gas", pa.string()),
    ("fossil_bio", pa.string()),
    ("year", pa.int32()),
    ("emissions_kt", pa.float64()),
])


@transient_retry(attempts=8, min_wait=8, max_wait=300)
def _download_zip(url: str) -> bytes:
    """Download one EDGAR ZIP with a long read timeout.

    The JRC FTP front-end can sit silent for long stretches and occasionally
    drops a connection mid-response. Streaming makes progress visible and lets
    the shared retry wrapper rerun the package-level fetch cleanly.
    """
    client = get_client()
    chunks: list[bytes] = []
    total = 0
    with client.stream("GET", url, timeout=(30.0, 1800.0)) as resp:
        resp.raise_for_status()
        for chunk in resp.iter_bytes(chunk_size=1024 * 1024):
            if not chunk:
                continue
            chunks.append(chunk)
            total += len(chunk)
            if total and total % (25 * 1024 * 1024) < len(chunk):
                print(f"    downloaded {total / (1024 * 1024):.0f} MiB ...")
    data = b"".join(chunks)
    if len(data) < 1024:
        raise RuntimeError(f"downloaded suspiciously small archive: {len(data)} bytes")
    return data


def _parse_workbook(zip_bytes: bytes) -> list[dict]:
    """Unpivot one substance workbook's 'IPCC 2006' sheet to long-form rows."""
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        xlsx_names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        if not xlsx_names:
            raise RuntimeError(f"no .xlsx in archive (saw {zf.namelist()})")
        xlsx_bytes = zf.read(xlsx_names[0])

    wb = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes), read_only=True, data_only=True
    )
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"sheet {SHEET!r} missing (sheets {wb.sheetnames})")
    ws = wb[SHEET]

    rows_iter = ws.iter_rows(values_only=True)
    header = None
    for i, row in enumerate(rows_iter):
        if i == HEADER_ROW:
            header = row
            break
    if header is None:
        raise RuntimeError("workbook ended before header row")

    # Year columns are labelled Y_####; everything else is an identifier col.
    year_cols = {
        idx: int(h[2:])
        for idx, h in enumerate(header)
        if isinstance(h, str) and h.startswith("Y_") and h[2:].isdigit()
    }
    if len(year_cols) < 10:
        raise RuntimeError(
            f"expected many Y_#### columns, found {len(year_cols)} in {header}"
        )

    def col(name: str) -> int:
        for idx, h in enumerate(header):
            if h == name:
                return idx
        raise RuntimeError(f"column {name!r} not in header {header}")

    i_annex = col("IPCC_annex")
    i_cgroup = col("C_group_IM24_sh")
    i_code = col("Country_code_A3")
    i_name = col("Name")
    i_sector = col("ipcc_code_2006_for_standard_report")
    i_sector_name = col("ipcc_code_2006_for_standard_report_name")
    i_sub = col("Substance")
    i_fb = col("fossil_bio")

    out: list[dict] = []
    for row in rows_iter:
        if row[i_code] is None:
            continue
        base = {
            "country_code": row[i_code],
            "country_name": row[i_name],
            "ipcc_annex": row[i_annex],
            "c_group": row[i_cgroup],
            "ipcc_sector_code": row[i_sector],
            "ipcc_sector_name": row[i_sector_name],
            "gas": row[i_sub],
            "fossil_bio": row[i_fb],
        }
        for idx, year in year_cols.items():
            val = row[idx]
            if val is None or val == "":
                continue
            rec = dict(base)
            rec["year"] = year
            rec["emissions_kt"] = float(val)
            out.append(rec)
    return out


def fetch_emissions(node_id: str) -> None:
    asset = node_id
    run_id = os.environ.get("RUN_ID", "unknown")
    fragments = list_raw_fragments(asset, "parquet")
    if os.environ.get("FORCE_REFRESH") == "1":
        done = {frag for frag, meta in fragments.items() if meta.get("run_id") == run_id}
    else:
        done = set(fragments)
    total_rows = 0

    for fragment, fname in GAS_PACKAGES:
        if fragment in done:
            print(f"  skipping {fname}: fragment already committed for run {run_id}")
            continue

        url = f"{BASE_URL}/{fname}"
        print(f"  fetching {fname} ...")
        zip_bytes = _download_zip(url)
        parsed = _parse_workbook(zip_bytes)
        print(f"    {fname}: {len(parsed):,} long rows")

        if not parsed:
            raise RuntimeError(f"EDGAR parse produced 0 rows for {fname}")

        # Cast string identifier columns explicitly; pyarrow keeps types honest.
        for r in parsed:
            for k in (
                "country_code", "country_name", "ipcc_annex", "c_group",
                "ipcc_sector_code", "ipcc_sector_name", "gas", "fossil_bio",
            ):
                v = r.get(k)
                r[k] = None if v is None else str(v)

        table = pa.Table.from_pylist(parsed, schema=SCHEMA)
        save_raw_parquet(table, asset, fragment=fragment)
        total_rows += table.num_rows

    if total_rows == 0 and not done:
        raise RuntimeError("EDGAR parse produced 0 rows across all substances")
    print(f"  total new rows this leg: {total_rows:,}")


def _has_complete_release(asset_id: str) -> bool:
    return {fragment for fragment, _ in GAS_PACKAGES}.issubset(
        set(list_raw_fragments(asset_id, "parquet"))
    )


DOWNLOAD_SPECS = [
    NodeSpec(
        id="edgar-ghg-emissions-by-country-sector",
        fn=fetch_emissions,
        kind="download",
    ),
]

MAINTAIN_SPECS = [
    MaintainSpec(
        asset_id="edgar-ghg-emissions-by-country-sector",
        description=(
            "EDGAR GHG is an annual versioned release; the pinned "
            "EDGAR_2025_GHG package set is fresh when all four CC-BY "
            "substance fragments are already committed. FORCE_REFRESH=1 "
            "forces a re-pull."
        ),
        check=_has_complete_release,
    ),
]
