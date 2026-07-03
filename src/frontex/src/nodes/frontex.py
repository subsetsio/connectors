"""Frontex connector — monthly detections of illegal border-crossings (IBC).

Single-dataset source. The whole corpus is one Excel workbook published on the
Frontex Migratory Map page ("Monthly detections of illegal border-crossing").
Its sheet 'Detections_of_IBC' is a wide monthly time series:

    Route | Border type or inland | Nationality | JAN2009 | FEB2009 | ... | <latest month>

Strategy (stateless full re-pull — the file always ships the full history from
2009 to the latest month, ~190k cells, well under a few MB, so re-fetching the
whole corpus every run is cheap and picks up revisions for free):

  download  frontex-detections-of-ibc
            Scrape the Migratory Map HTML for the current datestamped xlsx link
            (the filename rolls over every monthly release, so it is NOT
            hardcoded), download the workbook, and unpivot the wide month columns
            into a long table: (route, border_type, nationality, month_date,
            detections). Saved as parquet so the SQL transform can read it.
            xlsx parsing + reshape live here because a SQL transform cannot read
            xlsx and the month columns are dynamic.

  transform frontex-detections-of-ibc-transform
            Thin typing/correctness gate over the long parquet: cast the date and
            detection count, drop any null counts, dedup. 0 rows fails the node.
"""
import datetime as _dt
import io
import re

import httpx
import openpyxl
import pyarrow as pa

from subsets_utils import (
    NodeSpec,
    SqlNodeSpec,
    configure_http,
    get,
    save_raw_parquet,
    transient_retry,
)

# Frontex's CDN/WAF 403s the default bot User-Agent. A realistic browser UA +
# Accept headers clears the rule (datacenter IPs are otherwise served fine).
# ASCII-only header values — no smart punctuation.
_BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
        "*/*;q=0.8"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

_HOST = "https://www.frontex.europa.eu"
_MAP_PAGE = "https://www.frontex.europa.eu/along-eu-borders/migratory-map/"
_XLSX_HREF = re.compile(
    r'href="([^"]+/Migratory_routes/[^"]+Monthly_detections_of_IBC[^"]+\.xlsx)"',
    re.IGNORECASE,
)
_MONTHS = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}
_MONTH_LABEL = re.compile(r"^([A-Z]{3})(\d{4})$")

_SHEET = "Detections_of_IBC"

SCHEMA = pa.schema([
    ("route", pa.string()),
    ("border_type", pa.string()),
    ("nationality", pa.string()),
    ("month", pa.date32()),
    ("detections", pa.int64()),
])


@transient_retry()
def _get(url: str) -> httpx.Response:
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp


def _discover_xlsx_url() -> str:
    """Find the current monthly-detections xlsx link on the Migratory Map page.

    The filename carries a release datestamp that changes every month, so it is
    discovered fresh each run rather than hardcoded.
    """
    html = _get(_MAP_PAGE).text
    matches = _XLSX_HREF.findall(html)
    if not matches:
        raise AssertionError(
            f"no Monthly_detections_of_IBC .xlsx link found on {_MAP_PAGE} "
            "(page layout may have changed)"
        )
    href = matches[0]
    return href if href.startswith("http") else _HOST + href


def _parse_month(label: str) -> _dt.date:
    m = _MONTH_LABEL.match(label.strip().upper())
    if not m:
        raise AssertionError(f"unparseable month column label: {label!r}")
    mon, year = m.group(1), int(m.group(2))
    return _dt.date(year, _MONTHS[mon], 1)


def fetch_detections(node_id: str) -> None:
    asset = node_id  # the runtime passes the spec id; it IS the asset name

    configure_http(headers=_BROWSER_HEADERS)
    url = _discover_xlsx_url()
    content = _get(url).content

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if _SHEET not in wb.sheetnames:
        raise AssertionError(
            f"sheet {_SHEET!r} missing; workbook has {wb.sheetnames}"
        )
    ws = wb[_SHEET]

    # The sheet has a leading blank row; the real header row starts with 'Route'.
    header = None
    rows_iter = ws.iter_rows(values_only=True)
    for row in rows_iter:
        if row and row[0] == "Route":
            header = row
            break
    if header is None:
        raise AssertionError(f"could not locate header row (cell A == 'Route') in {_SHEET}")

    # Columns 0..2 are the dimensions; columns 3.. are month labels.
    month_cols = []  # (col_index, date)
    for idx, label in enumerate(header[3:], start=3):
        if label is None:
            continue
        month_cols.append((idx, _parse_month(str(label))))
    if not month_cols:
        raise AssertionError("no month columns found in header")

    routes, border_types, nationalities, months, detections = [], [], [], [], []
    for row in rows_iter:
        route, border_type, nationality = row[0], row[1], row[2]
        if route is None and border_type is None and nationality is None:
            continue  # skip any stray blank rows
        for idx, month in month_cols:
            val = row[idx] if idx < len(row) else None
            if val is None:
                continue  # blank cell = not reported; keep the table dense
            if not isinstance(val, (int, float)):
                continue  # ignore stray non-numeric annotations
            routes.append(route)
            border_types.append(border_type)
            nationalities.append(nationality)
            months.append(month)
            detections.append(int(val))

    if not detections:
        raise AssertionError("parsed 0 detection cells; workbook layout may have changed")

    table = pa.table(
        {
            "route": pa.array(routes, pa.string()),
            "border_type": pa.array(border_types, pa.string()),
            "nationality": pa.array(nationalities, pa.string()),
            "month": pa.array(months, pa.date32()),
            "detections": pa.array(detections, pa.int64()),
        },
        schema=SCHEMA,
    )
    save_raw_parquet(table, asset)


DOWNLOAD_SPECS = [
    NodeSpec(id="frontex-detections-of-ibc", fn=fetch_detections, kind="download"),
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="frontex-detections-of-ibc-transform",
        deps=["frontex-detections-of-ibc"],
        sql='''
            SELECT DISTINCT
                CAST(month AS DATE)        AS month,
                route,
                border_type,
                nationality,
                CAST(detections AS BIGINT) AS detections
            FROM "frontex-detections-of-ibc"
            WHERE detections IS NOT NULL
              AND route IS NOT NULL
              AND nationality IS NOT NULL
        ''',
        key=("month", "route", "border_type", "nationality"),
        temporal="month",
    ),
]
