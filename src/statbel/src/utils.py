"""Shared helpers for the Statbel connector.

Statbel publishes its whole open-data corpus through the official DCAT-BE
catalogue (Turtle/RDF). We resolve each dataset's current download URL from a
fresh read of that catalogue (filenames are point-in-time, so they must not be
hardcoded), download the best tabular distribution — delimited text, else
.xlsx — and normalise it to rows of strings that the SQL transform can publish.
"""

import csv
import datetime as dt
import gzip
import io
import re
import zipfile
from functools import lru_cache

import openpyxl

from subsets_utils import get, transient_retry

DCAT_URL = "https://doc.statbel.be/publications/DCAT/DCAT_opendata_datasets.ttl"

# Tabular distributions, in preference order (best first). Delimited text
# first — it is the widest and cheapest to parse; .xlsx is the last resort,
# used only where a dataset ships no delimited mirror at all (or where the
# delimited one is broken upstream). Anything not listed here — geospatial
# (.shp/.geojson/.gml) and binary DB (.mdb/.sqlite) mirrors — is not tabular
# and those datasets are deferred at the accept stage, not downloaded.
TABULAR_EXT = (".csv", ".txt.zip", ".csv.zip", ".zip", ".gz", ".xlsx")


@transient_retry()
def _fetch_bytes(url: str) -> bytes:
    resp = get(url, timeout=(10.0, 180.0))
    resp.raise_for_status()
    return resp.content


def _ext_of(filename: str) -> str:
    fn = filename.lower()
    for ext in (".shp.zip", ".geojson.zip", ".gml.zip", ".kml.zip",
                ".sqlite.tar.gz", ".sqlite.zip", ".mdb.zip", ".tar.gz",
                ".csv.zip", ".txt.zip"):
        if fn.endswith(ext):
            return ext
    m = re.search(r"(\.[a-z0-9]+)$", fn)
    return m.group(1) if m else ""


def _split_records(text: str):
    """Yield (subject_uri, body) for each top-level Turtle record. Records start
    at column 0 with `<uri> ...`; continuations are indented."""
    cur, buf = None, []
    subj_re = re.compile(r"^<([^>]+)>\s")
    for ln in text.splitlines():
        if ln and not ln[0].isspace() and ln.startswith("<"):
            if cur is not None:
                yield cur, "\n".join(buf)
            m = subj_re.match(ln)
            cur = m.group(1) if m else None
            buf = [ln]
        else:
            buf.append(ln)
    if cur is not None:
        yield cur, "\n".join(buf)


@lru_cache(maxsize=1)
def _catalog() -> dict:
    """Parse the DCAT catalogue once per process. Returns
    {identifier: [download_url, ...]} ordered by the dataset's listing."""
    text = _fetch_bytes(DCAT_URL).decode("utf-8", errors="replace")
    out = {}
    for subj, body in _split_records(text):
        if "a dcat:Dataset" not in body:
            continue
        ident_m = re.search(r'dct:identifier\s+"([^"]+)"', body)
        node_m = re.search(r"/node/(\d+)", subj)
        if ident_m:
            key = ident_m.group(1)
        elif node_m:
            key = "NodeID" + node_m.group(1)
        else:
            continue
        dist_block = re.search(
            r"dcat:distribution\s+(.*?)(?:;\s*\n\s*[a-z]+:|\.\s*\n|\Z)", body, re.DOTALL
        )
        urls = []
        if dist_block:
            for uri in re.findall(r"<([^>]+)>", dist_block.group(1)):
                urls.append(uri.split("#", 1)[0])
        out[key] = urls
    return out


def resolve_download_urls(entity_id: str) -> list[str]:
    """The dataset node's tabular distribution URLs, best first.

    More than one is returned so a distribution that is broken upstream (a
    truncated .zip, say) can fall through to the next-best mirror instead of
    killing the spec.
    """
    urls = _catalog().get(entity_id)
    if not urls:
        raise RuntimeError(f"{entity_id}: not present in DCAT catalogue")
    by_ext = {}
    for u in urls:
        fn = u.rstrip("/").rsplit("/", 1)[-1]
        by_ext.setdefault(_ext_of(fn), u)
    candidates = [by_ext[ext] for ext in TABULAR_EXT if ext in by_ext]
    if not candidates:
        raise RuntimeError(
            f"{entity_id}: no tabular distribution; have {sorted(by_ext)}"
        )
    return candidates


def _decode(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace")


def _extract_text(url: str, raw: bytes) -> str:
    """Return the decoded delimited-text body for a distribution."""
    low = url.lower()
    if low.endswith(".gz") and not low.endswith(".tar.gz"):
        return _decode(gzip.decompress(raw))
    if low.endswith(".zip"):  # .zip, .txt.zip, .csv.zip
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            members = [n for n in zf.namelist() if not n.endswith("/")]
            # Prefer a .txt / .csv member; else the largest file.
            members.sort(key=lambda n: (
                not n.lower().endswith((".txt", ".csv")),
                -zf.getinfo(n).file_size,
            ))
            if not members:
                raise RuntimeError(f"empty zip: {url}")
            return _decode(zf.read(members[0]))
    return _decode(raw)


def _sniff_delimiter(header: str) -> str:
    counts = {d: header.count(d) for d in ("|", ";", "\t", ",")}
    best = max(counts, key=counts.get)
    return best if counts[best] > 0 else ","


def _sanitize_columns(names):
    out, seen = [], {}
    for i, n in enumerate(names):
        clean = re.sub(r"\s+", "_", (n or "").strip())
        clean = re.sub(r"[^0-9A-Za-z_]+", "_", clean).strip("_")
        if not clean:
            clean = f"col_{i}"
        if clean in seen:
            seen[clean] += 1
            clean = f"{clean}_{seen[clean]}"
        else:
            seen[clean] = 0
        out.append(clean)
    return out


def _cell_str(value) -> str:
    """Render an .xlsx cell as the same string a delimited mirror would carry."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, dt.datetime):
        return (value.date().isoformat() if value.time() == dt.time()
                else value.isoformat(sep=" "))
    if isinstance(value, (dt.date, dt.time)):
        return value.isoformat()
    return str(value)


def _rows_from_xlsx(url: str, raw: bytes) -> list[dict]:
    """Parse the first worksheet of an .xlsx distribution.

    Statbel's workbooks declare a sheet dimension far wider and longer than the
    data (trailing formatted-but-empty cells), so the header's last non-empty
    column defines the real width and all-empty rows are dropped.
    """
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        # Sheet order is the source's own: the first is the dataset proper,
        # any later sheet is a higher-level aggregate of it.
        rows_iter = wb[wb.sheetnames[0]].iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            raise RuntimeError(f"empty worksheet: {url}")
        width = max((i + 1 for i, v in enumerate(header)
                     if v not in (None, "")), default=0)
        if not width:
            raise RuntimeError(f"no header row: {url}")
        cols = _sanitize_columns([_cell_str(v) for v in header[:width]])
        rows = []
        for rec in rows_iter:
            vals = list(rec[:width]) + [None] * (width - len(rec[:width]))
            if all(v in (None, "") for v in vals):
                continue
            rows.append({cols[i]: _cell_str(vals[i]) for i in range(width)})
        return rows
    finally:
        wb.close()


def _rows_from_delimited(url: str, raw: bytes) -> list[dict]:
    text = _extract_text(url, raw)
    # Normalise newlines and drop a leading BOM if any survived.
    text = text.lstrip("﻿").replace("\r\n", "\n").replace("\r", "\n")
    first_nl = text.find("\n")
    header_line = text[:first_nl] if first_nl != -1 else text
    delim = _sniff_delimiter(header_line)
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    try:
        header = next(reader)
    except StopIteration:
        raise RuntimeError(f"empty file at {url}")
    cols = _sanitize_columns(header)
    ncols = len(cols)
    rows = []
    for rec in reader:
        if not rec or (len(rec) == 1 and rec[0] == ""):
            continue
        if len(rec) < ncols:
            rec = rec + [""] * (ncols - len(rec))
        elif len(rec) > ncols:
            rec = rec[:ncols]
        rows.append({cols[i]: rec[i] for i in range(ncols)})
    return rows


def fetch_rows(entity_id: str):
    """Resolve, download and parse a Statbel dataset into a list of dict rows
    (all values are strings — typing is left to consumers).

    Distributions are tried best-first; a mirror that fails to download or
    parse falls through to the next one, so a single corrupt file upstream
    costs a worse format rather than the whole dataset.
    """
    candidates = resolve_download_urls(entity_id)
    failures = []
    for url in candidates:
        parse = _rows_from_xlsx if url.lower().endswith(".xlsx") else _rows_from_delimited
        try:
            rows = parse(url, _fetch_bytes(url))
        except Exception as exc:
            failures.append(f"{url.rsplit('/', 1)[-1]}: {type(exc).__name__}: {exc}")
            continue
        if failures:
            print(f"  !! {entity_id}: fell back to {url.rsplit('/', 1)[-1]} "
                  f"after {len(failures)} failed distribution(s): {failures[0]}")
        return rows
    raise RuntimeError(
        f"{entity_id}: every tabular distribution failed; {'; '.join(failures)}"
    )
