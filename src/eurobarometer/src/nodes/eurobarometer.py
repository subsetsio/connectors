"""Eurobarometer connector — aggregated survey results + survey catalog.

Source: the EU Open Data Portal DCAT catalog (data.europa.eu) for enumeration,
and the EC webgate ODP download endpoint for the actual "Volume" workbooks. See
research mechanism `odp_bulk_excel` + `data_europa_dcat`.

Two published subsets (the rank-accepted entity union):

- ``eurobarometer-surveys`` — one row per Eurobarometer survey dataset
  (publisher = Directorate-General for Communication / corporate-body COMMU),
  enumerated from the DCAT catalog. Reference catalog: id, title, dates,
  distribution count, whether an aggregated-results ("Volume A") workbook is
  available. Stateless full re-pull (~1000 rows, cheap).

- ``eurobarometer-responses`` — the long-format aggregated survey results,
  harvested from each survey's "Volume A" (results-by-question) workbook. Each
  workbook is a multi-sheet bilingual (FR/EN) banner cross-tab, one sheet per
  question; every sheet tidies to (survey, question, country, answer,
  weighted_n, share). Written per-dataset as batch parquet
  ``eurobarometer-responses-<dataset_id>`` and consumed by the transform via the
  ``eurobarometer-responses-*`` batch glob. Incremental per-dataset (shape c,
  release/version): a dataset is re-pulled only when its catalog ``modified``/
  ``issued`` version changes, so the full corpus is crawled once and refreshed
  cheaply. Resumable — raw+state are written after every dataset.

Scope note: ``responses`` parses the "Volume A" aggregated-results workbook
(direct XLSX distribution, or the ``volume_A`` ZIP that older waves ship, which
contains an old-binary ``.xls``). Surveys whose results are not published as a
Volume-A workbook (data-not-yet-released / microdata-only / EP-published) yield
no response rows and are logged, not silently dropped. Both subsets are scoped
to the canonical DG-Communication (COMMU) publisher.
"""
from __future__ import annotations

import io
import re
import time
import zipfile

import pyarrow as pa

from subsets_utils import (
    NodeSpec,
    SqlNodeSpec,
    get,
    transient_retry,
    save_raw_parquet,
    load_state,
    save_state,
)

STATE_VERSION = 1

# --- enumeration (DCAT) -----------------------------------------------------

_SEARCH_URL = "https://data.europa.eu/api/hub/search/search"
_COMMU_PUBLISHER = "Directorate-General for Communication"
_PAGE_SIZE = 100

# "Volume A" = results by question. Match "volume A", "volume_A", "volume A.xlsx",
# "ebs_474_volume_A_xls.zip" — but NOT volume AA / AP / AAP / B / C (the 'a' must
# not be followed by another letter or digit).
_VOLA_RE = re.compile(r"volume[ _]?a(?![a-z0-9])", re.I)
# A '<<Back to content' / 'Index' workbook also carries a per-sheet stub; the
# real data sheets are the ones with a TOTAL base row (detected structurally).
_NULL_TOKENS = {"-", "", ":", "n.a.", "na", "n/a", "*"}


@transient_retry()
def _get_json(url, **params):
    resp = get(url, params=params, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.json()


@transient_retry()
def _get_bytes(url):
    resp = get(url, timeout=(15.0, 300.0))
    resp.raise_for_status()
    return resp.content


def _enumerate_commu_datasets():
    """Page the DCAT catalog and yield COMMU Eurobarometer dataset records
    (with their inline distributions). Pagination is pinned on the first
    response's ``count``; a safety ceiling raises rather than truncating."""
    first = _get_json(_SEARCH_URL, q="eurobarometer", filter="dataset",
                      limit=_PAGE_SIZE, page=0)["result"]
    count = first["count"]
    pages = (count + _PAGE_SIZE - 1) // _PAGE_SIZE
    if pages > 200:
        raise RuntimeError(f"DCAT returned {count} datasets ({pages} pages) — "
                           "far beyond the expected ~1200; refusing to crawl "
                           "blindly (source shape changed).")
    out = []
    for p in range(pages):
        res = first if p == 0 else _get_json(
            _SEARCH_URL, q="eurobarometer", filter="dataset",
            limit=_PAGE_SIZE, page=p)["result"]
        for it in res.get("results", []):
            if (it.get("publisher") or {}).get("name") == _COMMU_PUBLISHER:
                out.append(it)
    return out


def _dist_url(d):
    u = d.get("download_url") or d.get("access_url")
    if isinstance(u, list):
        return u[0] if u else None
    return u


def _find_volume_a(dists):
    """Return (url, format_id) of the best Volume-A distribution, or None.
    Prefer a direct (X)LS(X) distribution; fall back to the volume_A ZIP."""
    excel, zip_ = None, None
    for d in dists:
        title_en = (d.get("title") or {}).get("en") or ""
        if not _VOLA_RE.search(title_en):
            continue
        fid = ((d.get("format") or {}).get("id") or "").upper()
        url = _dist_url(d)
        if not url:
            continue
        if fid in ("XLSX", "XLS") and excel is None:
            excel = (url, fid)
        elif fid == "ZIP" and zip_ is None:
            zip_ = (url, fid)
    return excel or zip_


# --- excel reading / banner parsing -----------------------------------------

def _rows_from_xlsx(content):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        return {sn: list(wb[sn].iter_rows(values_only=True)) for sn in wb.sheetnames}
    finally:
        wb.close()


def _rows_from_xls(content):
    import pandas as pd
    xl = pd.ExcelFile(io.BytesIO(content))  # engine inferred (xlrd for .xls)
    out = {}
    for sn in xl.sheet_names:
        df = xl.parse(sn, header=None)
        df = df.astype(object).where(pd.notna(df), None)
        out[sn] = [tuple(r) for r in df.itertuples(index=False, name=None)]
    return out


def _workbook_sheets(content, fid):
    """Return {sheet_name: [row_tuple, ...]} for a Volume-A payload."""
    if fid == "ZIP":
        z = zipfile.ZipFile(io.BytesIO(content))
        members = [m for m in z.namelist()
                   if not m.endswith("/") and m.lower().endswith((".xls", ".xlsx"))]
        if not members:
            return {}
        # prefer a member whose own name says volume A
        members.sort(key=lambda m: (0 if _VOLA_RE.search(m) else 1, m))
        m = members[0]
        raw = z.read(m)
        return _rows_from_xlsx(raw) if m.lower().endswith(".xlsx") else _rows_from_xls(raw)
    if fid == "XLSX":
        return _rows_from_xlsx(content)
    return _rows_from_xls(content)


def _clean_country(v):
    """Banner header cells carry a bilingual code like 'UE27\\nEU27'; take the
    last segment. Reject obvious non-codes (long, lowercase prose)."""
    tok = str(v).split("\n")[-1].strip()
    if not tok or len(tok) > 12:
        return None
    low = tok.lower()
    if "eighted" in low or "olume" in low or "base" in low:
        return None
    return tok


def _num(x):
    if x is None:
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s.lower() in _NULL_TOKENS:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _is_share(x):
    return isinstance(x, (int, float)) and not isinstance(x, bool) and 0.0 <= float(x) <= 1.0


def _parse_question_sheet(rows):
    """Parse one banner cross-tab sheet into (country, answer, weighted_n,
    share) tuples. Returns None if the sheet is not a question banner (e.g. the
    Content/Index sheet, which has no TOTAL base row)."""
    base_idx = None
    for i, r in enumerate(rows):
        if len(r) > 1 and r[1] is not None and str(r[1]).strip().upper() == "TOTAL":
            base_idx = i
            break
    if base_idx is None:
        return None
    # header = nearest preceding row with a non-null cell in the data columns
    header = None
    for i in range(base_idx - 1, -1, -1):
        r = rows[i]
        if len(r) > 2 and any(r[ci] is not None for ci in range(2, len(r))):
            header = r
            break
    if header is None:
        return None
    countries = {}
    for ci in range(2, len(header)):
        if header[ci] is None:
            continue
        code = _clean_country(header[ci])
        if code:
            countries[ci] = code
    if not countries:
        return None

    data_rows = [r for r in rows[base_idx + 1:]
                 if len(r) > 1 and r[1] is not None and str(r[1]).strip() != ""]
    out = []
    i, n = 0, len(data_rows)
    while i + 1 < n + 1 and i + 1 <= n - 1:
        cnt_row, pct_row = data_rows[i], data_rows[i + 1]
        # sanity: the percentage row must look like shares (fractions or null
        # tokens). If not, pairing has drifted — stop parsing this sheet.
        pct_cells = [pct_row[ci] for ci in countries if ci < len(pct_row)]
        looks_pct = any(
            _is_share(c) or (isinstance(c, str) and c.strip().lower() in _NULL_TOKENS)
            for c in pct_cells
        )
        if not looks_pct:
            break
        answer = str(pct_row[1]).strip()
        is_subtotal = answer == "Total" or answer.startswith(("Total '", 'Total "'))
        if answer and not is_subtotal:
            for ci, code in countries.items():
                wn = _num(cnt_row[ci]) if ci < len(cnt_row) else None
                sh = _num(pct_row[ci]) if ci < len(pct_row) else None
                if wn is None and sh is None:
                    continue
                out.append((code, answer, wn, sh))
        i += 2
    return out


# --- responses (subset 1) ---------------------------------------------------

RESPONSES_SCHEMA = pa.schema([
    ("survey_id", pa.string()),
    ("question_code", pa.string()),
    ("country", pa.string()),
    ("answer", pa.string()),
    ("weighted_n", pa.float64()),
    ("share", pa.float64()),
])


def _dataset_version(it):
    return it.get("modified") or it.get("issued") or ""


def fetch_responses(node_id: str) -> None:
    """Crawl the COMMU catalog, parse each survey's Volume-A workbook into long
    format, write one parquet batch per dataset. Incremental by catalog version
    (release shape): a dataset is reprocessed only when its version changes."""
    asset_prefix = node_id  # "eurobarometer-responses"
    state = load_state(asset_prefix)
    if state.get("schema_version") != STATE_VERSION:
        state = {"schema_version": STATE_VERSION, "processed": {}, "skipped": {}}
    processed = state.setdefault("processed", {})
    skipped = state.setdefault("skipped", {})
    now = int(time.time())
    # expire stale skip markers so source recovery needs no human
    for did in [d for d, m in skipped.items() if m.get("expires_at", 0) < now]:
        skipped.pop(did, None)

    datasets = _enumerate_commu_datasets()
    print(f"[responses] {len(datasets)} COMMU datasets enumerated")
    n_written = n_skip_novola = n_empty = 0
    for it in datasets:
        did = it["id"]
        version = _dataset_version(it)
        if processed.get(did) == version:
            continue  # unchanged since last successful pull
        asset = f"{asset_prefix}-{did}"
        try:
            vola = _find_volume_a(it.get("distributions", []))
            if vola is None:
                processed[did] = version  # no results workbook for this version
                n_skip_novola += 1
                continue
            url, fid = vola
            content = _get_bytes(url)
            sheets = _workbook_sheets(content, fid)
            rows = []
            for sheet_name, sheet_rows in sheets.items():
                parsed = _parse_question_sheet(sheet_rows)
                if not parsed:
                    continue
                qcode = str(sheet_name).strip()
                for country, answer, wn, sh in parsed:
                    rows.append({
                        "survey_id": did,
                        "question_code": qcode,
                        "country": country,
                        "answer": answer,
                        "weighted_n": wn,
                        "share": sh,
                    })
            if not rows:
                # workbook present but nothing parseable — record + log, don't retry forever
                processed[did] = version
                n_empty += 1
                print(f"[responses] {did}: no parseable rows from {fid} workbook")
            else:
                table = pa.Table.from_pylist(rows, schema=RESPONSES_SCHEMA)
                save_raw_parquet(table, asset)  # raw FIRST
                processed[did] = version         # then advance state
                n_written += 1
            save_state(asset_prefix, state)
        except Exception as exc:  # noqa: BLE001 — isolate per-dataset, log loudly
            print(f"[responses] {did}: FAILED url={vola} {type(exc).__name__}: {exc}")
            skipped[did] = {"reason": f"{type(exc).__name__}: {exc}"[:200],
                            "expires_at": now + 14 * 86400}
            save_state(asset_prefix, state)
    print(f"[responses] done: {n_written} datasets written, "
          f"{n_skip_novola} without Volume-A, {n_empty} empty, "
          f"{len(skipped)} skipped(transient)")


# --- surveys (subset 2) -----------------------------------------------------

SURVEYS_SCHEMA = pa.schema([
    ("survey_id", pa.string()),
    ("title", pa.string()),
    ("publisher", pa.string()),
    ("issued", pa.string()),
    ("modified", pa.string()),
    ("num_distributions", pa.int64()),
    ("has_volume_a", pa.bool_()),
])


def fetch_surveys(node_id: str) -> None:
    """One row per COMMU Eurobarometer survey dataset from the DCAT catalog.
    Stateless full re-pull (snapshot of the catalog)."""
    asset = node_id  # "eurobarometer-surveys"
    datasets = _enumerate_commu_datasets()
    rows = []
    for it in datasets:
        dists = it.get("distributions", [])
        title = (it.get("title") or {}).get("en")
        rows.append({
            "survey_id": it["id"],
            "title": title,
            "publisher": (it.get("publisher") or {}).get("name"),
            "issued": it.get("issued"),
            "modified": it.get("modified"),
            "num_distributions": len(dists),
            "has_volume_a": _find_volume_a(dists) is not None,
        })
    table = pa.Table.from_pylist(rows, schema=SURVEYS_SCHEMA)
    save_raw_parquet(table, asset)
    print(f"[surveys] wrote {len(rows)} survey rows")


# --- specs ------------------------------------------------------------------

DOWNLOAD_SPECS = [
    NodeSpec(id="eurobarometer-responses", fn=fetch_responses, kind="download"),
    NodeSpec(id="eurobarometer-surveys", fn=fetch_surveys, kind="download"),
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="eurobarometer-responses-transform",
        deps=["eurobarometer-responses"],
        sql='''
            SELECT
                survey_id,
                question_code,
                country,
                answer,
                CAST(weighted_n AS DOUBLE) AS weighted_n,
                CAST(share AS DOUBLE)      AS share
            FROM "eurobarometer-responses"
            WHERE country IS NOT NULL
              AND answer IS NOT NULL
              AND (weighted_n IS NOT NULL OR share IS NOT NULL)
        ''',
    ),
    SqlNodeSpec(
        id="eurobarometer-surveys-transform",
        deps=["eurobarometer-surveys"],
        sql='''
            SELECT
                survey_id,
                title,
                publisher,
                TRY_CAST(issued AS TIMESTAMP)   AS issued,
                TRY_CAST(modified AS TIMESTAMP)  AS modified,
                num_distributions,
                has_volume_a
            FROM "eurobarometer-surveys"
            WHERE survey_id IS NOT NULL
        ''',
    ),
]
