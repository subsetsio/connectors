"""Eurobarometer helpers: catalog enumeration, Volume-A selection, workbook parsing.

Kept out of the node module because the banner-cross-tab parser dwarfs the specs.

Two upstream surfaces:

* ``data.europa.eu`` CKAN-compatible ``package_search`` — the dataset catalog.
  Enumeration MUST pass ``sort=id asc``: the default relevance sort makes pages
  overlap and silently drops records.
* ``europa.eu/eurobarometer`` ``survey/get/one`` — per-survey metadata (fieldwork
  dates, series, themes). The dataset id embeds the survey id as its leading
  ``s<digits>_`` token, which is how the two surfaces join.

The "Volume A" workbook (weighted results by question) is the only aggregated
results artefact. Its layout drifted across 30 years of publication; see
``parse_sheet`` for the invariants the parser keys on rather than fixed offsets.
"""
from __future__ import annotations

import io
import re
import zipfile

from subsets_utils import get, transient_retry

CKAN_SEARCH = "https://data.europa.eu/api/hub/search/ckan/package_search"
SURVEY_ONE = "https://europa.eu/eurobarometer/api/survey/get/one"
COMMU = "/COMMU"          # publisher.resource suffix: DG Communication
PAGE_SIZE = 100
MAX_PAGES = 60            # safety ceiling; ~1132 records at 100/page today

# The volume code lives inside the resource filename, whose spelling drifted:
#   OP_VolumeAEB764HOMEhomeaffairs_v1.zip   VolumeAFlash336DGECFIN.zip
#   VolumeASP410EB802AGRIpacEN.zip          EB86A.zip / E76A.zip
#   FL382a_Volume A_xls.zip                 OP_EB73 VOLA.zip
#   Standard Eurobarometer 104_Autumn 2025_volume A.xlsx
# Anchor on a closed vocabulary so "VolumeAEB843" reads A and "VolumeAAEB764"
# reads AA. Longest alternatives first.
_CODES = r"(AAP|AA|AP|BP|SC|A|B|C|D|S)"
_AFTER = r"(?=EB\d|FL\d|EP\d|SP\d|Flash\s*\d|E\d|_|\s|\.|\(|$)"
_VOL_RE = re.compile(r"(?:volumes?|tables?|vol)\s*_?\s*" + _CODES + _AFTER, re.I)
_LEGACY_RE = re.compile(r"\b(?:EB|E)\d+(?:\.\d+)?[._]?" + _CODES + r"\.zip", re.I)

TOC_NAMES = {"index", "content", "contents", "table of contents", "toc", "sommaire",
             "note", "notes"}
BASE_LABELS = {"TOTAL", "WEIGHTED TOTAL", "TOTAL WEIGHTED", "BASE",
               "UNWEIGHTED BASE", "WEIGHTED BASE", "TOTAL N"}
NULL_TOKENS = {"-", "", ":", "n.a.", "na", "n/a", "*", "nan", "none"}
SUBTOTAL_RE = re.compile(r"^total\s*['\"‘’]", re.I)
CODE_RE = re.compile(r"^[A-Z][A-Z0-9\-\+/]{0,11}$")
MAX_BANNER_LEN = 24
SURVEY_ID_RE = re.compile(r"^s(\d+)_", re.I)


# --------------------------------------------------------------------------
# catalog
# --------------------------------------------------------------------------

@transient_retry()
def _get_json(url, **params):
    resp = get(url, params=params, timeout=(10.0, 180.0))
    resp.raise_for_status()
    return resp.json()


@transient_retry()
def _get_bytes(url):
    resp = get(url, timeout=(15.0, 300.0))
    resp.raise_for_status()
    return resp.content


def enumerate_commu_datasets():
    """Every DG-Communication Eurobarometer dataset record, deterministically paged."""
    out, start, count = {}, 0, None
    for page in range(MAX_PAGES + 1):
        if page > MAX_PAGES:
            raise RuntimeError(
                f"catalog exceeded {MAX_PAGES} pages (count={count}) — the source "
                "shape changed; refusing to crawl blindly")
        res = _get_json(CKAN_SEARCH, q="eurobarometer", rows=PAGE_SIZE,
                        start=start, sort="id asc")["result"]
        count = res["count"]
        batch = res.get("results") or []
        if not batch:
            break
        for rec in batch:
            out[rec["id"]] = rec
        start += PAGE_SIZE
        if start >= count:
            break
    if count and len(out) < count * 0.95:
        raise RuntimeError(f"enumerated {len(out)} of {count} catalog records — "
                           "pagination lost records")
    return {k: v for k, v in out.items()
            if ((v.get("publisher") or {}).get("resource") or "").endswith(COMMU)}


def resource_title(res):
    return ((res.get("translation") or {}).get("en") or {}).get("title") or ""


def volume_code(res):
    """The Volume letter this resource carries ('A', 'AA', 'C', ...) or None."""
    title = resource_title(res)
    for rx in (_VOL_RE, _LEGACY_RE):
        m = rx.search(title)
        if m:
            return m.group(1).upper()
    return None


def resource_url(res):
    """access_url arrives wrapped in literal square brackets."""
    u = res.get("access_url")
    if isinstance(u, list):
        u = u[0] if u else None
    if not u:
        return None
    return str(u).strip().lstrip("[").rstrip("]").strip() or None


def volume_a_resources(rec):
    """Every Volume-A resource of a catalog record (a few waves ship EU27 and
    EU28 restatements, or an a/b split, as separate Volume-A files)."""
    out = []
    for res in rec.get("resources") or []:
        if volume_code(res) != "A":
            continue
        url = resource_url(res)
        if url:
            out.append((resource_title(res), url))
    return sorted(out)


def dataset_title(rec):
    return ((rec.get("translation") or {}).get("en") or {}).get("title")


def dataset_notes(rec):
    return ((rec.get("translation") or {}).get("en") or {}).get("notes")


def survey_id_of(dataset_id):
    m = SURVEY_ID_RE.match(dataset_id)
    return int(m.group(1)) if m else None


def fetch_survey_metadata(survey_id):
    """Per-survey metadata from the EC portal API, or None when it has no record.

    ``survey/get/all`` was retired (404) and ``survey/search`` is a PDF document
    index; ``get/one`` is the only surface that returns clean survey metadata.
    """
    resp = get(SURVEY_ONE, params={"id": survey_id}, timeout=(10.0, 60.0))
    if resp.status_code == 404:
        return None
    resp.raise_for_status()
    body = resp.json()
    return body if isinstance(body, dict) and body.get("id") else None


# --------------------------------------------------------------------------
# workbook reading
# --------------------------------------------------------------------------

def workbook_members(payload):
    """yield (member_name, bytes) for each workbook in a downloaded payload.

    A payload is either a bare .xls, a bare .xlsx (itself a zip), or a zip
    holding one or more workbooks. Some older waves ship a zip of .txt/.pdf/.doc
    instead — those yield nothing.
    """
    if payload[:2] != b"PK":
        yield "__xls__", payload
        return
    try:
        zf = zipfile.ZipFile(io.BytesIO(payload))
    except zipfile.BadZipFile:
        return
    names = zf.namelist()
    books = [n for n in names
             if n.lower().endswith((".xls", ".xlsx")) and not n.startswith("__MACOSX")]
    if books:
        for name in sorted(books):
            yield name, zf.read(name)
    elif "[Content_Types].xml" in names:
        yield "__xlsx__", payload          # the payload IS an xlsx


def sheets_of(data):
    """{sheet_name: [row_tuple, ...]} — xlrd for old binary .xls, openpyxl for .xlsx."""
    if data[:2] == b"PK":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            return {sn: [tuple(r) for r in wb[sn].iter_rows(values_only=True)]
                    for sn in wb.sheetnames}
        finally:
            wb.close()
    import xlrd
    wb = xlrd.open_workbook(file_contents=data)
    return {
        sn: [tuple(c.value if c.ctype != 0 else None for c in wb.sheet_by_name(sn).row(r))
             for r in range(wb.sheet_by_name(sn).nrows)]
        for sn in wb.sheet_names()
    }


# --------------------------------------------------------------------------
# banner cross-tab parsing
# --------------------------------------------------------------------------

def _s(v):
    return "" if v is None else str(v).strip()


def _num(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = _s(v).replace("\u00a0", " ").replace(",", ".").replace("%", "").strip()
    if t.lower() in NULL_TOKENS:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _is_null_token(v):
    return _s(v).lower() in NULL_TOKENS


def _base_row_at(rows, i):
    """(label_col, mode) when row i is a base ('TOTAL') row, else None."""
    row = rows[i]
    for lab_col in (1, 0, 2):
        if lab_col >= len(row) or _s(row[lab_col]).upper() not in BASE_LABELS:
            continue
        # A real base row carries numbers in its data columns. A bare "TOTAL"
        # cell can also be a banner COLUMN header (FL176).
        if not any(_num(row[c]) is not None for c in range(lab_col + 1, len(row))):
            continue
        stats = {_s(r[lab_col]).upper() for r in rows[i + 1:i + 8] if lab_col < len(r)}
        mode = "triplet" if ("N" in stats and any(s.startswith("TOTAL %") for s in stats)) \
            else "pair"
        return lab_col, mode
    return None


def find_bases(rows):
    """Every base row in a sheet: [(row_index, label_col, mode), ...].

    Most workbooks put one question per sheet, but the Flash "Table A" family
    stacks every question of the survey down a single sheet, each with its own
    title, banner header and base row. Parsing only the first would silently
    drop the rest of the survey.
    """
    out = []
    for i in range(len(rows)):
        hit = _base_row_at(rows, i)
        if hit:
            out.append((i, hit[0], hit[1]))
    return out


def _banner_token(v):
    """Header cells are bilingual ('UE28\\nEU28' -> 'EU28') or multi-line labels
    ('Day 1\\n28 Dec.'). Keep the last line only when it reads as a code."""
    raw = _s(v)
    if not raw:
        return None
    lines = [x.strip() for x in raw.split("\n") if x.strip()]
    if not lines:
        return None
    tok = lines[-1] if CODE_RE.match(lines[-1]) else " ".join(lines)
    tok = re.sub(r"\s+", " ", tok).strip(" .:")
    if not tok or len(tok) > MAX_BANNER_LEN:
        return None
    if tok.upper() in BASE_LABELS or tok.lower() in NULL_TOKENS:
        return None
    return tok


def banner_columns(rows, base_idx, data_start, floor=0):
    """{column_index: banner_label} from the nearest labelled row above the base.

    Banner columns are usually countries / EU aggregates, but a few waves break
    results out by fieldwork day or country group instead — accept any short label.
    """
    for i in range(base_idx - 1, floor - 1, -1):
        row = rows[i]
        if len(row) <= data_start or not any(_s(row[c]) for c in range(data_start, len(row))):
            continue
        out = {c: _banner_token(row[c]) for c in range(data_start, len(row))}
        return ({c: t for c, t in out.items() if t}, i)
    return {}, base_idx


def _share_scale(samples, raw_max):
    """Are shares stored as fractions (0..1) or percents (0..100)?

    Decided from the ratio between the printed share and the share implied by
    the row's own count and the question's base N — a signal that does not care
    how small the values happen to be. A block whose every share is under 1%%
    (printed as `0.1`) is otherwise indistinguishable from fractions.
    """
    ratios = sorted(s for s in samples if s is not None)
    if ratios:
        median = ratios[len(ratios) // 2]
        return 100.0 if median > 10.0 else 1.0
    if raw_max is None:
        return 1.0
    return 100.0 if raw_max > 1.0001 else 1.0


def _question_title(rows, start, header_idx, lab_col):
    """The longest text cell in the rows between a block's start and its banner
    header — the question wording, as printed above the cross-tab."""
    best = ""
    for i in range(start, header_idx):
        for cell in rows[i]:
            text = _s(cell)
            if len(text) <= len(best) or len(text) < 4:
                continue
            low = text.lower()
            if low.startswith(("volume", "base:", "base :")) or low in BASE_LABELS:
                continue
            best = text
    return re.sub(r"\s+", " ", best) or None


_CODE_PREFIX_RE = re.compile(r"^([A-Za-z][A-Za-z0-9_.\-]{0,15}?)[.\s)]")


def _code_from_title(title):
    m = _CODE_PREFIX_RE.match(title or "")
    return m.group(1).strip(".") if m else None


def _parse_block(rows, base_idx, lab_col, mode, end, block_start):
    data_start = lab_col + 1
    banners, header_idx = banner_columns(rows, base_idx, data_start, floor=block_start)
    if not banners:
        return None, None
    cols = sorted(banners)
    base_row = rows[base_idx]
    base_n = {c: (_num(base_row[c]) if c < len(base_row) else None) for c in cols}

    body = [r for r in rows[base_idx + 1:end] if any(_s(x) for x in r)]
    step = 3 if mode == "triplet" else 2

    # First pass: pair the rows up and learn this block's share scale.
    groups = []
    for i in range(0, len(body) - step + 1, step):
        count_row, share_row = body[i], body[i + 1]
        if mode == "triplet":
            if lab_col >= len(count_row) or _s(count_row[lab_col]).upper() != "N":
                break
            answer, answer_fr = (_s(count_row[lab_col - 1]) if lab_col >= 1 else ""), None
        else:
            answer_fr = _s(count_row[lab_col]) if lab_col < len(count_row) else ""
            answer = (_s(share_row[lab_col]) if lab_col < len(share_row) else "") or answer_fr
        if not answer:
            continue
        shares = {c: (_num(share_row[c]) if c < len(share_row) else None) for c in cols}
        if all(v is None for v in shares.values()) and not any(
                _is_null_token(share_row[c]) if c < len(share_row) else False for c in cols):
            break  # the pairing has drifted off the answer rows
        groups.append((answer, answer_fr or None, count_row, shares))

    if not groups:
        return None, None

    ratios, raw_max = [], None
    for _, _, count_row, shares in groups:
        for c in cols:
            raw = shares[c]
            if raw is None or raw <= 0:
                continue
            raw_max = raw if raw_max is None else max(raw_max, raw)
            count, base = (_num(count_row[c]) if c < len(count_row) else None), base_n.get(c)
            if count and base and base > 0:
                ratios.append(raw / (count / base))
    scale = _share_scale(ratios, raw_max)

    title = _question_title(rows, block_start, header_idx, lab_col)
    out = []
    for answer_index, (answer, answer_fr, count_row, shares) in enumerate(groups):
        subtotal = bool(SUBTOTAL_RE.match(answer)) or answer.upper() in BASE_LABELS
        for c in cols:
            weighted_n = _num(count_row[c]) if c < len(count_row) else None
            share = shares[c]
            share = None if share is None else share / scale
            if share is not None and not 0.0 <= share <= 1.0001:
                share = None      # never publish a count as if it were a fraction
            if weighted_n is None and share is None:
                continue
            label = banners[c]
            out.append({
                "banner": label,
                "country": label if CODE_RE.match(label) else None,
                "answer": answer,
                "answer_fr": answer_fr,
                "answer_index": answer_index,
                "is_subtotal": subtotal,
                "weighted_n": weighted_n,
                "share": share,
                "base_n": base_n.get(c),
            })
    return out, title


def parse_sheet(rows):
    """A sheet -> [(question_code_hint, question_title, rows), ...], one per block."""
    bases = find_bases(rows)
    if not bases:
        return []
    blocks = []
    for k, (base_idx, lab_col, mode) in enumerate(bases):
        end = bases[k + 1][0] if k + 1 < len(bases) else len(rows)
        block_start = 0 if k == 0 else bases[k - 1][0] + 1
        parsed, title = _parse_block(rows, base_idx, lab_col, mode, end, block_start)
        if parsed:
            blocks.append((_code_from_title(title), title, parsed))
    return blocks


# --------------------------------------------------------------------------
# table of contents (question wording)
# --------------------------------------------------------------------------

_SHEET_REF_RE = re.compile(r"^'?(.+?)'?\s*!")


def toc_questions(sheets):
    """{sheet_name: (question_en, question_fr)} from the workbook's TOC sheet.

    Layouts differ by vintage — an `Index` sheet whose first cell is a
    ``'QA1a.1'!A1`` link, or a `Content` sheet with a plain sheet-name column —
    so rows are matched by resolving their first cell against the real sheet
    names. The remaining text cells are the wording, French half first, English
    half second; a lone text cell is taken as-is.
    """
    names = {n.strip(): n for n in sheets}
    toc = next((n for n in sheets if n.strip().lower() in TOC_NAMES), None)
    if not toc:
        return {}
    out = {}
    for row in sheets[toc]:
        cells = [_s(c) for c in row]
        if not cells or not cells[0]:
            continue
        ref = cells[0]
        m = _SHEET_REF_RE.match(ref)
        if m:
            ref = m.group(1)
        ref = ref.strip().strip("'")
        if ref not in names:
            continue
        texts = [c for c in cells[1:] if len(c) >= 2]
        if not texts:
            continue
        if len(texts) >= 2 and len(texts) % 2 == 0:
            half = len(texts) // 2
            fr, en = texts[0], texts[half]
        else:
            fr, en = texts[0], texts[-1]
        out[names[ref]] = (en, fr)
    return out


def parse_volume_a(payload, source_file_hint):
    """Every question block of every sheet of a Volume-A payload, as flat rows."""
    rows = []
    for member, data in workbook_members(payload):
        member_name = source_file_hint if member.startswith("__") else member
        try:
            sheets = sheets_of(data)
        except Exception as exc:  # a corrupt or unsupported workbook is not a bug
            print(f"    [warn] cannot open member {member!r}: {type(exc).__name__}: {exc}")
            continue
        wording = toc_questions(sheets)
        for sheet_name, sheet_rows in sheets.items():
            if sheet_name.strip().lower() in TOC_NAMES:
                continue
            blocks = parse_sheet(sheet_rows)
            single = len(blocks) == 1
            for code_hint, title, parsed in blocks:
                question_en, question_fr = wording.get(sheet_name, (None, None))
                code = sheet_name.strip() if single else (code_hint or sheet_name.strip())
                for row in parsed:
                    row["sheet_name"] = sheet_name.strip()
                    row["question_code"] = code
                    # en/fr come only from the TOC, where the two languages are
                    # positionally identified. question_title is the wording as
                    # printed above the cross-tab, in whatever language that is.
                    row["question_en"] = question_en
                    row["question_fr"] = question_fr
                    row["question_title"] = title
                    row["source_file"] = member_name
                    rows.append(row)
    return rows


def download(url):
    return _get_bytes(url)
