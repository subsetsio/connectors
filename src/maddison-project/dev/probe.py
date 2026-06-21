import openpyxl, io
from subsets_utils import get
r = get("https://dataverse.nl/api/access/datafile/421302", timeout=(10,120))
print("status", r.status_code, "bytes", len(r.content))
wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
print("sheets:", wb.sheetnames)

# Full data
ws = wb["Full data"]
rows = ws.iter_rows(values_only=True)
header = next(rows)
print("\nFull data header:", header)
sample = [next(rows) for _ in range(4)]
for s in sample: print("  ", s)

# Regional data — dump first ~5 rows fully
ws = wb["Regional data"]
print("\nRegional data dims:", ws.max_row, "x", ws.max_column)
rr = list(ws.iter_rows(values_only=True))
for i,row in enumerate(rr[:5]):
    print(f"  r{i}:", row)
print("  ...")
print(f"  last r{len(rr)-1}:", rr[-1])
