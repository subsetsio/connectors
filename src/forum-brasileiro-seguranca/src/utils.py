"""Shared helpers for the Fórum Brasileiro de Segurança Pública connector.

The source's flagship product is the annual *Anuário Brasileiro de Segurança
Pública*, published in the FBSP DSpace repository as one multi-sheet xlsx
"PLANILHA" workbook. Each entity in this connector is one data table (a sheet)
inside that workbook. The download stage resolves the latest edition, downloads
the workbook, and parses the requested sheet into a tidy long table.

Sheet codes (T01, T02, ...) shift between annual editions, so an entity is keyed
by a slug of its (stable) Índice title. At download time we re-parse the Índice,
slug each table title the same way collect did, and match the entity to its
current sheet — so the connector tracks the latest edition automatically.
"""

import io
import re
import unicodedata

from openpyxl import load_workbook

from subsets_utils import get, transient_retry

API = "https://publicacoes.forumseguranca.org.br/server/api"

# Brazilian geographic labels (normalized: ascii, lowercase) that mark a data row.
_UF = {
    "acre", "alagoas", "amapa", "amazonas", "bahia", "ceara", "distrito federal",
    "espirito santo", "goias", "maranhao", "mato grosso", "mato grosso do sul",
    "minas gerais", "para", "paraiba", "parana", "pernambuco", "piaui",
    "rio de janeiro", "rio grande do norte", "rio grande do sul", "rondonia",
    "roraima", "santa catarina", "sao paulo", "sergipe", "tocantins",
}
_REGIONS = {
    "regiao norte", "regiao nordeste", "regiao sudeste", "regiao sul",
    "regiao centro-oeste", "regiao centro oeste", "norte", "nordeste",
    "sudeste", "sul", "centro-oeste", "centro oeste",
}
_GEO = _UF | _REGIONS | {"brasil"}
_UF_ABBR = {
    "ac", "al", "am", "ap", "ba", "ce", "df", "es", "go", "ma", "mg", "ms",
    "mt", "pa", "pb", "pe", "pi", "pr", "rj", "rn", "ro", "rr", "rs", "sc",
    "se", "sp", "to",
}
_NULLISH = {"", "-", "--", "..", "...", "n/d", "nd", "s/i", "s/d", "x"}


def slug(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s[:70].strip("-")


def _norm(v) -> str:
    if v is None:
        return ""
    s = unicodedata.normalize("NFKD", str(v)).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s).strip().lower()


def _as_year(v):
    if isinstance(v, str):
        m = re.match(r"^\s*((?:19|20)\d{2})(?:\D|$)", v)
        if not m:
            return None
        n = int(m.group(1))
    else:
        try:
            n = int(float(v))
        except (TypeError, ValueError):
            return None
    return n if 1995 <= n <= 2026 else None


def _as_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return f
    s = str(v).strip()
    if _norm(s) in _NULLISH:
        return None
    s = s.replace("%", "").strip()
    # Brazilian decimals occasionally arrive as text "1.234,5"
    if re.fullmatch(r"-?\d{1,3}(\.\d{3})*(,\d+)?", s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _geo_level(norm_label: str) -> str:
    if norm_label == "brasil":
        return "brasil"
    if norm_label in _REGIONS:
        return "regiao"
    return "uf"


def _clean_label(v) -> str | None:
    if v in (None, ""):
        return None
    s = re.sub(r"\s+", " ", str(v)).strip()
    return s or None


@transient_retry()
def _get_json(url: str):
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.json()


@transient_retry()
def _get_bytes(url: str) -> bytes:
    resp = get(url, timeout=(10.0, 240.0))
    resp.raise_for_status()
    return resp.content


def resolve_latest_anuario() -> dict:
    """Newest 'Anuário Brasileiro de Segurança Pública' edition that ships a
    PLANILHA xlsx, via the public DSpace discover API."""
    url = (
        API + "/discover/search/objects?query="
        "Anu%C3%A1rio%20Brasileiro%20de%20Seguran%C3%A7a%20P%C3%BAblica"
        "&dsoType=item&sort=dc.date.issued,DESC&size=40&page=0"
    )
    objs = _get_json(url)["_embedded"]["searchResult"]["_embedded"]["objects"]
    for o in objs:
        ind = o["_embedded"]["indexableObject"]
        md = ind.get("metadata", {})
        title = md.get("dc.title", [{}])[0].get("value", "")
        if "Anuário Brasileiro de Segurança Pública" not in title:
            continue
        date = md.get("dc.date.issued", [{}])[0].get("value", "")
        bundles = _get_json(f"{API}/core/items/{ind['uuid']}/bundles")
        for bn in bundles["_embedded"]["bundles"]:
            if bn["name"] != "ORIGINAL":
                continue
            bits = _get_json(f"{API}/core/bundles/{bn['uuid']}/bitstreams")
            for x in bits["_embedded"]["bitstreams"]:
                nm = x["name"].lower()
                if nm.endswith(".xlsx") and "planilha" in nm:
                    return {
                        "title": title,
                        "date": date,
                        "item_uuid": ind["uuid"],
                        "bitstream_uuid": x["uuid"],
                    }
    raise RuntimeError("no Anuário PLANILHA xlsx found via DSpace discover")


def fetch_workbook() -> bytes:
    src = resolve_latest_anuario()
    return _get_bytes(f"{API}/core/bitstreams/{src['bitstream_uuid']}/content")


def index_slug_to_code(xlsx: bytes) -> dict:
    """Map slug(table title) -> sheet code (T01, ...) from the workbook Índice,
    matching exactly how collect keyed its entities."""
    import xml.etree.ElementTree as ET

    M = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    ns = {
        "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    rid = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    z = __import__("zipfile").ZipFile(io.BytesIO(xlsx))
    shared = []
    r = ET.fromstring(z.read("xl/sharedStrings.xml"))
    for si in r:
        shared.append("".join(t.text or "" for t in si.iter(M + "t")))
    wb = ET.fromstring(z.read("xl/workbook.xml"))
    sheet_els = wb.find("m:sheets", ns)
    names = {s.get("name") for s in sheet_els}
    rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    relmap = {x.get("Id"): x.get("Target") for x in rels}
    idx_target = next(
        relmap[s.get(rid)] for s in sheet_els if s.get("name") == "Índice"
    )

    def cell(c):
        v = c.find("m:v", ns)
        if v is None:
            return ""
        if c.get("t") == "s":
            return shared[int(v.text)]
        return v.text or ""

    ridx = ET.fromstring(z.read("xl/" + idx_target))
    out = {}
    seen = set()
    for row in ridx.iter(M + "row"):
        cells = [cell(c).strip() for c in row.findall("m:c", ns)]
        c0 = cells[0] if cells else ""
        c1 = cells[1] if len(cells) > 1 else ""
        mt = re.match(r"Tabela\s+(\d+)", c0, re.I)
        if not mt:
            continue
        code = f"T{int(mt.group(1)):02d}"
        if code not in names:
            continue
        title = c1 or code
        key = slug(title) or code.lower()
        if key in seen:
            key = f"{key}-{code.lower()}"
        seen.add(key)
        out[key] = code
    return out


def _measure_for(grid, year_row, ci, top, data_start, row_label=None):
    """Best label describing the metric of column `ci`, searched in the header
    band: nearest non-empty, non-year cell at the column, then leftward along
    each header row. Grouped tables also place sub-measures below the year row,
    so include those rows until the first data row."""
    parts = []
    header_rows = list(range(data_start - 1, year_row, -1))
    header_rows.extend(range(year_row - 1, top - 1, -1))
    for r in header_rows:
        row = grid[r]
        val = row[ci] if ci < len(row) else None
        if val in (None, "") or _as_year(val) is not None:
            # climb left within this row for a spanning header
            val = None
            for cc in range(min(ci, len(row) - 1), -1, -1):
                cand = row[cc]
                if cand not in (None, "") and _as_year(cand) is None:
                    val = cand
                    break
        if val not in (None, ""):
            t = re.sub(r"\s+", " ", str(val)).strip()
            if t and t not in parts:
                parts.append(t)
        if len(parts) >= 2:
            break
    if row_label:
        label = _clean_label(row_label)
        if label and label not in parts:
            parts.insert(0, label)
    return " / ".join(reversed(parts)) if parts else None


def _expanded_year_columns(row):
    markers = []
    for ci, v in enumerate(row):
        y = _as_year(v)
        if y is not None:
            markers.append((ci, y))
    if len(markers) < 2:
        return {}

    colyear = {}
    row_width = len(row)
    for i, (ci, year) in enumerate(markers):
        next_ci = markers[i + 1][0] if i + 1 < len(markers) else ci + 1
        if next_ci == ci + 1:
            cols = [ci]
        else:
            cols = range(ci, next_ci)
        for c in cols:
            if c < row_width:
                colyear[c] = year
    return colyear


def _sheet_years(grid):
    years = []
    for row in grid[:4]:
        for v in row:
            if isinstance(v, str):
                for m in re.finditer(r"(?:19|20)\d{2}", v):
                    y = int(m.group(0))
                    if 1995 <= y <= 2026:
                        years.append(y)
    return years


def _row_labels(row, first_value_col):
    return [
        _clean_label(v)
        for v in row[:first_value_col]
        if _clean_label(v) is not None and _as_year(v) is None
    ]


def _row_geography(labels):
    if not labels:
        return None, None, None
    first_norm = _norm(labels[0])
    if first_norm in _GEO:
        return labels[0], _geo_level(first_norm), None
    if first_norm in _UF_ABBR and len(labels) > 1 and _norm(labels[1]) not in _NULLISH:
        return labels[1], "capital", labels[0]
    return labels[0], "categoria", " / ".join(labels)


def _looks_like_data(row, colyear):
    if not colyear:
        return False
    first_col = min(colyear)
    if not _row_labels(row, first_col):
        return False
    return any(
        _as_float(row[ci] if ci < len(row) else None) is not None
        for ci in colyear
    )


def _header_label(grid, ci, top, bottom, row_label=None):
    parts = []
    for r in range(bottom - 1, top - 1, -1):
        row = grid[r]
        val = row[ci] if ci < len(row) else None
        if val in (None, "") or _as_year(val) is not None:
            val = None
            for cc in range(min(ci, len(row) - 1), -1, -1):
                cand = row[cc]
                if cand not in (None, "") and _as_year(cand) is None:
                    val = cand
                    break
        label = _clean_label(val)
        if label and label not in parts:
            parts.append(label)
        if len(parts) >= 2:
            break
    if row_label:
        label = _clean_label(row_label)
        if label and label not in parts:
            parts.insert(0, label)
    return " / ".join(reversed(parts)) if parts else _clean_label(row_label)


def _parse_rowwise_year_table(grid):
    out = []
    for data_start, row in enumerate(grid):
        year = _as_year(row[0] if row else None)
        if year is None:
            continue
        value_cols = [
            ci for ci in range(1, len(row))
            if _as_float(row[ci] if ci < len(row) else None) is not None
        ]
        if not value_cols:
            continue
        for r in range(data_start, len(grid)):
            row = grid[r]
            year = _as_year(row[0] if row else None)
            if year is None:
                continue
            for ci in value_cols:
                val = _as_float(row[ci] if ci < len(row) else None)
                if val is None:
                    continue
                out.append({
                    "geography": "Brasil",
                    "geo_level": "brasil",
                    "year": int(year),
                    "measure": _header_label(grid, ci, 0, data_start),
                    "value": float(val),
                })
        return out
    return out


def _parse_single_year_wide_table(grid, year):
    out = []
    for data_start, row in enumerate(grid):
        value_cols = [
            ci for ci in range(1, len(row))
            if _as_float(row[ci] if ci < len(row) else None) is not None
        ]
        if not value_cols:
            continue
        first_value_col = min(value_cols)
        labels = _row_labels(row, first_value_col)
        geography, geo_level, row_measure = _row_geography(labels)
        if geography is None:
            continue
        for r in range(data_start, len(grid)):
            row = grid[r]
            labels = _row_labels(row, first_value_col)
            geography, geo_level, row_measure = _row_geography(labels)
            if geography is None:
                continue
            for ci in value_cols:
                val = _as_float(row[ci] if ci < len(row) else None)
                if val is None:
                    continue
                out.append({
                    "geography": geography,
                    "geo_level": geo_level,
                    "year": int(year),
                    "measure": _header_label(grid, ci, 0, data_start, row_measure),
                    "value": float(val),
                })
        return out
    return out


def parse_table(xlsx: bytes, sheet_code: str) -> list[dict]:
    """Parse one Anuário table sheet into long rows.

    Output rows: geography, geo_level, year, measure, value. Handles multiple
    stacked metric blocks (each its own year-header row) and side-by-side
    metric blocks (absolutos / taxas) within one header row. Non-geographic
    Brazil-only tables use their row category as geography with geo_level
    "categoria" so the raw shape remains stable for the compiled transforms.
    """
    wb = load_workbook(io.BytesIO(xlsx), read_only=True, data_only=True)
    if sheet_code not in wb.sheetnames:
        raise RuntimeError(f"sheet {sheet_code} not present in workbook")
    ws = wb[sheet_code]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    blocks = []  # (year_row_index, {col: year})
    for ri, row in enumerate(grid):
        if _as_year(row[0] if row else None) is not None:
            continue
        colyear = _expanded_year_columns(row)
        if len(set(colyear.values())) >= 2:
            blocks.append((ri, colyear))
    if not blocks:
        out = _parse_rowwise_year_table(grid)
        if out:
            return out
        years = _sheet_years(grid)
        if years:
            return _parse_single_year_wide_table(grid, max(years))
        return []

    out = []
    for bi, (ri, colyear) in enumerate(blocks):
        top = blocks[bi - 1][0] + 1 if bi > 0 else 0
        next_ri = blocks[bi + 1][0] if bi + 1 < len(blocks) else len(grid)
        data_start = next(
            (r for r in range(ri + 1, next_ri) if _looks_like_data(grid[r], colyear)),
            ri + 1,
        )
        for r in range(ri + 1, next_ri):
            row = grid[r]
            labels = _row_labels(row, min(colyear))
            geography, geo_level, row_measure = _row_geography(labels)
            if geography is None:
                continue
            for ci, year in colyear.items():
                val = _as_float(row[ci] if ci < len(row) else None)
                if val is None:
                    continue
                measure = _measure_for(grid, ri, ci, top, data_start, row_measure)
                out.append({
                    "geography": geography,
                    "geo_level": geo_level,
                    "year": int(year),
                    "measure": measure,
                    "value": float(val),
                })
    return out
