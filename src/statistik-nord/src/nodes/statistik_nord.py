"""Statistik Nord CKAN connector.

Each accepted CKAN package becomes one raw asset. Statistik Nord publishes many
workbook layouts, so the download layer preserves machine-readable files as a
generic cell table instead of hard-coding publication-specific semantics.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import zipfile
from decimal import Decimal
from urllib.parse import urlparse

import openpyxl
import pyarrow as pa
import xlrd
from lxml import etree
from openpyxl.utils import get_column_letter

from subsets_utils import MaintainSpec, NodeSpec, get, raw_asset_exists, save_raw_parquet

from constants import ENTITY_IDS

CKAN_PACKAGE_SHOW = "https://opendata.schleswig-holstein.de/api/3/action/package_show"
PREFIX = "statistik-nord-"

PREFERRED_FORMATS = ("xlsx", "xls", "csv", "gml", "shp")
MAINTAIN_MAX_AGE_DAYS = 27

SCHEMA = pa.schema(
    [
        ("entity_id", pa.string()),
        ("package_id", pa.string()),
        ("package_name", pa.string()),
        ("package_title", pa.string()),
        ("resource_id", pa.string()),
        ("resource_name", pa.string()),
        ("resource_format", pa.string()),
        ("resource_url", pa.string()),
        ("sheet_name", pa.string()),
        ("row_number", pa.int32()),
        ("column_number", pa.int32()),
        ("column_label", pa.string()),
        ("value_text", pa.string()),
        ("value_number", pa.float64()),
        ("value_date", pa.date32()),
        ("value_bool", pa.bool_()),
    ]
)


def _entity_from_node_id(node_id: str) -> str:
    return node_id.removeprefix(PREFIX)


def _package_show(entity_id: str) -> dict:
    resp = get(CKAN_PACKAGE_SHOW, params={"id": entity_id}, timeout=(10.0, 60.0))
    resp.raise_for_status()
    payload = resp.json()
    if not payload.get("success"):
        raise RuntimeError(f"CKAN package_show failed for {entity_id}: {payload!r}")
    return payload["result"]


def _resource_format(resource: dict) -> str:
    fmt = str(resource.get("format") or "").strip().lower()
    if fmt:
        return fmt.lstrip(".")
    suffix = urlparse(str(resource.get("url") or "")).path.rsplit(".", 1)[-1].lower()
    if suffix == "zip" and str(resource.get("name") or "").lower().endswith(".zip"):
        return "shp"
    return suffix if suffix in PREFERRED_FORMATS else ""


def _machine_resource(package: dict) -> dict:
    resources = [r for r in package.get("resources", []) if r.get("url")]
    for fmt in PREFERRED_FORMATS:
        for resource in resources:
            if _resource_format(resource) == fmt:
                return resource
    formats = sorted({_resource_format(r) or str(r.get("format") or "") for r in resources})
    raise RuntimeError(f"{package.get('name')}: no supported XLSX/XLS/CSV/GML/SHP resource; formats={formats!r}")


def _fallback_url(url: str) -> str | None:
    if "statistik-nord.de/fileadmin/Dokumente/" not in url:
        return None
    if not url.lower().endswith(".xlsx"):
        return None
    stem = url[:-5]
    if stem.endswith("-endg"):
        return None
    return f"{stem}-endg.xlsx"


def _download_resource(resource: dict) -> tuple[dict, bytes]:
    resp = get(resource["url"], timeout=(10.0, 240.0))
    if resp.status_code == 404:
        fallback = _fallback_url(resource["url"])
        if fallback:
            fallback_resp = get(fallback, timeout=(10.0, 240.0))
            if fallback_resp.status_code < 400:
                updated = {**resource, "url": fallback}
                return updated, fallback_resp.content
    resp.raise_for_status()
    return resource, resp.content


def _base_row(entity_id: str, package: dict, resource: dict) -> dict:
    return {
        "entity_id": entity_id,
        "package_id": package.get("id"),
        "package_name": package.get("name"),
        "package_title": package.get("title"),
        "resource_id": resource.get("id"),
        "resource_name": resource.get("name"),
        "resource_format": _resource_format(resource),
        "resource_url": resource.get("url"),
    }


def _typed_value(value) -> tuple[str | None, float | None, dt.date | None, bool | None]:
    if value is None:
        return None, None, None, None
    if isinstance(value, bool):
        return str(value), None, None, value
    if isinstance(value, dt.datetime):
        return value.isoformat(), None, value.date(), None
    if isinstance(value, dt.date):
        return value.isoformat(), None, value, None
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        return str(value), float(value), None, None
    text = str(value).strip()
    return (text or None), None, None, None


def _append_cell(rows: list[dict], base: dict, sheet: str, row_num: int, col_num: int, value) -> None:
    value_text, value_number, value_date, value_bool = _typed_value(value)
    if value_text is None and value_number is None and value_date is None and value_bool is None:
        return
    rows.append(
        {
            **base,
            "sheet_name": sheet,
            "row_number": row_num,
            "column_number": col_num,
            "column_label": get_column_letter(col_num),
            "value_text": value_text,
            "value_number": value_number,
            "value_date": value_date,
            "value_bool": value_bool,
        }
    )


def _rows_from_xlsx(content: bytes, base: dict) -> list[dict]:
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows: list[dict] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for row_num, cells in enumerate(worksheet.iter_rows(values_only=True), start=1):
            for col_num, value in enumerate(cells, start=1):
                _append_cell(rows, base, sheet_name, row_num, col_num, value)
    return rows


def _rows_from_xls(content: bytes, base: dict) -> list[dict]:
    workbook = xlrd.open_workbook(file_contents=content)
    rows: list[dict] = []
    for sheet in workbook.sheets():
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(value, workbook.datemode).date()
                _append_cell(rows, base, sheet.name, row_idx + 1, col_idx + 1, value)
    return rows


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("latin-1", errors="replace")


def _rows_from_csv(content: bytes, base: dict) -> list[dict]:
    text = _decode_csv(content)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows: list[dict] = []
    for row_num, cells in enumerate(csv.reader(io.StringIO(text), dialect), start=1):
        for col_num, value in enumerate(cells, start=1):
            _append_cell(rows, base, "csv", row_num, col_num, value)
    return rows


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _rows_from_gml(content: bytes, base: dict) -> list[dict]:
    root = etree.fromstring(content)
    rows: list[dict] = []
    row_num = 0
    for element in root.iter():
        if not isinstance(element.tag, str):
            continue
        children = [child for child in element if isinstance(child.tag, str)]
        text = (element.text or "").strip()
        if children or not text:
            continue
        row_num += 1
        path_parts = [_local_name(ancestor.tag) for ancestor in element.iterancestors() if isinstance(ancestor.tag, str)]
        path = "/".join([*reversed(path_parts), _local_name(element.tag)])
        _append_cell(rows, base, "gml", row_num, 1, path)
        _append_cell(rows, base, "gml", row_num, 2, text)
    return rows


def _rows_from_shp_zip(content: bytes, base: dict) -> list[dict]:
    import shapefile

    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        names = archive.namelist()
        shp_name = next((name for name in names if name.lower().endswith(".shp")), None)
        shx_name = next((name for name in names if name.lower().endswith(".shx")), None)
        dbf_name = next((name for name in names if name.lower().endswith(".dbf")), None)
        if not shp_name or not dbf_name:
            raise RuntimeError(f"{base['entity_id']}: SHP archive lacks .shp/.dbf members")
        reader = shapefile.Reader(
            shp=io.BytesIO(archive.read(shp_name)),
            shx=io.BytesIO(archive.read(shx_name)) if shx_name else None,
            dbf=io.BytesIO(archive.read(dbf_name)),
        )
        field_names = [field[0] for field in reader.fields[1:]]
        rows: list[dict] = []
        for row_num, shape_record in enumerate(reader.iterShapeRecords(), start=1):
            record = shape_record.record.as_dict()
            for col_num, field_name in enumerate(field_names, start=1):
                _append_cell(rows, base, "shp", row_num, col_num, record.get(field_name))
            _append_cell(rows, base, "shp", row_num, len(field_names) + 1, shape_record.shape.shapeTypeName)
        return rows


def fetch_one(node_id: str) -> None:
    entity_id = _entity_from_node_id(node_id)
    package = _package_show(entity_id)
    resource = _machine_resource(package)
    resource, content = _download_resource(resource)
    base = _base_row(entity_id, package, resource)
    fmt = base["resource_format"]

    if fmt == "xlsx":
        rows = _rows_from_xlsx(content, base)
    elif fmt == "xls":
        rows = _rows_from_xls(content, base)
    elif fmt == "csv":
        rows = _rows_from_csv(content, base)
    elif fmt == "gml":
        rows = _rows_from_gml(content, base)
    elif fmt == "shp":
        rows = _rows_from_shp_zip(content, base)
    else:
        raise RuntimeError(f"{entity_id}: unsupported resource format {fmt!r}")

    if not rows:
        raise RuntimeError(f"{entity_id}: parsed 0 non-empty cells from {resource.get('url')}")
    save_raw_parquet(pa.Table.from_pylist(rows, schema=SCHEMA), node_id)


DOWNLOAD_SPECS = [
    NodeSpec(id=f"{PREFIX}{entity_id.lower().replace('_', '-')}", fn=fetch_one, kind="download")
    for entity_id in ENTITY_IDS
]


MAINTAIN_SPECS = [
    MaintainSpec(
        asset_id=spec.id,
        description=(
            f"Full CKAN package re-pull when raw is older than {MAINTAIN_MAX_AGE_DAYS}d. "
            "Statistik Nord exposes no incremental CKAN query for workbook contents; "
            "raw presence is also the resume marker for large catalog backfills. "
            "Monthly cadence is documented in maintenance.json."
        ),
        check=lambda aid: raw_asset_exists(aid, "parquet", max_age_days=MAINTAIN_MAX_AGE_DAYS),
    )
    for spec in DOWNLOAD_SPECS
]
