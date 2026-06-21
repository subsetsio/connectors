"""Baker Hughes — Worldwide monthly rig count (current report).

Discovers the most-recently-dated 'Month-Year WorldWide Rig Count Report' file on
the /intl-rig-count listing page and parses the long block at the bottom of the
'WW Monthly' sheet. Stateless full re-pull every run; the transform is a thin
cast/projection over parquet.
"""

import io

import pyarrow as pa
from openpyxl import load_workbook

from subsets_utils import NodeSpec, SqlNodeSpec, configure_http, save_raw_parquet

from utils import (
    INTL_PAGE, UA, MONTHS,
    discover, download, num, colmap, idx,
)

import re
from datetime import date

WW_SCHEMA = pa.schema([
    ("region", pa.string()), ("country", pa.string()), ("drill_for", pa.string()),
    ("location", pa.string()), ("rig_status", pa.string()),
    ("year", pa.int64()), ("month", pa.int64()), ("rig_count", pa.float64()),
])


def _pick_ww_current(files: list[tuple[str, str]]) -> str:
    """Most recently dated 'Month-Year WorldWide Rig Count Report' file."""
    best, best_dt = None, None
    for u, fn in files:
        low = fn.lower()
        if "worldwide rig count report" not in low.replace("  ", " "):
            continue
        m = re.match(r"([a-z]+)-(\d{4})", low.strip())
        if not m or m.group(1)[:3] not in MONTHS:
            continue
        dt = date(int(m.group(2)), MONTHS[m.group(1)[:3]], 1)
        if best_dt is None or dt > best_dt:
            best, best_dt = u, dt
    if best is None:
        raise RuntimeError("could not find current worldwide rig count report on page")
    return best


def _parse_ww_long(content: bytes) -> list[dict]:
    """Long block at the bottom of the current WW Monthly sheet.

    Header row's col 0 == 'Region' (pivot block above has 'Region' in a later
    column)."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb["WW Monthly"]
    it = ws.iter_rows(values_only=True)
    cm = None
    for row in it:
        c0 = row[0]
        if isinstance(c0, str) and c0.strip() == "Region":
            cm = colmap(row)
            break
    if cm is None:
        wb.close()
        raise RuntimeError("WW Monthly: long-format header (col0=='Region') not found")

    i_region = idx(cm, "Region")
    i_country = idx(cm, "Country")
    i_drillfor = idx(cm, "DrillFor", "Drill For")
    i_location = idx(cm, "Location")
    i_status = idx(cm, "Rig Status", "RigStatus")
    i_year = idx(cm, "Year")
    i_month = idx(cm, "Month")
    i_count = idx(cm, "Rig Count Value", "RigCountValue")
    if i_count is None or i_region is None or i_year is None or i_month is None:
        wb.close()
        raise RuntimeError("WW Monthly: required columns missing")

    def g(row, j):
        return row[j] if (j is not None and j < len(row)) else None

    out = []
    for row in it:
        region = g(row, i_region)
        if region is None or str(region).strip() == "":
            continue
        cnt = num(g(row, i_count))
        year, month = num(g(row, i_year)), num(g(row, i_month))
        if cnt is None or year is None or month is None:
            continue
        out.append({
            "region": str(region).strip(),
            "country": (str(g(row, i_country)).strip() if g(row, i_country) is not None else None),
            "drill_for": (str(g(row, i_drillfor)).strip() if g(row, i_drillfor) is not None else None),
            "location": (str(g(row, i_location)).strip() if g(row, i_location) is not None else None),
            "rig_status": (str(g(row, i_status)).strip() if g(row, i_status) is not None else None),
            "year": int(year),
            "month": int(month),
            "rig_count": cnt,
        })
    wb.close()
    return out


def fetch_worldwide_monthly(node_id: str) -> None:
    configure_http(headers={"User-Agent": UA})
    uuid = _pick_ww_current(discover(INTL_PAGE))
    rows = _parse_ww_long(download(uuid))
    if not rows:
        raise RuntimeError("WW Monthly produced 0 rows")
    save_raw_parquet(pa.Table.from_pylist(rows, schema=WW_SCHEMA), node_id)


DOWNLOAD_SPECS = [
    NodeSpec(id="baker-hughes-worldwide-rig-count-monthly", fn=fetch_worldwide_monthly, kind="download"),
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="baker-hughes-worldwide-rig-count-monthly-transform",
        deps=["baker-hughes-worldwide-rig-count-monthly"],
        sql='''
            SELECT DISTINCT
                make_date(CAST(year AS INTEGER), CAST(month AS INTEGER), 1) AS date,
                region, country, drill_for, location, rig_status,
                CAST(rig_count AS INTEGER)   AS rig_count
            FROM "baker-hughes-worldwide-rig-count-monthly"
            WHERE year IS NOT NULL AND month IS NOT NULL AND rig_count IS NOT NULL
        ''',
    ),
]
