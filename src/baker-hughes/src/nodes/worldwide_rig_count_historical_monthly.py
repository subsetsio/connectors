"""Baker Hughes — Worldwide frozen historical archive (monthly, from 2007).

Matched by its stable filename on the /intl-rig-count listing page. Stacked
year-blocks on one sheet: a year header row (col1==year, col2.. == region names)
followed by up to 12 month rows (col1 == month abbrev). Stateless full re-pull;
the transform is a thin cast over parquet.
"""

import io

import pyarrow as pa
from openpyxl import load_workbook

from subsets_utils import NodeSpec, SqlNodeSpec, configure_http, save_raw_parquet

from utils import INTL_PAGE, UA, MONTHS, discover, download, pick_match, num

WW_HIST_SCHEMA = pa.schema([
    ("date", pa.string()), ("region", pa.string()), ("rig_count", pa.float64()),
])


def _parse_ww_hist(content: bytes) -> list[dict]:
    """Frozen worldwide archive: stacked year-blocks on one sheet -- a year
    header row (col1==year, col2.. == region names) followed by up to 12 month
    rows (col1 == month abbrev)."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb["Worldwide_Rigcount"]
    data = list(ws.iter_rows(values_only=True))
    wb.close()
    out = []
    i = 0
    n = len(data)
    while i < n:
        row = data[i]
        yr = num(row[1]) if len(row) > 1 else None
        if yr is not None and 1990 <= int(yr) <= 2100 and len(row) > 2 and isinstance(row[2], str):
            yr = int(yr)
            regions = {j: str(row[j]).strip() for j in range(2, len(row))
                       if isinstance(row[j], str) and str(row[j]).strip()}
            i += 1
            while i < n:
                mrow = data[i]
                mcell = mrow[1] if len(mrow) > 1 else None
                key = str(mcell).strip().lower()[:3] if mcell is not None else ""
                if key in MONTHS:
                    for j, reg in regions.items():
                        val = num(mrow[j]) if j < len(mrow) else None
                        if val is None:
                            continue
                        out.append({
                            "date": f"{yr:04d}-{MONTHS[key]:02d}-01",
                            "region": reg,
                            "rig_count": val,
                        })
                    i += 1
                else:
                    break
            continue
        i += 1
    return out


def fetch_worldwide_historical(node_id: str) -> None:
    configure_http(headers={"User-Agent": UA})
    uuid = pick_match(discover(INTL_PAGE), "worldwide rig count", "2007")
    rows = _parse_ww_hist(download(uuid))
    if not rows:
        raise RuntimeError("Worldwide historical produced 0 rows")
    save_raw_parquet(pa.Table.from_pylist(rows, schema=WW_HIST_SCHEMA), node_id)


DOWNLOAD_SPECS = [
    NodeSpec(id="baker-hughes-worldwide-rig-count-historical-monthly", fn=fetch_worldwide_historical, kind="download"),
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="baker-hughes-worldwide-rig-count-historical-monthly-transform",
        deps=["baker-hughes-worldwide-rig-count-historical-monthly"],
        sql='''
            SELECT DISTINCT
                CAST(date AS DATE)           AS date,
                region,
                CAST(rig_count AS INTEGER)   AS rig_count
            FROM "baker-hughes-worldwide-rig-count-historical-monthly"
            WHERE date IS NOT NULL AND rig_count IS NOT NULL
        ''',
    ),
]
