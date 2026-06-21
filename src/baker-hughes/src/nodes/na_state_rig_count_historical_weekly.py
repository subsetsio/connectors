"""Baker Hughes — North America 'Rigs by State' frozen historical archive (weekly).

Matched by its stable filename on the /na-rig-count listing page. One sheet per
month, weekly date columns, 'STATE-LAND/INL WAT/OFFSHOR', 'TOTAL STATE' and
'GRAND TOTAL' rows. Stateless full re-pull; the transform is a thin cast over parquet.
"""

import io
from datetime import datetime

import pyarrow as pa
from openpyxl import load_workbook

from subsets_utils import NodeSpec, SqlNodeSpec, configure_http, save_raw_parquet

from utils import NA_PAGE, UA, discover, download, pick_match, num

STATE_HIST_SCHEMA = pa.schema([
    ("date", pa.string()), ("area", pa.string()), ("category", pa.string()),
    ("location_label", pa.string()), ("rig_count", pa.float64()),
])


def _split_state_label(lab: str) -> tuple[str, str]:
    u = lab.upper()
    for suf, cat in (("-INLAND WATERS", "Inland Waters"), ("-INL WATER", "Inland Waters"),
                     ("-INL WAT", "Inland Waters"),
                     ("-OFFSHORE", "Offshore"), ("-OFFSHOR", "Offshore"), ("-LAND", "Land")):
        if u.endswith(suf):
            return lab[: -len(suf)].strip(), cat
    if u == "GRAND TOTAL":
        return "GRAND TOTAL", "Total"
    if u.startswith("TOTAL "):
        return lab[6:].strip(), "Total"
    # A bare state name with no land/offshore/inland-waters split is that state's
    # single (total) line -- comparable to a 'TOTAL <state>' row.
    return lab, "Total"


def _parse_state_hist(content: bytes) -> list[dict]:
    """Frozen 'Rigs by State' archive: one sheet per month, weekly date columns,
    'STATE-LAND/INL WAT/OFFSHOR', 'TOTAL STATE' and 'GRAND TOTAL' rows."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        data = list(ws.iter_rows(values_only=True))
        date_row = None
        for i, row in enumerate(data[:10]):
            if sum(1 for c in row if isinstance(c, datetime)) >= 2:
                date_row = i
                break
        if date_row is None:
            continue
        date_cols = [(j, c) for j, c in enumerate(data[date_row]) if isinstance(c, datetime)]
        for row in data[date_row + 1:]:
            if not row or row[0] is None:
                continue
            lab = str(row[0]).strip()
            if lab == "" or lab.upper().startswith("NOTE"):
                continue
            area, category = _split_state_label(lab)
            for j, dt in date_cols:
                if j >= len(row):
                    continue
                val = num(row[j])
                if val is None:
                    continue
                out.append({
                    "date": dt.strftime("%Y-%m-%d"),
                    "area": area,
                    "category": category,
                    "location_label": lab,
                    "rig_count": val,
                })
    wb.close()
    return out


def fetch_na_state_historical(node_id: str) -> None:
    configure_http(headers={"User-Agent": UA})
    uuid = pick_match(discover(NA_PAGE), "rigs by state")
    rows = _parse_state_hist(download(uuid))
    if not rows:
        raise RuntimeError("Rigs by State historical produced 0 rows")
    save_raw_parquet(pa.Table.from_pylist(rows, schema=STATE_HIST_SCHEMA), node_id)


DOWNLOAD_SPECS = [
    NodeSpec(id="baker-hughes-na-state-rig-count-historical-weekly", fn=fetch_na_state_historical, kind="download"),
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="baker-hughes-na-state-rig-count-historical-weekly-transform",
        deps=["baker-hughes-na-state-rig-count-historical-weekly"],
        sql='''
            SELECT DISTINCT
                CAST(date AS DATE)           AS date,
                area, category, location_label,
                CAST(rig_count AS INTEGER)   AS rig_count
            FROM "baker-hughes-na-state-rig-count-historical-weekly"
            WHERE date IS NOT NULL AND rig_count IS NOT NULL
        ''',
    ),
]
