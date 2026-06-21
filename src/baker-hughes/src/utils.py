"""Shared discovery / HTTP / Excel-parse helpers for the Baker Hughes connector.

Source: https://rigcount.bakerhughes.com/ (North America weekly + Worldwide
monthly rig-count reports, plus frozen historical archives).

Access: bulk Excel downloads from https://rigcount.bakerhughes.com/static-files/<uuid>.
The CURRENT report files have ROTATING UUIDs keyed to publish date, so each run
re-discovers them by scraping the two listing pages (/na-rig-count, /intl-rig-count,
plain server-rendered HTML) and picking the most-recently-dated matching filename.
The frozen historical archives are matched by their stable filename patterns.

WAF NOTE: the CDN's bot manager blocks the default 'DataIntegrations/1.0' (and
browser 'Mozilla/*') user-agents with an instant connection drop, but allows a
curl-style UA. We set a curl UA via configure_http() at the top of every fetch fn.

This module is `_`-prefixed so the node loader (orchestrator.py:1123) skips it; it
holds ONLY genuinely-shared helpers — no NodeSpec definitions.
"""

import io
import re
from datetime import datetime, date

import pyarrow as pa
from openpyxl import load_workbook

from subsets_utils import get, transient_retry

BASE = "https://rigcount.bakerhughes.com/static-files/"
NA_PAGE = "https://rigcount.bakerhughes.com/na-rig-count"
INTL_PAGE = "https://rigcount.bakerhughes.com/intl-rig-count"
# Akamai bot manager allows curl-style UAs; blocks the default and browser UAs.
UA = "curl/8.4.0"

MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


@transient_retry()
def _http_get(url: str):
    resp = get(url, timeout=(10.0, 180.0))
    resp.raise_for_status()
    return resp


# --------------------------------------------------------------------------- #
# Discovery
# --------------------------------------------------------------------------- #
def discover(page_url: str) -> list[tuple[str, str]]:
    """Return [(uuid, filename)] for every static-file linked on a listing page."""
    html = _http_get(page_url).text
    uuids = sorted(set(re.findall(r"static-files/([0-9a-f-]{36})", html)))
    out = []
    for u in uuids:
        # HEAD-equivalent: a normal GET; the content-disposition carries the name.
        resp = _http_get(BASE + u)
        cd = resp.headers.get("content-disposition", "")
        m = re.search(r'filename="?([^"]+?)"?\s*$', cd)
        out.append((u, m.group(1) if m else ""))
    return out


def download(uuid: str) -> bytes:
    return _http_get(BASE + uuid).content


def pick_na_current(files: list[tuple[str, str]]) -> str:
    """Most recently dated 'MM-DD-YYYY North America Rig Count Report' file."""
    best, best_dt = None, None
    for u, fn in files:
        low = fn.lower()
        if not all(t in low for t in ("north", "america", "rig", "count", "report")):
            continue
        m = re.match(r"(\d{2})-(\d{2})-(\d{4})", fn.strip())
        if not m:
            continue
        dt = date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
        if best_dt is None or dt > best_dt:
            best, best_dt = u, dt
    if best is None:
        raise RuntimeError("could not find current North America rig count report on page")
    return best


def pick_match(files: list[tuple[str, str]], *needles: str) -> str:
    for u, fn in files:
        low = fn.lower()
        if all(n in low for n in needles):
            return u
    raise RuntimeError(f"no file matching {needles!r} found on page")


# --------------------------------------------------------------------------- #
# Parsing helpers
# --------------------------------------------------------------------------- #
def num(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def iso(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    return None


def colmap(header_row) -> dict:
    """name(normalized) -> column index, from a header tuple."""
    out = {}
    for j, c in enumerate(header_row):
        if c is None:
            continue
        out[str(c).strip().lower().replace(" ", "")] = j
    return out


def idx(cm: dict, *aliases: str):
    for a in aliases:
        k = a.strip().lower().replace(" ", "")
        if k in cm:
            return cm[k]
    return None


# --------------------------------------------------------------------------- #
# NAM long-format parse (shared by the weekly + monthly current reports)
# --------------------------------------------------------------------------- #
# Shared dimension columns for both NAM current-report schemas.
NA_DIMS = [
    ("country", pa.string()), ("county", pa.string()), ("basin", pa.string()),
    ("gom", pa.string()), ("drill_for", pa.string()), ("location", pa.string()),
    ("state_province", pa.string()), ("trajectory", pa.string()),
    ("year", pa.int64()), ("month", pa.int64()),
]


def parse_na_long(content: bytes, sheet: str, weekly: bool) -> list[dict]:
    """Parse the long/tidy block of a NAM Weekly / NAM Monthly sheet.

    The real header row is the first row whose column 0 == 'Country' (a separate
    pivot block above it has 'Country' in a later column, so anchoring on col 0
    is unambiguous)."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    cm = None
    for row in it:
        c0 = row[0]
        if isinstance(c0, str) and c0.strip() == "Country":
            cm = colmap(row)
            break
    if cm is None:
        wb.close()
        raise RuntimeError(f"{sheet}: long-format header (col0=='Country') not found")

    i_country = idx(cm, "Country")
    i_county = idx(cm, "County")
    i_basin = idx(cm, "Basin")
    i_gom = idx(cm, "GOM")
    i_drillfor = idx(cm, "DrillFor", "Drill For")
    i_location = idx(cm, "Location")
    i_state = idx(cm, "State/Province", "StateProvince")
    i_traj = idx(cm, "Trajectory")
    i_year = idx(cm, "Year")
    i_month = idx(cm, "Month")
    i_pub = idx(cm, "US_PublishDate", "PublishDate")
    i_count = idx(cm, "Rig Count Value", "RigCountValue")
    if i_count is None or i_country is None:
        wb.close()
        raise RuntimeError(f"{sheet}: required columns missing (count/country)")

    def g(row, j):
        return row[j] if (j is not None and j < len(row)) else None

    out = []
    for row in it:
        country = g(row, i_country)
        if country is None or str(country).strip() == "":
            continue
        cnt = num(g(row, i_count))
        if cnt is None:
            continue
        year = g(row, i_year)
        month = g(row, i_month)
        rec = {
            "country": str(country).strip(),
            "county": (str(g(row, i_county)).strip() if g(row, i_county) is not None else None),
            "basin": (str(g(row, i_basin)).strip() if g(row, i_basin) is not None else None),
            "gom": (str(g(row, i_gom)).strip() if g(row, i_gom) is not None else None),
            "drill_for": (str(g(row, i_drillfor)).strip() if g(row, i_drillfor) is not None else None),
            "location": (str(g(row, i_location)).strip() if g(row, i_location) is not None else None),
            "state_province": (str(g(row, i_state)).strip() if g(row, i_state) is not None else None),
            "trajectory": (str(g(row, i_traj)).strip() if g(row, i_traj) is not None else None),
            "year": int(year) if num(year) is not None else None,
            "month": int(month) if num(month) is not None else None,
            "rig_count": cnt,
        }
        if weekly:
            rec["publish_date"] = iso(g(row, i_pub))
        out.append(rec)
    wb.close()
    return out
