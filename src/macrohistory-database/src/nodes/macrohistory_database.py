"""Macrohistory Database connector — the Jordà-Schularick-Taylor (JST)
Macrohistory Database, Release 6.

The entire source is ONE bulk Excel workbook (JSTdatasetR6.xlsx): a single flat
long-run panel, one row per country-year, 18 advanced economies, annual since
1870, with ~55 macroeconomic and financial variable columns. There is no
queryable API and no incremental filter — the dataset is republished as whole
new releases, so the fetch is a stateless full re-pull: download the workbook,
parse the single sheet, write one typed parquet. The transform publishes the
panel wide (one Delta table, year+country keys + all variables as columns).
"""
import io

import openpyxl
import pyarrow as pa

from subsets_utils import (
    NodeSpec,
    get,
    save_raw_parquet,
    transient_retry,
)

# Stable handle for the R.6 workbook. The numeric id has been stable across the
# release; the trailing ?t=<epoch> is a Jimdo cache-buster that changes when the
# file is re-uploaded — harmless to keep, the server serves the current file.
DATASET_URL = (
    "https://www.macrohistory.net/app/download/9834512569/JSTdatasetR6.xlsx?t=1763503850"
)

# Column -> arrow type. Identifier columns, the integer-valued regime/crisis/
# interpolation flags, and the ~40 real-valued series. Everything is nullable
# (early years are sparse). Verified against the live R.6 workbook.
_INT_COLS = (
    "year", "ifs", "crisisJST", "crisisJST_old", "peg", "peg_strict",
    "rent_ipolated", "housing_capgain_ipolated", "eq_capgain_interp",
    "eq_tr_interp", "eq_dp_interp",
)
_STR_COLS = ("country", "iso", "peg_type", "peg_base")
_FLOAT_COLS = (
    "pop", "rgdpmad", "rgdpbarro", "rconsbarro", "gdp", "iy", "cpi", "ca",
    "imports", "exports", "narrowm", "money", "stir", "ltrate", "hpnom",
    "unemp", "wage", "debtgdp", "revenue", "expenditure", "xrusd", "tloans",
    "tmort", "thh", "tbus", "bdebt", "lev", "ltd", "noncore", "JSTtrilemmaIV",
    "eq_tr", "housing_tr", "bond_tr", "bill_rate", "housing_capgain",
    "housing_rent_rtn", "housing_rent_yd", "eq_capgain", "eq_dp", "bond_rate",
    "eq_div_rtn", "capital_tr", "risky_tr", "safe_tr",
)

# Column order as published by the workbook (also the published-table order).
_COLUMNS = (
    "year", "country", "iso", "ifs", "pop", "rgdpmad", "rgdpbarro",
    "rconsbarro", "gdp", "iy", "cpi", "ca", "imports", "exports", "narrowm",
    "money", "stir", "ltrate", "hpnom", "unemp", "wage", "debtgdp", "revenue",
    "expenditure", "xrusd", "tloans", "tmort", "thh", "tbus", "bdebt", "lev",
    "ltd", "noncore", "crisisJST", "crisisJST_old", "peg", "peg_strict",
    "peg_type", "peg_base", "JSTtrilemmaIV", "eq_tr", "housing_tr", "bond_tr",
    "bill_rate", "rent_ipolated", "housing_capgain_ipolated", "housing_capgain",
    "housing_rent_rtn", "housing_rent_yd", "eq_capgain", "eq_dp",
    "eq_capgain_interp", "eq_tr_interp", "eq_dp_interp", "bond_rate",
    "eq_div_rtn", "capital_tr", "risky_tr", "safe_tr",
)


def _arrow_type(col: str) -> pa.DataType:
    if col in _INT_COLS:
        return pa.int64()
    if col in _STR_COLS:
        return pa.string()
    return pa.float64()


SCHEMA = pa.schema([(c, _arrow_type(c)) for c in _COLUMNS])


@transient_retry()
def _download_workbook() -> bytes:
    resp = get(DATASET_URL, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.content


def fetch_panel(node_id: str) -> None:
    asset = node_id  # the runtime passes the spec id; it IS the asset name
    content = _download_workbook()

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))
    if header != list(_COLUMNS):
        raise AssertionError(
            f"workbook header drifted: got {header!r}, expected {list(_COLUMNS)!r}"
        )

    columns = {c: [] for c in _COLUMNS}
    for row in rows_iter:
        if all(v is None for v in row):  # trailing blank line guard
            continue
        for col, val in zip(_COLUMNS, row):
            columns[col].append(val)

    arrays = [pa.array(columns[c], type=_arrow_type(c)) for c in _COLUMNS]
    table = pa.Table.from_arrays(arrays, schema=SCHEMA)
    save_raw_parquet(table, asset)


DOWNLOAD_SPECS = [
    NodeSpec(
        id="macrohistory-database-jst-macrohistory-panel",
        fn=fetch_panel,
        kind="download",
    ),
]
