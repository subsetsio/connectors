"""Shared infrastructure for the Bangko Sentral ng Pilipinas (BSP) connector.

HTTP-with-retry wrappers and the generic Excel -> long-format "melt" machinery
live here so the per-dataset node files can import them without duplication.
Contains NO NodeSpec definitions.
"""

import io
from datetime import date, datetime

from subsets_utils import get, transient_retry

BASE = "https://www.bsp.gov.ph"


# ---------------------------------------------------------------------------
# HTTP with retry/backoff
# ---------------------------------------------------------------------------


@transient_retry()
def _get_bytes(url: str) -> bytes:
    resp = get(url, timeout=(10.0, 180.0))
    resp.raise_for_status()
    return resp.content


@transient_retry()
def _get_json(url: str) -> dict:
    resp = get(
        url,
        headers={"Accept": "application/json;odata=nometadata"},
        timeout=(10.0, 120.0),
    )
    resp.raise_for_status()
    return resp.json()


def _excel_url(rel: str) -> str:
    rel = rel.split("?")[0].strip()
    if not rel.startswith("/"):
        rel = "/" + rel
    # encode literal spaces, leave existing %xx escapes and slashes intact.
    return BASE + "".join("%20" if ch == " " else ch for ch in rel)


# ---------------------------------------------------------------------------
# Generic Excel -> long-format melt
# ---------------------------------------------------------------------------
def _clean_cell(v):
    """Normalise a raw cell to None / str / number, dates to ISO strings."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None  # booleans are not data values in these tables
    if isinstance(v, (datetime, date)):
        return v.isoformat()[:10]
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v  # int / float


def _as_number(v):
    """Coerce a cleaned cell to float, or None if not numeric."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace("−", "-")  # unicode minus
        if s in ("", "-", "--", "...", "n.a.", "n.a", "na", "NA", "N/A", "*"):
            return None
        neg = s.startswith("(") and s.endswith(")")
        if neg:
            s = s[1:-1]
        s = s.rstrip("%")
        try:
            f = float(s)
        except ValueError:
            return None
        return -f if neg else f
    return None


def _read_xls(content: bytes):
    """Return [(sheet_name, grid)] for a legacy .xls via xlrd."""
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    out = []
    for sh in book.sheets():
        grid = []
        for r in range(sh.nrows):
            row = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        row.append(
                            xlrd.xldate_as_datetime(cell.value, book.datemode)
                            .date()
                            .isoformat()
                        )
                    except Exception:
                        row.append(cell.value)
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    row.append(None)
                else:
                    row.append(_clean_cell(cell.value))
            grid.append(row)
        out.append((sh.name, grid))
    return out


def _read_xlsx(content: bytes):
    """Return [(sheet_name, grid)] for an .xlsx via openpyxl."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        grid = [[_clean_cell(v) for v in row] for row in ws.iter_rows(values_only=True)]
        out.append((ws.title, grid))
    wb.close()
    return out


def _is_year_header(row):
    """True if a row looks like a row of bare year labels (a header)."""
    nums = [_as_number(c) for c in row[1:]]
    nz = [n for n in nums if n is not None]
    return len(nz) >= 2 and all(n.is_integer() and 1900 <= n <= 2100 for n in nz)


def _col_header(header_rows, c):
    parts = []
    for hr in header_rows:
        if c < len(hr):
            v = hr[c]
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                v = str(int(v)) if float(v).is_integer() else str(v)
            if isinstance(v, str) and v.strip():
                parts.append(v.strip())
    return " ".join(parts)[:300] if parts else f"col{c}"


def _label_text(v):
    """Render a stub-label cell as text; integer-valued floats drop the '.0'."""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _is_year(v):
    return v is not None and v.is_integer() and 1900 <= v <= 2100


def _first_data_col(row):
    """First column (>=1) holding a real numeric VALUE. Leading bare years are
    treated as stub labels (e.g. a 'Year' column), not values, so they are
    skipped while scanning for the first genuine data cell."""
    for c in range(1, len(row)):
        v = _as_number(row[c])
        if v is None:
            continue
        if _is_year(v):
            continue
        return c
    return None


def _melt_grid(sheet_name, grid):
    """Melt one sheet into long rows: (sheet, row_label, column_header, value)."""
    rows_out = []

    # locate the first genuine data row (label + numeric, not a year header)
    first = None
    for i, row in enumerate(grid):
        fidx = _first_data_col(row)
        if fidx is None:
            continue
        if _is_year_header(row):
            continue
        if not any(row[c] not in (None, "") for c in range(fidx)):
            continue  # no stub label anywhere to the left
        first = i
        break
    if first is None:
        return rows_out

    header_rows = grid[:first]
    for ri in range(first, len(grid)):
        row = grid[ri]
        fidx = _first_data_col(row)
        if fidx is None:
            continue
        label_parts = [
            _label_text(row[c]) for c in range(fidx) if row[c] not in (None, "")
        ]
        row_label = " ".join(label_parts).strip()
        if not row_label:
            continue
        for c in range(fidx, len(row)):
            val = _as_number(row[c])
            if val is None:
                continue
            rows_out.append(
                {
                    "sheet": sheet_name,
                    "row_index": ri,
                    "col_index": c,
                    "row_label": row_label[:300],
                    "column_header": _col_header(header_rows, c),
                    "value": val,
                }
            )
    return rows_out


def _melt_workbook(content: bytes, rel_path: str):
    # Dispatch on the file's real magic bytes, not its extension — BSP serves a
    # few .xls URLs whose bytes are actually OOXML (.xlsx) and vice versa.
    if content[:4] == b"PK\x03\x04":  # ZIP container -> OOXML xlsx
        sheets = _read_xlsx(content)
    elif content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":  # OLE2 -> legacy xls
        sheets = _read_xls(content)
    else:  # fall back to the extension
        is_xlsx = rel_path.lower().split("?")[0].rstrip().endswith(".xlsx")
        sheets = _read_xlsx(content) if is_xlsx else _read_xls(content)
    out = []
    for name, grid in sheets:
        out.extend(_melt_grid(name, grid))
    return out
