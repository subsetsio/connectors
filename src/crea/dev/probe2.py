import io, re, zipfile
import openpyxl
from subsets_utils import get

html = get("https://www.crea.ca/housing-market-stats/mls-home-price-index/hpi-tool/", timeout=60).text
url = re.search(r"https://www\.crea\.ca/files/mls-hpi-data/MLS_HPI_[^\"']+\.zip", html).group(0)
data = get(url, timeout=120).content
zf = zipfile.ZipFile(io.BytesIO(data))

wb = openpyxl.load_workbook(io.BytesIO(zf.read("Not Seasonally Adjusted (M).xlsx")), read_only=True)
from collections import Counter
variants = {}
for s in wb.sheetnames:
    h = tuple(next(wb[s].iter_rows(max_row=1, values_only=True)))
    variants.setdefault(h, []).append(s)
print("num header variants:", len(variants))
for h, sheets in variants.items():
    print("\n LEN", len(h), "count", len(sheets), "e.g.", sheets[:5])
    print("  cols:", h)

# null handling: count real data rows in a sheet vs max_row
ws = wb["AGGREGATE"]
real = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None)
print("\nAGGREGATE real rows:", real, "max_row:", ws.max_row)
# show a board that might have nulls / missing housing types
ws2 = wb[wb.sheetnames[-1]]
print("last sheet:", wb.sheetnames[-1])
for r in list(ws2.iter_rows(min_row=2, values_only=True))[:3]:
    print("  ", r)
