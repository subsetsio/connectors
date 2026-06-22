import io, re, zipfile
import openpyxl
from subsets_utils import get

html = get("https://www.crea.ca/housing-market-stats/mls-home-price-index/hpi-tool/", timeout=60).text
url = re.search(r"https://www\.crea\.ca/files/mls-hpi-data/MLS_HPI_[^\"']+\.zip", html).group(0)
print("ZIP URL:", url)
data = get(url, timeout=120).content
print("zip bytes:", len(data))
zf = zipfile.ZipFile(io.BytesIO(data))
print("members:", zf.namelist())

for fname in zf.namelist():
    wb = openpyxl.load_workbook(io.BytesIO(zf.read(fname)), read_only=True)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    header = [c for c in next(ws.iter_rows(max_row=1, values_only=True))]
    r2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    rlast = None
    for rlast in ws.iter_rows(min_row=2, values_only=True):
        pass
    print("\n===", fname, "| sheets:", len(sheets))
    print("  header:", header)
    print("  row2:", r2[:3], "... dtype date:", type(r2[0]))
    print("  last row date:", rlast[0])
    # check column consistency across a few sheets
    colsets = set()
    for s in sheets[:8] + sheets[-4:]:
        h = tuple(next(wb[s].iter_rows(max_row=1, values_only=True)))
        colsets.add(h)
    print("  distinct header tuples across sampled sheets:", len(colsets))
    wb.close()
