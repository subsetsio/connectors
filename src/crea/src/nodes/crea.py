"""CREA (Canadian Real Estate Association) — MLS Home Price Index.

One published subset: the MLS HPI, a long-format panel of benchmark prices and
index values by geography, housing type, frequency and seasonal adjustment.

The whole corpus is one ~3MB monthly ZIP holding 4 Excel workbooks:
  - Seasonally Adjusted (M)      → monthly,   seasonally adjusted
  - Not Seasonally Adjusted (M)  → monthly,   not adjusted
  - Not Seasonally Adjusted (Q)  → quarterly, not adjusted
  - Not Seasonally Adjusted (A)  → annual,    not adjusted

Each workbook has ~60 sheets, one per geography (national AGGREGATE + provinces
+ local real-estate boards). Sheets carry a Date column plus, per housing type,
a `<Type>_HPI[_SA]` index column and a `<Type>_Benchmark[_SA]` price column.
Boards differ in which housing types they publish (sheets have 9/11/13 columns),
so columns are paired by parsing the header names, never by position.

Fetch shape: stateless full re-pull. Each ZIP is the complete history back to
2005, so we download it whole, reshape to long format, and overwrite. No
watermark/cursor — revisions are picked up for free. The download URL is
month-stamped with inconsistent casing, so the current link is scraped from the
HPI tool page rather than constructed.

xlsx is read with openpyxl (in the connector venv); the SQL transform reads the
clean parquet we write here.
"""

import io
import re
import zipfile
from datetime import date, datetime

import openpyxl
import pyarrow as pa

from subsets_utils import NodeSpec, SqlNodeSpec, get, save_raw_parquet, transient_retry

HPI_TOOL_URL = "https://www.crea.ca/housing-market-stats/mls-home-price-index/hpi-tool/"
_ZIP_RE = re.compile(r"https://www\.crea\.ca/files/mls-hpi-data/MLS_HPI_[^\"']+\.zip")

# (workbook filename, frequency, seasonally_adjusted)
_WORKBOOKS = [
    ("Seasonally Adjusted (M).xlsx", "monthly", True),
    ("Not Seasonally Adjusted (M).xlsx", "monthly", False),
    ("Not Seasonally Adjusted (Q).xlsx", "quarterly", False),
    ("Not Seasonally Adjusted (A).xlsx", "annual", False),
]

_PROVINCES = {
    "BRITISH_COLUMBIA", "ALBERTA", "SASKATCHEWAN", "MANITOBA", "ONTARIO",
    "QUEBEC", "NEW_BRUNSWICK", "NOVA_SCOTIA", "PRINCE_EDWARD_ISLAND",
    "NEWFOUNDLAND_AND_LABRADOR",
}

SCHEMA = pa.schema([
    ("date", pa.date32()),
    ("geography", pa.string()),
    ("geography_name", pa.string()),
    ("level", pa.string()),
    ("housing_type", pa.string()),
    ("frequency", pa.string()),
    ("seasonally_adjusted", pa.bool_()),
    ("hpi_index", pa.float64()),
    ("benchmark_price", pa.float64()),
])


@transient_retry()
def _fetch_bytes(url: str) -> bytes:
    resp = get(url, timeout=(10.0, 180.0))
    resp.raise_for_status()
    return resp.content


@transient_retry()
def _fetch_text(url: str) -> str:
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.text


def _zip_url() -> str:
    html = _fetch_text(HPI_TOOL_URL)
    m = _ZIP_RE.search(html)
    if not m:
        raise RuntimeError("could not locate the MLS HPI zip link on the HPI tool page")
    return m.group(0)


def _level(sheet: str) -> str:
    if sheet == "AGGREGATE":
        return "national"
    if sheet in _PROVINCES:
        return "provincial"
    return "board"


def _parse_col(name):
    """('Single_Family_HPI_SA') -> ('single_family', 'hpi'); Date/unknown -> None."""
    if not isinstance(name, str):
        return None
    core = name[:-3] if name.endswith("_SA") else name
    if core.endswith("_Benchmark"):
        return core[: -len("_Benchmark")].lower(), "benchmark"
    if core.endswith("_HPI"):
        return core[: -len("_HPI")].lower(), "hpi"
    return None


def _to_date(value, freq):
    if value is None:
        return None
    if freq == "monthly":
        return value.date() if isinstance(value, datetime) else None
    if freq == "annual":
        return date(int(value), 1, 1)
    if freq == "quarterly":
        # e.g. "2005Q1" -> 2005-01-01, Q2 -> 04-01, Q3 -> 07-01, Q4 -> 10-01
        year, q = str(value).upper().split("Q")
        return date(int(year), (int(q) - 1) * 3 + 1, 1)
    return None


def _rows_from_sheet(ws, sheet, freq, sa):
    geography_name = sheet.replace("_", " ").title()
    level = _level(sheet)
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return
    # column index -> (housing_type, metric)
    colmap = {}
    for idx, name in enumerate(header):
        parsed = _parse_col(name)
        if parsed:
            colmap[idx] = parsed
    for row in rows:
        if not row or row[0] is None:
            continue
        d = _to_date(row[0], freq)
        if d is None:
            continue
        # accumulate {housing_type: {"hpi": x, "benchmark": y}}
        acc = {}
        for idx, (ht, metric) in colmap.items():
            if idx >= len(row):
                continue
            v = row[idx]
            if v is None:
                continue
            acc.setdefault(ht, {})[metric] = float(v)
        for ht, vals in acc.items():
            yield {
                "date": d,
                "geography": sheet,
                "geography_name": geography_name,
                "level": level,
                "housing_type": ht,
                "frequency": freq,
                "seasonally_adjusted": sa,
                "hpi_index": vals.get("hpi"),
                "benchmark_price": vals.get("benchmark"),
            }


def fetch_hpi(node_id: str) -> None:
    asset = node_id  # the spec id IS the asset name
    data = _fetch_bytes(_zip_url())
    zf = zipfile.ZipFile(io.BytesIO(data))
    members = set(zf.namelist())

    records = []
    for fname, freq, sa in _WORKBOOKS:
        if fname not in members:
            raise RuntimeError(f"expected workbook missing from ZIP: {fname!r}")
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(fname)), read_only=True)
        try:
            for sheet in wb.sheetnames:
                records.extend(_rows_from_sheet(wb[sheet], sheet, freq, sa))
        finally:
            wb.close()

    if not records:
        raise RuntimeError("parsed zero HPI records — workbook layout may have changed")

    table = pa.Table.from_pylist(records, schema=SCHEMA)
    save_raw_parquet(table, asset)


DOWNLOAD_SPECS = [
    NodeSpec(id="crea-hpi", fn=fetch_hpi, kind="download"),
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="crea-hpi-transform",
        deps=["crea-hpi"],
        sql='''
            SELECT
                CAST(date AS DATE)            AS date,
                geography,
                geography_name,
                level,
                housing_type,
                frequency,
                seasonally_adjusted,
                CAST(hpi_index AS DOUBLE)       AS hpi_index,
                CAST(benchmark_price AS DOUBLE) AS benchmark_price
            FROM "crea-hpi"
            WHERE hpi_index IS NOT NULL OR benchmark_price IS NOT NULL
        ''',
        key=("date", "geography", "housing_type", "frequency", "seasonally_adjusted"),
        temporal="date",
    ),
]
