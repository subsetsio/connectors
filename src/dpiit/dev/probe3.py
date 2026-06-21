import re
from subsets_utils import get
base = "https://eaindustry.nic.in/"
page = get(base+"download_data_1112.asp", timeout=(10,60)).text
cfile = sorted(set(re.findall(r'eight_core_infra/Core_Industries_2011_12_\d{8}\.xlsx', page)))[-1]
print("core file:", cfile)
import openpyxl, io
content = get(base+cfile, timeout=(10,120)).content
print("bytes:", len(content))
wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
print("sheets:", wb.sheetnames)
for sn in wb.sheetnames:
    sh = wb[sn]
    print(f"--- sheet {sn}: {sh.max_row} x {sh.max_column} ---")
    rows = list(sh.iter_rows(values_only=True))
    for i,r in enumerate(rows[:8]):
        print(i, r[:14])
    print("  ... tail ...")
    for i,r in enumerate(rows[-3:]):
        print(len(rows)-3+i, r[:14])
