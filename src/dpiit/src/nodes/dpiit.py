"""DPIIT / Office of the Economic Adviser (OEA) connector.

Source: eaindustry.nic.in — direct static Excel files, no auth (mechanism
'eaindustry_static_xls'). Three published subsets:

  - dpiit-wpi-series        WPI commodity catalog (item_code, item_name, weight)
  - dpiit-wpi-values        WPI monthly index, long format (item x month)
  - dpiit-core-industries   Index of Eight Core Industries, long format

Strategy is stateless full re-pull: the whole corpus is a handful of small
Excel files (~1.5MB total) re-downloadable in seconds, so every run re-fetches
and overwrites — revisions are picked up for free. The latest-monthly WPI file
name embeds YYYYMM and the Core Industries file name embeds YYYYMMDD and both
roll forward each release, so the current file names are discovered by parsing
the WPI download page rather than hardcoded (stale URLs return a tiny
placeholder stub, not data).
"""

import io
import re
import datetime as dt

import pyarrow as pa
from subsets_utils import get, transient_retry, save_raw_parquet, NodeSpec, SqlNodeSpec

BASE = "https://eaindustry.nic.in/"
DOWNLOAD_PAGE = BASE + "download_data_1112.asp"
BASE_YEAR = "2011-12"

CORE_SECTORS = [
    ("Overall", "Overall Index"),
    ("Coal", "Index of Coal"),
    ("Crude Oil", "Index of Crude Oil"),
    ("Natural Gas", "Index of Natural Gas"),
    ("Petroleum Refinery Products", "Index of Petroleum Refinery Products"),
    ("Fertilizers", "Index of Fertilizers"),
    ("Steel", "Index of Steel"),
    ("Cement", "Index of Cement"),
    ("Electricity", "Index of Electricity"),
]

WPI_SERIES_SCHEMA = pa.schema([
    ("item_code", pa.string()),
    ("item_name", pa.string()),
    ("weight", pa.float64()),
    ("base_year", pa.string()),
])

WPI_VALUES_SCHEMA = pa.schema([
    ("item_code", pa.string()),
    ("item_name", pa.string()),
    ("weight", pa.float64()),
    ("base_year", pa.string()),
    ("date", pa.date32()),
    ("index_value", pa.float64()),
])

CORE_SCHEMA = pa.schema([
    ("sector", pa.string()),
    ("base_year", pa.string()),
    ("date", pa.date32()),
    ("index_value", pa.float64()),
    ("growth_rate", pa.float64()),
])


@transient_retry()
def _fetch_bytes(url: str) -> bytes:
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.content


@transient_retry()
def _fetch_text(url: str) -> str:
    resp = get(url, timeout=(10.0, 60.0))
    resp.raise_for_status()
    return resp.text


def _discover(pattern: str) -> str:
    """Find the current file path matching `pattern` on the WPI download page.

    Returns the absolute URL of the lexicographically-greatest match (the file
    names embed a date stamp, so max == latest). Raises if none found — a
    silent fallback to a stale URL would fetch a placeholder stub, not data.
    """
    page = _fetch_text(DOWNLOAD_PAGE)
    matches = sorted(set(re.findall(pattern, page)))
    if not matches:
        raise RuntimeError(f"no file matching {pattern!r} on {DOWNLOAD_PAGE}")
    return BASE + matches[-1]


def _load_wpi_monthly() -> tuple[list, list]:
    """Parse the latest WPI monthly index .xls (xlrd).

    Returns (header_row, data_rows). Wide format: columns are
    COMM_NAME, COMM_CODE, COMM_WT, then one INDXmmyyyy column per month.
    """
    import xlrd

    url = _discover(r"indx_download_1112/monthly_index_\d{6}\.xls")
    wb = xlrd.open_workbook(file_contents=_fetch_bytes(url))
    sh = wb.sheet_by_index(0)
    rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    if not rows:
        raise RuntimeError(f"empty WPI monthly workbook at {url}")
    header = rows[0]
    if header[:3] != ["COMM_NAME", "COMM_CODE", "COMM_WT"]:
        raise RuntimeError(f"unexpected WPI header {header[:3]!r} at {url}")
    return header, rows[1:]


def _wpi_records(header: list, data_rows: list):
    """Yield (item_code, item_name, weight, date|None, index_value|None).

    For the series catalog, date/index_value are None (one row per item). For
    values, one record per (item, month) with a real index value.
    """
    # month columns: header like INDX042012 -> April 2012
    month_cols = []
    for idx, h in enumerate(header):
        m = re.fullmatch(r"INDX(\d{2})(\d{4})", str(h).strip())
        if m:
            month_cols.append((idx, dt.date(int(m.group(2)), int(m.group(1)), 1)))

    for row in data_rows:
        name = str(row[0]).strip()
        code = str(row[1]).strip()
        if not code or not name:
            continue
        try:
            weight = float(row[2]) if row[2] not in ("", None) else None
        except (TypeError, ValueError):
            weight = None
        yield (code, name, weight, None, None)
        for ci, d in month_cols:
            v = row[ci]
            if v in ("", None):
                continue
            try:
                val = float(v)
            except (TypeError, ValueError):
                continue
            yield (code, name, weight, d, val)


def fetch_wpi_series(node_id: str) -> None:
    asset = node_id
    header, data_rows = _load_wpi_monthly()
    out = {}
    for code, name, weight, d, _val in _wpi_records(header, data_rows):
        if d is None:
            out[code] = (code, name, weight)
    recs = list(out.values())
    if not recs:
        raise RuntimeError("WPI series catalog parsed 0 items")
    table = pa.table({
        "item_code": pa.array([r[0] for r in recs], pa.string()),
        "item_name": pa.array([r[1] for r in recs], pa.string()),
        "weight": pa.array([r[2] for r in recs], pa.float64()),
        "base_year": pa.array([BASE_YEAR] * len(recs), pa.string()),
    }, schema=WPI_SERIES_SCHEMA)
    save_raw_parquet(table, asset)


def fetch_wpi_values(node_id: str) -> None:
    asset = node_id
    header, data_rows = _load_wpi_monthly()
    codes, names, weights, dates, vals = [], [], [], [], []
    for code, name, weight, d, val in _wpi_records(header, data_rows):
        if d is None:
            continue
        codes.append(code)
        names.append(name)
        weights.append(weight)
        dates.append(d)
        vals.append(val)
    if not codes:
        raise RuntimeError("WPI values parsed 0 observations")
    table = pa.table({
        "item_code": pa.array(codes, pa.string()),
        "item_name": pa.array(names, pa.string()),
        "weight": pa.array(weights, pa.float64()),
        "base_year": pa.array([BASE_YEAR] * len(codes), pa.string()),
        "date": pa.array(dates, pa.date32()),
        "index_value": pa.array(vals, pa.float64()),
    }, schema=WPI_VALUES_SCHEMA)
    save_raw_parquet(table, asset)


def _parse_core_sheet(wb, sheet_name: str) -> dict:
    """Return {(date, sector): value} from a Core Industries sheet.

    Both the 'Index' and 'Growth (%)' sheets share the layout: row 0 is a
    header, col 0 is a month datetime, cols 1..9 are the nine sectors in the
    fixed CORE_SECTORS order.
    """
    sh = wb[sheet_name]
    out = {}
    for r in sh.iter_rows(min_row=2, values_only=True):
        d = r[0]
        if not isinstance(d, dt.datetime):
            continue
        date = dt.date(d.year, d.month, 1)
        for col, (sector, _label) in enumerate(CORE_SECTORS, start=1):
            if col >= len(r):
                continue
            v = r[col]
            if v in ("", None):
                continue
            try:
                out[(date, sector)] = float(v)
            except (TypeError, ValueError):
                continue
    return out


def fetch_core_industries(node_id: str) -> None:
    import openpyxl

    asset = node_id
    url = _discover(r"eight_core_infra/Core_Industries_2011_12_\d{8}\.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(_fetch_bytes(url)), read_only=True, data_only=True)
    if "Index" not in wb.sheetnames:
        raise RuntimeError(f"no 'Index' sheet in {url}; sheets={wb.sheetnames}")
    index = _parse_core_sheet(wb, "Index")
    growth = _parse_core_sheet(wb, "Growth (%)") if "Growth (%)" in wb.sheetnames else {}
    if not index:
        raise RuntimeError(f"Core Industries 'Index' sheet parsed 0 rows at {url}")

    sectors, base_years, dates, idx_vals, grw_vals = [], [], [], [], []
    for (date, sector), val in sorted(index.items(), key=lambda kv: (kv[0][0], kv[0][1])):
        sectors.append(sector)
        base_years.append(BASE_YEAR)
        dates.append(date)
        idx_vals.append(val)
        grw_vals.append(growth.get((date, sector)))
    table = pa.table({
        "sector": pa.array(sectors, pa.string()),
        "base_year": pa.array(base_years, pa.string()),
        "date": pa.array(dates, pa.date32()),
        "index_value": pa.array(idx_vals, pa.float64()),
        "growth_rate": pa.array(grw_vals, pa.float64()),
    }, schema=CORE_SCHEMA)
    save_raw_parquet(table, asset)


DOWNLOAD_SPECS = [
    NodeSpec(id="dpiit-wpi-series", fn=fetch_wpi_series, kind="download"),
    NodeSpec(id="dpiit-wpi-values", fn=fetch_wpi_values, kind="download"),
    NodeSpec(id="dpiit-core-industries", fn=fetch_core_industries, kind="download"),
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="dpiit-wpi-series-transform",
        deps=["dpiit-wpi-series"],
        sql='''
            SELECT
                item_code,
                item_name,
                CAST(weight AS DOUBLE) AS weight,
                base_year
            FROM "dpiit-wpi-series"
            WHERE item_code IS NOT NULL
        ''',
    ),
    SqlNodeSpec(
        id="dpiit-wpi-values-transform",
        deps=["dpiit-wpi-values"],
        sql='''
            SELECT
                item_code,
                item_name,
                CAST(weight AS DOUBLE) AS weight,
                base_year,
                CAST(date AS DATE) AS date,
                CAST(index_value AS DOUBLE) AS index_value
            FROM "dpiit-wpi-values"
            WHERE index_value IS NOT NULL
        ''',
    ),
    SqlNodeSpec(
        id="dpiit-core-industries-transform",
        deps=["dpiit-core-industries"],
        sql='''
            SELECT
                sector,
                base_year,
                CAST(date AS DATE) AS date,
                CAST(index_value AS DOUBLE) AS index_value,
                CAST(growth_rate AS DOUBLE) AS growth_rate
            FROM "dpiit-core-industries"
            WHERE index_value IS NOT NULL
        ''',
    ),
]
