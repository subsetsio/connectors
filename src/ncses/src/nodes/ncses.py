"""NCSES (National Center for Science and Engineering Statistics, NSF) connector.

Mechanism: per-table .xlsx bulk download. The REST metadata API
(ncsesdata.nsf.gov/api/v1/metadata/datatables) enumerated 513 published data
tables at collect time. This module fetches each table's canonical .xlsx,
parses the publication-formatted sheet (title/subtitle/unit rows, one-to-three
merged header rows, data block, trailing footnotes) into a tidy typed grid,
and publishes one Delta table per data table via a passthrough SQL transform.

Fetch shape: stateless full re-pull. The whole corpus is ~513 small xlsx files
(<1 MB each) with no upstream incremental filter, so every refresh re-fetches
and overwrites. URLs are CONSTRUCTED from the data-table id
(pub_id = id.split('-')[0]) rather than the catalog's dataTableUrl, because the
source mis-states that URL for 4 tables (e.g. id nsf25321-tab007 -> .../tab07.xlsx
which 404s); the id-based URL is correct for all 513.
"""
import io
import re

import httpx
import openpyxl
from tenacity import (
    retry,
    retry_if_exception,
    stop_after_attempt,
    wait_exponential,
)

from subsets_utils import NodeSpec, SqlNodeSpec, get, save_raw_ndjson

BASE = "https://ncses.nsf.gov/pubs"

# Cell value classification.
_NUM_RE = re.compile(r"^-?[\d,]*\.?\d+$")
_LEAD_NUM_RE = re.compile(r"^-?[\d,]*\.?\d+")
# Tokens that mean "no value" in NCSES tables (suppression / not-available flags).
_NA_TOK = {
    "na", "n/a", "nan", "-", "--", "(x)", "x", "s", "d", "(d)", "(s)",
    "z", "(z)", "", ".", "..",
}

from constants import ENTITY_IDS

_TRANSIENT = (
    httpx.ConnectError, httpx.ConnectTimeout, httpx.ReadTimeout,
    httpx.WriteTimeout, httpx.PoolTimeout, httpx.RemoteProtocolError,
    httpx.ProxyError,
)


def _is_transient(exc: BaseException) -> bool:
    if isinstance(exc, _TRANSIENT):
        return True
    if isinstance(exc, httpx.HTTPStatusError):
        code = exc.response.status_code
        return code == 429 or 500 <= code < 600
    return False


@retry(
    retry=retry_if_exception(_is_transient),
    stop=stop_after_attempt(6),
    wait=wait_exponential(min=4, max=120),
    reraise=True,
)
def _fetch_xlsx(url: str) -> bytes:
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp.content


def _norm(cell) -> str:
    if cell is None:
        return ""
    return str(cell).replace("\xa0", " ").strip()


def _is_num(s: str) -> bool:
    return bool(_NUM_RE.match(s.strip().replace(",", "")))


def _is_year(s: str) -> bool:
    """A 4-digit calendar year. NCSES tables routinely use years as the column
    header row (e.g. 'State | 2009 | 2010 | ...'); such a row must be read as a
    header, not mistaken for the first data row."""
    s = s.strip().replace(",", "")
    if not _NUM_RE.match(s):
        return False
    v = float(s)
    return v == int(v) and 1900 <= v <= 2100


def _hdr_label(s: str) -> str:
    """Render a header cell as a column-name fragment: an integer-valued float
    ('2009.0') collapses to '2009'."""
    t = s.strip().replace(",", "")
    if _NUM_RE.match(t):
        v = float(t)
        if v == int(v):
            return str(int(v))
    return s.strip()


def _to_num(s: str):
    s = s.strip().replace(",", "")
    if _NUM_RE.match(s):
        return float(s)
    m = _LEAD_NUM_RE.match(s)  # recover "2020 (preliminary)" -> 2020, "1978a" -> 1978
    return float(m.group().replace(",", "")) if m else None


def _clean_name(nm: str, idx: int) -> str:
    nm = re.sub(r'[\x00-\x1f"`]', " ", nm)        # drop control chars / quotes
    nm = nm.replace("&", " and ")
    nm = re.sub(r"[{}();,=]", " ", nm)            # drop Delta-unsafe chars
    nm = re.sub(r"\s+", " ", nm).strip()
    return nm or f"col{idx}"


def parse_xlsx(content: bytes):
    """Parse one NCSES data-table xlsx into (column_names, records).

    Strips the title/subtitle/unit prefix rows, flattens 1-3 merged header
    rows into ' - '-joined column names, keeps the data block (section labels
    included), drops trailing footnote rows and fully-empty rows/columns, and
    coerces majority-numeric columns to floats (NA/suppression tokens -> None).
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    grid = [[_norm(c) for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    if not grid:
        return [], []
    ncol = max(len(r) for r in grid)
    grid = [r + [""] * (ncol - len(r)) for r in grid]
    # drop fully-empty columns
    keep = [j for j in range(ncol) if any(grid[i][j] for i in range(len(grid)))]
    if not keep:
        return [], []
    grid = [[r[j] for j in keep] for r in grid]
    ncol = len(keep)

    def nfilled(r):
        return sum(1 for x in r if x)

    def has_data_num(r):
        # a value cell that is numeric but NOT a bare year -> this row is data,
        # not a year-header row (which carries only years across cols >= 1)
        return any(_is_num(r[j]) and not _is_year(r[j]) for j in range(1, ncol))

    # header block starts at the first row with >1 filled cell (skip title rows)
    hstart = 0
    while hstart < len(grid) and nfilled(grid[hstart]) <= 1:
        hstart += 1
    if hstart >= len(grid):
        return [], []
    # data starts at the first row (bounded) carrying real (non-year) numeric data
    data_start = None
    for i in range(hstart, min(hstart + 8, len(grid))):
        if has_data_num(grid[i]):
            data_start = i
            break
    if data_start is None:
        data_start = hstart + 1  # fallback: single header row, rest is data
    header_rows = grid[hstart:data_start]
    # data ends at the last data-bearing row -> everything after is footnotes
    last_data = data_start
    for i in range(len(grid) - 1, data_start - 1, -1):
        if has_data_num(grid[i]):
            last_data = i
            break
    data = grid[data_start:last_data + 1]
    if not data:
        return [], []

    # drop footnote-superscript "spacer" columns: many wide NCSES tables insert
    # an empty column after each data column to hold a footnote letter. Such a
    # column has NO header text of its own (only a borrowed/merged name) AND is
    # almost entirely empty in the data block. Real merged-subheader data columns
    # always carry their own label in some header row, so they're kept.
    nrows = len(data)
    spacer = []
    for j in range(ncol):
        own_header = any(hr[j] for hr in header_rows)
        nonempty = sum(1 for i in range(nrows) if data[i][j])
        if not own_header and nonempty <= 0.2 * nrows:
            spacer.append(j)
    if spacer:
        keep2 = [j for j in range(ncol) if j not in set(spacer)]
        header_rows = [[hr[j] for j in keep2] for hr in header_rows]
        data = [[r[j] for j in keep2] for r in data]
        ncol = len(keep2)
    if ncol == 0:
        return [], []

    # flatten header: forward-fill merged spans across each header row, then
    # join the distinct labels top-to-bottom per column
    filled_rows = []
    for hr in header_rows:
        out, last = [], ""
        for x in hr:
            if x:
                last = _hdr_label(x)
            out.append(last)
        filled_rows.append(out)
    raw_names = []
    for j in range(ncol):
        parts = []
        for fr in filled_rows:
            v = fr[j].strip()
            if v and (not parts or parts[-1] != v):
                parts.append(v)
        raw_names.append(" - ".join(parts) if parts else f"col{j}")

    seen, names = {}, []
    for j, nm in enumerate(raw_names):
        nm = _clean_name(nm, j)
        if nm in seen:
            seen[nm] += 1
            nm = f"{nm}_{seen[nm]}"
        else:
            seen[nm] = 1
        names.append(nm)

    # per-column numeric detection over the data block
    is_numeric = []
    for j in range(ncol):
        vals = [data[i][j].strip() for i in range(len(data))]
        nonempty = [v for v in vals if v and v.lower() not in _NA_TOK]
        is_numeric.append(
            bool(nonempty)
            and sum(_is_num(v) for v in nonempty) / len(nonempty) >= 0.8
        )

    # build records, then drop columns that ended up entirely null
    recs = []
    for i in range(len(data)):
        rec = {}
        for j in range(ncol):
            v = data[i][j].strip()
            if is_numeric[j]:
                rec[names[j]] = _to_num(v) if v and v.lower() not in _NA_TOK else None
            else:
                rec[names[j]] = v or None
        recs.append(rec)
    live = [c for c in names if any(r.get(c) is not None for r in recs)]
    if len(live) < len(names):
        recs = [{c: r.get(c) for c in live} for r in recs]
    return live, recs


def fetch_one(node_id: str) -> None:
    asset = node_id  # the runtime passes the spec id; it IS the asset name
    table_id = node_id[len("ncses-"):]
    pub_id = table_id.split("-", 1)[0]
    url = f"{BASE}/{pub_id}/assets/data-tables/tables/{table_id}.xlsx"
    content = _fetch_xlsx(url)
    _, records = parse_xlsx(content)
    if not records:
        raise ValueError(f"{table_id}: parsed 0 data rows from {url}")
    save_raw_ndjson(records, asset)


DOWNLOAD_SPECS = [
    NodeSpec(
        id=f"ncses-{eid.lower().replace('_', '-')}",
        fn=fetch_one,
        kind="download",
    )
    for eid in ENTITY_IDS
]

# Passthrough publish: parsing/typing already happened in the fetch fn, so each
# subset is its raw grid as-is (DuckDB infers column types from the ndjson).
TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"{s.id}-transform",
        deps=[s.id],
        sql=f'SELECT * FROM "{s.id}"',
    )
    for s in DOWNLOAD_SPECS
]
